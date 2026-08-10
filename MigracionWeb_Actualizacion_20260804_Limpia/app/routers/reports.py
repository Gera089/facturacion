import io, os, tempfile, datetime, unicodedata
import matplotlib
matplotlib.use("Agg")
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional

from app.dependencies import require_user
from app.legacy_db import get_legacy_connection


def _auto_width(ws, factor=1.1, min_width=8):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value or "")
            length = sum(2 if ord(c) > 127 else 1 for c in val)
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = max(max_len * factor + 1, min_width)


router = APIRouter(prefix="/api/reports", tags=["reports"])


def _empresa_key(value):
    text = unicodedata.normalize("NFD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    return " ".join(text.replace("_", " ").split()).upper()


def _empresa_visible(value):
    key = _empresa_key(value)
    return key and not (key.startswith("TEST ") or key.startswith("PRUEBA ") or " CODEX" in key)


def _empresa_display(value):
    aliases = {
        "GOURMET ESPANA": "Gourmet España",
        "IBERSUR": "Ibersur",
        "REMISION": "Remisiones",
        "REMISIONES": "Remisiones",
    }
    key = _empresa_key(value)
    return aliases.get(key, str(value or "").strip())


def _consolidar_por_empresa(rows):
    grouped = {}
    for row in rows or []:
        if not _empresa_visible(row.get("empresa")):
            continue
        empresa = _empresa_display(row.get("empresa"))
        key = _empresa_key(empresa)
        if key not in grouped:
            grouped[key] = {"empresa": empresa, "facturas": 0, "total": 0}
        grouped[key]["facturas"] += int(row.get("facturas") or 0)
        grouped[key]["total"] += row.get("total") or 0
    return sorted(grouped.values(), key=lambda item: item["total"], reverse=True)


def _build_facturas_where(empresa=None, cliente=None, producto=None, desde=None, hasta=None, anio=0, mes=0):
    wheres = ["f.estatus = 'Activa'"]
    params = []
    if empresa:
        wheres.append("f.empresa = %s")
        params.append(empresa)
    if cliente:
        wheres.append("f.numero_cliente = %s")
        params.append(cliente)
    if desde:
        wheres.append("f.fecha >= %s")
        params.append(desde)
    if hasta:
        wheres.append("f.fecha <= %s")
        params.append(hasta)
    if anio:
        wheres.append("YEAR(f.fecha) = %s")
        params.append(anio)
    if mes:
        wheres.append("MONTH(f.fecha) = %s")
        params.append(mes)
    producto_factura_ids = None
    if producto:
        prod_conn = get_legacy_connection()
        try:
            prod_cur = prod_conn.cursor(dictionary=True)
            prod_cur.execute("SELECT DISTINCT fd.factura_id FROM factura_detalle fd WHERE fd.descripcion = %s", [producto])
            rows = prod_cur.fetchall()
            producto_factura_ids = [r["factura_id"] for r in rows]
        finally:
            prod_conn.close()
        if not producto_factura_ids:
            return None, None, True
        placeholders = ",".join(["%s"] * len(producto_factura_ids))
        wheres.append(f"f.id IN ({placeholders})")
        params.extend(producto_factura_ids)
    where_sql = " AND ".join(wheres)
    return where_sql, params, False


@router.get("/dashboard-general")
def dashboard_general(
    empresa: Optional[str] = None,
    cliente: Optional[str] = None,
    producto: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    anio: Optional[int] = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        where_sql, params, empty = _build_facturas_where(empresa, cliente, producto, desde, hasta, anio=anio or 0)
        if empty:
            return {"top_clientes": [], "top_productos": [], "frecuencia": [], "por_empresa": []}

        cur.execute(f"""
            SELECT f.numero_cliente, f.empresa,
                   COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(f.total), 0) AS total
            FROM facturas f
            LEFT JOIN clientes c
              ON c.numero = f.numero_cliente
             AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(f.empresa))
            WHERE {where_sql}
            GROUP BY f.numero_cliente, f.empresa
            ORDER BY total DESC
            LIMIT 10
        """, params)
        top_clientes = cur.fetchall()

        cur.execute(f"""
            SELECT fd.descripcion AS producto,
                   SUM(fd.piezas) AS piezas,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                FROM factura_detalle fd2
                GROUP BY fd2.factura_id
            ) inv ON inv.factura_id = f.id
            WHERE {where_sql}
            GROUP BY fd.descripcion
            ORDER BY total DESC
            LIMIT 10
        """, params)
        top_productos = cur.fetchall()

        cur.execute(f"""
            SELECT t.compras AS n, COUNT(*) AS clientes
            FROM (
                SELECT f.numero_cliente, f.empresa, COUNT(DISTINCT f.id) AS compras
                FROM facturas f
                WHERE {where_sql}
                GROUP BY f.numero_cliente, f.empresa
            ) t
            GROUP BY t.compras
            ORDER BY t.compras
        """, params)
        frecuencia = cur.fetchall()

        cur.execute(f"""
            SELECT f.empresa,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(f.total), 0) AS total
            FROM facturas f
            WHERE {where_sql}
            GROUP BY f.empresa
            ORDER BY total DESC
        """, params)
        por_empresa = _consolidar_por_empresa(cur.fetchall())

        return {
            "top_clientes": top_clientes,
            "top_productos": top_productos,
            "frecuencia": frecuencia,
            "por_empresa": por_empresa,
        }
    except Exception as e:
        raise HTTPException(500, f"dashboard_general: {e}")
    finally:
        conn.close()


@router.get("/frecuencia-detalle")
def frecuencia_detalle(
    compras_min: int = Query(),
    compras_max: int = Query(),
    empresa: Optional[str] = None,
    cliente: Optional[str] = None,
    producto: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        where_sql, params, empty = _build_facturas_where(empresa, cliente, producto, desde, hasta)
        if empty:
            return {"top_productos": [], "clientes": []}

        # Clients with purchase count in range
        cur.execute(f"""
            SELECT t.numero_cliente, t.empresa,
                   COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   t.compras
            FROM (
                SELECT f.numero_cliente, f.empresa, COUNT(DISTINCT f.id) AS compras
                FROM facturas f
                WHERE {where_sql}
                GROUP BY f.numero_cliente, f.empresa
                HAVING compras >= %s AND compras <= %s
            ) t
            LEFT JOIN clientes c
              ON c.numero = t.numero_cliente
             AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(t.empresa))
            ORDER BY t.compras DESC, t.numero_cliente
        """, params + [compras_min, compras_max])
        clientes = cur.fetchall()
        if not clientes:
            return {"top_productos": [], "clientes": []}

        # Build client filter for product queries
        or_clauses = []
        or_params = []
        for c in clientes:
            or_clauses.append("(f.numero_cliente = %s AND f.empresa = %s)")
            or_params.extend([c["numero_cliente"], c["empresa"]])
        client_filter = "(" + " OR ".join(or_clauses) + ")"
        freq_where = f"{where_sql} AND {client_filter}"
        freq_params = params + or_params

        # Top products (proportional factor)
        cur.execute(f"""
            SELECT fd.descripcion AS producto,
                   SUM(fd.piezas) AS piezas,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                FROM factura_detalle fd2
                GROUP BY fd2.factura_id
            ) inv ON inv.factura_id = f.id
            WHERE {freq_where}
            GROUP BY fd.descripcion
            ORDER BY total DESC
            LIMIT 8
        """, freq_params)
        top_productos = cur.fetchall()

        # Client detail – top 5 products for web display
        # NOTE: {freq_where} appears twice so params must be duplicated
        cur.execute(f"""
            SELECT p.cliente, p.empresa,
                   COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   COALESCE(fc.total_compras, 0) AS compras,
                   COALESCE(fc.facturas, '') AS facturas,
                   SUM(p.piezas) AS total_piezas,
                   SUBSTRING_INDEX(GROUP_CONCAT(CONCAT('(', p.piezas, ' pzas) ', p.descripcion) ORDER BY p.total DESC SEPARATOR ' ||| '), ' ||| ', 5) AS productos_recomprados
            FROM (
                SELECT f.numero_cliente AS cliente, f.empresa,
                       fd.descripcion,
                       SUM(fd.piezas) AS piezas,
                       COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
                FROM facturas f
                JOIN factura_detalle fd ON fd.factura_id = f.id
                JOIN (
                    SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                    FROM factura_detalle fd2
                    GROUP BY fd2.factura_id
                ) inv ON inv.factura_id = f.id
                WHERE {freq_where}
                GROUP BY f.numero_cliente, f.empresa, fd.descripcion
            ) p
            LEFT JOIN clientes c ON c.numero = p.cliente AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(p.empresa))
            LEFT JOIN (
                SELECT f.numero_cliente, f.empresa,
                       COUNT(DISTINCT f.id) AS total_compras,
                       SUBSTRING_INDEX(GROUP_CONCAT(DISTINCT f.factura ORDER BY f.fecha DESC SEPARATOR ', '), ', ', 5) AS facturas
                FROM facturas f
                WHERE {freq_where}
                GROUP BY f.numero_cliente, f.empresa
            ) fc ON fc.numero_cliente = p.cliente AND fc.empresa = p.empresa
            GROUP BY p.cliente, p.empresa, fc.total_compras, fc.facturas
            ORDER BY compras DESC, p.cliente
        """, freq_params + freq_params)
        clientes_detalle = cur.fetchall()

        # Also get total_ventas per client within this group
        cur.execute(f"""
            SELECT f.numero_cliente AS cliente, f.empresa,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total_ventas
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                FROM factura_detalle fd2
                GROUP BY fd2.factura_id
            ) inv ON inv.factura_id = f.id
            WHERE {freq_where}
            GROUP BY f.numero_cliente, f.empresa
        """, freq_params)
        ventas_cliente = {f"{r['cliente']}::{r['empresa']}": r['total_ventas'] for r in cur.fetchall()}

        # Merge total_ventas into clientes_detalle
        for c in clientes_detalle:
            key = f"{c['cliente']}::{c['empresa']}"
            c['total_ventas'] = ventas_cliente.get(key, 0)

        return {
            "compras_min": compras_min,
            "compras_max": compras_max,
            "top_productos": top_productos,
            "clientes": clientes_detalle,
        }
    except Exception as e:
        raise HTTPException(500, f"frecuencia_detalle: {e}")
    finally:
        conn.close()


@router.get("/indicadores")
def indicadores_gerenciales(
    empresa: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'"]
        params = []
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if desde:
            wheres.append("f.fecha >= %s")
            params.append(desde)
        if hasta:
            wheres.append("f.fecha <= %s")
            params.append(hasta)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT
                COUNT(DISTINCT f.id) AS total_facturas,
                COUNT(DISTINCT CONCAT(f.numero_cliente, '|', f.empresa)) AS clientes_unicos,
                COALESCE(SUM(f.total), 0) AS total_ventas,
                COALESCE(AVG(f.total), 0) AS ticket_promedio
            FROM facturas f
            WHERE {where_sql}
        """, params)
        kpi = cur.fetchone()

        cur.execute(f"""
            SELECT f.empresa,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(f.total), 0) AS total
            FROM facturas f
            WHERE {where_sql}
            GROUP BY f.empresa
            ORDER BY total DESC
        """, params)
        por_empresa = _consolidar_por_empresa(cur.fetchall())

        cur.execute(f"""
            SELECT DATE_FORMAT(f.fecha, '%Y-%m') AS mes,
                   COALESCE(SUM(f.total), 0) AS total
            FROM facturas f
            WHERE {where_sql}
            GROUP BY DATE_FORMAT(f.fecha, '%Y-%m')
            ORDER BY mes
        """, params)
        ventas_mensuales = cur.fetchall()

        # Convertir "YYYY-MM" a nombre de mes en español
        MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        for row in ventas_mensuales:
            try:
                y, m = row["mes"].split("-")
                row["mes"] = f"{MESES[int(m)]} {y}"
            except (ValueError, KeyError, IndexError):
                pass

        growth = 0
        if desde and hasta:
            prev_params = [empresa] if empresa else []
            prev_where = ["f.estatus = 'Activa'"]
            if empresa:
                prev_where.append("f.empresa = %s")
            period_days = f"DATEDIFF(%s, %s)"
            prev_where.append(f"f.fecha >= DATE_SUB(%s, INTERVAL {period_days} DAY)")
            prev_where.append("f.fecha < %s")
            prev_params.extend([desde, hasta, desde, desde])
            prev_sql = " AND ".join(prev_where)
            cur.execute(f"""
                SELECT COALESCE(SUM(f.total), 0) AS total
                FROM facturas f
                WHERE {prev_sql}
            """, prev_params)
            prev = cur.fetchone()
            if prev and prev["total"] > 0:
                growth = ((kpi["total_ventas"] - prev["total"]) / prev["total"]) * 100

        return {
            "kpi": kpi,
            "por_empresa": por_empresa,
            "ventas_mensuales": ventas_mensuales,
            "crecimiento": growth,
        }
    except Exception as e:
        raise HTTPException(500, f"indicadores: {e}")
    finally:
        conn.close()


@router.get("/top-clientes")
def top_clientes(
    empresa: Optional[str] = None,
    limite: int = 10,
    anio: int = Query(default=0),
    mes: int = Query(default=0),
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'"]
        params = []
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if anio:
            wheres.append("YEAR(f.fecha) = %s")
            params.append(anio)
        if mes:
            wheres.append("MONTH(f.fecha) = %s")
            params.append(mes)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT f.numero_cliente, f.empresa,
                   COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(f.total), 0) AS total
            FROM facturas f
            LEFT JOIN clientes c
              ON c.numero = f.numero_cliente
             AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(f.empresa))
            WHERE {where_sql}
            GROUP BY f.numero_cliente, f.empresa
            ORDER BY total DESC
            LIMIT %s
        """, params + [limite])
        return cur.fetchall()
    except Exception as e:
        raise HTTPException(500, f"top_clientes: {e}")
    finally:
        conn.close()


@router.get("/top-productos")
def top_productos(
    empresa: Optional[str] = None,
    limite: int = 10,
    anio: int = Query(default=0),
    mes: int = Query(default=0),
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'"]
        params = []
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if anio:
            wheres.append("YEAR(f.fecha) = %s")
            params.append(anio)
        if mes:
            wheres.append("MONTH(f.fecha) = %s")
            params.append(mes)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT fd.descripcion AS producto,
                   SUM(fd.piezas) AS piezas,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(fd.piezas * fd.precio), 0) AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            WHERE {where_sql}
            GROUP BY fd.descripcion
            ORDER BY total DESC
            LIMIT %s
        """, params + [limite])
        return cur.fetchall()
    except Exception as e:
        raise HTTPException(500, f"top_productos: {e}")
    finally:
        conn.close()


@router.get("/resumen-empresa")
def resumen_empresa(
    anio: int = 0,
    mes: int = 0,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'"]
        params = []
        if anio:
            wheres.append("YEAR(f.fecha) = %s")
            params.append(anio)
        if mes:
            wheres.append("MONTH(f.fecha) = %s")
            params.append(mes)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT f.empresa,
                   COUNT(DISTINCT f.id) AS facturas,
                   COUNT(DISTINCT CONCAT(f.numero_cliente, '|', f.empresa)) AS clientes,
                   COALESCE(SUM(f.total), 0) AS total,
                   COALESCE(AVG(f.total), 0) AS ticket_promedio
            FROM facturas f
            WHERE {where_sql}
            GROUP BY f.empresa
            ORDER BY total DESC
        """, params)
        return cur.fetchall()
    except Exception as e:
        raise HTTPException(500, f"resumen_empresa: {e}")
    finally:
        conn.close()


@router.get("/reporte-cliente/{numero}")
def reporte_cliente(
    numero: str,
    empresa: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'", "f.numero_cliente = %s"]
        params = [numero]
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if desde:
            wheres.append("f.fecha >= %s")
            params.append(desde)
        if hasta:
            wheres.append("f.fecha <= %s")
            params.append(hasta)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT
                COUNT(DISTINCT f.id) AS facturas,
                COALESCE(SUM(f.total), 0) AS total,
                COALESCE(AVG(f.total), 0) AS ticket_promedio,
                MAX(f.fecha) AS ultima_factura
            FROM facturas f
            WHERE {where_sql}
        """, params)
        resumen = cur.fetchone()
        if not resumen or resumen["facturas"] == 0:
            return None

        cur.execute(f"""
            SELECT fd.descripcion AS producto,
                   SUM(fd.piezas) AS piezas,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                FROM factura_detalle fd2
                GROUP BY fd2.factura_id
            ) inv ON inv.factura_id = f.id
            WHERE {where_sql}
            GROUP BY fd.descripcion
            ORDER BY total DESC
            LIMIT 8
        """, params)
        productos = cur.fetchall()

        cur.execute(f"""
            SELECT f.id, f.factura AS folio, f.fecha, f.empresa, f.estatus, f.total, f.subtotal, f.descuento, f.iva
            FROM facturas f
            WHERE {where_sql}
            ORDER BY f.fecha DESC
        """, params)
        facturas = cur.fetchall()

        cur.execute(f"""
            SELECT COALESCE(c.nombre, '') AS nombre, c.rfc
            FROM clientes c
            WHERE c.numero = %s
            ORDER BY c.empresa LIMIT 1
        """, [numero])
        cliente_info = cur.fetchone()

        return {
            "numero": numero,
            "empresa": empresa or "",
            "nombre": (cliente_info or {}).get("nombre", ""),
            "rfc": (cliente_info or {}).get("rfc", ""),
            "resumen": resumen,
            "top_productos": productos,
            "facturas": facturas,
        }
    except Exception as e:
        raise HTTPException(500, f"reporte_cliente: {e}")
    finally:
        conn.close()


@router.get("/reporte-producto/{producto}")
def reporte_producto(
    producto: str,
    empresa: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'", "fd.descripcion = %s"]
        params = [producto]
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if desde:
            wheres.append("f.fecha >= %s")
            params.append(desde)
        if hasta:
            wheres.append("f.fecha <= %s")
            params.append(hasta)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT
                COUNT(DISTINCT f.id) AS facturas,
                COALESCE(SUM(fd.piezas), 0) AS piezas_vendidas,
                COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                FROM factura_detalle fd2
                GROUP BY fd2.factura_id
            ) inv ON inv.factura_id = f.id
            WHERE {where_sql}
        """, params)
        resumen = cur.fetchone()

        if not resumen or resumen["facturas"] == 0:
            return None

        # Apply proportional discount factor (like desktop app)
        cli_params = [producto]
        cli_wheres = ["fd.descripcion = %s", "f.estatus = 'Activa'"]
        if empresa:
            cli_wheres.append("f.empresa = %s")
            cli_params.append(empresa)
        if desde:
            cli_wheres.append("f.fecha >= %s")
            cli_params.append(desde)
        if hasta:
            cli_wheres.append("f.fecha <= %s")
            cli_params.append(hasta)
        cli_where_sql = " AND ".join(cli_wheres)
        cur.execute(f"""
            SELECT t.cliente, t.empresa,
                   COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   SUM(t.monto_real) AS total_ventas
            FROM (
                SELECT f.numero_cliente AS cliente, f.empresa,
                       fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0) AS monto_real
                FROM facturas f
                JOIN factura_detalle fd ON fd.factura_id = f.id
                JOIN (
                    SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                    FROM factura_detalle fd2
                    GROUP BY fd2.factura_id
                ) inv ON inv.factura_id = f.id
                WHERE {cli_where_sql}
            ) t
            LEFT JOIN clientes c ON c.numero = t.cliente AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(t.empresa))
            GROUP BY t.cliente, t.empresa
            ORDER BY total_ventas DESC
            LIMIT 8
        """, cli_params)
        clientes = cur.fetchall()

        cur.execute(f"""
            SELECT f.id, f.factura AS folio, f.fecha, f.empresa, f.numero_cliente,
                   COALESCE(c.nombre, '') AS cliente_nombre,
                   f.estatus, f.total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            LEFT JOIN clientes c ON c.numero = f.numero_cliente AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(f.empresa))
            WHERE {where_sql}
            ORDER BY f.fecha DESC
        """, params)
        facturas = cur.fetchall()

        return {
            "producto": producto,
            "resumen": resumen,
            "top_clientes": clientes,
            "facturas": facturas,
        }
    except Exception as e:
        raise HTTPException(500, f"reporte_producto: {e}")
    finally:
        conn.close()


@router.get("/productos")
def productos_facturados(
    empresa: Optional[str] = None,
    q: Optional[str] = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'"]
        params = []
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if q:
            wheres.append("(fd.descripcion LIKE %s OR p.cip LIKE %s)")
            params.append(f"%{q}%")
            params.append(f"%{q}%")
        where_sql = " AND ".join(wheres)
        cur.execute(f"""
            SELECT DISTINCT fd.descripcion, IFNULL(p.cip, '') AS cip
            FROM factura_detalle fd
            JOIN facturas f ON f.id = fd.factura_id
            LEFT JOIN productos p ON p.descripcion = fd.descripcion OR p.cip = fd.descripcion
            WHERE {where_sql}
            ORDER BY fd.descripcion
            LIMIT 500
        """, params)
        return [{"descripcion": r["descripcion"], "cip": r["cip"]} for r in cur.fetchall()]
    except Exception as e:
        raise HTTPException(500, f"productos: {e}")
    finally:
        conn.close()


@router.get("/clientes-reporte")
def clientes_reporte(
    empresa: Optional[str] = None,
    q: Optional[str] = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'"]
        params = []
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if q:
            wheres.append("(f.numero_cliente LIKE %s OR c.nombre LIKE %s)")
            params.append(f"%{q}%")
            params.append(f"%{q}%")
        where_sql = " AND ".join(wheres)
        cur.execute(f"""
            SELECT f.numero_cliente, COALESCE(ANY_VALUE(c.nombre), '') AS nombre, f.empresa
            FROM facturas f
            LEFT JOIN clientes c ON c.numero = f.numero_cliente AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(f.empresa))
            WHERE {where_sql}
            GROUP BY f.numero_cliente, f.empresa
            ORDER BY nombre
            LIMIT 500
        """, params)
        return cur.fetchall()
    except Exception as e:
        raise HTTPException(500, f"clientes_reporte: {e}")
    finally:
        conn.close()


def _fetch_frecuencia_data(compras_min, compras_max, empresa, cliente, producto, desde, hasta):
    """Shared helper for frecuencia-detalle JSON + exports."""
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        where_sql, params, empty = _build_facturas_where(empresa, cliente, producto, desde, hasta)
        if empty:
            return None, None, None, None, None

        cur.execute(f"""
            SELECT t.numero_cliente, t.empresa,
                   COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   t.compras
            FROM (
                SELECT f.numero_cliente, f.empresa, COUNT(DISTINCT f.id) AS compras
                FROM facturas f
                WHERE {where_sql}
                GROUP BY f.numero_cliente, f.empresa
                HAVING compras >= %s AND compras <= %s
            ) t
            LEFT JOIN clientes c
              ON c.numero = t.numero_cliente
             AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(t.empresa))
            ORDER BY t.compras DESC, t.numero_cliente
        """, params + [compras_min, compras_max])
        clientes = cur.fetchall()
        if not clientes:
            return None, None, None, None, None

        or_clauses = []
        or_params = []
        for c in clientes:
            or_clauses.append("(f.numero_cliente = %s AND f.empresa = %s)")
            or_params.extend([c["numero_cliente"], c["empresa"]])
        client_filter = "(" + " OR ".join(or_clauses) + ")"
        freq_where = f"{where_sql} AND {client_filter}"
        freq_params = params + or_params

        # Top products
        cur.execute(f"""
            SELECT fd.descripcion AS producto,
                   SUM(fd.piezas) AS piezas,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                FROM factura_detalle fd2
                GROUP BY fd2.factura_id
            ) inv ON inv.factura_id = f.id
            WHERE {freq_where}
            GROUP BY fd.descripcion
            ORDER BY total DESC
            LIMIT 8
        """, freq_params)
        top_productos = cur.fetchall()

        # Client detail – all products for export
        cur.execute("SET SESSION group_concat_max_len = 100000")
        try: cur.fetchall()
        except: pass
        cur.execute(f"""
            SELECT p.cliente, p.empresa,
                   COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   COALESCE(fc.total_compras, 0) AS compras,
                   COALESCE(fc.facturas, '') AS facturas,
                   SUM(p.piezas) AS total_piezas,
                   GROUP_CONCAT(CONCAT('(', p.piezas, ' pzas) ', p.descripcion) ORDER BY p.total DESC SEPARATOR ' ||| ') AS productos_recomprados
            FROM (
                SELECT f.numero_cliente AS cliente, f.empresa,
                       fd.descripcion,
                       SUM(fd.piezas) AS piezas,
                       COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
                FROM facturas f
                JOIN factura_detalle fd ON fd.factura_id = f.id
                JOIN (
                    SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                    FROM factura_detalle fd2
                    GROUP BY fd2.factura_id
                ) inv ON inv.factura_id = f.id
                WHERE {freq_where}
                GROUP BY f.numero_cliente, f.empresa, fd.descripcion
            ) p
            LEFT JOIN clientes c ON c.numero = p.cliente AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(p.empresa))
            LEFT JOIN (
                SELECT f.numero_cliente, f.empresa,
                       COUNT(DISTINCT f.id) AS total_compras,
                       GROUP_CONCAT(DISTINCT f.factura ORDER BY f.fecha DESC SEPARATOR ', ') AS facturas
                FROM facturas f
                WHERE {freq_where}
                GROUP BY f.numero_cliente, f.empresa
            ) fc ON fc.numero_cliente = p.cliente AND fc.empresa = p.empresa
            GROUP BY p.cliente, p.empresa, fc.total_compras, fc.facturas
            ORDER BY compras DESC, p.cliente
        """, freq_params + freq_params)
        clientes_detalle = cur.fetchall()

        # Total ventas per client
        cur.execute(f"""
            SELECT f.numero_cliente AS cliente, f.empresa,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total_ventas
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                FROM factura_detalle fd2
                GROUP BY fd2.factura_id
            ) inv ON inv.factura_id = f.id
            WHERE {freq_where}
            GROUP BY f.numero_cliente, f.empresa
        """, freq_params)
        ventas_cliente = {f"{r['cliente']}::{r['empresa']}": r['total_ventas'] for r in cur.fetchall()}
        for c in clientes_detalle:
            key = f"{c['cliente']}::{c['empresa']}"
            c['total_ventas'] = ventas_cliente.get(key, 0)

        return clientes, top_productos, clientes_detalle, freq_where, freq_params
    finally:
        conn.close()


def _gen_freq_chart(clientes, compras_min, compras_max):
    """Generate a frequency bar chart using matplotlib, return PNG bytes."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    counts = {}
    for c in clientes:
        n = c["compras"]
        counts[n] = counts.get(n, 0) + 1

    labels = sorted(counts.keys())
    values = [counts[k] for k in labels]

    fig, ax = plt.subplots(figsize=(6, 2.5))
    bars = ax.bar([str(l) for l in labels], values, color="#F5A623", edgecolor="black")
    ax.set_title(f"Frecuencia de recompra ({compras_min}–{compras_max} compras)", fontsize=10)
    ax.set_xlabel("Compras", fontsize=8)
    ax.set_ylabel("Clientes", fontsize=8)
    ax.tick_params(labelsize=7)
    for bar, v in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3, str(v),
                ha="center", va="bottom", fontsize=8)
    plt.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


@router.get("/frecuencia-detalle/export-excel")
def frecuencia_detalle_export_excel(
    compras_min: int = Query(),
    compras_max: int = Query(),
    empresa: Optional[str] = None,
    cliente: Optional[str] = None,
    producto: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    clientes, top_productos, clientes_detalle, freq_where, freq_params = \
        _fetch_frecuencia_data(compras_min, compras_max, empresa, cliente, producto, desde, hasta)
    if not clientes:
        raise HTTPException(404, "No hay datos para exportar.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle Recompra"

    hdr_fill = PatternFill(start_color="0A74D7", end_color="0A74D7", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    total_font = Font(bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = f"Detalle de Recompra · {compras_min}–{compras_max} compras"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary
    total_clientes = len(clientes)
    total_productos = len(set(p["producto"] for p in top_productos))
    total_monto = sum(c["total_ventas"] for c in clientes_detalle)

    summary = [
        ("Rango de compras", f"{compras_min}–{compras_max}"),
        ("Clientes en el grupo", total_clientes),
        ("Total productos distintos", total_productos),
        ("Monto total recomprado ($)", f"${total_monto:,.2f}"),
    ]
    ws.cell(row=3, column=1, value="Métrica").font = hdr_font
    ws.cell(row=3, column=1).fill = hdr_fill
    ws.cell(row=3, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=1).border = thin_border
    ws.cell(row=3, column=2, value="Valor").font = hdr_font
    ws.cell(row=3, column=2).fill = hdr_fill
    ws.cell(row=3, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=2).border = thin_border
    for i, (label, val) in enumerate(summary):
        r = 4 + i
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=2, value=val).font = value_font
        for col in (1, 2):
            ws.cell(row=r, column=col).border = thin_border

    # Chart image
    chart_png = _gen_freq_chart(clientes, compras_min, compras_max)
    tmp_chart = os.path.join(tempfile.gettempdir(), f"freq_chart_{compras_min}_{compras_max}.png")
    with open(tmp_chart, "wb") as f:
        f.write(chart_png)
    ws.add_image(XLImage(tmp_chart), "A9")

    # Client detail table
    table_start = 35
    cols = ["Cliente", "Empresa", "Compras", "Total Ventas", "Total Piezas", "Productos Recomprados", "Facturas"]
    for ci, col_name in enumerate(cols, 1):
        cell = ws.cell(row=table_start, column=ci, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, c in enumerate(clientes_detalle):
        r = table_start + 1 + ri
        vals = [
            c.get("nombre") or c.get("cliente", ""),
            c.get("empresa", ""),
            c["compras"],
            c["total_ventas"],
            c.get("total_piezas", 0),
            c.get("productos_recomprados", ""),
            c.get("facturas", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = value_font
            cell.border = thin_border
            if ci in (3, 5):
                cell.alignment = Alignment(horizontal="center")
            elif ci == 4:
                cell.number_format = '#,##0.00'
            elif ci in (6, 7):
                cell.alignment = Alignment(wrap_text=True)

    # Total row
    total_row = table_start + 1 + len(clientes_detalle)
    total_vals = ["Total", "", sum(c["compras"] for c in clientes_detalle),
                  sum(c["total_ventas"] for c in clientes_detalle),
                  sum(c.get("total_piezas", 0) for c in clientes_detalle), "", ""]
    for ci, v in enumerate(total_vals, 1):
        cell = ws.cell(row=total_row, column=ci, value=v)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        if ci == 4:
            cell.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 70
    ws.column_dimensions["G"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    try:
        os.remove(tmp_chart)
    except Exception:
        pass

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Recompra_{compras_min}_{compras_max}.xlsx"'},
    )


@router.get("/frecuencia-detalle/export-pdf")
def frecuencia_detalle_export_pdf(
    compras_min: int = Query(),
    compras_max: int = Query(),
    empresa: Optional[str] = None,
    cliente: Optional[str] = None,
    producto: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    clientes, top_productos, clientes_detalle, freq_where, freq_params = \
        _fetch_frecuencia_data(compras_min, compras_max, empresa, cliente, producto, desde, hasta)
    if not clientes:
        raise HTTPException(404, "No hay datos para exportar.")

    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=36, leftMargin=36,
                             topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    flow = []

    # Title
    empresa_text = empresa or "Todas"
    title = Paragraph(
        f"<b>Detalle de Recompra</b><br/>{compras_min}–{compras_max} compras · Empresa: {empresa_text}",
        styles["Title"]
    )
    flow.append(title)
    flow.append(Spacer(1, 12))

    # Summary
    total_clientes = len(clientes)
    total_productos = len(set(p["producto"] for p in top_productos))
    total_monto = sum(c["total_ventas"] for c in clientes_detalle)

    resumen_text = f"""
        <b>Clientes en el grupo:</b> {total_clientes}<br/>
        <b>Total productos distintos:</b> {total_productos}<br/>
        <b>Monto total recomprado:</b> ${total_monto:,.2f}
    """
    flow.append(Paragraph(resumen_text, styles["Normal"]))
    flow.append(Spacer(1, 12))

    # Chart
    chart_png = _gen_freq_chart(clientes, compras_min, compras_max)
    tmp_chart = os.path.join(tempfile.gettempdir(), f"freq_pdf_chart_{compras_min}_{compras_max}.png")
    with open(tmp_chart, "wb") as f:
        f.write(chart_png)
    flow.append(Image(tmp_chart, width=500, height=200))
    flow.append(Spacer(1, 16))

    # Client table
    cell_left = ParagraphStyle(name="CellLeft", fontSize=7, alignment=0)
    cell_center = ParagraphStyle(name="CellCenter", fontSize=7, alignment=1)

    data = [["Cliente", "Empresa", "Compras", "Total Piezas", "Productos Recomprados", "Facturas"]]
    for c in clientes_detalle:
        data.append([
            Paragraph(str(c.get("nombre") or c.get("cliente", "")), cell_left),
            Paragraph(str(c.get("empresa", "")), cell_center),
            Paragraph(str(c["compras"]), cell_center),
            Paragraph(str(int(c.get("total_piezas", 0))), cell_center),
            Paragraph(str(c.get("productos_recomprados", "")), cell_left),
            Paragraph(str(c.get("facturas", "")), cell_left),
        ])

    total_compras = sum(c["compras"] for c in clientes_detalle)
    total_piezas = sum(c.get("total_piezas", 0) for c in clientes_detalle)
    data.append([
        Paragraph("<b>Total</b>", cell_center),
        "", Paragraph(f"<b>{int(total_compras)}</b>", cell_center),
        Paragraph(f"<b>{int(total_piezas)}</b>", cell_center), "", ""
    ])

    table = Table(data, colWidths=[90, 40, 35, 35, 170, 190])
    total_filas = len(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0A74D7")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (2, 1), (3, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, total_filas - 2), colors.whitesmoke),
        ('BACKGROUND', (0, total_filas - 1), (-1, total_filas - 1), colors.lightgrey),
        ('FONTNAME', (0, total_filas - 1), (-1, total_filas - 1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    flow.append(table)

    doc.build(flow)

    try:
        os.remove(tmp_chart)
    except Exception:
        pass

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Recompra_{compras_min}_{compras_max}.pdf"'},
    )


def _fetch_dashboard_data(empresa, cliente, producto, desde, hasta):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        where_sql, params, empty = _build_facturas_where(empresa, cliente, producto, desde, hasta)
        top_clientes = []
        top_productos = []
        if not empty:
            cur.execute(f"""
                SELECT f.numero_cliente, f.empresa,
                       COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                       COUNT(DISTINCT f.id) AS facturas,
                       COALESCE(SUM(f.total), 0) AS total
                FROM facturas f
                LEFT JOIN clientes c
                  ON c.numero = f.numero_cliente
                 AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(f.empresa))
                WHERE {where_sql}
                GROUP BY f.numero_cliente, f.empresa
                ORDER BY total DESC
                LIMIT 10
            """, params)
            top_clientes = cur.fetchall()

            cur.execute(f"""
                SELECT fd.descripcion AS producto,
                       SUM(fd.piezas) AS piezas,
                       COUNT(DISTINCT f.id) AS facturas,
                       COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
                FROM facturas f
                JOIN factura_detalle fd ON fd.factura_id = f.id
                JOIN (
                    SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto
                    FROM factura_detalle fd2
                    GROUP BY fd2.factura_id
                ) inv ON inv.factura_id = f.id
                WHERE {where_sql}
                GROUP BY fd.descripcion
                ORDER BY total DESC
                LIMIT 10
            """, params)
            top_productos = cur.fetchall()
        return top_clientes, top_productos, where_sql, params
    finally:
        conn.close()


@router.get("/dashboard-general/export-excel")
def dashboard_export_excel(
    empresa: Optional[str] = None,
    cliente: Optional[str] = None,
    producto: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    top_clientes, top_productos, where_sql, params = \
        _fetch_dashboard_data(empresa, cliente, producto, desde, hasta)
    if not top_clientes and not top_productos:
        raise HTTPException(404, "No hay datos para exportar.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet: Clientes Top ---
    ws1 = wb.active
    ws1.title = "Clientes Top"

    hdr_fill = PatternFill(start_color="0A74D7", end_color="0A74D7", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    ws1.merge_cells("A1:E1")
    ws1["A1"] = "Clientes Top · Dashboard General"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A1"].alignment = Alignment(horizontal="center")

    cols1 = ["Cliente", "Empresa", "Nombre", "Facturas", "Total"]
    for ci, col_name in enumerate(cols1, 1):
        cell = ws1.cell(row=3, column=ci, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, c in enumerate(top_clientes):
        r = 4 + ri
        vals = [
            c.get("numero_cliente", ""),
            c.get("empresa", ""),
            c.get("nombre", ""),
            c["facturas"],
            c["total"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=ci, value=v)
            cell.font = value_font
            cell.border = thin_border
            if ci == 5:
                cell.number_format = '#,##0.00'

    ws1.column_dimensions["A"].width = 16
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 15

    # --- Productos más vendidos (same sheet) ---
    prod_start = 4 + len(top_clientes) + 2
    ws1.merge_cells(start_row=prod_start, start_column=1, end_row=prod_start, end_column=4)
    ws1.cell(row=prod_start, column=1, value="Productos más vendidos · Dashboard General").font = Font(bold=True, size=14)
    ws1.cell(row=prod_start, column=1).alignment = Alignment(horizontal="center")

    cols2 = ["Producto", "Piezas", "Facturas", "Total"]
    for ci, col_name in enumerate(cols2, 1):
        cell = ws1.cell(row=prod_start + 2, column=ci, value=col_name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border

    for ri, p in enumerate(top_productos):
        r = prod_start + 3 + ri
        vals = [p.get("producto", ""), int(p.get("piezas", 0)), p["facturas"], p["total"]]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=ci, value=v)
            cell.font = value_font; cell.border = thin_border
            if ci == 4:
                cell.number_format = '#,##0.00'

    # --- Charts (same sheet) ---
    from matplotlib import pyplot as plt
    from openpyxl.drawing.image import Image as XlImage

    chart_row = prod_start + 3 + len(top_productos) + 2

    if top_clientes:
        labels = [f"{c.get('nombre','')[:15]}" for c in reversed(top_clientes)]
        values = [float(c['total']) for c in reversed(top_clientes)]
        fig, ax = plt.subplots(figsize=(7, 3.5))
        ax.barh(range(len(labels)), values, color="#2563eb", zorder=3)
        ax.set_yticks(range(len(labels)))
        ax.set_yticklabels(labels, fontsize=8)
        ax.set_title("Clientes Top por monto", fontsize=12, fontweight="bold")
        ax.set_xlabel("Monto ($)"); ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
        fig.subplots_adjust(left=0.20, right=0.95, top=0.88, bottom=0.12)
        buf1 = io.BytesIO()
        fig.savefig(buf1, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig); buf1.seek(0)
        ws1.add_image(XlImage(buf1), f"A{chart_row}")

    if top_productos:
        labels = [p.get('producto','')[:30] for p in reversed(top_productos)]
        values = [float(p['total']) for p in reversed(top_productos)]
        fig, ax = plt.subplots(figsize=(7, 3.5))
        ax.barh(range(len(labels)), values, color="#10b981", zorder=3)
        ax.set_yticks(range(len(labels)))
        ax.set_yticklabels(labels, fontsize=8)
        ax.set_title("Productos más vendidos por monto", fontsize=12, fontweight="bold")
        ax.set_xlabel("Monto ($)"); ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
        fig.subplots_adjust(left=0.25, right=0.95, top=0.88, bottom=0.12)
        buf2 = io.BytesIO()
        fig.savefig(buf2, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig); buf2.seek(0)
        anchor = f"A{chart_row + 20 if top_clientes else chart_row}"
        ws1.add_image(XlImage(buf2), anchor)

    _auto_width(ws1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Dashboard_General.xlsx"'},
    )


@router.get("/dashboard-general/export-pdf")
def dashboard_export_pdf(
    empresa: Optional[str] = None,
    cliente: Optional[str] = None,
    producto: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    top_clientes, top_productos, where_sql, params = \
        _fetch_dashboard_data(empresa, cliente, producto, desde, hasta)
    if not top_clientes and not top_productos:
        raise HTTPException(404, "No hay datos para exportar.")

    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=36, leftMargin=36,
                             topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    flow = []

    empresa_text = empresa or "Todas"
    title = Paragraph(
        f"<b>Dashboard General</b><br/>Empresa: {empresa_text}",
        styles["Title"]
    )
    flow.append(title)
    flow.append(Spacer(1, 12))

    # --- Clientes Top ---
    flow.append(Paragraph("<b>Clientes Top</b>", styles["Heading2"]))
    flow.append(Spacer(1, 6))

    cell_left = ParagraphStyle(name="CellLeft", fontSize=8, alignment=0)
    cell_center = ParagraphStyle(name="CellCenter", fontSize=8, alignment=1)

    data1 = [["Cliente", "Empresa", "Nombre", "Facturas", "Total"]]
    for c in top_clientes:
        data1.append([
            Paragraph(str(c.get("numero_cliente", "")), cell_center),
            Paragraph(str(c.get("empresa", "")), cell_center),
            Paragraph(str(c.get("nombre", "")), cell_left),
            Paragraph(str(c["facturas"]), cell_center),
            Paragraph(f"${c['total']:,.2f}", cell_center),
        ])

    table1 = Table(data1, colWidths=[55, 90, 230, 48, 80])
    table1.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0A74D7")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    flow.append(table1)
    flow.append(Spacer(1, 20))

    # --- Productos más vendidos ---
    flow.append(Paragraph("<b>Productos más vendidos</b>", styles["Heading2"]))
    flow.append(Spacer(1, 6))

    data2 = [["Producto", "Piezas", "Facturas", "Total"]]
    for p in top_productos:
        data2.append([
            Paragraph(str(p.get("producto", "")), cell_left),
            Paragraph(str(int(p.get("piezas", 0))), cell_center),
            Paragraph(str(p["facturas"]), cell_center),
            Paragraph(f"${p['total']:,.2f}", cell_center),
        ])

    table2 = Table(data2, colWidths=[310, 50, 50, 80])
    table2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0891b2")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    flow.append(table2)

    # --- Charts ---
    if top_clientes or top_productos:
        import matplotlib
        matplotlib.use("Agg")
        from matplotlib import pyplot as plt
        from reportlab.platypus import Image as RlImage

        if top_clientes:
            labels = [f"{c.get('nombre','')[:15]}" for c in reversed(top_clientes)]
            values = [float(c['total']) for c in reversed(top_clientes)]
            fig, ax = plt.subplots(figsize=(7, 3))
            ax.barh(range(len(labels)), values, color="#2563eb", zorder=3)
            ax.set_yticks(range(len(labels)))
            ax.set_yticklabels(labels, fontsize=8)
            
            ax.set_title("Clientes Top por monto", fontsize=12, fontweight="bold")
            ax.set_xlabel("Monto ($)")
            ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
            fig.subplots_adjust(left=0.20, right=0.95, top=0.88, bottom=0.15)
            buf1 = io.BytesIO()
            fig.savefig(buf1, format="png", dpi=120, bbox_inches="tight")
            plt.close(fig); buf1.seek(0)
            flow.append(Spacer(1, 12))
            flow.append(RlImage(buf1, width=480, height=220))

        if top_productos:
            labels = [p.get('producto','')[:30] for p in reversed(top_productos)]
            values = [float(p['total']) for p in reversed(top_productos)]
            fig, ax = plt.subplots(figsize=(7, 3))
            ax.barh(range(len(labels)), values, color="#10b981", zorder=3)
            ax.set_yticks(range(len(labels)))
            ax.set_yticklabels(labels, fontsize=8)
            
            ax.set_title("Productos más vendidos por monto", fontsize=12, fontweight="bold")
            ax.set_xlabel("Monto ($)")
            ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
            fig.subplots_adjust(left=0.25, right=0.95, top=0.88, bottom=0.15)
            buf2 = io.BytesIO()
            fig.savefig(buf2, format="png", dpi=120, bbox_inches="tight")
            plt.close(fig); buf2.seek(0)
            flow.append(Spacer(1, 12))
            flow.append(RlImage(buf2, width=480, height=220))

    doc.build(flow)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="Dashboard_General.pdf"'},
    )


def _fetch_reporte_cliente(numero, empresa, desde, hasta):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'", "f.numero_cliente = %s"]
        params = [numero]
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if desde:
            wheres.append("f.fecha >= %s")
            params.append(desde)
        if hasta:
            wheres.append("f.fecha <= %s")
            params.append(hasta)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(f.total), 0) AS total,
                   COALESCE(AVG(f.total), 0) AS ticket_promedio,
                   MAX(f.fecha) AS ultima_factura
            FROM facturas f WHERE {where_sql}
        """, params)
        resumen = cur.fetchone()
        if not resumen or resumen["facturas"] == 0:
            return None, None, None, None

        cur.execute(f"""
            SELECT fd.descripcion AS producto,
                   SUM(fd.piezas) AS piezas,
                   COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
            FROM facturas f JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto FROM factura_detalle fd2 GROUP BY fd2.factura_id) inv ON inv.factura_id = f.id
            WHERE {where_sql} GROUP BY fd.descripcion ORDER BY total DESC
        """, params)
        productos = cur.fetchall()

        cur.execute(f"""
            SELECT f.id, f.factura AS folio, f.fecha, f.empresa, f.estatus, f.total
            FROM facturas f WHERE {where_sql} ORDER BY f.fecha DESC
        """, params)
        facturas = cur.fetchall()

        cur.execute("SELECT COALESCE(c.nombre,'') AS nombre, c.rfc FROM clientes c WHERE c.numero = %s ORDER BY c.empresa LIMIT 1", [numero])
        cliente_info = cur.fetchone()
        return resumen, productos, facturas, cliente_info
    finally:
        conn.close()


@router.get("/reporte-cliente/{numero}/export-excel")
def reporte_cliente_export_excel(
    numero: str,
    empresa: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    resumen, productos, facturas, cliente_info = _fetch_reporte_cliente(numero, empresa, desde, hasta)
    if resumen is None:
        raise HTTPException(404, "No hay datos para exportar.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    nombre_cliente = (cliente_info or {}).get("nombre", "")

    # --- Sheet 1: Reporte (Resumen + Productos + Gráfico) ---
    ws = wb.active
    ws.title = "Reporte"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Reporte de Cliente — {numero} {nombre_cliente}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    resumen_data = [
        ("Cliente", nombre_cliente),
        ("Número", numero),
        ("RFC", (cliente_info or {}).get("rfc", "")),
        ("", ""),
        ("KPI", "Valor"),
        ("Total vendido", f"${resumen['total']:,.2f}"),
        ("Facturas activas", str(resumen['facturas'])),
        ("Última compra", str(resumen.get('ultima_factura', '') or '')),
        ("Ticket Promedio", f"${resumen['ticket_promedio']:,.2f}"),
    ]
    for i, (k, v) in enumerate(resumen_data):
        r = 3 + i
        cell_k = ws.cell(row=r, column=1, value=k)
        cell_v = ws.cell(row=r, column=2, value=v)
        if k in ("KPI",):
            cell_k.font = hdr_font; cell_k.fill = hdr_fill
            cell_v.font = hdr_font; cell_v.fill = hdr_fill
        elif k:
            cell_k.font = label_font
            cell_v.font = value_font
        cell_k.border = thin_border; cell_v.border = thin_border

    from matplotlib import pyplot as plt
    from openpyxl.drawing.image import Image as XlImage
    chart_row = 3 + len(resumen_data) + 1
    top10 = sorted(productos, key=lambda p: p['total'], reverse=True)[:10]
    if top10:
        labels = [p['producto'][:35] for p in reversed(top10)]
        values = [float(p['total']) for p in reversed(top10)]
        fig, ax = plt.subplots(figsize=(7, 3.5))
        ax.barh(range(len(labels)), values, color="#2563eb", zorder=3)
        ax.set_yticks(range(len(labels))); ax.set_yticklabels(labels, fontsize=8)
        ax.set_title("Top 10 productos por monto vendido", fontsize=12, fontweight="bold")
        ax.set_xlabel("Monto ($)"); ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
        fig.subplots_adjust(left=0.25, right=0.95, top=0.88, bottom=0.12)
        chart_buf = io.BytesIO()
        fig.savefig(chart_buf, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig); chart_buf.seek(0)
        ws.add_image(XlImage(chart_buf), f"A{chart_row}")

    prod_header_row = chart_row + 20
    ws.merge_cells(start_row=prod_header_row, start_column=1, end_row=prod_header_row, end_column=4)
    ws.cell(row=prod_header_row, column=1, value="Productos más comprados").font = Font(bold=True, size=12)

    cols_p = ["Producto", "Piezas", "Facturas", "Total"]
    for ci, col_name in enumerate(cols_p, 1):
        cell = ws.cell(row=prod_header_row + 1, column=ci, value=col_name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
    for ri, p in enumerate(productos):
        r = prod_header_row + 2 + ri
        ws.cell(row=r, column=1, value=p['producto']).font = value_font; ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=int(p['piezas'])).font = value_font; ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=3, value=p['facturas']).font = value_font; ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=p['total']).font = value_font; ws.cell(row=r, column=4).number_format = '#,##0.00'; ws.cell(row=r, column=4).border = thin_border
    _auto_width(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 50)

    # --- Sheet 2: Facturas ---
    ws2 = wb.create_sheet("Facturas")
    cols_f = ["Folio", "Fecha", "Estatus", "Total"]
    for ci, col_name in enumerate(cols_f, 1):
        cell = ws2.cell(row=1, column=ci, value=col_name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
    for ri, f in enumerate(facturas):
        r = 2 + ri
        ws2.cell(row=r, column=1, value=f.get('folio', '')).font = value_font; ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=2, value=str(f.get('fecha', '') or '')).font = value_font; ws2.cell(row=r, column=2).border = thin_border
        ws2.cell(row=r, column=3, value=f.get('estatus', '')).font = value_font; ws2.cell(row=r, column=3).border = thin_border
        ws2.cell(row=r, column=4, value=f['total']).font = value_font; ws2.cell(row=r, column=4).number_format = '#,##0.00'; ws2.cell(row=r, column=4).border = thin_border
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reporte_Cliente_{numero}.xlsx"'},
    )


@router.get("/reporte-cliente/{numero}/export-pdf")
def reporte_cliente_export_pdf(
    numero: str,
    empresa: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    resumen, productos, facturas, cliente_info = _fetch_reporte_cliente(numero, empresa, desde, hasta)
    if resumen is None:
        raise HTTPException(404, "No hay datos para exportar.")

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.platypus import Table as RlTable, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=letter)
    width, height = letter

    nombre_cliente = (cliente_info or {}).get("nombre", "")

    def encabezado():
        c.setFillColorRGB(37/255, 99/255, 235/255)
        c.rect(0, height - 40, width, 40, fill=1, stroke=0)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, height - 28, "Reporte de Cliente")
        c.setFont("Helvetica", 11)
        c.drawRightString(width - 40, height - 28, f"{numero} — {nombre_cliente}")

    def draw_table(data, col_widths, y_pos):
        t = RlTable(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f0f4f8")]),
        ]))
        w, h = t.wrapOn(c, width - 80, 400)
        t.drawOn(c, 40, y_pos - h)
        return y_pos - h - 25

    # --- Page 1: Datos + KPIs + Chart ---
    encabezado()
    y = height - 80

    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(0, 0, 0)
    c.drawString(40, y, "Datos del cliente"); y -= 20
    c.setFont("Helvetica", 11)
    for d in [f"Cliente: {nombre_cliente}", f"Número: {numero}", f"RFC: {(cliente_info or {}).get('rfc', '')}"]:
        c.drawString(50, y, d); y -= 16
    y -= 10

    c.setFont("Helvetica-Bold", 12)
    c.drawString(40, y, "KPIs"); y -= 8
    card_w, card_h, gap = 125, 55, 8
    x_start = 40
    ultima = resumen.get('ultima_factura', '')
    if isinstance(ultima, (datetime.date, datetime.datetime)):
        ultima = ultima.strftime('%d/%m/%Y')
    else:
        ultima = str(ultima or '')
    kpi_data = [
        ("Ventas Totales", f"${resumen['total']:,.2f}", (32/255, 84/255, 147/255)),
        ("Facturas", str(resumen['facturas']), (22/255, 163/255, 74/255)),
        ("Ticket Promedio", f"${resumen['ticket_promedio']:,.2f}", (245/255, 166/255, 35/255)),
        ("Ultima Factura", ultima, (124/255, 58/255, 237/255)),
    ]
    for i, (label, value, rgb) in enumerate(kpi_data):
        cx = x_start + i * (card_w + gap)
        c.setFillColorRGB(1, 1, 1)
        c.setStrokeColorRGB(0.85, 0.85, 0.85)
        c.rect(cx, y - card_h, card_w, card_h, fill=1, stroke=1)
        c.setFillColorRGB(*rgb)
        c.rect(cx, y - card_h, 4, card_h, fill=1, stroke=0)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.setFont("Helvetica", 7)
        c.drawString(cx + 10, y - 14, label)
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 13)
        c.drawString(cx + 10, y - 34, value)
    y = y - card_h - 16

    if productos:
        import matplotlib
        matplotlib.use("Agg")
        from matplotlib import pyplot as plt

        top_prod = sorted(productos, key=lambda p: p['total'], reverse=True)[:10]
        labels = [p['producto'][:40] for p in reversed(top_prod)]
        values = [p['total'] for p in reversed(top_prod)]

        fig, ax = plt.subplots(figsize=(7.5, 3.5))
        ax.barh(range(len(labels)), values, color="#2563eb", zorder=3)
        ax.set_yticks(range(len(labels)))
        ax.set_yticklabels(labels, fontsize=7)
        ax.set_title("Top 10 productos por monto vendido", fontsize=12, fontweight="bold", pad=12)
        ax.set_xlabel("Monto ($)")
        ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
        fig.subplots_adjust(left=0.01, right=0.99, top=0.88, bottom=0.18)
        chart_buf = io.BytesIO()
        fig.savefig(chart_buf, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig)
        chart_buf.seek(0)

        img = ImageReader(chart_buf)
        c.drawImage(img, 40, y - 170, width=width - 80, height=170, preserveAspectRatio=True)

    # --- Page 2: Productos table + Facturas table ---
    c.showPage()
    encabezado()
    y = height - 80

    if productos:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, "Productos"); y -= 20
        p_data = [["Producto", "Piezas", "Facturas", "Total"]]
        for p in productos:
            p_data.append([str(p['producto']), str(int(p['piezas'])), str(p['facturas']), f"${p['total']:,.2f}"])
        y = draw_table(p_data, [300, 50, 50, 80], y)
        if y < 150: c.showPage(); encabezado(); y = height - 80

    if facturas:
        if y < 200: c.showPage(); encabezado(); y = height - 80
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, "Facturas"); y -= 20
        f_data = [["Folio", "Fecha", "Estatus", "Total"]]
        for f in facturas:
            f_data.append([str(f.get('folio', '')), str(f.get('fecha', '') or ''), str(f.get('estatus', '')), f"${f['total']:,.2f}"])
        y = draw_table(f_data, [100, 100, 100, 120], y)

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Reporte_Cliente_{numero}.pdf"'},
    )


def _fetch_reporte_producto(producto, empresa, desde, hasta):
    conn = get_legacy_connection()
    try:
        cur = conn.cursor(dictionary=True)
        wheres = ["f.estatus = 'Activa'", "fd.descripcion = %s"]
        params = [producto]
        if empresa:
            wheres.append("f.empresa = %s")
            params.append(empresa)
        if desde:
            wheres.append("f.fecha >= %s")
            params.append(desde)
        if hasta:
            wheres.append("f.fecha <= %s")
            params.append(hasta)
        where_sql = " AND ".join(wheres)

        cur.execute(f"""
            SELECT COUNT(DISTINCT f.id) AS facturas,
                   COALESCE(SUM(fd.piezas), 0) AS piezas_vendidas,
                   COALESCE(SUM(fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0)), 0) AS total
            FROM facturas f JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto FROM factura_detalle fd2 GROUP BY fd2.factura_id) inv ON inv.factura_id = f.id
            WHERE {where_sql}
        """, params)
        resumen = cur.fetchone()
        if not resumen or resumen["facturas"] == 0:
            return None, None, None

        cli_params = [producto]
        cli_wheres = ["fd.descripcion = %s", "f.estatus = 'Activa'"]
        if empresa:
            cli_wheres.append("f.empresa = %s")
            cli_params.append(empresa)
        if desde:
            cli_wheres.append("f.fecha >= %s")
            cli_params.append(desde)
        if hasta:
            cli_wheres.append("f.fecha <= %s")
            cli_params.append(hasta)
        cli_where = " AND ".join(cli_wheres)
        cur.execute(f"""
            SELECT t.cliente, t.empresa, COALESCE(ANY_VALUE(c.nombre), '') AS nombre,
                   SUM(t.monto_real) AS total_ventas
            FROM (SELECT f.numero_cliente AS cliente, f.empresa, fd.piezas * fd.precio * f.total / NULLIF(inv.bruto, 0) AS monto_real
                  FROM facturas f JOIN factura_detalle fd ON fd.factura_id = f.id
                  JOIN (SELECT fd2.factura_id, SUM(fd2.piezas * fd2.precio) AS bruto FROM factura_detalle fd2 GROUP BY fd2.factura_id) inv ON inv.factura_id = f.id
                  WHERE {cli_where}) t
            LEFT JOIN clientes c ON c.numero = t.cliente AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(t.empresa))
            GROUP BY t.cliente, t.empresa ORDER BY total_ventas DESC
        """, cli_params)
        clientes = cur.fetchall()

        cur.execute(f"""
            SELECT f.id, f.factura AS folio, f.fecha, f.empresa, f.numero_cliente,
                   COALESCE(c.nombre, '') AS cliente_nombre, f.estatus, f.total
            FROM facturas f JOIN factura_detalle fd ON fd.factura_id = f.id
            LEFT JOIN clientes c ON c.numero = f.numero_cliente AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(f.empresa))
            WHERE {where_sql} ORDER BY f.fecha DESC
        """, params)
        facturas = cur.fetchall()
        return resumen, clientes, facturas
    finally:
        conn.close()


@router.get("/reporte-producto/{producto}/export-excel")
def reporte_producto_export_excel(
    producto: str,
    empresa: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    resumen, clientes, facturas = _fetch_reporte_producto(producto, empresa, desde, hasta)
    if resumen is None:
        raise HTTPException(404, "No hay datos para exportar.")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # --- Sheet 1: Reporte (Resumen + Clientes + Gráfico) ---
    ws = wb.active
    ws.title = "Reporte"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Reporte de Producto — {producto}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    resumen_data = [
        ("Producto", producto),
        ("", ""),
        ("KPI", "Valor"),
        ("Total vendido", f"${resumen['total']:,.2f}"),
        ("Piezas vendidas", str(int(resumen['piezas_vendidas']))),
        ("Facturas", str(resumen['facturas'])),
    ]
    for i, (k, v) in enumerate(resumen_data):
        r = 3 + i
        cell_k = ws.cell(row=r, column=1, value=k)
        cell_v = ws.cell(row=r, column=2, value=v)
        if k in ("KPI",):
            cell_k.font = hdr_font; cell_k.fill = hdr_fill
            cell_v.font = hdr_font; cell_v.fill = hdr_fill
        elif k:
            cell_k.font = label_font
            cell_v.font = value_font
        cell_k.border = thin_border; cell_v.border = thin_border

    from matplotlib import pyplot as plt
    from openpyxl.drawing.image import Image as XlImage
    chart_row = 3 + len(resumen_data) + 1
    top10 = sorted(clientes, key=lambda c: c['total_ventas'], reverse=True)[:10]
    if top10:
        labels = [f"{c.get('nombre','')[:20]} ({c.get('cliente','')})" for c in reversed(top10)]
        values = [float(c['total_ventas']) for c in reversed(top10)]
        fig, ax = plt.subplots(figsize=(7, 3.5))
        ax.barh(range(len(labels)), values, color="#2563eb", zorder=3)
        ax.set_yticks(range(len(labels))); ax.set_yticklabels(labels, fontsize=8)
        ax.set_title("Top 10 clientes por monto", fontsize=12, fontweight="bold")
        ax.set_xlabel("Monto ($)"); ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
        fig.subplots_adjust(left=0.25, right=0.95, top=0.88, bottom=0.12)
        chart_buf = io.BytesIO()
        fig.savefig(chart_buf, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig); chart_buf.seek(0)
        ws.add_image(XlImage(chart_buf), f"A{chart_row}")

    cli_header_row = chart_row + 20
    ws.merge_cells(start_row=cli_header_row, start_column=1, end_row=cli_header_row, end_column=4)
    ws.cell(row=cli_header_row, column=1, value="Clientes principales").font = Font(bold=True, size=12)

    cols_c = ["Cliente", "Empresa", "Nombre", "Total Ventas"]
    for ci, col_name in enumerate(cols_c, 1):
        cell = ws.cell(row=cli_header_row + 1, column=ci, value=col_name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
    for ri, c in enumerate(clientes):
        r = cli_header_row + 2 + ri
        ws.cell(row=r, column=1, value=c.get('cliente', '')).font = value_font; ws.cell(row=r, column=1).border = thin_border
        ws.cell(row=r, column=2, value=c.get('empresa', '')).font = value_font; ws.cell(row=r, column=2).border = thin_border
        ws.cell(row=r, column=3, value=c.get('nombre', '')).font = value_font; ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=4, value=c['total_ventas']).font = value_font; ws.cell(row=r, column=4).number_format = '#,##0.00'; ws.cell(row=r, column=4).border = thin_border
    _auto_width(ws)

    # --- Sheet 2: Facturas ---
    ws2 = wb.create_sheet("Facturas")
    cols_f = ["Folio", "Fecha", "Cliente", "Total"]
    for ci, col_name in enumerate(cols_f, 1):
        cell = ws2.cell(row=1, column=ci, value=col_name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = Alignment(horizontal="center"); cell.border = thin_border
    for ri, f in enumerate(facturas):
        r = 2 + ri
        ws2.cell(row=r, column=1, value=f.get('folio', '')).font = value_font; ws2.cell(row=r, column=1).border = thin_border
        ws2.cell(row=r, column=2, value=str(f.get('fecha', '') or '')).font = value_font; ws2.cell(row=r, column=2).border = thin_border
        ws2.cell(row=r, column=3, value=f.get('cliente_nombre', '')).font = value_font; ws2.cell(row=r, column=3).border = thin_border
        ws2.cell(row=r, column=4, value=f['total']).font = value_font; ws2.cell(row=r, column=4).number_format = '#,##0.00'; ws2.cell(row=r, column=4).border = thin_border
    _auto_width(ws2)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Reporte_Producto_{producto[:30]}.xlsx"'},
    )


@router.get("/reporte-producto/{producto}/export-pdf")
def reporte_producto_export_pdf(
    producto: str,
    empresa: Optional[str] = None,
    desde: Optional[str] = None,
    hasta: Optional[str] = None,
    user=Depends(require_user),
):
    resumen, clientes, facturas = _fetch_reporte_producto(producto, empresa, desde, hasta)
    if resumen is None:
        raise HTTPException(404, "No hay datos para exportar.")

    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas as pdfcanvas
    from reportlab.platypus import Table as RlTable, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader

    buf = io.BytesIO()
    c = pdfcanvas.Canvas(buf, pagesize=letter)
    width, height = letter

    def encabezado():
        c.setFillColorRGB(37/255, 99/255, 235/255)
        c.rect(0, height - 40, width, 40, fill=1, stroke=0)
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 16)
        c.drawString(40, height - 28, "Reporte de Producto")
        c.setFont("Helvetica", 11)
        c.drawRightString(width - 40, height - 28, f"{producto}")

    def draw_table(data, col_widths, y_pos):
        t = RlTable(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563eb")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f0f4f8")]),
        ]))
        w, h = t.wrapOn(c, width - 80, 400)
        t.drawOn(c, 40, y_pos - h)
        return y_pos - h - 25

    # --- Page 1: KPIs + Chart ---
    encabezado()
    y = height - 80

    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(0, 0, 0)
    c.drawString(40, y, "KPIs"); y -= 8
    card_w, card_h, gap = 125, 55, 8
    x_start = 40
    tp = f"${resumen['total'] / resumen['facturas']:,.2f}" if resumen['facturas'] else "$0.00"
    kpi_data = [
        ("Ventas Totales", f"${resumen['total']:,.2f}", (32/255, 84/255, 147/255)),
        ("Facturas", str(resumen['facturas']), (22/255, 163/255, 74/255)),
        ("Piezas Vendidas", str(int(resumen['piezas_vendidas'])), (8/255, 145/255, 178/255)),
        ("Ticket Promedio", tp, (245/255, 166/255, 35/255)),
    ]
    for i, (label, value, rgb) in enumerate(kpi_data):
        cx = x_start + i * (card_w + gap)
        c.setFillColorRGB(1, 1, 1)
        c.setStrokeColorRGB(0.85, 0.85, 0.85)
        c.rect(cx, y - card_h, card_w, card_h, fill=1, stroke=1)
        c.setFillColorRGB(*rgb)
        c.rect(cx, y - card_h, 4, card_h, fill=1, stroke=0)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.setFont("Helvetica", 7)
        c.drawString(cx + 10, y - 14, label)
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica-Bold", 13)
        c.drawString(cx + 10, y - 34, value)
    y = y - card_h - 16

    if clientes:
        import matplotlib
        matplotlib.use("Agg")
        from matplotlib import pyplot as plt

        top = sorted(clientes, key=lambda cl: cl['total_ventas'], reverse=True)[:10]
        labels = [f"{cl.get('nombre','')[:20]} ({cl.get('cliente','')})" for cl in reversed(top)]
        values = [cl['total_ventas'] for cl in reversed(top)]

        fig, ax = plt.subplots(figsize=(7.5, 3.5))
        ax.barh(range(len(labels)), values, color="#2563eb", zorder=3)
        ax.set_yticks(range(len(labels)))
        ax.set_yticklabels(labels, fontsize=7)
        ax.set_title("Top 10 clientes por monto", fontsize=12, fontweight="bold", pad=12)
        ax.set_xlabel("Monto ($)")
        ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
        fig.subplots_adjust(left=0.01, right=0.99, top=0.88, bottom=0.18)
        chart_buf = io.BytesIO()
        fig.savefig(chart_buf, format="png", dpi=120, bbox_inches="tight")
        plt.close(fig)
        chart_buf.seek(0)

        img = ImageReader(chart_buf)
        c.drawImage(img, 40, y - 170, width=width - 80, height=170, preserveAspectRatio=True)

    # --- Page 2: Clientes table + Facturas table ---
    c.showPage()
    encabezado()
    y = height - 80

    if clientes:
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, "Clientes"); y -= 20
        c_data = [["Cliente", "Empresa", "Nombre", "Total Ventas"]]
        for cl in clientes:
            c_data.append([str(cl.get('cliente', '')), str(cl.get('empresa', '')), str(cl.get('nombre', '')), f"${cl['total_ventas']:,.2f}"])
        y = draw_table(c_data, [55, 90, 220, 100], y)
        if y < 150: c.showPage(); encabezado(); y = height - 80

    if facturas:
        if y < 200: c.showPage(); encabezado(); y = height - 80
        c.setFont("Helvetica-Bold", 12)
        c.drawString(40, y, "Facturas"); y -= 20
        f_data = [["Folio", "Fecha", "Cliente", "Total"]]
        for f in facturas:
            f_data.append([str(f.get('folio', '')), str(f.get('fecha', '') or ''), str(f.get('cliente_nombre', '')), f"${f['total']:,.2f}"])
        y = draw_table(f_data, [80, 80, 220, 100], y)

    c.save()
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="Reporte_Producto_{producto[:30]}.pdf"'},
    )
