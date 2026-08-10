from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
import io
from datetime import date, datetime

from app.dependencies import require_user
from app.legacy_db import get_legacy_connection

router = APIRouter(prefix="/api/cadenas", tags=["cadenas"])


def _dict_rows(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _norm(s):
    return (s or "").strip()


def _cmp_empresa(a_sql: str, b_sql: str) -> str:
    return (
        f"UPPER(TRIM(REPLACE(REPLACE(REPLACE("
        f"CAST({a_sql} AS CHAR),' ','_'),'-','_'),'Ñ','N')) = "
        f"UPPER(TRIM(REPLACE(REPLACE(REPLACE("
        f"CAST({b_sql} AS CHAR),' ','_'),'-','_'),'Ñ','N'))"
    )


def _cmp(a_sql: str, b_sql: str) -> str:
    return f"UPPER(TRIM(CAST({a_sql} AS CHAR))) = UPPER(TRIM(CAST({b_sql} AS CHAR)))"


def _fmt_date(d):
    if isinstance(d, (date, datetime)):
        return d.strftime("%d/%m/%Y")
    if isinstance(d, str) and d:
        try:
            return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
        except Exception:
            return d
    return str(d) if d else ""


def _auto_width(ws, factor=1.1, min_width=8):
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode("utf-8")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len * factor + 2, 60)


def _asegurar_columna_cobranza(conn):
    try:
        cursor = conn.cursor()
        cursor.execute(
            "ALTER TABLE cadenas ADD COLUMN activa_cobranza TINYINT(1) NOT NULL DEFAULT 0"
        )
        conn.commit()
        cursor.close()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass


@router.get("")
def listar_cadenas(
    solo_cobranza: int | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        _asegurar_columna_cobranza(conn)
        cursor = conn.cursor(dictionary=True)
        sql = """
            SELECT id, nombre, descripcion, fecha_creacion,
                   COALESCE(activa_cobranza, 0) AS activa_cobranza
            FROM cadenas
        """
        if solo_cobranza:
            sql += " WHERE COALESCE(activa_cobranza, 0) = 1"
        sql += " ORDER BY COALESCE(activa_cobranza, 0) DESC, nombre"
        cursor.execute(sql)
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(500, f"listar_cadenas: {e}")
    finally:
        conn.close()


@router.post("")
def crear_cadena(
    nombre: str,
    descripcion: str = "",
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        _asegurar_columna_cobranza(conn)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO cadenas (nombre, descripcion, activa_cobranza) VALUES (%s, %s, 0)",
            (_norm(nombre), _norm(descripcion)),
        )
        conn.commit()
        return {"mensaje": "Cadena creada"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, str(e))
    finally:
        conn.close()


@router.put("/{cadena_id}")
def actualizar_cadena(
    cadena_id: int,
    nombre: str | None = None,
    descripcion: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor()
        campos = []
        params = []
        if nombre is not None:
            campos.append("nombre = %s")
            params.append(_norm(nombre))
        if descripcion is not None:
            campos.append("descripcion = %s")
            params.append(_norm(descripcion))
        if not campos:
            raise HTTPException(400, "Sin campos para actualizar")
        params.append(cadena_id)
        cursor.execute(
            f"UPDATE cadenas SET {', '.join(campos)} WHERE id = %s", params
        )
        if cursor.rowcount == 0:
            raise HTTPException(404, "Cadena no encontrada")
        conn.commit()
        return {"mensaje": "Cadena actualizada"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"actualizar_cadena: {e}")
    finally:
        conn.close()


@router.delete("/{cadena_id}")
def eliminar_cadena(cadena_id: int, user=Depends(require_user)):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM cadenas_clientes WHERE cadena_id = %s", (cadena_id,))
        cursor.execute("DELETE FROM cadenas WHERE id = %s", (cadena_id,))
        if cursor.rowcount == 0:
            raise HTTPException(404, "Cadena no encontrada")
        conn.commit()
        return {"mensaje": "Cadena eliminada"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"eliminar_cadena: {e}")
    finally:
        conn.close()


@router.put("/{cadena_id}/cobranza")
def actualizar_cadena_cobranza(
    cadena_id: int,
    activa: int,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        _asegurar_columna_cobranza(conn)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE cadenas SET activa_cobranza = %s WHERE id = %s",
            (1 if activa else 0, cadena_id),
        )
        if cursor.rowcount == 0:
            raise HTTPException(404, "Cadena no encontrada")
        conn.commit()
        return {"mensaje": "Configuracion de cobranza actualizada"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"actualizar_cadena_cobranza: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/clientes")
def obtener_clientes_cadena(cadena_id: int, user=Depends(require_user)):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        fn = "REPLACE(cc.cliente_numero, '.0', '')"
        cursor.execute(f"""
            SELECT DISTINCT
                cc.cliente_numero, cc.empresa,
                COALESCE(c.nombre, '') AS nombre
            FROM cadenas_clientes cc
            LEFT JOIN clientes c
              ON CAST(c.numero AS CHAR) = {fn}
             AND UPPER(TRIM(c.empresa)) = UPPER(TRIM(cc.empresa))
            WHERE cc.cadena_id = %s
            ORDER BY COALESCE(c.nombre, '')
        """, (cadena_id,))
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(500, f"obtener_clientes_cadena: {e}")
    finally:
        conn.close()


@router.post("/{cadena_id}/clientes")
def agregar_cliente_cadena(
    cadena_id: int,
    cliente_numero: str,
    empresa: str,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM cadenas WHERE id = %s", (cadena_id,))
        if not cursor.fetchone():
            raise HTTPException(404, "Cadena no encontrada")
        cursor.execute(
            "INSERT IGNORE INTO cadenas_clientes (cadena_id, cliente_numero, empresa) VALUES (%s, %s, %s)",
            (cadena_id, _norm(cliente_numero), _norm(empresa)),
        )
        conn.commit()
        return {"mensaje": "Cliente agregado a la cadena"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"agregar_cliente_cadena: {e}")
    finally:
        conn.close()


@router.post("/{cadena_id}/clientes/bulk")
def agregar_clientes_cadena_bulk(
    cadena_id: int,
    payload: dict,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM cadenas WHERE id = %s", (cadena_id,))
        if not cursor.fetchone():
            raise HTTPException(404, "Cadena no encontrada")
        clientes = payload.get("clientes", [])
        if not clientes or not isinstance(clientes, list):
            raise HTTPException(400, "Se requiere una lista 'clientes' no vacia")
        total = 0
        insertados = 0
        for c in clientes:
            num = _norm(c.get("cliente_numero") or c.get("numero") or "")
            emp = _norm(c.get("empresa") or "")
            if num and emp:
                total += 1
                cursor.execute(
                    "INSERT IGNORE INTO cadenas_clientes (cadena_id, cliente_numero, empresa) VALUES (%s, %s, %s)",
                    (cadena_id, num, emp),
                )
                if cursor.rowcount > 0:
                    insertados += 1
        conn.commit()
        omitidos = total - insertados
        return {"mensaje": f"{insertados} agregados" + (f", {omitidos} omitidos (ya existian)" if omitidos else ""), "insertados": insertados, "omitidos": omitidos}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"agregar_clientes_cadena_bulk: {e}")
    finally:
        conn.close()


@router.delete("/{cadena_id}/clientes/{cliente}/{empresa}")
def quitar_cliente_cadena(
    cadena_id: int,
    cliente: str,
    empresa: str,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor()
        cursor.execute(f"""
            DELETE FROM cadenas_clientes
            WHERE cadena_id = %s
              AND cliente_numero = %s
              AND {_cmp("empresa", "%s")}
        """, (cadena_id, _norm(cliente), _norm(empresa)))
        if cursor.rowcount == 0:
            raise HTTPException(404, "Cliente no estaba en la cadena")
        conn.commit()
        return {"mensaje": "Cliente eliminado de la cadena"}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"quitar_cliente_cadena: {e}")
    finally:
        conn.close()


@router.get("/ventas")
def ventas_por_cadena(
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        params = []
        filtros = ""
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])
        cursor.execute(f"""
            SELECT
                c.id AS cadena_id, c.nombre AS cadena,
                COUNT(DISTINCT CONCAT(f.numero_cliente, '|', f.empresa)) AS clientes,
                COUNT(DISTINCT f.id) AS facturas,
                IFNULL(SUM(f.total), 0) AS total_ventas
            FROM cadenas c
            LEFT JOIN cadenas_clientes cc ON cc.cadena_id = c.id
            LEFT JOIN facturas f
              ON f.numero_cliente = cc.cliente_numero
             AND {_cmp("f.empresa", "cc.empresa")}
             AND f.estatus = 'Activa'
            {filtros}
            GROUP BY c.id, c.nombre
            ORDER BY total_ventas DESC
        """, params)
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(500, f"ventas_por_cadena: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/ventas_clientes")
def ventas_por_cliente_cadena(
    cadena_id: int,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        params = [cadena_id]
        filtros = ""
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])
        cursor.execute(f"""
            SELECT cc.cliente_numero, cc.empresa, cl.nombre AS cliente,
                   IFNULL(SUM(f.total), 0) AS total_ventas
            FROM cadenas_clientes cc
            LEFT JOIN facturas f
              ON f.numero_cliente = cc.cliente_numero
             AND {_cmp("f.empresa", "cc.empresa")}
             AND f.estatus = 'Activa'
            LEFT JOIN clientes cl
              ON cl.numero = cc.cliente_numero
             AND {_cmp("cl.empresa", "cc.empresa")}
            WHERE cc.cadena_id = %s
            {filtros}
            GROUP BY cc.cliente_numero, cc.empresa, cl.nombre
            HAVING total_ventas > 0
            ORDER BY total_ventas DESC
        """, params)
        rows = cursor.fetchall()

        if not rows:
            return rows

        # Fetch product summary per client
        sub_where = ""
        sub_params = [cadena_id]
        if fecha_inicio and fecha_fin:
            sub_where = " AND DATE(f2.fecha) BETWEEN %s AND %s "
            sub_params.extend([fecha_inicio, fecha_fin])

        # We reorder the results in a dict keyed by (cliente_numero, empresa)
        cursor.execute(f"""
            SELECT f2.numero_cliente, f2.empresa, fd.descripcion, SUM(fd.piezas) AS piezas
            FROM facturas f2
            JOIN factura_detalle fd ON fd.factura_id = f2.id
            WHERE (f2.numero_cliente, f2.empresa) IN (
                SELECT cc2.cliente_numero, cc2.empresa
                FROM cadenas_clientes cc2
                WHERE cc2.cadena_id = %s
            )
              AND f2.estatus = 'Activa'
              {sub_where}
            GROUP BY f2.numero_cliente, f2.empresa, fd.descripcion
            ORDER BY f2.numero_cliente, f2.empresa, piezas DESC
        """, sub_params)

        prod_rows = cursor.fetchall()
        prod_map = {}
        for pr in prod_rows:
            key = (pr["numero_cliente"], pr["empresa"])
            if key not in prod_map:
                prod_map[key] = []
            prod_map[key].append(f"{pr['descripcion']} ({pr['piezas']})")

        # Merge product summary into each row
        for r in rows:
            key = (r["cliente_numero"], r["empresa"])
            r["productos"] = ", ".join(prod_map.get(key, []))

        return rows
    except Exception as e:
        raise HTTPException(500, f"ventas_por_cliente_cadena: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/ventas_productos")
def ventas_por_producto_cadena(
    cadena_id: int,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        filtros = ""
        params = []
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])
        params.append(cadena_id)
        cursor.execute(f"""
            SELECT fd.descripcion AS producto,
                   SUM(fd.piezas) AS piezas,
                   SUM(((fd.piezas * fd.precio) / NULLIF(t.subtotal_factura, 0)) * f.total) AS total_ventas
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT factura_id, SUM(piezas * precio) AS subtotal_factura
                FROM factura_detalle GROUP BY factura_id
            ) t ON t.factura_id = f.id
            WHERE f.estatus = 'Activa'
            {filtros}
            AND IFNULL(t.subtotal_factura, 0) > 0
            AND EXISTS (
                SELECT 1 FROM cadenas_clientes cc
                WHERE cc.cliente_numero = f.numero_cliente
                  AND {_cmp("cc.empresa", "f.empresa")}
                  AND cc.cadena_id = %s
            )
            GROUP BY fd.descripcion
            HAVING piezas > 0
            ORDER BY total_ventas DESC
        """, params)
        rows = cursor.fetchall()
        for r in rows:
            if r.get("total_ventas") is None:
                r["total_ventas"] = 0
        return rows
    except Exception as e:
        raise HTTPException(500, f"ventas_por_producto_cadena: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/producto/{producto}/clientes")
def clientes_por_producto_cadena(
    cadena_id: int,
    producto: str,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        filtros = ""
        params = [cadena_id, producto]
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])
        cursor.execute(f"""
            SELECT cl.nombre AS cliente, f.numero_cliente, f.empresa,
                   SUM(fd.piezas) AS piezas,
                   SUM(((fd.piezas * fd.precio) / NULLIF(t.subtotal_factura, 0)) * f.total) AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT factura_id, SUM(piezas * precio) AS subtotal_factura
                FROM factura_detalle GROUP BY factura_id
            ) t ON t.factura_id = f.id
            JOIN cadenas_clientes cc
              ON cc.cliente_numero = f.numero_cliente
             AND cc.cadena_id = %s
             AND {_cmp("cc.empresa", "f.empresa")}
            LEFT JOIN clientes cl
              ON cl.numero = f.numero_cliente
             AND {_cmp("cl.empresa", "f.empresa")}
            WHERE fd.descripcion = %s
              AND f.estatus = 'Activa'
              AND IFNULL(t.subtotal_factura, 0) > 0
              {filtros}
            GROUP BY cl.nombre, f.numero_cliente, f.empresa
            HAVING piezas > 0
            ORDER BY piezas DESC
        """, params)
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(500, f"clientes_por_producto_cadena: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/producto/{producto}/facturas")
def facturas_por_producto_cadena(
    cadena_id: int,
    producto: str,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        filtros = ""
        params = [cadena_id, producto]
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])
        cursor.execute(f"""
            SELECT f.id, f.factura, f.fecha, f.numero_cliente, cl.nombre AS cliente, f.empresa,
                   fd.piezas,
                   ((fd.piezas * fd.precio) / NULLIF(t.subtotal_factura, 0)) * f.total AS total
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (
                SELECT factura_id, SUM(piezas * precio) AS subtotal_factura
                FROM factura_detalle GROUP BY factura_id
            ) t ON t.factura_id = f.id
            JOIN cadenas_clientes cc
              ON cc.cliente_numero = f.numero_cliente
             AND cc.cadena_id = %s
             AND {_cmp("cc.empresa", "f.empresa")}
            LEFT JOIN clientes cl
              ON cl.numero = f.numero_cliente
             AND {_cmp("cl.empresa", "f.empresa")}
            WHERE fd.descripcion = %s
              AND f.estatus = 'Activa'
              AND IFNULL(t.subtotal_factura, 0) > 0
              {filtros}
            ORDER BY f.fecha DESC
        """, params)
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(500, f"facturas_por_producto_cadena: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/facturas")
def facturas_cadena(
    cadena_id: int,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        filtros = ""
        params = []
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])
        params.append(cadena_id)
        cursor.execute(f"""
            SELECT f.factura, f.fecha, f.numero_cliente,
                   cl.nombre AS cliente, f.empresa, f.total
            FROM facturas f
            JOIN clientes cl
              ON cl.numero = f.numero_cliente
             AND {_cmp("cl.empresa", "f.empresa")}
            WHERE f.estatus = 'Activa'
              {filtros}
              AND EXISTS (
                  SELECT 1 FROM cadenas_clientes cc
                  WHERE cc.cadena_id = %s
                    AND cc.cliente_numero = f.numero_cliente
                    AND {_cmp("cc.empresa", "f.empresa")}
              )
            ORDER BY f.fecha
        """, params)
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(500, f"facturas_cadena: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/cliente/{cliente_numero}/{empresa}/facturas")
def facturas_cliente_cadena(
    cadena_id: int,
    cliente_numero: str,
    empresa: str,
    producto: str | None = None,
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        params = [cadena_id, _norm(cliente_numero), _norm(empresa)]
        filtros = ""
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])
        prod_select = "f.total, NULL AS piezas"
        prod_join = ""
        prod_where = ""
        if producto:
            prod_select = "((fd.piezas * fd.precio) / NULLIF(t.subtotal_factura, 0)) * f.total AS total, fd.piezas"
            prod_join = """ JOIN factura_detalle fd ON fd.factura_id = f.id
                            JOIN (SELECT factura_id, SUM(piezas * precio) AS subtotal_factura FROM factura_detalle GROUP BY factura_id) t ON t.factura_id = f.id """
            prod_where = " AND fd.descripcion = %s AND IFNULL(t.subtotal_factura, 0) > 0 "
            params.append(producto)
        cursor.execute(f"""
            SELECT f.id, f.factura, f.fecha, {prod_select}
            FROM cadenas_clientes cc
            JOIN facturas f
              ON f.numero_cliente = cc.cliente_numero
             AND {_cmp("f.empresa", "cc.empresa")}
             AND f.estatus = 'Activa'
            {prod_join}
            WHERE cc.cadena_id = %s
              AND cc.cliente_numero = %s
              AND {_cmp("cc.empresa", "%s")}
            {prod_where}
            {filtros}
            ORDER BY f.fecha DESC
        """, params)
        return cursor.fetchall()
    except Exception as e:
        raise HTTPException(500, f"facturas_cliente_cadena: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/export-excel")
def export_cadena_excel(
    cadena_id: int,
    tipo: str = Query("monto"),
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        params = [cadena_id]
        filtros = ""
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])

        cursor.execute("SELECT nombre FROM cadenas WHERE id = %s", (cadena_id,))
        row_cad = cursor.fetchone()
        cadena_nombre = row_cad["nombre"] if row_cad else f"Cadena {cadena_id}"

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()

        hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        hdr_font = Font(bold=True, color="FFFFFF", size=10)
        label_font = Font(bold=True, size=10)
        value_font = Font(size=10)
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        ws = wb.active
        ws.title = "Reporte"
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Reporte Cadena — {cadena_nombre}"
        ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Fetch summary + detail data for both tipos
        params_det = [cadena_id]
        filtros_det = ""
        if fecha_inicio and fecha_fin:
            filtros_det = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params_det.extend([fecha_inicio, fecha_fin])

        if tipo == "producto":
            cursor.execute(f"""
                SELECT fp.descripcion AS producto, CAST(SUM(fp.piezas) AS UNSIGNED) AS piezas,
                       COUNT(DISTINCT f.id) AS facturas,
                       COALESCE(SUM(fp.piezas * fp.precio), 0) AS total
                FROM cadenas_clientes cc
                JOIN facturas f ON f.numero_cliente = cc.cliente_numero
                  AND {_cmp("f.empresa", "cc.empresa")} AND f.estatus = 'Activa'
                JOIN factura_detalle fp ON fp.factura_id = f.id
                WHERE cc.cadena_id = %s {filtros}
                GROUP BY fp.descripcion ORDER BY total DESC
            """, params)
            rows = cursor.fetchall()
            if not rows:
                raise HTTPException(404, "Sin datos para exportar")

            total_ventas = sum(r["total"] for r in rows)
            total_suma = len(rows)
            total_facturas = sum(r["facturas"] for r in rows)
            prom = total_ventas / total_suma if total_suma else 0

            r = 3
            for k, v in [("Cadena", cadena_nombre), ("Productos", str(total_suma)), ("", ""),
                         ("KPI", "Valor"), ("Ventas Totales", f"${total_ventas:,.2f}"),
                         ("Productos", str(total_suma)), ("Facturas", str(total_facturas)),
                         ("Promedio por Producto", f"${prom:,.2f}")]:
                ck = ws.cell(row=r, column=1, value=k); cv = ws.cell(row=r, column=2, value=v)
                if k == "KPI":
                    ck.font = hdr_font; ck.fill = hdr_fill; cv.font = hdr_font; cv.fill = hdr_fill
                elif k:
                    ck.font = label_font; cv.font = value_font
                ck.border = thin_border; cv.border = thin_border; r += 1
            info_end = r
        else:
            cursor.execute(f"""
                SELECT cc.cliente_numero, cc.empresa, cl.nombre AS cliente,
                       COALESCE(SUM(f.total), 0) AS total_ventas
                FROM cadenas_clientes cc
                LEFT JOIN facturas f ON f.numero_cliente = cc.cliente_numero
                  AND {_cmp("f.empresa", "cc.empresa")} AND f.estatus = 'Activa'
                LEFT JOIN clientes cl ON cl.numero = cc.cliente_numero
                  AND {_cmp("cl.empresa", "cc.empresa")}
                WHERE cc.cadena_id = %s {filtros}
                GROUP BY cc.cliente_numero, cc.empresa, cl.nombre
                HAVING total_ventas > 0 ORDER BY total_ventas DESC
            """, params)
            rows = cursor.fetchall()
            if not rows:
                raise HTTPException(404, "Sin datos para exportar")

            total_ventas = sum(r["total_ventas"] for r in rows)
            total_suma = len(rows)
            prom = total_ventas / total_suma if total_suma else 0

            r = 3
            for k, v in [("Cadena", cadena_nombre), ("Clientes", str(total_suma)), ("", ""),
                         ("KPI", "Valor"), ("Ventas Totales", f"${total_ventas:,.2f}"),
                         ("Clientes", str(total_suma)),
                         ("Promedio por Cliente", f"${prom:,.2f}")]:
                ck = ws.cell(row=r, column=1, value=k); cv = ws.cell(row=r, column=2, value=v)
                if k == "KPI":
                    ck.font = hdr_font; ck.fill = hdr_fill; cv.font = hdr_font; cv.fill = hdr_fill
                elif k:
                    ck.font = label_font; cv.font = value_font
                ck.border = thin_border; cv.border = thin_border; r += 1
            info_end = r

        # Matplotlib chart
        from matplotlib import pyplot as plt
        from openpyxl.drawing.image import Image as XlImage
        top_n = sorted(rows, key=lambda x: x.get("total", x.get("total_ventas", 0)), reverse=True)[:10]
        if top_n:
            if tipo == "producto":
                lbl = [r["producto"][:35] for r in reversed(top_n)]
                val = [float(r["total"]) for r in reversed(top_n)]
                ct = "Top 10 Productos por Monto Vendido"; xl = "Monto ($)"
            else:
                lbl = [r.get("cliente", "")[:35] for r in reversed(top_n)]
                val = [float(r["total_ventas"]) for r in reversed(top_n)]
                ct = "Top 10 Clientes por Ventas"; xl = "Total Ventas ($)"
            fig, ax = plt.subplots(figsize=(7, 3.5))
            ax.barh(range(len(lbl)), val, color="#2563eb", zorder=3)
            ax.set_yticks(range(len(lbl))); ax.set_yticklabels(lbl, fontsize=8)
            ax.set_title(ct, fontsize=12, fontweight="bold"); ax.set_xlabel(xl)
            ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
            fig.subplots_adjust(left=0.25, right=0.95, top=0.88, bottom=0.12)
            cbuf = io.BytesIO(); fig.savefig(cbuf, format="png", dpi=120, bbox_inches="tight")
            plt.close(fig); cbuf.seek(0)
            ws.add_image(XlImage(cbuf), f"A{info_end + 2}")

        # Summary table
        tbl_start = info_end + 22
        ws.merge_cells(start_row=tbl_start, start_column=1, end_row=tbl_start, end_column=4)
        ws.cell(row=tbl_start, column=1, value="Productos más comprados" if tipo == "producto" else "Clientes").font = Font(bold=True, size=12)

        if tipo == "producto":
            cols = ["Producto", "Piezas", "Facturas", "Total"]
            for ci, cn in enumerate(cols, 1):
                c = ws.cell(row=tbl_start + 1, column=ci, value=cn)
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center"); c.border = thin_border
            for ri, p in enumerate(rows):
                rr = tbl_start + 2 + ri
                ws.cell(row=rr, column=1, value=p["producto"]).font = value_font; ws.cell(row=rr, column=1).border = thin_border
                ws.cell(row=rr, column=2, value=int(p["piezas"])).font = value_font; ws.cell(row=rr, column=2).border = thin_border
                ws.cell(row=rr, column=3, value=p["facturas"]).font = value_font; ws.cell(row=rr, column=3).border = thin_border
                c4 = ws.cell(row=rr, column=4, value=float(p["total"])); c4.font = value_font; c4.number_format = '#,##0.00'; c4.border = thin_border
        else:
            cols = ["Cliente", "Número", "Empresa", "Total Ventas"]
            for ci, cn in enumerate(cols, 1):
                c = ws.cell(row=tbl_start + 1, column=ci, value=cn)
                c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center"); c.border = thin_border
            for ri, p in enumerate(rows):
                rr = tbl_start + 2 + ri
                ws.cell(row=rr, column=1, value=p.get("cliente", "")).font = value_font; ws.cell(row=rr, column=1).border = thin_border
                ws.cell(row=rr, column=2, value=p["cliente_numero"]).font = value_font; ws.cell(row=rr, column=2).border = thin_border
                ws.cell(row=rr, column=3, value=p["empresa"]).font = value_font; ws.cell(row=rr, column=3).border = thin_border
                c4 = ws.cell(row=rr, column=4, value=float(p["total_ventas"])); c4.font = value_font; c4.number_format = '#,##0.00'; c4.border = thin_border

        # Detail sheet
        cursor.execute(f"""
            SELECT f.factura, f.fecha, f.numero_cliente, cl.nombre AS cliente, f.empresa,
                   fd.descripcion AS producto, fd.piezas,
                   ((fd.piezas * fd.precio) / NULLIF(t.subtotal_factura, 0)) * f.total AS total_producto
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (SELECT factura_id, SUM(piezas * precio) AS subtotal_factura FROM factura_detalle GROUP BY factura_id) t ON t.factura_id = f.id
            JOIN cadenas_clientes cc ON cc.cliente_numero = f.numero_cliente
              AND cc.cadena_id = %s AND {_cmp("cc.empresa", "f.empresa")}
            LEFT JOIN clientes cl ON cl.numero = f.numero_cliente AND {_cmp("cl.empresa", "f.empresa")}
            WHERE f.estatus = 'Activa' AND IFNULL(t.subtotal_factura, 0) > 0
              {filtros_det}
            ORDER BY f.fecha DESC
        """, params_det)
        det = cursor.fetchall()
        ws2 = wb.create_sheet("Facturas")
        cols_f = ["Factura", "Fecha", "Cliente #", "Cliente", "Empresa", "Producto", "Piezas", "Total Producto"]
        for ci, cn in enumerate(cols_f, 1):
            c = ws2.cell(row=1, column=ci, value=cn)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = Alignment(horizontal="center"); c.border = thin_border
        for ri, f in enumerate(det):
            rr = 2 + ri
            ws2.cell(row=rr, column=1, value=f["factura"]).font = value_font; ws2.cell(row=rr, column=1).border = thin_border
            ws2.cell(row=rr, column=2, value=_fmt_date(f["fecha"])).font = value_font; ws2.cell(row=rr, column=2).border = thin_border
            ws2.cell(row=rr, column=3, value=f["numero_cliente"]).font = value_font; ws2.cell(row=rr, column=3).border = thin_border
            ws2.cell(row=rr, column=4, value=f.get("cliente", "")).font = value_font; ws2.cell(row=rr, column=4).border = thin_border
            ws2.cell(row=rr, column=5, value=f["empresa"]).font = value_font; ws2.cell(row=rr, column=5).border = thin_border
            ws2.cell(row=rr, column=6, value=f["producto"]).font = value_font; ws2.cell(row=rr, column=6).border = thin_border
            ws2.cell(row=rr, column=7, value=f["piezas"]).font = value_font; ws2.cell(row=rr, column=7).border = thin_border
            c8 = ws2.cell(row=rr, column=8, value=float(f["total_producto"])); c8.font = value_font; c8.number_format = '#,##0.00'; c8.border = thin_border

        _auto_width(ws); _auto_width(ws2)
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 50)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  headers={"Content-Disposition": f'attachment; filename="Cadena_{cadena_id}_{tipo}.xlsx"'})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"export_cadena_excel: {e}")
    finally:
        conn.close()


@router.get("/{cadena_id}/export-pdf")
def export_cadena_pdf(
    cadena_id: int,
    tipo: str = Query("monto"),
    fecha_inicio: str | None = None,
    fecha_fin: str | None = None,
    user=Depends(require_user),
):
    conn = get_legacy_connection()
    try:
        cursor = conn.cursor(dictionary=True)
        params = [cadena_id]
        filtros = ""
        if fecha_inicio and fecha_fin:
            filtros = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params.extend([fecha_inicio, fecha_fin])

        cursor.execute("SELECT nombre FROM cadenas WHERE id = %s", (cadena_id,))
        row_cad = cursor.fetchone()
        cadena_nombre = row_cad["nombre"] if row_cad else f"Cadena {cadena_id}"

        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import Table as RlTable, TableStyle, Paragraph
        from reportlab.pdfgen import canvas as pdfcanvas
        from reportlab.lib.utils import ImageReader
        from reportlab.lib.styles import getSampleStyleSheet
        _pstyles = getSampleStyleSheet()
        _cell_style = _pstyles["Normal"]
        _cell_style.fontSize = 6
        _cell_style.leading = 7.5
        _cell_style.spaceBefore = 0
        _cell_style.spaceAfter = 0

        buf = io.BytesIO()
        c = pdfcanvas.Canvas(buf, pagesize=landscape(letter))
        width, height = landscape(letter)

        def encabezado():
            c.setFillColorRGB(37/255, 99/255, 235/255)
            c.rect(0, height - 40, width, 40, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            c.setFont("Helvetica-Bold", 16)
            c.drawString(40, height - 28, "Reporte de Cadena")
            c.setFont("Helvetica", 11)
            c.drawRightString(width - 40, height - 28, cadena_nombre)

        def _table_style(left_cols=None):
            style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2563eb")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.grey),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.HexColor("#f0f4f8")]),
            ]
            if left_cols:
                for lc in left_cols:
                    style.append(('ALIGN', (lc, 1), (lc, -1), 'LEFT'))
            return style

        def _center_x(table_width):
            return (width - table_width) / 2

        def draw_table(data, col_widths, y_pos, left_cols=None):
            t = RlTable(data, colWidths=col_widths)
            t.setStyle(TableStyle(_table_style(left_cols)))
            w, h = t.wrapOn(c, width - 80, 400)
            avail = y_pos - 40
            total_data_rows = len(data) - 1
            if h > avail and total_data_rows > 1:
                rows_per = max(1, int(total_data_rows * avail / h))
                y = y_pos
                for i in range(0, total_data_rows, rows_per):
                    if i > 0:
                        c.showPage(); encabezado(); y = height - 80
                    chunk = [data[0]] + data[i+1:i+1+rows_per]
                    t2 = RlTable(chunk, colWidths=col_widths)
                    t2.setStyle(TableStyle(_table_style(left_cols)))
                    w2, h2 = t2.wrapOn(c, width - 80, 400)
                    t2.drawOn(c, _center_x(w2), y - h2)
                    y = y - h2 - 25
                return y
            else:
                t.drawOn(c, _center_x(w), y_pos - h)
                return y_pos - h - 25

        # Fetch data
        params_det = [cadena_id]
        filtros_det = ""
        if fecha_inicio and fecha_fin:
            filtros_det = " AND DATE(f.fecha) BETWEEN %s AND %s "
            params_det.extend([fecha_inicio, fecha_fin])

        if tipo == "producto":
            cursor.execute(f"""
                SELECT fp.descripcion AS producto, CAST(SUM(fp.piezas) AS UNSIGNED) AS piezas,
                       COUNT(DISTINCT f.id) AS facturas,
                       COALESCE(SUM(fp.piezas * fp.precio), 0) AS total
                FROM cadenas_clientes cc
                JOIN facturas f ON f.numero_cliente = cc.cliente_numero
                  AND {_cmp("f.empresa", "cc.empresa")} AND f.estatus = 'Activa'
                JOIN factura_detalle fp ON fp.factura_id = f.id
                WHERE cc.cadena_id = %s {filtros}
                GROUP BY fp.descripcion ORDER BY total DESC
            """, params)
            rows = cursor.fetchall()
            if not rows:
                raise HTTPException(404, "Sin datos para exportar")
            total_ventas = sum(r["total"] for r in rows)
            total_sum = len(rows)
            total_facturas = sum(r["facturas"] for r in rows)
            prom = total_ventas / total_sum if total_sum else 0
        else:
            cursor.execute(f"""
                SELECT cc.cliente_numero, cc.empresa, cl.nombre AS cliente,
                       COALESCE(SUM(f.total), 0) AS total_ventas
                FROM cadenas_clientes cc
                LEFT JOIN facturas f ON f.numero_cliente = cc.cliente_numero
                  AND {_cmp("f.empresa", "cc.empresa")} AND f.estatus = 'Activa'
                LEFT JOIN clientes cl ON cl.numero = cc.cliente_numero
                  AND {_cmp("cl.empresa", "cc.empresa")}
                WHERE cc.cadena_id = %s {filtros}
                GROUP BY cc.cliente_numero, cc.empresa, cl.nombre
                HAVING total_ventas > 0 ORDER BY total_ventas DESC
            """, params)
            rows = cursor.fetchall()
            if not rows:
                raise HTTPException(404, "Sin datos para exportar")
            total_ventas = sum(r["total_ventas"] for r in rows)
            total_sum = len(rows)
            prom = total_ventas / total_sum if total_sum else 0

        # --- Page 1: Datos + KPIs + Chart ---
        encabezado()
        y = height - 80

        c.setFont("Helvetica-Bold", 12)
        c.setFillColorRGB(0, 0, 0)
        c.drawCentredString(width / 2, y, "Datos de la Cadena"); y -= 20
        c.setFont("Helvetica", 11)
        for d in [f"Cadena: {cadena_nombre}", f"{'Productos' if tipo == 'producto' else 'Clientes'}: {total_sum}"]:
            c.drawCentredString(width / 2, y, d); y -= 16
        y -= 10

        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(width / 2, y, "KPIs"); y -= 8
        card_w, card_h, gap = 135, 55, 8
        total_kpi = 3 * card_w + 2 * gap
        x_start = (width - total_kpi) / 2
        kpi_data = [
            ("Ventas Totales", f"${total_ventas:,.2f}", (32/255, 84/255, 147/255)),
            ("Facturas", str(total_facturas if tipo == "producto" else total_sum), (22/255, 163/255, 74/255)),
            ("Promedio", f"${prom:,.2f}", (245/255, 166/255, 35/255)),
        ]
        for i, (label, value, rgb) in enumerate(kpi_data):
            cx = x_start + i * (card_w + gap)
            c.setFillColorRGB(1, 1, 1)
            c.setStrokeColorRGB(0.85, 0.85, 0.85)
            c.rect(cx, y - card_h, card_w, card_h, fill=1, stroke=1)
            c.setFillColorRGB(*rgb)
            c.rect(cx, y - card_h, 4, card_h, fill=1, stroke=0)
            c.setFillColorRGB(0.4, 0.4, 0.4)
            c.setFont("Helvetica", 7)
            c.drawString(cx + 10, y - 14, label)
            c.setFillColorRGB(0, 0, 0)
            c.setFont("Helvetica-Bold", 13)
            c.drawString(cx + 10, y - 34, value)
        y = y - card_h - 16

        # Chart
        if rows:
            import matplotlib
            matplotlib.use("Agg")
            from matplotlib import pyplot as plt

            top_n = sorted(rows, key=lambda x: x.get("total", x.get("total_ventas", 0)), reverse=True)[:10]
            if tipo == "producto":
                lbl = [r["producto"][:40] for r in reversed(top_n)]
                val = [float(r["total"]) for r in reversed(top_n)]
                ct = "Top 10 Productos por Monto Vendido"; xl = "Monto ($)"
            else:
                lbl = [r.get("cliente", "")[:40] for r in reversed(top_n)]
                val = [float(r["total_ventas"]) for r in reversed(top_n)]
                ct = "Top 10 Clientes por Ventas"; xl = "Total Ventas ($)"

            fig, ax = plt.subplots(figsize=(7.5, 3.5))
            ax.barh(range(len(lbl)), val, color="#2563eb", zorder=3)
            ax.set_yticks(range(len(lbl))); ax.set_yticklabels(lbl, fontsize=7)
            ax.set_title(ct, fontsize=12, fontweight="bold", pad=12); ax.set_xlabel(xl)
            ax.grid(axis="x", linestyle="--", alpha=0.35, zorder=0)
            fig.subplots_adjust(left=0.01, right=0.99, top=0.88, bottom=0.18)
            cbuf = io.BytesIO(); fig.savefig(cbuf, format="png", dpi=120, bbox_inches="tight")
            plt.close(fig); cbuf.seek(0)
            img = ImageReader(cbuf)
            chart_w = width - 140
            c.drawImage(img, (width - chart_w) / 2, y - 170, width=chart_w, height=170, preserveAspectRatio=True)
            y -= 180

        # --- Page 2+: Tables ---
        c.showPage()
        encabezado()
        y = height - 80

        if tipo == "producto":
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y, "Productos"); y -= 20
            p_data = [["Producto", "Piezas", "Facturas", "Total"]]
            for r in rows:
                p_data.append([r["producto"], str(int(r["piezas"])), str(r["facturas"]), f"${r['total']:,.2f}"])
            y = draw_table(p_data, [4*inch, 0.8*inch, 0.8*inch, 1.4*inch], y, left_cols=[0])
            if y < 150:
                c.showPage(); encabezado(); y = height - 80
        else:
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y, "Clientes"); y -= 20
            p_data = [["Cliente", "Número", "Empresa", "Total Ventas"]]
            for r in rows:
                p_data.append([r.get("cliente", ""), r["cliente_numero"], r["empresa"], f"${r['total_ventas']:,.2f}"])
            y = draw_table(p_data, [3.5*inch, 0.8*inch, 0.8*inch, 1.5*inch], y, left_cols=[0])
            if y < 150:
                c.showPage(); encabezado(); y = height - 80

        # Detail facturas
        cursor.execute(f"""
            SELECT f.factura, f.fecha, f.numero_cliente, cl.nombre AS cliente, f.empresa,
                   fd.descripcion AS producto, fd.piezas,
                   ((fd.piezas * fd.precio) / NULLIF(t.subtotal_factura, 0)) * f.total AS total_producto
            FROM facturas f
            JOIN factura_detalle fd ON fd.factura_id = f.id
            JOIN (SELECT factura_id, SUM(piezas * precio) AS subtotal_factura FROM factura_detalle GROUP BY factura_id) t ON t.factura_id = f.id
            JOIN cadenas_clientes cc ON cc.cliente_numero = f.numero_cliente
              AND cc.cadena_id = %s AND {_cmp("cc.empresa", "f.empresa")}
            LEFT JOIN clientes cl ON cl.numero = f.numero_cliente AND {_cmp("cl.empresa", "f.empresa")}
            WHERE f.estatus = 'Activa' AND IFNULL(t.subtotal_factura, 0) > 0
              {filtros_det}
            ORDER BY f.fecha DESC
        """, params_det)
        det = cursor.fetchall()
        if det:
            if y < 200:
                c.showPage(); encabezado(); y = height - 80
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y, "Facturas"); y -= 20
            f_data = [["Factura", "Fecha", "Cliente #", "Cliente", "Empresa", "Producto", "Piezas", "Total"]]
            _trunc = lambda s, n: s if len(s) <= n else s[:n-3] + "..."
            for f in det:
                f_data.append([
                    f["factura"], _fmt_date(f["fecha"]), f["numero_cliente"],
                    _trunc(f.get("cliente", ""), 28), f["empresa"],
                    _trunc(f["producto"], 45), str(f["piezas"]),
                    f"${f['total_producto']:,.2f}"
                ])
            y = draw_table(f_data, [0.7*inch, 0.65*inch, 0.65*inch, 2.0*inch, 0.6*inch, 3.2*inch, 0.45*inch, 1.0*inch], y, left_cols=[3, 4, 5])

        c.save()
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
                                  headers={"Content-Disposition": f'attachment; filename="Cadena_{cadena_id}_{tipo}.pdf"'})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"export_cadena_pdf: {e}")
    finally:
        conn.close()
