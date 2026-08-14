import io
import re
import asyncio
import threading
import time
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import mysql.connector
import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from pydantic import BaseModel, Field

from app.db import list_companies
from app.dependencies import require_user
from app.legacy_db import get_legacy_connection
from app.routers.timbrado_core import (
    get_timbrado_connection,
    obtener_config_timbrado,
    procesar_siguiente_timbrado,
    _resolver_codigo_barras_timbrado,
    sincronizar_factura_para_timbrado,
)


router = APIRouter(prefix="/api/billing", tags=["billing"])

_IMPORT_JOBS: dict[str, dict] = {}
_IMPORT_JOBS_LOCK = threading.Lock()
_IMPORT_CONTEXT = threading.local()


def _actualizar_progreso_importacion(**changes) -> None:
    job_id = getattr(_IMPORT_CONTEXT, "job_id", "")
    if not job_id:
        return
    with _IMPORT_JOBS_LOCK:
        job = _IMPORT_JOBS.get(job_id)
        if job:
            job.update(changes)
            job["updated_at"] = time.time()


def _detalle_error_importacion(exc: Exception) -> str:
    if isinstance(exc, HTTPException):
        return str(exc.detail)
    return str(exc) or exc.__class__.__name__


def _facturacion_automatica_post_guardado(factura_id: int, empresa: str) -> dict:
    with get_timbrado_connection() as conn_timb:
        cfg = obtener_config_timbrado(conn_timb, empresa)
        if not cfg or not cfg.get("timbrado_activo") or not cfg.get("facturacion_automatica"):
            return {"activa": False}
        conn_legacy = get_legacy_connection()
        try:
            sync = sincronizar_factura_para_timbrado(conn_timb, conn_legacy, factura_id, motivo="FACTURACION_AUTOMATICA")
            if not sync.get("timbrado_requerido"):
                return {"activa": True, "enviada": False, "detalle": sync}
            procesado = procesar_siguiente_timbrado(conn_timb, conn_legacy)
            return {"activa": True, "enviada": True, "procesado": procesado, "detalle": sync}
        finally:
            conn_legacy.close()


class BillingLine(BaseModel):
    cip: str = ""
    descripcion: str = ""
    cantidad: float = 0
    # Algunos artículos se venden por fracción de pieza (por ejemplo 0.5).
    # Debe conservarse como decimal desde la captura hasta el CFDI.
    piezas: float = 0
    precio: float = 0
    precio_base: float = 0
    precio_otro: float = 0
    rebanado: str = ""
    importe: float = 0
    # Se guardan en la partida y no solamente en el encabezado. Esto conserva
    # el descuento original aunque después cambie el catálogo de productos.
    descuento_pct: float = 0
    descuento_importe: float = 0


class BillingCreate(BaseModel):
    factura: str = Field(..., min_length=1)
    fecha: date | None = None
    numero_cliente: str = Field(..., min_length=1)
    cliente_nombre: str = ""
    consignatario: str = ""
    rfc: str = ""
    empresa: str = Field(..., min_length=1)
    numero_salida: str = ""
    comanda: str = ""
    vendedor: str = ""
    lista_precios: str = ""
    descuento_pct: float = 0
    cargo_rebanado_pct: float = 0
    subtotal: float = 0
    descuento: float = 0
    iva: float = 0
    total: float = 0
    sae_codigo: str | None = None
    productos: list[BillingLine] = Field(default_factory=list)


class BillingObservacionesUpdate(BaseModel):
    observaciones: str = ""


def _dict_rows(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _table_columns(cursor, table: str) -> set[str]:
    cursor.execute(f"SHOW COLUMNS FROM {table}")
    columns = set()
    for row in cursor.fetchall():
        if isinstance(row, dict):
            columns.add(str(row.get("Field") or row.get("field") or ""))
        else:
            columns.add(str(row[0]))
    return {name for name in columns if name}


def _ensure_invoice_detail_discount_columns(cursor) -> set[str]:
    """Asegura el histórico de descuento por partida en la base legado."""
    columns = _table_columns(cursor, "factura_detalle")
    definitions = {
        "descuento_pct": "DECIMAL(12,4) NOT NULL DEFAULT 0",
        "descuento_importe": "DECIMAL(14,4) NOT NULL DEFAULT 0",
    }
    for name, definition in definitions.items():
        if name not in columns:
            cursor.execute(f"ALTER TABLE factura_detalle ADD COLUMN {name} {definition}")
            columns.add(name)
    return columns


def _ensure_invoice_observaciones_column(cursor) -> set[str]:
    columns = _table_columns(cursor, "facturas")
    if "observaciones_mio" not in columns:
        cursor.execute("ALTER TABLE facturas ADD COLUMN observaciones_mio TEXT NULL")
        columns.add("observaciones_mio")
    return columns


def _insert_dynamic(cursor, table: str, columns: set[str], payload: dict):
    insert_data = {key: value for key, value in payload.items() if key in columns}
    if not insert_data:
        raise ValueError(f"No hay columnas compatibles para insertar en {table}.")
    names = list(insert_data.keys())
    placeholders = ", ".join(["%s"] * len(names))
    cursor.execute(
        f"INSERT INTO {table} ({', '.join(names)}) VALUES ({placeholders})",
        tuple(insert_data[name] for name in names),
    )


def _update_dynamic(cursor, table: str, columns: set[str], payload: dict, where: str, params: tuple):
    update_data = {key: value for key, value in payload.items() if key in columns}
    if not update_data:
        raise ValueError(f"No hay columnas compatibles para actualizar en {table}.")
    assignments = ", ".join([f"{name}=%s" for name in update_data])
    cursor.execute(
        f"UPDATE {table} SET {assignments} WHERE {where}",
        tuple(update_data.values()) + params,
    )


def _to_float(value) -> float:
    try:
        if isinstance(value, (int, float, Decimal)):
            return float(value)
        text = str(value or "").replace("$", "").replace("%", "").replace(" ", "").strip()
        if not text or text.lower() == "nan":
            return 0.0
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")
        return float(text)
    except Exception:
        return 0.0


def _decimal_excel(value) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    text = str(value).replace("$", "").replace(",", "").strip()
    if not text or text.lower() == "nan":
        return Decimal("0.00")
    try:
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0.00")


def _to_int(value) -> int:
    try:
        return int(round(_to_float(value)))
    except Exception:
        return 0


def _clean_text(value) -> str:
    text = str(value or "").strip()
    return "" if text.lower() == "nan" else text


def _norm_header(value) -> str:
    text = _clean_text(value).lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "n°": "numero",
        "nº": "numero",
        "#": "numero",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    text = text.replace(".", "").replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_import_folio(folio: str) -> str:
    text = str(folio or "").strip().upper()
    if re.match(r"^A0{4,}\d+$", text):
        return re.sub(r"^A0{4,}", "A00", text)
    return text


MONTH_NAMES = {
    1: "ENE",
    2: "FEB",
    3: "MAR",
    4: "ABR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AGO",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DIC",
}


def _parse_mes_date(year_value, day_value, month_value, folio: str) -> datetime:
    month_lookup = {
        "ENE": 1, "JAN": 1,
        "FEB": 2,
        "MAR": 3,
        "ABR": 4, "APR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AGO": 8, "AUG": 8,
        "SEP": 9,
        "OCT": 10,
        "NOV": 11,
        "DIC": 12, "DEC": 12,
    }
    year = _to_int(year_value)
    day = _to_int(day_value)
    month_text = str(month_value or "").strip().upper()
    month = month_lookup.get(month_text) or _to_int(month_value)
    if not year or not day or not month:
        raise ValueError(f"La factura {folio} no tiene una fecha válida en Año/Día/Mes.")
    return datetime(year, month, day)


def _colnum_to_colname(n: int) -> str:
    name = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(65 + r) + name
    return name


def _resolve_company_aliases(raw_company: str) -> list[str]:
    company = str(raw_company or "").strip()
    if not company:
        return []

    aliases = {company}
    for item in list_companies():
        name = str(item.get("name") or "").strip()
        code = str(item.get("code") or "").strip()
        if company.lower() == name.lower() or company.lower() == code.lower():
            if name:
                aliases.add(name)
            if code:
                aliases.add(code)
    return list(aliases)


def _folio_prefix_for_company(company: str) -> str:
    text = str(company or "").strip().lower()
    if "gourmet" in text:
        return "00A"
    if "ibersur" in text:
        return "A00"
    if "eza2007" in text:
        return "EZ"
    if "remision" in text or "remisión" in text:
        return "R00"
    return "000"


def _display_invoice_folio(folio: str) -> str:
    text = _clean_text(folio).upper()
    match = re.match(r"^([A-Z]+)(0+)(\d+)$", text)
    if not match:
        return _clean_text(folio)
    prefix, _zeros, number = match.groups()
    if prefix == "A":
        return f"A00{str(int(number)).zfill(3)}"
    return _clean_text(folio)


def _invoice_order_sql() -> str:
    return """
        ORDER BY
            REGEXP_REPLACE(UPPER(TRIM(COALESCE(f.factura, ''))), '[0-9]', '') DESC,
            CAST(REGEXP_REPLACE(UPPER(TRIM(COALESCE(f.factura, ''))), '[^0-9]', '') AS UNSIGNED) DESC,
            f.fecha DESC,
            f.id DESC
    """


def _company_from_folio(folio: str) -> str:
    text = str(folio or "").strip().upper()
    if text.startswith("A00"):
        return "IBERSUR"
    if text.startswith("R00"):
        return "REMISION"
    # Las series anteriores siguen identificándose para importaciones históricas.
    if text.startswith("FE") or text.startswith("00A"):
        return "GOURMET ESPAÑA"
    if text.startswith("EZ") or text.startswith("CFDI"):
        return "EZA2007"
    return ""


def _increment_folio(last_folio: str, prefix: str) -> str:
    match = re.search(r"(\d+)$", str(last_folio or "").strip())
    next_number = int(match.group(1)) + 1 if match else 1
    # Las nuevas series internas EZ (EZA2007) y FE (Gourmet España)
    # no llevan relleno de ceros: EZ1, FE1, etc.
    if str(prefix or "").upper() == "EZ":
        return f"{prefix}{next_number}"
    min_width = 3 if str(prefix or "").upper() == "A00" else 5
    width = max(min_width, len(str(next_number)))
    return f"{prefix}{str(next_number).zfill(width)}"


def _logo_filenames_for_company(company: str) -> list[str]:
    text = str(company or "").strip().lower()
    if "remision" in text or "remisión" in text:
        return ["Remision.png", "default2.png", "default.png"]
    if "ibersur" in text:
        return ["ibersur.png", "default2.png", "default.png"]
    if "eza2007" in text or "eza 2007" in text:
        return ["eza2007.png", "eza 2007 logo blanco.png", "default.png"]
    if "gourmet" in text:
        return ["gourmet.png", "gourmet(2).png", "default2.png", "default.png"]
    return ["default2.png", "default.png"]


def _logo_bases() -> list[Path]:
    project_dir = Path(__file__).resolve().parents[2]
    return [
        # Carpeta administrable por Soporte (incluye logos subidos en servidor).
        project_dir.parent / "logos",
        project_dir.parent / "AspelAPI" / "logos",
        project_dir / "AspelAPI" / "logos",
        project_dir / "app" / "comandas_legacy" / "logos",
        project_dir / "comandas_legacy" / "logos",
        project_dir / "logos",
    ]


def _logo_path_for_company(company: str, logo_archivo: str = ""):
    filenames = _logo_filenames_for_company(company)
    seleccionado = Path(str(logo_archivo or "").strip()).name
    if seleccionado:
        filenames = [seleccionado, *filenames]
    bases = _logo_bases()
    for filename in filenames:
        for base in bases:
            path = base / filename
            if path.exists():
                return path
    return None


@router.get("/logos")
def list_invoice_logos(user=Depends(require_user)):
    vistos: set[str] = set()
    logos = []
    for base in _logo_bases():
        if not base.is_dir():
            continue
        for path in sorted(base.iterdir(), key=lambda p: p.name.lower()):
            if path.is_file() and path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".gif"} and path.name not in vistos:
                vistos.add(path.name)
                logos.append({"nombre": path.name, "url": f"/api/billing/logo?file={path.name}"})
    return {"items": logos}


def _invoice_where(
    company: str | None,
    q: str | None,
    date_from: str | None,
    date_to: str | None,
    month: int | None = None,
    year: int | None = None,
):
    where = []
    params = []

    if company:
        aliases = _resolve_company_aliases(company)
        normalized_company = (
            str(company or "").strip().upper()
            .replace("Á", "A").replace("É", "E").replace("Í", "I")
            .replace("Ó", "O").replace("Ú", "U").replace("Ü", "U").replace("Ñ", "N")
        )
        is_remisiones = normalized_company in {"REMISION", "REMISIONES"}
        if is_remisiones:
            # La tabla histórica usa ambas variantes, pero el filtro expone una sola empresa.
            aliases = list({*aliases, "REMISION"})
        if aliases:
            placeholders = ", ".join(["UPPER(TRIM(%s) COLLATE utf8mb4_unicode_ci)"] * len(aliases))
            company_sql = "UPPER(TRIM(f.empresa) COLLATE utf8mb4_unicode_ci) IN " f"({placeholders})"
            if is_remisiones:
                # Algunas remisiones antiguas quedaron con empresa EZA2007; el folio R00
                # o puramente numérico es su identificador operativo y las reúne aquí.
                where.append(
                    f"({company_sql} "
                    "OR UPPER(TRIM(COALESCE(f.factura, ''))) LIKE 'R00%' "
                    "OR (UPPER(TRIM(COALESCE(f.empresa, ''))) = 'EZA2007' "
                    "AND TRIM(COALESCE(f.factura, '')) REGEXP '^[0-9]+$'))"
                )
                params.extend(aliases)
            else:
                where.append(company_sql)
                params.extend(aliases)
                if normalized_company == "EZA2007":
                    # En esta empresa los CFDI fiscales son exclusivamente folios CFDI.
                    # R00 y los folios numéricos son remisiones aunque hayan quedado con EZA2007.
                    where.append(
                        "UPPER(TRIM(COALESCE(f.factura, ''))) NOT LIKE 'R00%' "
                        "AND TRIM(COALESCE(f.factura, '')) NOT REGEXP '^[0-9]+$'"
                    )

    query_text = (q or "").strip()
    if query_text:
        where.append(
            "("
            "TRIM(COALESCE(f.factura, '')) LIKE %s "
            "OR TRIM(COALESCE(f.numero_cliente, '')) LIKE %s "
            "OR UPPER(COALESCE(c.nombre, '')) LIKE UPPER(%s) "
            "OR TRIM(COALESCE(f.comanda, '')) LIKE %s "
            "OR TRIM(COALESCE(f.numero_salida, '')) LIKE %s"
            ")"
        )
        like = f"%{query_text}%"
        params.extend([like, like, like, like, like])

    if date_from:
        where.append("DATE(f.fecha) >= %s")
        params.append(date_from)
    if date_to:
        where.append("DATE(f.fecha) <= %s")
        params.append(date_to)

    if month and 1 <= month <= 12:
        where.append("MONTH(f.fecha) = %s")
        params.append(month)

    if year:
        where.append("YEAR(f.fecha) = %s")
        params.append(year)

    return where, params


@router.get("/reservar-folio")
def reserve_invoice_folio(company: str = Query(..., min_length=1), user=Depends(require_user)):
    prefix = _folio_prefix_for_company(company)
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        if prefix.upper() == "A00":
            cursor.execute(
                """
                SELECT factura
                FROM facturas
                WHERE UPPER(TRIM(factura)) REGEXP '^A0+[0-9]+$'
                ORDER BY CAST(REGEXP_REPLACE(UPPER(TRIM(factura)), '[^0-9]', '') AS UNSIGNED) DESC
                LIMIT 1
                """
            )
        else:
            cursor.execute(
                """
                SELECT factura
                FROM facturas
                WHERE factura LIKE %s
                ORDER BY CAST(SUBSTRING(factura, %s) AS UNSIGNED) DESC
                LIMIT 1
                """,
                (f"{prefix}%", len(prefix) + 1),
            )
        row = cursor.fetchone()
        last_folio = row[0] if row else ""
        return {
            "folio": _increment_folio(last_folio, prefix),
            "prefix": prefix,
            "company": company,
            "last_folio": last_folio or "",
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/logo")
def get_invoice_logo(company: str = Query(default=""), file: str = Query(default="")):
    seleccionado = file
    if not seleccionado and company:
        try:
            with get_timbrado_connection() as conn:
                seleccionado = str((obtener_config_timbrado(conn, company) or {}).get("logo_archivo") or "")
        except Exception:
            seleccionado = ""
    path = _logo_path_for_company(company, seleccionado)
    if not path:
        raise HTTPException(status_code=404, detail="Logo no encontrado.")
    return FileResponse(path, media_type="image/png")


@router.get("/logo-debug")
def get_invoice_logo_debug(company: str = Query(default="")):
    """Diagnóstico de despliegue: muestra el archivo real seleccionado para la empresa."""
    seleccionado = ""
    try:
        with get_timbrado_connection() as conn:
            seleccionado = str((obtener_config_timbrado(conn, company) or {}).get("logo_archivo") or "")
    except Exception as exc:
        return {"empresa": company, "error": str(exc), "logo_configurado": "", "ruta_resuelta": ""}
    path = _logo_path_for_company(company, seleccionado)
    return {"empresa": company, "logo_configurado": seleccionado, "ruta_resuelta": str(path or ""), "existe": bool(path and path.exists())}


@router.post("", status_code=status.HTTP_201_CREATED)
def create_invoice(payload: BillingCreate, user=Depends(require_user)):
    factura_numero = payload.factura.strip()
    numero_cliente = payload.numero_cliente.strip()
    empresa = payload.empresa.strip()
    detail = [
        row for row in payload.productos
        if row.cip.strip() or row.descripcion.strip() or row.importe > 0
    ]

    if not detail:
        raise HTTPException(status_code=400, detail="Agrega al menos un producto a la factura.")

    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT id FROM facturas WHERE TRIM(COALESCE(factura, '')) = TRIM(%s) LIMIT 1",
            (factura_numero,),
        )
        if cursor.fetchone():
            raise HTTPException(status_code=409, detail=f"El folio {factura_numero} ya existe.")

        facturas_columns = _table_columns(cursor, "facturas")
        detalle_columns = _ensure_invoice_detail_discount_columns(cursor)
        now_value = datetime.now()
        fecha_value = datetime.combine(payload.fecha, now_value.time()) if payload.fecha else now_value
        consignatario = payload.consignatario.strip() or payload.cliente_nombre.strip()

        header_payload = {
            "fecha": fecha_value,
            "numero_cliente": numero_cliente,
            "consignatario": consignatario,
            "cliente_nombre": payload.cliente_nombre.strip(),
            "factura": factura_numero,
            "subtotal": payload.subtotal,
            "descuento_pct": payload.descuento_pct,
            "descuento": payload.descuento,
            "iva": payload.iva,
            "total": payload.total,
            "sae_codigo": payload.sae_codigo,
            "estatus": "Activa",
            "empresa": empresa,
            "numero_salida": payload.numero_salida.strip(),
            "comanda": payload.comanda.strip() or payload.numero_salida.strip(),
            "rfc": payload.rfc.strip(),
            "vendedor": payload.vendedor.strip(),
            "lista_precios": payload.lista_precios.strip(),
            "cargo_rebanado_pct": payload.cargo_rebanado_pct,
        }

        _insert_dynamic(cursor, "facturas", facturas_columns, header_payload)
        factura_id = cursor.lastrowid
        if not factura_id:
            cursor.execute(
                "SELECT id FROM facturas WHERE TRIM(COALESCE(factura, '')) = TRIM(%s) LIMIT 1",
                (factura_numero,),
            )
            row = cursor.fetchone()
            factura_id = row[0] if row else None
        if not factura_id:
            raise RuntimeError("No se pudo obtener el ID de la factura guardada.")

        for row in detail:
            selected_price = row.precio_otro if row.precio_otro > 0 else row.precio
            detail_payload = {
                "factura_id": factura_id,
                "cip": row.cip.strip(),
                "descripcion": row.descripcion.strip(),
                "cantidad": row.cantidad,
                "piezas": row.piezas,
                "precio": selected_price,
                "precio_base": row.precio_base,
                "precio_real": row.precio,
                "precio_otro": row.precio_otro,
                "rebanado": row.rebanado.strip(),
                "importe": row.importe,
                "descuento_pct": row.descuento_pct,
                "descuento_importe": row.descuento_importe,
            }
            _insert_dynamic(cursor, "factura_detalle", detalle_columns, detail_payload)

        conn.commit()
        auto_timbrado = {"activa": False}
        try:
            auto_timbrado = _facturacion_automatica_post_guardado(factura_id, empresa)
        except Exception as exc:
            auto_timbrado = {"activa": True, "error": str(exc)}
        return {
            "id": factura_id,
            "folio": factura_numero,
            "detail_count": len(detail),
            "total": payload.total,
            "auto_timbrado": auto_timbrado,
        }
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except mysql.connector.errors.IntegrityError as exc:
        if conn:
            conn.rollback()
        if getattr(exc, "errno", None) == 1062:
            raise HTTPException(status_code=409, detail=f"El folio {factura_numero} ya existe.")
        raise HTTPException(status_code=500, detail=str(exc))
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("")
def list_invoices(
    user=Depends(require_user),
    company: str | None = Query(default=None),
    q: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    month: int | None = Query(default=None, ge=0, le=12),
    year: int | None = Query(default=None, ge=2020, le=2100),
    limit: int = Query(default=200, ge=1, le=10000),
):
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        _ensure_invoice_observaciones_column(cursor)
        where, params = _invoice_where(company, q, date_from, date_to, month, year)
        sql = """
            SELECT
                f.id,
                f.fecha,
                DAY(f.fecha) AS dia,
                MONTH(f.fecha) AS mes_num,
                f.empresa,
                f.factura,
                f.numero_cliente,
                COALESCE(NULLIF(TRIM(c.nombre), ''), NULLIF(TRIM(f.cliente_nombre), ''), NULLIF(TRIM(f.consignatario), ''), '') AS cliente_nombre,
                COALESCE(NULLIF(TRIM(c.nombre), ''), NULLIF(TRIM(f.cliente_nombre), ''), NULLIF(TRIM(f.consignatario), ''), '') AS tienda,
                f.comanda,
                f.numero_salida,
                CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.subtotal END AS subtotal,
                CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.descuento END AS descuento,
                CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.iva END AS iva,
                CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.total END AS total,
                CASE
                    WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA')
                    THEN 'CANCELADO'
                    WHEN COALESCE(ce.serie, '') <> '' OR COALESCE(ce.folio_cfdi, '') <> ''
                    THEN CONCAT(COALESCE(ce.serie, ''), COALESCE(ce.folio_cfdi, ''))
                    ELSE COALESCE(f.sae_codigo, '')
                END AS sae_codigo,
                f.estatus,
                f.timbrado_estatus,
                f.cfdi_uuid,
                COALESCE(ce.uuid, '') AS uuid_emitido,
                COALESCE(ce.estatus_cfdi, '') AS estatus_cfdi,
                COALESCE(ce.factura, '') AS factura_cfdi_emitida,
                COALESCE(ce.serie, '') AS serie_cfdi_emitida,
                COALESCE(ce.folio_cfdi, '') AS folio_cfdi_emitido,
                COALESCE(f.observaciones_mio, '') AS observaciones_mio
            FROM facturas f
            LEFT JOIN clientes c
              ON TRIM(CAST(c.numero AS CHAR)) = TRIM(COALESCE(f.numero_cliente, ''))
             AND UPPER(TRIM(c.empresa) COLLATE utf8mb4_unicode_ci) =
                 UPPER(TRIM(f.empresa) COLLATE utf8mb4_unicode_ci)
            LEFT JOIN (
                SELECT ce1.* FROM cfdi_emitidos ce1
                INNER JOIN (
                    SELECT factura_id, MAX(id) AS id FROM cfdi_emitidos
                    WHERE COALESCE(UPPER(TRIM(estatus_cfdi)), '') NOT IN ('CANCELADO', 'CANCELADA')
                    GROUP BY factura_id
                ) ultimo
                  ON ultimo.id = ce1.id
            ) ce ON (
                ce.factura_id = f.id
                OR EXISTS (
                    SELECT 1 FROM cfdi_consolidacion_facturas ccf
                    WHERE ccf.cfdi_emitido_id = ce.id AND ccf.factura_id = f.id
                )
                OR EXISTS (
                    SELECT 1 FROM timbrado_queue tq
                    WHERE tq.factura_id = f.id
                      AND COALESCE(tq.uuid, '') <> ''
                      AND tq.uuid = ce.uuid
                )
            )
        """
        visibles = [
            "UPPER(TRIM(COALESCE(f.factura, ''))) NOT LIKE 'TEST%'",
            "UPPER(TRIM(COALESCE(f.factura, ''))) NOT LIKE 'PRUEBA%'",
        ]
        sql += " WHERE " + " AND ".join(where + visibles)
        sql += _invoice_order_sql() + " LIMIT %s"
        params.append(limit)
        cursor.execute(sql, tuple(params))
        items = _dict_rows(cursor)
        for item in items:
            item["factura_original"] = item.get("factura") or ""
            item["factura"] = _display_invoice_folio(item.get("factura"))

        totals = {
            "count": len(items),
            "subtotal": float(sum(item.get("subtotal") or 0 for item in items)),
            "iva": float(sum(item.get("iva") or 0 for item in items)),
            "total": float(sum(item.get("total") or 0 for item in items)),
        }

        # Compute overall totals (unlimited)
        count_params = params[:-1]  # exclude limit
        count_sql = f"""
            SELECT COUNT(*) AS total_count,
                   COALESCE(SUM(CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.total END), 0) AS total_sum,
                   COALESCE(SUM(CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.subtotal END), 0) AS subtotal_sum,
                   COALESCE(SUM(CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.iva END), 0) AS iva_sum
            FROM facturas f
            LEFT JOIN clientes c
              ON TRIM(CAST(c.numero AS CHAR)) = TRIM(COALESCE(f.numero_cliente, ''))
             AND UPPER(TRIM(c.empresa) COLLATE utf8mb4_unicode_ci) =
                 UPPER(TRIM(f.empresa) COLLATE utf8mb4_unicode_ci)
            LEFT JOIN (
                SELECT ce1.* FROM cfdi_emitidos ce1
                INNER JOIN (
                    SELECT factura_id, MAX(id) AS id FROM cfdi_emitidos
                    WHERE COALESCE(UPPER(TRIM(estatus_cfdi)), '') NOT IN ('CANCELADO', 'CANCELADA')
                    GROUP BY factura_id
                ) ultimo
                  ON ultimo.id = ce1.id
            ) ce ON (
                ce.factura_id = f.id
                OR EXISTS (
                    SELECT 1 FROM cfdi_consolidacion_facturas ccf
                    WHERE ccf.cfdi_emitido_id = ce.id AND ccf.factura_id = f.id
                )
                OR EXISTS (
                    SELECT 1 FROM timbrado_queue tq
                    WHERE tq.factura_id = f.id
                      AND COALESCE(tq.uuid, '') <> ''
                      AND tq.uuid = ce.uuid
                )
            )
        """
        count_sql += " WHERE " + " AND ".join(where + visibles)
        cursor.execute(count_sql, tuple(count_params))
        row = cursor.fetchone()
        overall = {
            "total_count": row[0],
            "total_sum": float(row[1] or 0),
            "subtotal_sum": float(row[2] or 0),
            "iva_sum": float(row[3] or 0),
        }
        return {
            "items": items,
            "count": len(items),
            "totals": totals,
            "overall": overall,
            "filters": {
                "company": company or "",
                "q": q or "",
                "date_from": date_from or "",
                "date_to": date_to or "",
                "month": month or 0,
                "year": year or "",
                "limit": limit,
            },
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/years")
def list_invoice_years(user=Depends(require_user)):
    """Años realmente disponibles en el histórico para el filtro de MIO."""
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT DISTINCT YEAR(fecha) AS anio
            FROM facturas
            WHERE fecha IS NOT NULL
              AND YEAR(fecha) IS NOT NULL
            ORDER BY anio DESC
            """
        )
        return {"years": [int(row[0]) for row in cursor.fetchall() if row and row[0] is not None]}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudieron obtener los años disponibles: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/export")
def export_invoices(
    user=Depends(require_user),
    company: str | None = Query(default=None),
    q: str | None = Query(default=None),
    date_from: str | None = Query(default=None),
    date_to: str | None = Query(default=None),
    month: int | None = Query(default=None, ge=0, le=12),
    year: int | None = Query(default=None, ge=2020, le=2100),
):
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_invoice_observaciones_column(cursor)
        where, params = _invoice_where(company, q, date_from, date_to, month, year)
        sql = """
            SELECT
                f.id,
                f.fecha,
                f.empresa,
                f.factura,
                f.numero_cliente AS cliente,
                COALESCE(NULLIF(TRIM(c.nombre), ''), NULLIF(TRIM(f.cliente_nombre), ''), NULLIF(TRIM(f.consignatario), ''), '') AS tienda,
                f.numero_salida,
                f.comanda,
                f.subtotal,
                f.descuento_pct,
                f.descuento,
                f.iva,
                f.total,
                CASE
                    WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA')
                    THEN 'CANCELADO'
                    ELSE COALESCE(f.sae_codigo, '')
                END AS sae_codigo,
                COALESCE(f.observaciones_mio, '') AS observaciones_mio,
                f.estatus,
                -- Las facturas historicas de Gourmet pueden conservar el codigo
                -- GOURMET_ESPANA mientras que el catalogo usa Gourmet España.
                -- Primero usamos el cliente de la empresa correspondiente y, si
                -- falta su tipo, recuperamos el tipo comercial compartido. El 0
                -- evita celdas vacias en la exportacion de Mío.
                COALESCE(NULLIF(TRIM(c.tipo), ''), NULLIF(TRIM(c_eza.tipo), ''), NULLIF(TRIM(c_ibe.tipo), ''), '0') AS tipo,
                COALESCE(c.agente, '') AS agente,
                COALESCE(c.vendedor, '') AS vendedor
            FROM facturas f
            LEFT JOIN clientes c
              ON TRIM(CAST(c.numero AS CHAR)) = TRIM(COALESCE(f.numero_cliente, ''))
             AND (
                    UPPER(TRIM(c.empresa) COLLATE utf8mb4_unicode_ci) =
                    UPPER(TRIM(f.empresa) COLLATE utf8mb4_unicode_ci)
                    OR (
                        UPPER(TRIM(f.empresa) COLLATE utf8mb4_unicode_ci) LIKE 'GOURMET%'
                        AND UPPER(TRIM(c.empresa) COLLATE utf8mb4_unicode_ci) LIKE 'GOURMET%'
                    )
                 )
            LEFT JOIN clientes c_eza
              ON TRIM(CAST(c_eza.numero AS CHAR)) = TRIM(COALESCE(f.numero_cliente, ''))
             AND UPPER(TRIM(c_eza.empresa) COLLATE utf8mb4_unicode_ci) = 'EZA2007'
            LEFT JOIN clientes c_ibe
              ON TRIM(CAST(c_ibe.numero AS CHAR)) = TRIM(COALESCE(f.numero_cliente, ''))
             AND UPPER(TRIM(c_ibe.empresa) COLLATE utf8mb4_unicode_ci) = 'IBERSUR'
        """
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += _invoice_order_sql() + " LIMIT 5000"
        cursor.execute(sql, tuple(params))
        invoices = cursor.fetchall()
        if not invoices:
            raise HTTPException(status_code=404, detail="No hay facturas para exportar.")

        invoice_ids = [row["id"] for row in invoices]
        placeholders = ",".join(["%s"] * len(invoice_ids))
        cursor.execute(
            f"""
            SELECT factura_id, cip, descripcion AS producto, cantidad, piezas, precio, importe
            FROM factura_detalle
            WHERE factura_id IN ({placeholders})
            ORDER BY factura_id, id
            """,
            tuple(invoice_ids),
        )
        product_rows = cursor.fetchall()
        products_by_invoice = {}
        for row in product_rows:
            products_by_invoice.setdefault(row["factura_id"], []).append(row)

        wb = Workbook()
        ws = wb.active
        ws.title = "Mes"
        headers = ["Año", "Día", "Mes", "Cliente", "Factura", "Descuento", "IEPS", "IVA%", "Total", "Status"]
        for index in range(1, 16):
            headers.append(f"SEP{index}")
            headers.extend([f"SKU{index}", f"DESC{index}", f"C{index}", f"P{index}", f"PR{index}"])
        headers.extend(["SEP16", "Empresa"])
        ws.append(headers)

        for invoice in invoices:
            fecha = invoice.get("fecha") or datetime.now()
            if not isinstance(fecha, datetime):
                fecha = pd.to_datetime(fecha, errors="coerce").to_pydatetime()
            subtotal_dec = _decimal_excel(invoice.get("subtotal"))
            descuento_pct_dec = _decimal_excel(invoice.get("descuento_pct"))
            descuento_dec = _decimal_excel(invoice.get("descuento"))
            iva_dec = _decimal_excel(invoice.get("iva"))
            total_dec = _decimal_excel(invoice.get("total"))
            base_gravable = subtotal_dec - descuento_dec
            iva_pct = ""
            if iva_dec > 0 and base_gravable > 0:
                iva_pct_calc = float(((iva_dec / base_gravable) * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
                if 15.5 <= iva_pct_calc <= 16.5:
                    iva_pct = 16
                elif 7.5 <= iva_pct_calc <= 8.5:
                    iva_pct = 8
                elif iva_pct_calc >= 0.5:
                    iva_pct = iva_pct_calc
            desc = float(descuento_pct_dec) if descuento_pct_dec > 0 else ""
            cancelada = (
                str(invoice.get("estatus") or "").lower() == "cancelada"
                or str(invoice.get("sae_codigo") or "").upper() == "CANCELADO"
            )
            row_data = [
                fecha.year,
                fecha.day,
                MONTH_NAMES.get(fecha.month, ""),
                _clean_text(invoice.get("cliente")),
                _display_invoice_folio(invoice.get("factura")),
                desc,
                "",
                iva_pct,
                float(total_dec),
                "CANCELADA" if cancelada else "ACTIVA",
            ]

            details = products_by_invoice.get(invoice["id"], [])
            pr_export = []
            for detail in details[:15]:
                amount_dec = _decimal_excel(detail.get("importe"))
                if amount_dec <= 0:
                    amount_dec = (_decimal_excel(detail.get("cantidad")) * _decimal_excel(detail.get("precio"))).quantize(
                        Decimal("0.01"), rounding=ROUND_HALF_UP
                    )
                pr_export.append(float(amount_dec))

            if pr_export and descuento_pct_dec > 0 and iva_pct in ("", None):
                factor_desc = Decimal("1.00") - (descuento_pct_dec / Decimal("100.00"))
                if factor_desc <= 0:
                    factor_desc = Decimal("0.0001")
                neto_estimado = (sum(Decimal(str(value)) for value in pr_export) * factor_desc).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                diferencia = (total_dec - neto_estimado).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                if abs(diferencia) >= Decimal("0.01") and abs(diferencia) <= Decimal("1.00"):
                    ajuste = (diferencia / factor_desc).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    pr_export[-1] = float((Decimal(str(pr_export[-1])) + ajuste).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

            for index in range(15):
                if index < len(details):
                    detail = details[index]
                    row_data.extend([
                        "|",
                        _clean_text(detail.get("cip")),
                        _clean_text(detail.get("producto")),
                        float(_decimal_excel(detail.get("cantidad"))),
                        float(_decimal_excel(detail.get("piezas"))),
                        pr_export[index],
                    ])
                else:
                    row_data.extend(["|", "", "", "", "", ""])
            row_data.extend(["|", _clean_text(invoice.get("empresa"))])
            ws.append(row_data)

            excel_row = ws.max_row
            if cancelada:
                ws.cell(row=excel_row, column=10).font = Font(color="FF0000", bold=True)
            for block in range(15):
                sep_col = 11 + (block * 6)
                cantidad_col = 14 + (block * 6)
                piezas_col = 15 + (block * 6)
                precio_col = 16 + (block * 6)
                ws.cell(row=excel_row, column=sep_col).font = Font(color="0000FF", bold=True)
                ws.cell(row=excel_row, column=cantidad_col).number_format = '#,##0.00'
                ws.cell(row=excel_row, column=piezas_col).number_format = '#,##0.00'
                ws.cell(row=excel_row, column=precio_col).number_format = '#,##0.00'
            ws.cell(row=excel_row, column=9).number_format = '#,##0.00'
            ws.cell(row=excel_row, column=101).font = Font(color="0000FF", bold=True)

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 6
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 14
        for col in range(11, 102, 6):
            ws.column_dimensions[_colnum_to_colname(col)].width = 3

        ws_mio = wb.create_sheet("Mio")
        ws_mio.append(["Factura", "Día", "Mes", "Cliente", "Importe", "Tienda", "Tipo", "N° Salida", "Agente", "Vendedor", "SAE", "Observaciones", "Empresa"])
        for invoice in invoices:
            fecha = invoice.get("fecha") or datetime.now()
            if not isinstance(fecha, datetime):
                fecha = pd.to_datetime(fecha, errors="coerce").to_pydatetime()
            ws_mio.append([
                _display_invoice_folio(invoice.get("factura")),
                fecha.day,
                MONTH_NAMES.get(fecha.month, ""),
                _clean_text(invoice.get("cliente")),
                float(_decimal_excel(invoice.get("total"))),
                _clean_text(invoice.get("tienda")),
                _clean_text(invoice.get("tipo")),
                _clean_text(invoice.get("numero_salida")),
                _clean_text(invoice.get("agente")),
                _clean_text(invoice.get("vendedor")),
                "CANCELADO" if cancelada else _clean_text(invoice.get("sae_codigo")),
                _clean_text(invoice.get("observaciones_mio")),
                _clean_text(invoice.get("empresa")),
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=facturas_mio.xlsx"},
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/{invoice_id}")
def get_invoice(invoice_id: int, user=Depends(require_user)):
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                f.*,
                COALESCE(NULLIF(TRIM(c.nombre), ''), NULLIF(TRIM(f.cliente_nombre), ''), NULLIF(TRIM(f.consignatario), ''), '') AS cliente_nombre,
                COALESCE(c.rfc, '') AS cliente_rfc,
                COALESCE(c.correo_electronico, '') AS cliente_correo,
                COALESCE(c.telefono, '') AS cliente_telefono,
                COALESCE(ce.uuid, '') AS uuid_emitido,
                COALESCE(ce.estatus_cfdi, '') AS estatus_cfdi,
                COALESCE(ce.xml_path, '') AS xml_path,
                COALESCE(ce.pdf_path, '') AS pdf_path
            FROM facturas f
            LEFT JOIN clientes c
              ON TRIM(CAST(c.numero AS CHAR)) = TRIM(COALESCE(f.numero_cliente, ''))
             AND UPPER(TRIM(c.empresa) COLLATE utf8mb4_unicode_ci) =
                 UPPER(TRIM(f.empresa) COLLATE utf8mb4_unicode_ci)
            LEFT JOIN (
                SELECT ce1.* FROM cfdi_emitidos ce1
                INNER JOIN (SELECT factura_id, MAX(id) AS id FROM cfdi_emitidos GROUP BY factura_id) ultimo
                  ON ultimo.id = ce1.id
            ) ce ON (
                ce.factura_id = f.id
                OR EXISTS (
                    SELECT 1 FROM cfdi_consolidacion_facturas ccf
                    WHERE ccf.cfdi_emitido_id = ce.id AND ccf.factura_id = f.id
                )
                OR EXISTS (
                    SELECT 1 FROM timbrado_queue tq
                    WHERE tq.factura_id = f.id
                      AND COALESCE(tq.uuid, '') <> ''
                      AND tq.uuid = ce.uuid
                )
            )
            WHERE f.id = %s
            LIMIT 1
            """,
            (invoice_id,),
        )
        rows = _dict_rows(cursor)
        if not rows:
            raise HTTPException(status_code=404, detail="Factura no encontrada.")
        invoice = rows[0]

        cursor.execute(
            """
            SELECT
                d.id,
                d.cip,
                d.descripcion,
                d.cantidad,
                d.piezas,
                d.precio,
                d.precio_base,
                d.precio_real,
                d.precio_otro,
                d.rebanado,
                d.importe,
                COALESCE(d.descuento_pct, 0) AS descuento_pct,
                COALESCE(d.descuento_importe, 0) AS descuento_importe,
                COALESCE(p.unidad, 'PZA') AS unidad,
                COALESCE(p.iva, 'No') AS iva,
                COALESCE(p.codigo_barras, '') AS codigo_barras,
                COALESCE(p.descuento, 'No') AS descuento
            FROM factura_detalle d
            LEFT JOIN productos p ON TRIM(CAST(p.cip AS CHAR)) = TRIM(COALESCE(d.cip, ''))
            WHERE d.factura_id = %s
            ORDER BY d.id
            """,
            (invoice_id,),
        )
        detail = _dict_rows(cursor)
        cur_codigos = conn.cursor(dictionary=True)
        try:
            for row in detail:
                if str(row.get("codigo_barras") or "").strip():
                    continue
                try:
                    row["codigo_barras"] = _resolver_codigo_barras_timbrado(
                        cur_codigos,
                        row.get("cip"),
                        invoice.get("empresa"),
                        invoice.get("lista_precios"),
                        invoice.get("cliente_nombre") or invoice.get("consignatario"),
                    )
                except Exception:
                    row["codigo_barras"] = str(row.get("codigo_barras") or "").strip()
        finally:
            cur_codigos.close()
        return {"item": invoice, "detail": detail}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.put("/{invoice_id}")
def update_invoice(invoice_id: int, payload: BillingCreate, user=Depends(require_user)):
    detail = [
        row for row in payload.productos
        if row.cip.strip() or row.descripcion.strip() or row.importe > 0
    ]
    if not detail:
        raise HTTPException(status_code=400, detail="Agrega al menos un producto a la factura.")

    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id, fecha, comanda, numero_salida FROM facturas WHERE id=%s LIMIT 1", (invoice_id,))
        existing_row = cursor.fetchone()
        if not existing_row:
            raise HTTPException(status_code=404, detail="Factura no encontrada.")

        cursor.execute(
            "SELECT id FROM facturas WHERE TRIM(COALESCE(factura, '')) = TRIM(%s) AND id <> %s LIMIT 1",
            (payload.factura.strip(), invoice_id),
        )
        if cursor.fetchone():
            raise HTTPException(status_code=409, detail=f"El folio {payload.factura.strip()} ya existe.")

        facturas_columns = _table_columns(cursor, "facturas")
        detalle_columns = _ensure_invoice_detail_discount_columns(cursor)
        existing_fecha = existing_row[1] if len(existing_row) > 1 else None
        existing_comanda = str(existing_row[2] or "").strip() if len(existing_row) > 2 else ""
        existing_numero_salida = str(existing_row[3] or "").strip() if len(existing_row) > 3 else ""
        if payload.fecha:
            existing_time = existing_fecha.time() if isinstance(existing_fecha, datetime) else datetime.min.time()
            fecha_value = datetime.combine(payload.fecha, existing_time)
        else:
            fecha_value = existing_fecha or datetime.now()
        consignatario = payload.consignatario.strip() or payload.cliente_nombre.strip()
        # Al editar una factura no se debe romper el vínculo a la comanda si
        # el formulario no envía esos campos (por ejemplo, al editar desde MIO).
        numero_salida = payload.numero_salida.strip() or existing_numero_salida
        comanda = payload.comanda.strip() or payload.numero_salida.strip() or existing_comanda or numero_salida

        header_payload = {
            "fecha": fecha_value,
            "numero_cliente": payload.numero_cliente.strip(),
            "consignatario": consignatario,
            "cliente_nombre": payload.cliente_nombre.strip(),
            "factura": payload.factura.strip(),
            "subtotal": payload.subtotal,
            "descuento_pct": payload.descuento_pct,
            "descuento": payload.descuento,
            "iva": payload.iva,
            "total": payload.total,
            "empresa": payload.empresa.strip(),
            "numero_salida": numero_salida,
            "comanda": comanda,
            "rfc": payload.rfc.strip(),
            "vendedor": payload.vendedor.strip(),
            "lista_precios": payload.lista_precios.strip(),
            "cargo_rebanado_pct": payload.cargo_rebanado_pct,
        }
        _update_dynamic(cursor, "facturas", facturas_columns, header_payload, "id=%s", (invoice_id,))

        cursor.execute("DELETE FROM factura_detalle WHERE factura_id=%s", (invoice_id,))
        for row in detail:
            selected_price = row.precio_otro if row.precio_otro > 0 else row.precio
            detail_payload = {
                "factura_id": invoice_id,
                "cip": row.cip.strip(),
                "descripcion": row.descripcion.strip(),
                "cantidad": row.cantidad,
                "piezas": row.piezas,
                "precio": selected_price,
                "precio_base": row.precio_base,
                "precio_real": row.precio,
                "precio_otro": row.precio_otro,
                "rebanado": row.rebanado.strip(),
                "importe": row.importe,
                "descuento_pct": row.descuento_pct,
                "descuento_importe": row.descuento_importe,
            }
            _insert_dynamic(cursor, "factura_detalle", detalle_columns, detail_payload)

        conn.commit()
        return {"id": invoice_id, "folio": payload.factura.strip(), "detail_count": len(detail), "total": payload.total}
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.put("/{invoice_id}/observaciones")
def update_invoice_observaciones(invoice_id: int, payload: BillingObservacionesUpdate, user=Depends(require_user)):
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        _ensure_invoice_observaciones_column(cursor)
        observaciones = (payload.observaciones or "").strip()
        if len(observaciones) > 2000:
            observaciones = observaciones[:2000]
        cursor.execute(
            "UPDATE facturas SET observaciones_mio=%s WHERE id=%s",
            (observaciones, invoice_id),
        )
        if cursor.rowcount == 0:
            cursor.execute("SELECT id FROM facturas WHERE id=%s LIMIT 1", (invoice_id,))
            if not cursor.fetchone():
                raise HTTPException(status_code=404, detail="Factura no encontrada.")
        conn.commit()
        return {"ok": True, "id": invoice_id, "observaciones_mio": observaciones}
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/{invoice_id}/cancel")
def cancel_invoice(invoice_id: int, datos: dict | None = None, user=Depends(require_user)):
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        limpiar_sae = bool((datos or {}).get("limpiar_sae"))
        sae_codigo = "" if limpiar_sae else "CANCELADO"
        cursor.execute(
            "UPDATE facturas SET estatus=%s, sae_codigo=%s WHERE id=%s",
            ("Cancelada", sae_codigo, invoice_id),
        )
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Factura no encontrada.")
        cursor.execute("SELECT factura, estatus, sae_codigo FROM facturas WHERE id=%s LIMIT 1", (invoice_id,))
        confirmado = cursor.fetchone()
        if not confirmado or str(confirmado[1] or "").strip().upper() != "CANCELADA":
            raise HTTPException(status_code=500, detail="La cancelación interna no pudo confirmarse en la base de datos.")
        conn.commit()
        return {"ok": True, "id": invoice_id, "factura": confirmado[0], "status": confirmado[1], "sae_codigo": confirmado[2]}
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.delete("/{invoice_id}")
def delete_invoice(invoice_id: int, user=Depends(require_user)):
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM facturas WHERE id=%s LIMIT 1", (invoice_id,))
        if not cursor.fetchone():
            raise HTTPException(status_code=404, detail="Factura no encontrada.")
        cursor.execute("SHOW TABLES LIKE 'cfdi_emitidos'")
        if cursor.fetchone():
            # También protege facturas que pertenezcan a un CFDI consolidado.
            cursor.execute(
                """SELECT ce.id
                   FROM cfdi_emitidos ce
                   WHERE (
                        ce.factura_id=%s
                        OR EXISTS (
                            SELECT 1 FROM cfdi_consolidacion_facturas ccf
                            WHERE ccf.cfdi_emitido_id=ce.id AND ccf.factura_id=%s
                        )
                   )
                   AND COALESCE(UPPER(TRIM(ce.estatus_cfdi)), '') NOT IN ('CANCELADO', 'CANCELADA')
                   LIMIT 1""",
                (invoice_id, invoice_id),
            )
            if cursor.fetchone():
                raise HTTPException(status_code=409, detail="No se puede eliminar una factura incluida en un CFDI vigente. Cancele primero el CFDI fiscal.")
        # Si no hay CFDI vigente, elimina los vínculos residuales antes de
        # borrar el folio, sin dejar asociaciones inválidas en Mío.
        cursor.execute("SHOW TABLES LIKE 'cfdi_consolidacion_facturas'")
        if cursor.fetchone():
            cursor.execute("DELETE FROM cfdi_consolidacion_facturas WHERE factura_id=%s", (invoice_id,))
        cursor.execute("SHOW TABLES LIKE 'timbrado_queue'")
        if cursor.fetchone():
            cursor.execute("DELETE FROM timbrado_queue WHERE factura_id=%s", (invoice_id,))
        cursor.execute("SHOW TABLES LIKE 'timbrado_factura_addenda_campos'")
        if cursor.fetchone():
            cursor.execute("DELETE FROM timbrado_factura_addenda_campos WHERE factura_id=%s", (invoice_id,))
        cursor.execute("DELETE FROM factura_detalle WHERE factura_id=%s", (invoice_id,))
        cursor.execute("DELETE FROM facturas WHERE id=%s", (invoice_id,))
        conn.commit()
        return {"id": invoice_id, "deleted": True}
    except HTTPException:
        if conn:
            conn.rollback()
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(exc))
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/actions/import/iniciar")
async def iniciar_importacion_facturas(
    file: UploadFile = File(...),
    empresa_predeterminada: str = Form(default=""),
    sobrescribir_fiscales: bool = Form(default=False),
    user=Depends(require_user),
):
    """Inicia una importación y permite consultar progreso real por factura."""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="El archivo de importación está vacío.")
    job_id = uuid.uuid4().hex
    with _IMPORT_JOBS_LOCK:
        # Conserva solo tareas recientes para no acumular memoria en el servicio.
        cutoff = time.time() - 3600
        for old_id in [key for key, value in _IMPORT_JOBS.items() if value.get("updated_at", 0) < cutoff]:
            _IMPORT_JOBS.pop(old_id, None)
        _IMPORT_JOBS[job_id] = {
            "id": job_id, "estado": "PREPARANDO", "total": 0, "procesadas": 0,
            "importadas": 0, "sobrescritas": 0, "omitidas": 0, "errores": [],
            "created_at": time.time(), "updated_at": time.time(),
        }

    def worker():
        _IMPORT_CONTEXT.job_id = job_id
        _actualizar_progreso_importacion(estado="IMPORTANDO")
        try:
            uploaded = UploadFile(filename=file.filename or "facturas.xlsx", file=io.BytesIO(content))
            result = asyncio.run(import_invoices(uploaded, empresa_predeterminada, sobrescribir_fiscales, user))
            _actualizar_progreso_importacion(
                estado="COMPLETADA",
                total=int(result.get("total") or result.get("imported", 0) + result.get("skipped", 0)),
                procesadas=int(result.get("total") or result.get("imported", 0) + result.get("skipped", 0)),
                importadas=int(result.get("imported") or 0),
                sobrescritas=int(result.get("overwritten") or 0),
                omitidas=int(result.get("skipped") or 0),
                errores=result.get("errors") or [],
                resultado=result,
            )
        except Exception as exc:
            _actualizar_progreso_importacion(estado="ERROR", error=_detalle_error_importacion(exc))
        finally:
            _IMPORT_CONTEXT.job_id = ""

    threading.Thread(target=worker, name=f"billing-import-{job_id[:8]}", daemon=True).start()
    return {"id": job_id, "estado": "PREPARANDO"}


@router.get("/actions/import/{job_id}")
def consultar_progreso_importacion_facturas(job_id: str, user=Depends(require_user)):
    with _IMPORT_JOBS_LOCK:
        job = _IMPORT_JOBS.get(str(job_id or ""))
        if not job:
            raise HTTPException(status_code=404, detail="No se encontró la tarea de importación.")
        return dict(job)


@router.post("/actions/import")
async def import_invoices(
    file: UploadFile = File(...),
    empresa_predeterminada: str = Form(default=""),
    sobrescribir_fiscales: bool = Form(default=False),
    user=Depends(require_user),
):
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {exc}")

    if "Mes" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="El archivo no contiene la hoja 'Mes'.")

    ws = wb["Mes"]
    headers = [cell.value for cell in ws[1]]
    idx = {_norm_header(header): pos for pos, header in enumerate(headers)}
    required = ["ano", "dia", "mes", "cliente", "factura", "descuento", "iva%", "total", "status"]
    missing = [name for name in required if name not in idx]
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan columnas requeridas en 'Mes': {', '.join(missing)}")

    mio_extra = {}
    if "Mio" in wb.sheetnames:
        ws_mio = wb["Mio"]
        headers_mio = [_norm_header(cell.value) for cell in ws_mio[1]]
        idx_mio = {header: pos for pos, header in enumerate(headers_mio)}
        sae_key = next((k for k in idx_mio if k.replace(" ", "") == "sae" or k.startswith("sae") or k.endswith("sae")), None)
        for row in ws_mio.iter_rows(min_row=2, values_only=True):
            folio = ""
            try:
                folio_col = next((k for k in idx_mio if "factur" in k or "folio" in k), None)
                if folio_col:
                    folio = _normalize_import_folio(_clean_text(row[idx_mio[folio_col]]))
            except (IndexError, TypeError):
                folio = ""
            if not folio:
                continue
            _sae_val = ""
            if sae_key:
                try:
                    _sae_val = _clean_text(row[idx_mio[sae_key]])
                except (IndexError, TypeError):
                    _sae_val = ""
            mio_extra[folio] = {
                "tienda": _clean_text(row[idx_mio.get("tienda", -1)] if "tienda" in idx_mio else ""),
                "numero_salida": _clean_text(row[idx_mio.get("numero salida", idx_mio.get("numero_salida", -1))] if ("numero salida" in idx_mio or "numero_salida" in idx_mio) else ""),
                "comanda": _clean_text(row[idx_mio.get("comanda", -1)] if "comanda" in idx_mio else ""),
                "sae_codigo": _sae_val,
                "empresa": _clean_text(row[idx_mio["empresa"]] if "empresa" in idx_mio else ""),
            }

    parsed = []
    seen = set()
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        folio = _normalize_import_folio(_clean_text(row[idx["factura"]]))
        if not folio:
            continue
        if folio in seen:
            raise HTTPException(status_code=400, detail=f"El archivo contiene la factura duplicada {folio}.")
        seen.add(folio)

        numero_cliente = _clean_text(row[idx["cliente"]])
        if not numero_cliente:
            errors.append(f"{folio}: sin cliente.")
            continue

        try:
            fecha_doc = _parse_mes_date(row[idx["ano"]], row[idx["dia"]], row[idx["mes"]], folio)
        except ValueError as exc:
            errors.append(str(exc))
            continue

        productos = []
        subtotal = 0.0
        sum_prices = 0.0
        for index in range(1, 16):
            cip_key = _norm_header(f"SKU{index}")
            desc_key = _norm_header(f"DESC{index}")
            qty_key = _norm_header(f"C{index}")
            pieces_key = _norm_header(f"P{index}")
            price_key = _norm_header(f"PR{index}")
            cip = _clean_text(row[idx[cip_key]] if cip_key in idx else "")
            desc = _clean_text(row[idx[desc_key]] if desc_key in idx else "")
            qty = _to_float(row[idx[qty_key]] if qty_key in idx else 0)
            pieces = _to_int(row[idx[pieces_key]] if pieces_key in idx else 0)
            price = _to_float(row[idx[price_key]] if price_key in idx else 0)
            if price <= 0 and pieces_key in idx:
                price = _to_float(row[idx[pieces_key]])
            if not cip and not desc:
                continue
            productos.append({
                "cip": cip,
                "descripcion": desc,
                "cantidad": qty,
                "piezas": pieces,
                "precio": price,
            })
            subtotal += qty * price
            sum_prices += price

        total = _to_float(row[idx["total"]])
        if productos:
            if abs(sum_prices - total) + 0.01 < abs(subtotal - total):
                subtotal = 0.0
                for product in productos:
                    qty = _to_float(product.get("cantidad"))
                    price = _to_float(product.get("precio"))
                    if qty > 0:
                        price = price / qty
                    product["precio"] = round(price, 6)
                    subtotal += qty * price

        extra = mio_extra.get(folio, {})
        empresa_excel = _clean_text(row[idx["empresa"]] if "empresa" in idx else "")
        empresa = empresa_excel or extra.get("empresa") or str(empresa_predeterminada or "").strip() or _company_from_folio(folio)
        if not empresa:
            errors.append(f"{folio}: sin empresa. Agrega la columna Empresa o selecciona una empresa predeterminada.")
            continue
        status = _clean_text(row[idx["status"]]).upper()
        sae_key = next((k for k in idx if k.replace(" ", "") == "sae" or k.startswith("sae") or k.endswith("sae")), None)
        sae_codigo = extra.get("sae_codigo", "") or (_clean_text(row[idx[sae_key]]) if sae_key else "")
        cancelada = ("CANCEL" in status) or (sae_codigo.strip().upper() == "CANCELADO")
        if cancelada and not sae_codigo:
            sae_codigo = "CANCELADO"
        descuento_pct = _to_float(row[idx["descuento"]])
        descuento = round(subtotal * (descuento_pct / 100), 2)
        iva_pct = _to_float(row[idx["iva%"]])
        iva = round(total - (subtotal - descuento), 2)
        if iva < 0:
            iva = 0.0
        if iva == 0 and iva_pct > 0:
            iva = round((subtotal - descuento) * (iva_pct / 100), 2)

        detail = [
            BillingLine(
                cip=product["cip"],
                descripcion=product["descripcion"],
                cantidad=product["cantidad"],
                piezas=product["piezas"],
                precio=product["precio"],
                precio_base=product["precio"],
                importe=product["cantidad"] * product["precio"],
            )
            for product in productos
        ]
        parsed.append({
            "payload": BillingCreate(
                factura=folio,
                fecha=fecha_doc.date(),
                numero_cliente=numero_cliente,
                cliente_nombre=extra.get("tienda", ""),
                consignatario=extra.get("tienda") or numero_cliente,
                empresa=empresa,
                numero_salida=extra.get("numero_salida", ""),
                comanda=extra.get("comanda") or extra.get("numero_salida", ""),
                subtotal=round(subtotal, 2),
                descuento_pct=descuento_pct,
                descuento=descuento,
                iva=iva,
                total=total,
                sae_codigo=sae_codigo or None,
                productos=detail,
            ),
            "fecha": fecha_doc,
            "estatus": "Cancelada" if cancelada else "Activa",
        })

    total_fuente = len(seen)
    if not parsed:
        _actualizar_progreso_importacion(
            estado="COMPLETADA", total=total_fuente, procesadas=total_fuente, importadas=0,
            omitidas=len(errors), errores=errors[:20],
        )
        return {"imported": 0, "overwritten": 0, "skipped": len(errors), "total": total_fuente, "errors": errors[:20]}

    imported = 0
    overwritten = 0
    total_importar = total_fuente
    # Las filas descartadas al leer el Excel también cuentan dentro del avance.
    procesadas = len(errors)
    _actualizar_progreso_importacion(total=total_importar, procesadas=procesadas, importadas=0, omitidas=len(errors), errores=errors[:20])
    conn = None
    cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor()
        folios = [item["payload"].factura for item in parsed]
        placeholders = ",".join(["%s"] * len(folios))
        cursor.execute(f"SELECT id, factura FROM facturas WHERE factura IN ({placeholders})", tuple(folios))
        existing = cursor.fetchall()
        existing_by_folio = {str(row[1] or ""): int(row[0]) for row in existing}
        existing_ids = list(existing_by_folio.values())
        # Una reimportación no puede borrar ni sustituir una factura que ya
        # tiene CFDI: perdería su UUID, XML, acuse y trazabilidad fiscal.
        # Esos folios se omiten de forma explícita; primero deben gestionarse
        # mediante la cancelación fiscal, no por el importador.
        folios_protegidos = set()
        protegidos_por_folio = {}
        if existing_ids:
            id_placeholders = ",".join(["%s"] * len(existing_ids))
            cursor.execute(f"SELECT DISTINCT factura_id FROM cfdi_emitidos WHERE factura_id IN ({id_placeholders})", tuple(existing_ids))
            protegidos_ids = {int(row[0]) for row in cursor.fetchall()}
            folios_protegidos = {folio for folio, invoice_id in existing_by_folio.items() if invoice_id in protegidos_ids}
            if folios_protegidos:
                protegidos_por_folio = {folio: existing_by_folio[folio] for folio in folios_protegidos}
                if not sobrescribir_fiscales:
                    parsed = [item for item in parsed if item["payload"].factura not in folios_protegidos]
                    for folio_protegido in sorted(folios_protegidos):
                        errors.append(f"{folio_protegido}: omitida porque ya tiene CFDI emitido; active la sobreescritura fiscal para actualizar sus datos internos sin borrar el CFDI.")
                existing_ids = [invoice_id for folio, invoice_id in existing_by_folio.items() if folio not in folios_protegidos]
        overwritten = len(existing_ids) + (len(protegidos_por_folio) if sobrescribir_fiscales else 0)
        if existing_ids:
            id_placeholders = ",".join(["%s"] * len(existing_ids))
            cursor.execute(f"DELETE FROM factura_detalle WHERE factura_id IN ({id_placeholders})", tuple(existing_ids))
            cursor.execute(f"DELETE FROM facturas WHERE id IN ({id_placeholders})", tuple(existing_ids))
        conn.commit()
        _actualizar_progreso_importacion(sobrescritas=overwritten)
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudieron limpiar folios existentes: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    for item in parsed:
        payload = item["payload"]
        conn = cursor = None
        try:
            protected_invoice_id = protegidos_por_folio.get(payload.factura) if sobrescribir_fiscales else None
            if protected_invoice_id:
                # Conserva UUID, XML, acuse, cola y estatus fiscal: sólo actualiza
                # encabezado y partidas con la fuente histórica.
                update_invoice(protected_invoice_id, payload, user)
                inv_id = protected_invoice_id
            elif payload.productos and len(payload.productos) > 0:
                result = create_invoice(payload, user)
                inv_id = result["id"]
            else:
                conn = get_legacy_connection()
                cursor = conn.cursor()
                now_value = datetime.now()
                fecha_value = datetime.combine(payload.fecha, now_value.time()) if payload.fecha else now_value
                columns = _table_columns(cursor, "facturas")
                header = {
                    "fecha": fecha_value, "numero_cliente": payload.numero_cliente.strip(),
                    "cliente_nombre": payload.cliente_nombre.strip(), "consignatario": payload.consignatario.strip() or payload.cliente_nombre.strip(),
                    "factura": payload.factura.strip(), "subtotal": payload.subtotal,
                    "descuento_pct": payload.descuento_pct, "descuento": payload.descuento,
                    "iva": payload.iva, "total": payload.total, "sae_codigo": payload.sae_codigo,
                    "estatus": item["estatus"], "empresa": payload.empresa.strip(),
                    "numero_salida": payload.numero_salida.strip(), "comanda": payload.comanda.strip() or payload.numero_salida.strip(),
                    "rfc": payload.rfc.strip(), "vendedor": payload.vendedor.strip(),
                    "lista_precios": payload.lista_precios.strip(), "cargo_rebanado_pct": payload.cargo_rebanado_pct,
                }
                _insert_dynamic(cursor, "facturas", columns, header)
                inv_id = cursor.lastrowid
                if not inv_id:
                    cursor.execute("SELECT id FROM facturas WHERE TRIM(COALESCE(factura,''))=TRIM(%s) LIMIT 1", (payload.factura.strip(),))
                    row = cursor.fetchone()
                    inv_id = row[0] if row else None
                conn.commit()
            imported += 1
        except HTTPException as exc:
            errors.append(f"{payload.factura}: {exc.detail}")
        except Exception as exc:
            errors.append(f"{payload.factura}: {exc}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
            procesadas += 1
            _actualizar_progreso_importacion(
                total=total_importar,
                procesadas=procesadas,
                importadas=imported,
                omitidas=len(errors),
                errores=errors[:20],
            )

    return {"imported": imported, "overwritten": overwritten, "skipped": len(errors), "total": total_importar, "errors": errors[:20]}
