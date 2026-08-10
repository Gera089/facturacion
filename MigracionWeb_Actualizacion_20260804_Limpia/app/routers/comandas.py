from datetime import datetime
from io import BytesIO
import json
from pathlib import Path
import unicodedata

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.comandas_legacy.api_vendedor import crear_jwt_web
from app.dependencies import require_user
from app.legacy_db import get_editor_connection, get_legacy_connection
from app.core.config import settings
from app.routers.billing import _invoice_where

router = APIRouter(prefix="/api/comandas", tags=["Comandas"])


class ComandaProductoIn(BaseModel):
    cip: str = Field(min_length=1, max_length=30)
    descripcion: str = Field(min_length=1, max_length=255)
    kgs: float = Field(default=0, ge=0)
    piezas: float = Field(default=0, ge=0)
    observaciones: str = Field(default="", max_length=200)


class ComandaGuardarIn(BaseModel):
    folio: str = Field(default="", max_length=50)
    vendedor: str = Field(default="", max_length=120)
    empresa: str = Field(min_length=1, max_length=120)
    cliente_numero: str = Field(min_length=1, max_length=60)
    cliente_nombre: str = Field(min_length=1, max_length=255)
    observaciones_pedido: str = Field(default="", max_length=500)
    productos: list[ComandaProductoIn] = Field(min_length=1)


class FoliosIn(BaseModel):
    folios: list[str] = Field(min_length=1, max_length=200)


class EstatusComandaIn(BaseModel):
    folio: str = Field(min_length=1, max_length=50)
    estatus: int = Field(ge=1, le=2)


class RevisionClienteIn(BaseModel):
    empresa: str = Field(min_length=1, max_length=60)
    numero: str = Field(min_length=1, max_length=255)
    revision: str = Field(default="", max_length=255)


class ClienteIn(BaseModel):
    numero: str = Field(min_length=1, max_length=60)
    nombre: str = Field(min_length=1, max_length=255)
    empresa: str = Field(min_length=1, max_length=120)
    direccion_entrega: str = Field(default="", max_length=500)
    observaciones: str = Field(default="", max_length=1000)
    dias_credito: str = Field(default="", max_length=100)
    contacto1: str = Field(default="", max_length=255)
    revision: str = Field(default="", max_length=255)
    numero_original: str = Field(default="", max_length=60)
    empresa_original: str = Field(default="", max_length=120)


class ProductoIn(BaseModel):
    cip: str = Field(min_length=1, max_length=50)
    descripcion: str = Field(min_length=1, max_length=255)
    unidad: str = Field(default="", max_length=50)
    cip_original: str = Field(default="", max_length=50)


class NombreCatalogoIn(BaseModel):
    id: int | None = None
    nombre: str = Field(min_length=1, max_length=120)


class PedidoVendedorIn(BaseModel):
    id: int = Field(gt=0)
    estado: str = Field(default="", max_length=20)


class RutaRespaldoIn(BaseModel):
    nombre: str = Field(default="rutas", max_length=120)
    datos: dict


class RutaExportarIn(BaseModel):
    fecha: str = Field(default="", max_length=30)
    entregas: list[dict] = Field(default_factory=list)
    cobranza: list[dict] = Field(default_factory=list)


class ReporteEntregaFacturasExportIn(BaseModel):
    empresa: str = Field(default="", max_length=120)
    fecha_inicio: str = Field(default="", max_length=30)
    fecha_fin: str = Field(default="", max_length=30)
    rows: list[dict] = Field(default_factory=list)


RUTAS_RESPALDOS_DIR = settings.storage_dir / "rutas_respaldos"


def _nombre_respaldo_rutas(nombre: str) -> str:
    limpio = "".join(caracter if (caracter.isalnum() or caracter in "-_ ") else "-" for caracter in str(nombre or ""))
    limpio = "-".join(limpio.strip().split()).strip("-_")[:80]
    return limpio or "rutas"


def _texto_excel(value) -> str:
    """Normaliza las celdas de una plantilla sin convertir códigos a notación científica."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _etiqueta_excel(value) -> str:
    texto = unicodedata.normalize("NFKD", _texto_excel(value)).encode("ascii", "ignore").decode("ascii")
    return " ".join(texto.upper().replace(".", " ").replace(":", " ").replace("°", " ").split())


def _encabezado_productos(filas: list[tuple]) -> tuple[int, dict[str, int]] | tuple[None, dict]:
    """Ubica la fila de encabezados aun si la plantilla tiene columnas adicionales."""
    for indice, fila in enumerate(filas[:60]):
        etiquetas = {_etiqueta_excel(valor): columna for columna, valor in enumerate(fila) if _etiqueta_excel(valor)}
        if "CIP" not in etiquetas:
            continue
        descripcion = next((columna for etiqueta, columna in etiquetas.items() if etiqueta in {"DESCRIPCION", "PRODUCTO"}), None)
        piezas = next((columna for etiqueta, columna in etiquetas.items() if etiqueta in {"PIEZAS", "PZS", "PZAS"}), None)
        kilos = next((columna for etiqueta, columna in etiquetas.items() if etiqueta in {"KGS", "KG", "KILOS"}), None)
        observaciones = next((columna for etiqueta, columna in etiquetas.items() if etiqueta.startswith("OBSERV")), None)
        if descripcion is not None and (piezas is not None or kilos is not None):
            return indice, {"cip": etiquetas["CIP"], "descripcion": descripcion, "piezas": piezas, "kgs": kilos, "observaciones": observaciones}
    return None, {}


def _hoja_importacion_comanda(workbook):
    """Prioriza la hoja de captura Comanda sobre reportes/formatos auxiliares del mismo archivo."""
    for sheet in workbook.worksheets:
        if _etiqueta_excel(sheet["A1"].value) == "VENDEDOR":
            return sheet
    for sheet in workbook.worksheets:
        if _etiqueta_excel(sheet.title) == "COMANDA":
            return sheet
    return workbook.active


def _hoja_pedido_retail(workbook):
    """Encuentra la hoja real de captura y evita el resumen con fórmulas."""
    for sheet in workbook.worksheets:
        filas = list(sheet.iter_rows(max_row=60, max_col=64, values_only=True))
        indice, columnas = _encabezado_productos(filas)
        if indice is None:
            continue
        encabezados = {_etiqueta_excel(valor) for valor in filas[indice]}
        superiores = {_etiqueta_excel(valor) for fila in filas[:indice] for valor in fila}
        if "PRODUCTO" in encabezados and ("NO CLIE" in superiores or "NO CLIENTE" in superiores):
            return sheet, filas, indice, columnas
    return None


def _observaciones_retail(filas: list[tuple], inicio: int) -> str:
    for indice, fila in enumerate(filas[inicio:], start=inicio):
        # En City/Fresko el bloque de observaciones está en la columna B.
        # No se usa toda la fila: contiene tablas auxiliares de búsqueda en N:V.
        if len(fila) < 2 or _etiqueta_excel(fila[1]) != "OBSERVACIONES":
            continue
        for siguiente in filas[indice + 1:indice + 5]:
            texto = _texto_excel(siguiente[1] if len(siguiente) > 1 else "")
            if texto:
                return texto
    return ""


def _catalogo_cip_retail(sheet) -> dict[str, str]:
    """City/Fresko guarda la descripción capturada en N y el CIP real en O."""
    catalogo: dict[str, str] = {}
    for descripcion, cip in sheet.iter_rows(min_col=14, max_col=15, values_only=True):
        if _texto_excel(descripcion) and _texto_excel(cip):
            catalogo[_etiqueta_excel(descripcion)] = _texto_excel(cip)
    return catalogo


def _asegurar_columna_revision(cur) -> None:
    """Migra el campo propio de revisión sin reutilizar ni modificar el vendedor."""
    cur.execute("SHOW COLUMNS FROM clientes LIKE 'revision'")
    if not cur.fetchone():
        cur.execute("ALTER TABLE clientes ADD COLUMN revision VARCHAR(255) NULL DEFAULT NULL AFTER dias_credito")


@router.get("/sso")
def iniciar_sesion_catalogo(user: dict = Depends(require_user)):
    """Emite una sesión firmada para el catálogo integrado de Comandas."""
    return {
        "token": crear_jwt_web({
            "id": user.get("id") or 0,
            "usuario": user.get("username") or user.get("full_name") or "usuario",
            "rol": user.get("role") or "consulta",
        })
    }


@router.get("/empresas")
def listar_empresas(user: dict = Depends(require_user)):
    conn = get_legacy_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT DISTINCT empresa FROM clientes WHERE TRIM(COALESCE(empresa, '')) <> '' ORDER BY empresa")
        return [str(row[0]).strip() for row in cur.fetchall() if str(row[0] or '').strip()]
    finally:
        cur.close()
        conn.close()


@router.get("/vendedores")
def listar_vendedores(user: dict = Depends(require_user)):
    conn = get_legacy_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT nombre FROM vendedores WHERE TRIM(COALESCE(nombre, '')) <> '' ORDER BY nombre")
        return [str(row[0]).strip() for row in cur.fetchall() if str(row[0] or '').strip()]
    finally:
        cur.close()
        conn.close()


@router.get("/siguiente-folio")
def siguiente_folio(user: dict = Depends(require_user)):
    conn = get_legacy_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COALESCE(MAX(CAST(folio AS UNSIGNED)), 0) + 1 FROM comandas WHERE folio REGEXP '^[0-9]+$'")
        return {"folio": str((cur.fetchone() or [1])[0] or 1)}
    finally:
        cur.close()
        conn.close()


@router.get("/plantilla")
def descargar_plantilla(user: dict = Depends(require_user)):
    """Genera la misma plantilla de captura que usa el proyecto original de Comandas."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Comanda"
    for row, (etiqueta, valor) in enumerate((
        ("Vendedor:", ""),
        ("Empresa:", ""),
        ("Número de Cliente:", ""),
        ("Observaciones de Logística:", ""),
    ), start=1):
        sheet.cell(row=row, column=1, value=etiqueta)
        sheet.cell(row=row, column=2, value=valor)
    headers = ("CIP", "Descripción", "Kgs", "Piezas", "Observaciones")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=6, column=column, value=header)
        cell.font = Font(bold=True)
    for row in range(7, 22):
        for column in range(1, 6):
            sheet.cell(row=row, column=column, value="")
    for column, width in {"A": 18, "B": 50, "C": 14, "D": 14, "E": 34}.items():
        sheet.column_dimensions[column].width = width
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_comanda.xlsx"'},
    )


@router.post("/importar")
async def importar_plantilla(archivo: UploadFile = File(...), user: dict = Depends(require_user)):
    """Lee una plantilla y devuelve la captura para revisión; todavía no guarda la comanda."""
    if not (archivo.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Selecciona un archivo Excel .xlsx.")
    contenido = await archivo.read()
    if not contenido or len(contenido) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="El archivo debe contener datos y no superar 5 MB.")
    try:
        workbook = load_workbook(BytesIO(contenido), data_only=True, read_only=True)
        retail = _hoja_pedido_retail(workbook)
        if retail:
            sheet, filas, indice_encabezado, columnas = retail
        else:
            sheet = _hoja_importacion_comanda(workbook)
            filas = list(sheet.iter_rows(max_col=64, values_only=True))
            indice_encabezado, columnas = _encabezado_productos(filas)
        if indice_encabezado is None:
            raise HTTPException(status_code=400, detail="No se encontró la tabla de productos (CIP, descripción y piezas/kgs).")

        # Plantilla Comanda / City / Ibersur: encabezado de captura fijo en B1:B4.
        es_formato_comanda = not retail and _etiqueta_excel(filas[0][0] if filas and filas[0] else "") == "VENDEDOR"
        vendedor = _texto_excel(sheet["B1"].value) if es_formato_comanda else ""
        empresa = _texto_excel(sheet["B2"].value) if es_formato_comanda else ""
        cliente_numero = _texto_excel(sheet["B3"].value) if es_formato_comanda else ""
        cliente_nombre = _texto_excel(sheet["E3"].value) if es_formato_comanda else ""
        observaciones_pedido = _texto_excel(sheet["B4"].value) if es_formato_comanda else ""

        # Formato Walmart: la cabecera de cliente está en una fila superior a su tabla CIP.
        if not es_formato_comanda:
            for indice, fila in enumerate(filas[:indice_encabezado]):
                etiquetas = [_etiqueta_excel(valor) for valor in fila]
                if "NO CLIE" in etiquetas or "NO CLIENTE" in etiquetas:
                    columna_cliente = etiquetas.index("NO CLIE") if "NO CLIE" in etiquetas else etiquetas.index("NO CLIENTE")
                    columna_nombre = next((i for i, etiqueta in enumerate(etiquetas) if etiqueta == "NOMBRE DEL CLIENTE"), None)
                    columna_vendedor = next((i for i, etiqueta in enumerate(etiquetas) if etiqueta == "VENDEDOR"), None)
                    datos = filas[indice + 1] if indice + 1 < len(filas) else ()
                    cliente_numero = _texto_excel(datos[columna_cliente] if columna_cliente < len(datos) else "")
                    cliente_nombre = _texto_excel(datos[columna_nombre] if columna_nombre is not None and columna_nombre < len(datos) else "")
                    vendedor = _texto_excel(datos[columna_vendedor] if columna_vendedor is not None and columna_vendedor < len(datos) else "")
                    empresa = "Gourmet España"
                    break

        if retail:
            observaciones_pedido = _observaciones_retail(filas, indice_encabezado)

        productos = []
        vacias_seguidas = 0
        catalogo_retail = _catalogo_cip_retail(sheet) if retail else {}
        observaciones_sin_partida = []
        for fila in filas[indice_encabezado + 1:]:
            descripcion = _texto_excel(fila[columnas["descripcion"]] if columnas["descripcion"] < len(fila) else "")
            kgs = _texto_excel(fila[columnas["kgs"]] if columnas.get("kgs") is not None and columnas["kgs"] < len(fila) else "")
            piezas = _texto_excel(fila[columnas["piezas"]] if columnas.get("piezas") is not None and columnas["piezas"] < len(fila) else "")
            observaciones = _texto_excel(fila[columnas["observaciones"]] if columnas.get("observaciones") is not None and columnas["observaciones"] < len(fila) else "")
            cip_guardado = _texto_excel(fila[columnas["cip"]] if columnas["cip"] < len(fila) else "")
            cip = catalogo_retail.get(_etiqueta_excel(descripcion), cip_guardado) if retail else cip_guardado
            if not cip and not descripcion:
                if observaciones:
                    observaciones_sin_partida.append(observaciones)
                vacias_seguidas += 1
                if vacias_seguidas >= 3:
                    break
                continue
            if retail and (not descripcion or (not kgs and not piezas)):
                if observaciones:
                    observaciones_sin_partida.append(observaciones)
                continue
            vacias_seguidas = 0
            productos.append({
                "cip": cip,
                "descripcion": descripcion,
                "kgs": kgs,
                "piezas": piezas,
                "observaciones": observaciones,
            })
        if observaciones_sin_partida:
            observaciones_pedido = "\n".join(filter(None, [observaciones_pedido, *observaciones_sin_partida]))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer la plantilla Excel: {exc}") from exc
    return {
        "formato": "retail directo" if retail else ("comanda" if es_formato_comanda else "walmart"),
        "vendedor": vendedor,
        "empresa": empresa,
        "cliente_numero": cliente_numero,
        "cliente_nombre": cliente_nombre,
        "observaciones_pedido": observaciones_pedido,
        "productos": productos,
    }


@router.get("/clientes")
def buscar_clientes(empresa: str, q: str = "", numero: str = "", nombre: str = "", limit: int = Query(30, ge=1, le=100), user: dict = Depends(require_user)):
    conn = get_legacy_connection()
    cur = conn.cursor(dictionary=True)
    try:
        _asegurar_columna_revision(cur)
        sql = """SELECT numero, nombre, empresa, COALESCE(observaciones, '') AS observaciones,
                       COALESCE(direccion_entrega, '') AS direccion_entrega,
                       COALESCE(dias_credito, '') AS pago,
                       COALESCE(revision, '') AS revision,
                       COALESCE(contacto1, '') AS contactos
                 FROM clientes WHERE empresa=%s"""
        params = [empresa.strip()]
        if numero.strip():
            sql += " AND numero LIKE %s"
            params.append(f"%{numero.strip()}%")
        if nombre.strip():
            sql += " AND nombre LIKE %s"
            params.append(f"%{nombre.strip()}%")
        if q.strip():
            sql += " AND (numero LIKE %s OR nombre LIKE %s)"
            termino = f"%{q.strip()}%"
            params.extend([termino, termino])
        # El número de cliente es el orden operativo; la empresa resuelve los
        # números repetidos entre compañías.  CAST evita el orden alfabético
        # (por ejemplo, 100000 antes que 99999).
        sql += " ORDER BY CAST(numero AS UNSIGNED), empresa, nombre LIMIT %s"
        params.append(int(limit))
        cur.execute(sql, tuple(params))
        return cur.fetchall() or []
    finally:
        cur.close()
        conn.close()


@router.put("/clientes/revision")
def guardar_revision_cliente(payload: RevisionClienteIn, user: dict = Depends(require_user)):
    """Guarda la revisión como dato propio del cliente; no modifica su vendedor."""
    conn = get_legacy_connection()
    cur = conn.cursor()
    try:
        _asegurar_columna_revision(cur)
        cur.execute(
            "UPDATE clientes SET revision=%s WHERE empresa=%s AND numero=%s",
            (payload.revision.strip(), payload.empresa.strip(), payload.numero.strip()),
        )
        if cur.rowcount < 1:
            raise HTTPException(status_code=404, detail="No se encontró el cliente para guardar la revisión.")
        conn.commit()
        return {"ok": True, "empresa": payload.empresa.strip(), "numero": payload.numero.strip(), "revision": payload.revision.strip()}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as exc:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo guardar la revisión: {exc}") from exc
    finally:
        cur.close()
        conn.close()


@router.get("/productos")
def buscar_productos(q: str = "", cip: str = "", descripcion: str = "", limit: int = Query(30, ge=1, le=100), user: dict = Depends(require_user)):
    conn = get_legacy_connection()
    cur = conn.cursor(dictionary=True)
    try:
        sql = """SELECT cip, descripcion, COALESCE(unidad, '') AS unidad,
                       COALESCE(badge_1, '') AS badge_1, COALESCE(badge_2, '') AS badge_2,
                       COALESCE(badge_3, '') AS badge_3, COALESCE(etiquetas_retail, '') AS etiquetas_retail,
                       COALESCE(premium_sort, 0) AS orden, COALESCE(premium_activo, 0) AS activo
                 FROM productos WHERE 1=1"""
        params = []
        if cip.strip(): sql += " AND cip LIKE %s"; params.append(f"%{cip.strip()}%")
        if descripcion.strip(): sql += " AND descripcion LIKE %s"; params.append(f"%{descripcion.strip()}%")
        if q.strip(): sql += " AND (cip LIKE %s OR descripcion LIKE %s)"; params.extend([f"%{q.strip()}%", f"%{q.strip()}%"])
        sql += " ORDER BY descripcion LIMIT %s"; params.append(int(limit))
        cur.execute(sql, tuple(params))
        return cur.fetchall() or []
    finally:
        cur.close()
        conn.close()


@router.get("/diario")
def listar_diario(empresa: str = "", q: str = "", fecha_inicio: str = "", fecha_fin: str = "", limit: int = Query(200, ge=1, le=500), user: dict = Depends(require_user)):
    conn = get_legacy_connection()
    cur = conn.cursor(dictionary=True)
    try:
        sql = """SELECT c.id, c.folio, c.fecha, c.vendedor, c.empresa, c.cliente_numero, c.cliente_nombre,
                         COALESCE(c.observaciones_pedido,'') AS observaciones_pedido,
                         COUNT(d.id) AS productos
                  FROM comandas c LEFT JOIN productos_comanda d ON d.comanda_id=c.id WHERE 1=1"""
        params = []
        if empresa.strip(): sql += " AND c.empresa=%s"; params.append(empresa.strip())
        if q.strip():
            sql += " AND (c.folio LIKE %s OR c.cliente_numero LIKE %s OR c.cliente_nombre LIKE %s)"
            params.extend([f"%{q.strip()}%"] * 3)
        if fecha_inicio.strip():
            sql += " AND DATE(c.fecha) >= %s"; params.append(fecha_inicio.strip())
        if fecha_fin.strip():
            sql += " AND DATE(c.fecha) <= %s"; params.append(fecha_fin.strip())
        sql += " GROUP BY c.id ORDER BY c.id DESC LIMIT %s"; params.append(int(limit))
        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []
        folios = [str(row.get("folio") or "").strip() for row in rows if str(row.get("folio") or "").strip()]
        enviados = set()
        if folios:
            editor_conn = get_editor_connection(); editor_cur = editor_conn.cursor()
            try:
                placeholders = ",".join(["%s"] * len(folios))
                editor_cur.execute(f"SELECT folio FROM comandas_editables WHERE folio IN ({placeholders})", tuple(folios))
                enviados = {str(row[0]).strip() for row in editor_cur.fetchall() if row and row[0] is not None}
            finally:
                editor_cur.close(); editor_conn.close()
        for row in rows:
            row["enviado_facturar"] = str(row.get("folio") or "").strip() in enviados
        return rows
    finally:
        cur.close()
        conn.close()


@router.get("/diario/{comanda_id}")
def detalle_diario(comanda_id: int, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM comandas WHERE id=%s LIMIT 1", (comanda_id,)); cab = cur.fetchone()
        if not cab: raise HTTPException(status_code=404, detail="Comanda no encontrada")
        cur.execute("SELECT COALESCE(observaciones,'') AS observaciones FROM clientes WHERE empresa=%s AND numero=%s LIMIT 1", (cab.get("empresa") or "", cab.get("cliente_numero") or ""))
        cliente = cur.fetchone() or {}
        cab["observaciones_cliente"] = cliente.get("observaciones", "")
        cur.execute("SELECT cip, descripcion, kgs, piezas, COALESCE(observaciones,'') AS observaciones FROM productos_comanda WHERE comanda_id=%s ORDER BY id", (comanda_id,))
        cab["productos"] = cur.fetchall() or []
        return cab
    finally:
        cur.close(); conn.close()


@router.get("/pedidos-vendedor")
def listar_pedidos_vendedor(estado: str = "PENDIENTE", limit: int = Query(300, ge=1, le=500), user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""SELECT id, fecha, vendedor, empresa, cliente_numero, cliente_nombre,
                              COALESCE(folio_usado,'') AS comanda, COALESCE(observaciones_pedido,'') AS observaciones,
                              estado
                       FROM pedidos_vendedor WHERE (%s='' OR estado=%s) ORDER BY id DESC LIMIT %s""", (estado.strip(), estado.strip(), int(limit)))
        return cur.fetchall() or []
    finally:
        cur.close(); conn.close()


@router.get("/repartidores")
def listar_repartidores(q: str = "", user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, nombre FROM repartidores WHERE nombre LIKE %s ORDER BY nombre", (f"%{q.strip()}%",))
        return cur.fetchall() or []
    finally:
        cur.close(); conn.close()


@router.get("/por-facturar")
def listar_por_facturar(empresa: str = "", q: str = "", fecha_inicio: str = "", fecha_fin: str = "", limit: int = Query(300, ge=1, le=500), user: dict = Depends(require_user)):
    conn = get_editor_connection(); cur = conn.cursor(dictionary=True)
    try:
        sql = """SELECT c.id, c.folio, c.cliente_numero, c.cliente_nombre, c.empresa, c.fecha,
                         COALESCE(c.observaciones_pedido,'') AS observaciones,
                         COALESCE(c.estatus_facturado, 0) AS estatus_facturado
                  FROM comandas_editables c WHERE 1=1"""
        params = []
        if empresa.strip():
            sql += " AND c.empresa=%s"; params.append(empresa.strip())
        if q.strip():
            sql += " AND (c.folio LIKE %s OR c.cliente_numero LIKE %s OR c.cliente_nombre LIKE %s)"
            params.extend([f"%{q.strip()}%"] * 3)
        if fecha_inicio.strip():
            sql += " AND DATE(c.fecha) >= %s"; params.append(fecha_inicio.strip())
        if fecha_fin.strip():
            sql += " AND DATE(c.fecha) <= %s"; params.append(fecha_fin.strip())
        sql += " ORDER BY c.id DESC LIMIT %s"; params.append(int(limit))
        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []
        folios = [str(row.get("folio") or "").strip() for row in rows if str(row.get("folio") or "").strip()]
        facturados = set()
        if folios:
            legacy_conn = get_legacy_connection(); legacy_cur = legacy_conn.cursor()
            try:
                marks = ",".join(["%s"] * len(folios))
                legacy_cur.execute(f"SELECT comanda FROM facturas WHERE comanda IS NOT NULL AND TRIM(CAST(comanda AS CHAR)) IN ({marks})", tuple(folios))
                facturados = {str(row[0]).strip() for row in legacy_cur.fetchall() if row and row[0] is not None}
            finally:
                legacy_cur.close(); legacy_conn.close()
        for row in rows:
            row["facturado_emitido"] = str(row.get("folio") or "").strip() in facturados
        return rows
    finally:
        cur.close(); conn.close()


@router.get("/rutas")
def listar_rutas(empresa: str = "", q: str = "", fecha_inicio: str = "", fecha_fin: str = "", limit: int = Query(300, ge=1, le=500), user: dict = Depends(require_user)):
    # La vista original parte de las comandas aún no facturadas para planear ruta.
    conn = get_editor_connection(); cur = conn.cursor(dictionary=True)
    try:
        sql = """SELECT c.id, c.folio, c.cliente_numero, c.cliente_nombre, c.empresa, c.fecha,
                         COALESCE(c.observaciones_pedido,'') AS observaciones,
                         COALESCE(c.estatus_facturado, 0) AS estatus_facturado
                  FROM comandas_editables c WHERE 1=1"""
        params = []
        if empresa.strip(): sql += " AND c.empresa=%s"; params.append(empresa.strip())
        if q.strip():
            sql += " AND (c.folio LIKE %s OR c.cliente_numero LIKE %s OR c.cliente_nombre LIKE %s)"
            params.extend([f"%{q.strip()}%"] * 3)
        if fecha_inicio.strip(): sql += " AND DATE(c.fecha) >= %s"; params.append(fecha_inicio.strip())
        if fecha_fin.strip(): sql += " AND DATE(c.fecha) <= %s"; params.append(fecha_fin.strip())
        sql += " ORDER BY CAST(c.folio AS UNSIGNED) DESC, c.id DESC LIMIT %s"; params.append(int(limit))
        cur.execute(sql, tuple(params))
        return cur.fetchall() or []
    finally:
        cur.close(); conn.close()


@router.get("/rutas/generar")
def datos_generar_rutas(empresa: str = "", q: str = "", fecha_inicio: str = "", fecha_fin: str = "", limit: int = Query(300, ge=1, le=500), user: dict = Depends(require_user)):
    """Entrega las comandas visibles y sus partidas para el planificador de rutas."""
    conn = get_editor_connection(); cur = conn.cursor(dictionary=True)
    try:
        sql = """SELECT c.id, c.folio, c.cliente_numero, c.cliente_nombre, c.empresa,
                         c.fecha, COALESCE(c.observaciones_pedido, '') AS observaciones
                  FROM comandas_editables c WHERE 1=1"""
        params = []
        if empresa.strip():
            sql += " AND c.empresa=%s"; params.append(empresa.strip())
        if q.strip():
            sql += " AND (c.folio LIKE %s OR c.cliente_numero LIKE %s OR c.cliente_nombre LIKE %s)"
            params.extend([f"%{q.strip()}%"] * 3)
        if fecha_inicio.strip():
            sql += " AND DATE(c.fecha) >= %s"; params.append(fecha_inicio.strip())
        if fecha_fin.strip():
            sql += " AND DATE(c.fecha) <= %s"; params.append(fecha_fin.strip())
        sql += " ORDER BY CAST(c.folio AS UNSIGNED) DESC, c.id DESC LIMIT %s"; params.append(int(limit))
        cur.execute(sql, tuple(params))
        rows = cur.fetchall() or []
        ids = [row["id"] for row in rows]
        productos_por_comanda: dict[int, list[str]] = {ident: [] for ident in ids}
        if ids:
            marcas = ",".join(["%s"] * len(ids))
            cur.execute(f"""SELECT comanda_id, COALESCE(piezas, 0) AS piezas,
                                   COALESCE(descripcion, '') AS descripcion
                            FROM comandas_editables_detalle
                            WHERE comanda_id IN ({marcas}) ORDER BY comanda_id, id""", tuple(ids))
            for detalle in cur.fetchall() or []:
                piezas = float(detalle.get("piezas") or 0)
                descripcion = str(detalle.get("descripcion") or "").strip()
                if not descripcion:
                    continue
                cantidad = str(int(piezas)) if piezas.is_integer() else str(piezas).rstrip("0").rstrip(".")
                productos_por_comanda.setdefault(detalle["comanda_id"], []).append(f"{cantidad} {descripcion}".strip() if piezas else descripcion)
        for row in rows:
            row["productos"] = ", ".join(productos_por_comanda.get(row["id"], []))
        return rows
    finally:
        cur.close(); conn.close()


@router.get("/rutas/respaldos")
def listar_respaldos_rutas(user: dict = Depends(require_user)):
    RUTAS_RESPALDOS_DIR.mkdir(parents=True, exist_ok=True)
    archivos = []
    for ruta in sorted(RUTAS_RESPALDOS_DIR.glob("*.json"), key=lambda item: item.stat().st_mtime, reverse=True):
        stat = ruta.stat()
        archivos.append({"archivo": ruta.name, "nombre": ruta.stem, "fecha": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"), "bytes": stat.st_size})
    return archivos


@router.post("/rutas/respaldos")
def guardar_respaldo_rutas(payload: RutaRespaldoIn, user: dict = Depends(require_user)):
    datos = payload.datos or {}
    if not isinstance(datos.get("entregas", {}).get("rows"), list) or not isinstance(datos.get("cobranza", {}).get("rows"), list):
        raise HTTPException(status_code=400, detail="El respaldo debe incluir las rutas de entregas y cobranza.")
    RUTAS_RESPALDOS_DIR.mkdir(parents=True, exist_ok=True)
    base = _nombre_respaldo_rutas(payload.nombre)
    marca = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = RUTAS_RESPALDOS_DIR / f"{base}_{marca}.json"
    datos["saved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    datos["nombre"] = base
    try:
        destino.write_text(json.dumps(datos, ensure_ascii=False, indent=2), encoding="utf-8")
    except OSError as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo guardar el respaldo en servidor: {exc}") from exc
    return {"ok": True, "archivo": destino.name, "nombre": base, "ruta": str(destino)}


@router.get("/rutas/respaldos/{archivo}")
def obtener_respaldo_rutas(archivo: str, user: dict = Depends(require_user)):
    nombre = Path(archivo).name
    if not nombre.lower().endswith(".json") or nombre != archivo:
        raise HTTPException(status_code=400, detail="Nombre de respaldo no válido.")
    ruta = RUTAS_RESPALDOS_DIR / nombre
    if not ruta.is_file():
        raise HTTPException(status_code=404, detail="No se encontró el respaldo solicitado.")
    try:
        return json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo leer el respaldo: {exc}") from exc


def _fecha_ruta_excel(fecha: str) -> str:
    dias = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sabado", "Domingo"]
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    valor = str(fecha or "").strip()
    parsed = None
    for formato in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            parsed = datetime.strptime(valor, formato)
            break
        except ValueError:
            pass
    if not parsed:
        parsed = datetime.now()
    return f"{dias[parsed.weekday()]} {parsed.day} {meses[parsed.month - 1]} {parsed.year}"


def _safe_sheet_title(nombre: str, usados: set[str]) -> str:
    base = "".join(" " if char in r'[]:*?/\\' else char for char in str(nombre or "").strip())[:31].strip() or "SIN_REPARTIDOR"
    titulo = base
    contador = 2
    while titulo in usados:
        sufijo = f"_{contador}"
        titulo = f"{base[:31 - len(sufijo)]}{sufijo}"
        contador += 1
    usados.add(titulo)
    return titulo


def _valor_ruta(row: dict, *keys: str) -> str:
    for key in keys:
        valor = row.get(key)
        if valor is not None and str(valor).strip():
            return str(valor).replace("\n", " ").strip()
    return ""


def _fecha_reporte_factura(fecha) -> tuple[str, str, str]:
    meses_es = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    fecha_obj = fecha
    if isinstance(fecha_obj, str):
        fecha_obj = fecha_obj.strip()
        for formato in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                fecha_obj = datetime.strptime(fecha_obj[:19], formato)
                break
            except Exception:
                continue
    if hasattr(fecha_obj, "strftime"):
        return fecha_obj.strftime("%d-%m-%Y"), str(fecha_obj.day), meses_es[fecha_obj.month - 1]
    return str(fecha or ""), "", ""


def _consultar_reporte_entrega_facturas(
    empresa: str = "",
    fecha_inicio: str = "",
    fecha_fin: str = "",
    q: str = "",
    month: int | None = None,
    year: int | None = None,
) -> list[dict]:
    where, params = _invoice_where(empresa or None, q or None, fecha_inicio or None, fecha_fin or None, month, year)
    visibles = [
        "UPPER(TRIM(COALESCE(f.factura, ''))) NOT LIKE 'TEST%'",
        "UPPER(TRIM(COALESCE(f.factura, ''))) NOT LIKE 'PRUEBA%'",
    ]
    query = """
        SELECT
            f.id,
            f.fecha,
            DAY(f.fecha) AS dia,
            MONTH(f.fecha) AS mes_num,
            TRIM(CAST(IFNULL(f.factura, '') AS CHAR)) AS factura,
            TRIM(CAST(IFNULL(f.numero_cliente, '') AS CHAR)) AS numero_cliente,
            COALESCE(NULLIF(TRIM(c.nombre), ''), NULLIF(TRIM(f.cliente_nombre), ''), NULLIF(TRIM(f.consignatario), ''), '') AS tienda,
            CASE WHEN UPPER(TRIM(COALESCE(f.estatus, ''))) IN ('CANCELADO', 'CANCELADA') THEN 0 ELSE f.total END AS total,
            CASE
                WHEN COALESCE(ce.serie, '') <> '' OR COALESCE(ce.folio_cfdi, '') <> ''
                THEN CONCAT(COALESCE(ce.serie, ''), COALESCE(ce.folio_cfdi, ''))
                ELSE COALESCE(f.sae_codigo, '')
            END AS sae_codigo,
            f.estatus,
            TRIM(CAST(IFNULL(f.empresa, '') AS CHAR)) AS empresa
        FROM facturas f
        LEFT JOIN clientes c
          ON TRIM(CAST(c.numero AS CHAR)) = TRIM(COALESCE(f.numero_cliente, ''))
         AND UPPER(TRIM(c.empresa) COLLATE utf8mb4_unicode_ci) =
             UPPER(TRIM(f.empresa) COLLATE utf8mb4_unicode_ci)
        LEFT JOIN (
            SELECT ce1.* FROM cfdi_emitidos ce1
            INNER JOIN (
                SELECT factura_id, MAX(id) AS id FROM cfdi_emitidos
                WHERE COALESCE(UPPER(TRIM(estatus_cfdi)), '') NOT IN ('CANCELADO', 'CANCELADA')
                GROUP BY factura_id
            ) ultimo
              ON ultimo.id = ce1.id
        ) ce ON (
            ce.factura_id = f.id
            OR EXISTS (
                SELECT 1 FROM cfdi_consolidacion_facturas ccf
                WHERE ccf.cfdi_emitido_id = ce.id AND ccf.factura_id = f.id
            )
            OR EXISTS (
                SELECT 1 FROM timbrado_queue tq
                WHERE tq.factura_id = f.id
                  AND COALESCE(tq.uuid, '') <> ''
                  AND tq.uuid = ce.uuid
            )
        )
    """
    query += " WHERE " + " AND ".join(where + visibles)
    query += " ORDER BY f.factura DESC, f.fecha DESC, f.id DESC LIMIT 10000"

    conn = get_legacy_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute(query, tuple(params))
        datos = []
        for row in cur.fetchall() or []:
            fecha_txt, dia, mes = _fecha_reporte_factura(row.get("fecha"))
            try:
                importe = float(row.get("total") or 0)
            except Exception:
                importe = 0.0
            estatus_raw = str(row.get("estatus") or "").strip().lower()
            datos.append({
                "estatus": "CANCELADA" if "cancel" in estatus_raw else "ACTIVA",
                "factura": str(row.get("factura") or "").strip(),
                "fecha": fecha_txt,
                "dia": str(row.get("dia") or dia or ""),
                "mes": mes,
                "cliente": str(row.get("numero_cliente") or "").strip(),
                "importe": importe,
                "tienda": str(row.get("tienda") or "").strip(),
                "sae": str(row.get("sae_codigo") or "").strip(),
                "empresa": str(row.get("empresa") or "").strip(),
            })
        return datos
    finally:
        cur.close()
        conn.close()


@router.get("/rutas/reporte-entrega-facturas")
def reporte_entrega_facturas(
    empresa: str = "",
    fecha_inicio: str = "",
    fecha_fin: str = "",
    q: str = "",
    month: int | None = Query(default=None, ge=0, le=12),
    year: int | None = Query(default=None, ge=2020, le=2100),
    user: dict = Depends(require_user),
):
    return _consultar_reporte_entrega_facturas(empresa=empresa, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, q=q, month=month, year=year)


@router.get("/rutas/reporte-entrega-facturas/exportar-excel")
def exportar_reporte_entrega_facturas(
    empresa: str = "",
    fecha_inicio: str = "",
    fecha_fin: str = "",
    q: str = "",
    month: int | None = Query(default=None, ge=0, le=12),
    year: int | None = Query(default=None, ge=2020, le=2100),
    user: dict = Depends(require_user),
):
    datos = _consultar_reporte_entrega_facturas(empresa=empresa, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, q=q, month=month, year=year)
    if not datos:
        raise HTTPException(status_code=404, detail="POR FAVOR PRIMERO FILTRA LOS DATOS A EXPORTAR")

    def _fmt_fecha_txt(txt: str) -> str:
        try:
            return datetime.strptime(str(txt or "").strip(), "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            return str(txt or "").strip()

    empresa_sel = empresa.strip() or "Todas"
    fecha_inicio_fmt = _fmt_fecha_txt(fecha_inicio)
    fecha_fin_fmt = _fmt_fecha_txt(fecha_fin)
    rango = fecha_inicio_fmt if fecha_inicio_fmt == fecha_fin_fmt or not fecha_fin_fmt else f"{fecha_inicio_fmt} al {fecha_fin_fmt}"

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Entrega Facturas"

    titulo = f'Entrega de Facturas de la empresa "{empresa_sel}", de la fecha "{rango}"'
    ws.merge_cells("A1:J1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    encabezados = ["estatus", "factura", "dia", "mes", "cliente", "importe", "Tienda", "F", "C", "L", "Observaciones"]
    for col_idx, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=2, column=col_idx, value=encabezado)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for fila_idx, dato in enumerate(datos, start=3):
        ws.cell(row=fila_idx, column=1, value=dato.get("estatus", ""))
        ws.cell(row=fila_idx, column=2, value=dato.get("factura", ""))
        ws.cell(row=fila_idx, column=3, value=dato.get("dia", ""))
        ws.cell(row=fila_idx, column=4, value=dato.get("mes", ""))
        ws.cell(row=fila_idx, column=5, value=dato.get("cliente", ""))
        ws.cell(row=fila_idx, column=6, value=float(dato.get("importe", 0) or 0))
        ws.cell(row=fila_idx, column=7, value=dato.get("tienda", ""))
        ws.cell(row=fila_idx, column=8, value="")
        ws.cell(row=fila_idx, column=9, value="")
        ws.cell(row=fila_idx, column=10, value="")
        ws.cell(row=fila_idx, column=11, value="")

    anchos = [14, 16, 10, 16, 14, 14, 34, 8, 8, 8, 26]
    for idx, ancho_col in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho_col
    for fila in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=11):
        for celda in fila:
            celda.alignment = Alignment(horizontal="center", vertical="center")
    for celda in ws["F"][2:]:
        celda.number_format = '$#,##0.00'

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"entrega_facturas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/rutas/reporte-entrega-facturas/exportar-excel")
def exportar_reporte_entrega_facturas_mio(payload: ReporteEntregaFacturasExportIn, user: dict = Depends(require_user)):
    datos = payload.rows or []
    if not datos:
        raise HTTPException(status_code=404, detail="POR FAVOR PRIMERO FILTRA LOS DATOS A EXPORTAR")

    def _fmt_fecha_txt(txt: str) -> str:
        try:
            return datetime.strptime(str(txt or "").strip(), "%Y-%m-%d").strftime("%d-%m-%Y")
        except Exception:
            return str(txt or "").strip()

    empresa_sel = payload.empresa.strip() or "Todas"
    fecha_inicio_fmt = _fmt_fecha_txt(payload.fecha_inicio)
    fecha_fin_fmt = _fmt_fecha_txt(payload.fecha_fin)
    rango = fecha_inicio_fmt if fecha_inicio_fmt == fecha_fin_fmt or not fecha_fin_fmt else f"{fecha_inicio_fmt} al {fecha_fin_fmt}"

    workbook = Workbook()
    ws = workbook.active
    ws.title = "Entrega Facturas"

    titulo = f'Entrega de Facturas de la empresa "{empresa_sel}", de la fecha "{rango}"'
    ws.merge_cells("A1:J1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    encabezados = ["estatus", "factura", "dia", "mes", "cliente", "importe", "Tienda", "F", "C", "L", "Observaciones"]
    for col_idx, encabezado in enumerate(encabezados, start=1):
        cell = ws.cell(row=2, column=col_idx, value=encabezado)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for fila_idx, dato in enumerate(datos, start=3):
        ws.cell(row=fila_idx, column=1, value=str(dato.get("estatus") or ""))
        ws.cell(row=fila_idx, column=2, value=str(dato.get("factura") or ""))
        ws.cell(row=fila_idx, column=3, value=str(dato.get("dia") or ""))
        ws.cell(row=fila_idx, column=4, value=str(dato.get("mes") or ""))
        ws.cell(row=fila_idx, column=5, value=str(dato.get("cliente") or ""))
        try:
            importe = float(dato.get("importe") or 0)
        except Exception:
            importe = 0.0
        ws.cell(row=fila_idx, column=6, value=importe)
        ws.cell(row=fila_idx, column=7, value=str(dato.get("tienda") or ""))
        ws.cell(row=fila_idx, column=8, value="")
        ws.cell(row=fila_idx, column=9, value="")
        ws.cell(row=fila_idx, column=10, value="")
        ws.cell(row=fila_idx, column=11, value="")

    anchos = [14, 16, 10, 16, 14, 14, 34, 8, 8, 8, 26]
    for idx, ancho_col in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho_col
    for fila in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=11):
        for celda in fila:
            celda.alignment = Alignment(horizontal="center", vertical="center")
    for celda in ws["F"][2:]:
        celda.number_format = '$#,##0.00'

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"entrega_facturas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/rutas/exportar-excel")
def exportar_rutas_excel(payload: RutaExportarIn, user: dict = Depends(require_user)):
    """Genera el Excel de rutas con el formato del proyecto Comandas06 original."""
    import math

    if not payload.entregas and not payload.cobranza:
        raise HTTPException(status_code=400, detail="No hay rutas para exportar.")

    columnas = ["Comanda", "Factura", "Cliente Nº", "Nombre Cliente", "Productos", "Repartidor", "Horario", "Observaciones"]
    anchos = {
        "Comanda": 12,
        "Factura": 12,
        "Cliente Nº": 20,
        "Nombre Cliente": 32,
        "Productos": 72,
        "Repartidor": 12,
        "Horario": 15,
        "Observaciones": 36,
    }
    unidades_repartidor = {
        "miguel": "TIDA PBJ9866",
        "francisco": "SAVEIRO PBJ8908",
        "jose": "MOTO 2771M9",
    }
    borde_negro = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    encabezado_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")
    encabezado_font = Font(color="FF0000", bold=True)
    bold_font = Font(bold=True)

    def repartidor_de(row: dict) -> str:
        return _valor_ruta(row, "repartidor") or "SIN_REPARTIDOR"

    rutas_entregas: dict[str, list[dict]] = {}
    rutas_cobranza: dict[str, list[dict]] = {}
    for row in payload.entregas:
        rutas_entregas.setdefault(repartidor_de(row), []).append(row)
    for row in payload.cobranza:
        rutas_cobranza.setdefault(repartidor_de(row), []).append(row)

    workbook = Workbook()
    workbook.remove(workbook.active)
    usados: set[str] = set()
    fecha_texto = _fecha_ruta_excel(payload.fecha)

    for repartidor in sorted(set(rutas_entregas) | set(rutas_cobranza), key=lambda item: item.lower()):
        ws = workbook.create_sheet(title=_safe_sheet_title(repartidor, usados))

        ws["A3"] = "Fecha:"; ws["A3"].font = bold_font
        ws["B3"] = fecha_texto; ws["B3"].font = bold_font
        ws["A4"] = "REPARTIDOR:"; ws["A4"].font = bold_font
        ws["B4"] = repartidor; ws["B4"].font = bold_font
        ws["A5"] = "UNIDAD:"; ws["A5"].font = bold_font
        ws["B5"] = unidades_repartidor.get(repartidor.lower(), ""); ws["B5"].font = bold_font
        ws["D3"] = "HORA SALIDA A RUTA"; ws["D3"].font = bold_font
        ws["D4"] = "KM INICIAL:________________________KM FINAL:_____________________"; ws["D4"].font = bold_font
        ws["D5"] = "ACOMODAR Y LAVAR LOS COCHES: LUNES MIERCOLES Y VIERNES"
        ws["D6"] = "REVISAR NIVELES DE ACEITE, ANTICONGELANTE Y PRESION DE LLANTAS"
        ws.merge_cells("D3:E3")
        ws.merge_cells("D4:E4")
        ws.merge_cells("D5:E5")
        ws.merge_cells("D6:E6")
        ws.merge_cells("F2:H3")
        ws.merge_cells("F6:H7")
        for rango in ("F2:H3", "F6:H7"):
            for fila in ws[rango]:
                for cell in fila:
                    cell.border = borde_negro
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        def escribir_encabezado(row_idx: int) -> None:
            for col_idx, col_name in enumerate(columnas, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=col_name)
                cell.font = encabezado_font
                cell.fill = encabezado_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fila_inicio = 9
        escribir_encabezado(fila_inicio)
        for index, row in enumerate(rutas_entregas.get(repartidor, []), start=fila_inicio + 1):
            valores = [
                _valor_ruta(row, "folio", "comanda"),
                _valor_ruta(row, "factura"),
                _valor_ruta(row, "cliente_numero"),
                _valor_ruta(row, "cliente_nombre"),
                _valor_ruta(row, "productos"),
                _valor_ruta(row, "repartidor"),
                _valor_ruta(row, "horario"),
                _valor_ruta(row, "observaciones"),
            ]
            for col_idx, valor in enumerate(valores, start=1):
                ws.cell(row=index, column=col_idx, value=valor)

        fila_cobranza = fila_inicio + len(rutas_entregas.get(repartidor, [])) + 2
        escribir_encabezado(fila_cobranza)
        for index, row in enumerate(rutas_cobranza.get(repartidor, []), start=fila_cobranza + 1):
            valores = [
                _valor_ruta(row, "folio", "comanda"),
                _valor_ruta(row, "factura"),
                _valor_ruta(row, "cliente_numero"),
                _valor_ruta(row, "cliente_nombre"),
                "",
                _valor_ruta(row, "repartidor"),
                _valor_ruta(row, "horario"),
                _valor_ruta(row, "observaciones"),
            ]
            for col_idx, valor in enumerate(valores, start=1):
                ws.cell(row=index, column=col_idx, value=valor)

        fila_nota = fila_cobranza + len(rutas_cobranza.get(repartidor, [])) + 2
        ws.merge_cells(start_row=fila_nota, start_column=1, end_row=fila_nota, end_column=8)
        nota = ws.cell(row=fila_nota, column=1)
        nota.value = (
            "NOTA: LLAMAR Y REPORTAR CUALQUIER INCIDENCIA QUE SE PRESENTE EN LA ENTREGA O COBRO "
            "(ANOTAR EN LAS OBSERVACIONES CUALQUIER DATO RELEVANTE RELACIONADO CON LOS CLIENTES, FACTURACION, COBRO, ETC.)"
        )
        nota.font = bold_font
        nota.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila_nota].height = 40

        for col_name, width in anchos.items():
            col_idx = columnas.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            for row_idx in range(fila_inicio, ws.max_row + 1):
                if row_idx == fila_nota:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is None:
                    continue
                horizontal = "center" if col_name in {"Comanda", "Factura", "Cliente Nº", "Nombre Cliente", "Repartidor", "Horario"} else "left"
                cell.alignment = Alignment(wrap_text=True, horizontal=horizontal, vertical="top")
                line_count = max(1, math.ceil(len(str(cell.value)) / (width or 1)))
                ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 15, line_count * 15)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = f"rutas_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/por-facturar/resumen")
def resumen_por_facturar(user: dict = Depends(require_user)):
    hoy = datetime.now().date()
    inicio = hoy.fromordinal(hoy.toordinal() - (3 if hoy.weekday() == 0 else 1))
    conn = get_editor_connection(); cur = conn.cursor()
    try:
        cur.execute("""SELECT folio FROM comandas_editables
                       WHERE COALESCE(estatus_facturado, 0)=0
                         AND DATE(COALESCE(creado_en, fecha)) BETWEEN %s AND %s""", (inicio, hoy))
        folios = [str(row[0]).strip() for row in cur.fetchall() if row and row[0] is not None]
    finally:
        cur.close(); conn.close()
    facturados = set()
    if folios:
        legacy_conn = get_legacy_connection(); legacy_cur = legacy_conn.cursor()
        try:
            marks = ",".join(["%s"] * len(folios))
            legacy_cur.execute(f"SELECT comanda FROM facturas WHERE comanda IS NOT NULL AND TRIM(CAST(comanda AS CHAR)) IN ({marks})", tuple(folios))
            facturados = {str(row[0]).strip() for row in legacy_cur.fetchall() if row and row[0] is not None}
        finally:
            legacy_cur.close(); legacy_conn.close()
    return {"pendientes": sum(1 for folio in folios if folio not in facturados), "fecha_inicio": inicio.isoformat(), "fecha_fin": hoy.isoformat()}


@router.get("/por-facturar/{folio}/detalle")
def detalle_por_facturar(folio: str, user: dict = Depends(require_user)):
    conn = get_editor_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM comandas_editables WHERE folio=%s ORDER BY id DESC LIMIT 1", (folio.strip(),))
        comanda = cur.fetchone()
        if not comanda: raise HTTPException(status_code=404, detail="Pedido por facturar no encontrado.")
        cur.execute("SELECT cip, descripcion, kgs, piezas, COALESCE(observaciones,'') AS observaciones FROM comandas_editables_detalle WHERE comanda_id=%s ORDER BY id", (comanda["id"],))
        comanda["productos"] = cur.fetchall() or []
        legacy_conn = get_legacy_connection(); legacy_cur = legacy_conn.cursor(dictionary=True)
        try:
            legacy_cur.execute("SELECT COALESCE(observaciones,'') AS observaciones FROM clientes WHERE empresa=%s AND numero=%s LIMIT 1", (comanda.get("empresa") or "", comanda.get("cliente_numero") or ""))
            cliente = legacy_cur.fetchone() or {}
            comanda["observaciones_cliente"] = cliente.get("observaciones", "")
        finally:
            legacy_cur.close(); legacy_conn.close()
        return comanda
    finally:
        cur.close(); conn.close()


@router.post("/por-facturar/estatus")
def cambiar_estatus_por_facturar(payload: EstatusComandaIn, user: dict = Depends(require_user)):
    conn = get_editor_connection(); cur = conn.cursor()
    try:
        cur.execute("UPDATE comandas_editables SET estatus_facturado=%s, actualizado_en=%s WHERE folio=%s", (payload.estatus, datetime.now(), payload.folio.strip()))
        if cur.rowcount != 1: raise HTTPException(status_code=404, detail="Pedido por facturar no encontrado.")
        cur.execute("SELECT id FROM comandas_editables WHERE folio=%s LIMIT 1", (payload.folio.strip(),))
        row = cur.fetchone()
        if row:
            etiqueta = "Facturado" if payload.estatus == 1 else "Cancelado"
            cur.execute("INSERT INTO comandas_editables_historial (comanda_id, usuario, accion, detalle, fecha) VALUES (%s,%s,%s,%s,%s)", (row[0], str(user.get("username") or "usuario"), "Cambio de estatus", etiqueta, datetime.now()))
        conn.commit()
        return {"ok": True, "folio": payload.folio.strip(), "estatus": payload.estatus}
    except HTTPException:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/por-facturar/eliminar")
def eliminar_por_facturar(payload: FoliosIn, user: dict = Depends(require_user)):
    folios = list(dict.fromkeys(f.strip() for f in payload.folios if f.strip()))
    if not folios: raise HTTPException(status_code=400, detail="No se indicaron pedidos.")
    conn = get_editor_connection(); cur = conn.cursor()
    try:
        marks = ",".join(["%s"] * len(folios))
        cur.execute(f"SELECT id FROM comandas_editables WHERE folio IN ({marks})", tuple(folios))
        ids = [row[0] for row in cur.fetchall()]
        if ids:
            detail_marks = ",".join(["%s"] * len(ids))
            cur.execute(f"DELETE FROM comandas_editables_detalle WHERE comanda_id IN ({detail_marks})", tuple(ids))
            cur.execute(f"DELETE FROM comandas_editables_historial WHERE comanda_id IN ({detail_marks})", tuple(ids))
            cur.execute(f"DELETE FROM comandas_editables WHERE id IN ({detail_marks})", tuple(ids))
        conn.commit()
        return {"ok": True, "eliminados": len(ids)}
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=500, detail=f"No se pudieron eliminar los pedidos: {exc}")
    finally:
        cur.close(); conn.close()


@router.post("/diario/eliminar")
def eliminar_diario(payload: FoliosIn, user: dict = Depends(require_user)):
    folios = list(dict.fromkeys(f.strip() for f in payload.folios if f.strip()))
    if not folios: raise HTTPException(status_code=400, detail="No se indicaron comandas.")
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        marks = ",".join(["%s"] * len(folios))
        cur.execute(f"SELECT id FROM comandas WHERE folio IN ({marks})", tuple(folios))
        ids = [row[0] for row in cur.fetchall()]
        if ids:
            detail_marks = ",".join(["%s"] * len(ids))
            cur.execute(f"DELETE FROM productos_comanda WHERE comanda_id IN ({detail_marks})", tuple(ids))
            cur.execute(f"DELETE FROM comandas WHERE id IN ({detail_marks})", tuple(ids))
        conn.commit()
        return {"ok": True, "eliminados": len(ids)}
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=500, detail=f"No se pudieron eliminar las comandas: {exc}")
    finally:
        cur.close(); conn.close()


@router.post("/guardar")
def guardar_comanda(payload: ComandaGuardarIn, user: dict = Depends(require_user)):
    productos = [p for p in payload.productos if p.cip.strip() and p.descripcion.strip() and (p.kgs > 0 or p.piezas > 0)]
    if not productos: raise HTTPException(status_code=400, detail="Agrega al menos un producto con kilos o piezas.")
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        folio = payload.folio.strip()
        if not folio:
            cur.execute("SELECT COALESCE(MAX(CAST(folio AS UNSIGNED)), 0) + 1 FROM comandas WHERE folio REGEXP '^[0-9]+$'")
            folio = str((cur.fetchone() or [1])[0] or 1)
        cur.execute("SELECT id FROM comandas WHERE folio=%s LIMIT 1", (folio,))
        if cur.fetchone(): raise HTTPException(status_code=409, detail=f"El folio {folio} ya existe.")
        vendedor = payload.vendedor.strip() or str(user.get('full_name') or user.get('username') or '')
        cur.execute("""INSERT INTO comandas (folio,vendedor,empresa,cliente_numero,cliente_nombre,fecha,observaciones_pedido)
                       VALUES (%s,%s,%s,%s,%s,%s,%s)""", (folio,vendedor,payload.empresa.strip(),payload.cliente_numero.strip(),payload.cliente_nombre.strip(),datetime.now(),payload.observaciones_pedido.strip()))
        comanda_id = cur.lastrowid
        for p in productos:
            cur.execute("INSERT INTO productos_comanda (comanda_id,cip,descripcion,kgs,piezas,observaciones) VALUES (%s,%s,%s,%s,%s,%s)", (comanda_id,p.cip.strip(),p.descripcion.strip(),p.kgs,p.piezas,p.observaciones.strip()))
        conn.commit()
    except HTTPException:
        conn.rollback(); raise
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=500, detail=f"No se pudo guardar la comanda: {exc}")
    finally:
        cur.close(); conn.close()
    return {"ok": True, "id": comanda_id, "folio": folio, "productos": len(productos)}


# --- Catálogos: mismas operaciones que el proyecto de escritorio Comandas06 ---
@router.post("/clientes")
def crear_cliente(payload: ClienteIn, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        _asegurar_columna_revision(cur)
        cur.execute("SELECT id FROM clientes WHERE empresa=%s AND numero=%s LIMIT 1", (payload.empresa.strip(), payload.numero.strip()))
        if cur.fetchone():
            raise HTTPException(status_code=409, detail="Ya existe un cliente con ese número en la empresa indicada.")
        cur.execute("""INSERT INTO clientes (numero,nombre,empresa,direccion_entrega,observaciones,dias_credito,contacto1,revision)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""", (payload.numero.strip(), payload.nombre.strip(), payload.empresa.strip(), payload.direccion_entrega.strip(), payload.observaciones.strip(), payload.dias_credito.strip(), payload.contacto1.strip(), payload.revision.strip()))
        conn.commit(); return {"ok": True, "id": cur.lastrowid}
    except HTTPException:
        conn.rollback(); raise
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=500, detail=f"No se pudo guardar el cliente: {exc}") from exc
    finally:
        cur.close(); conn.close()


@router.put("/clientes")
def editar_cliente(payload: ClienteIn, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        _asegurar_columna_revision(cur)
        anterior_numero = payload.numero_original.strip() or payload.numero.strip()
        anterior_empresa = payload.empresa_original.strip() or payload.empresa.strip()
        cur.execute("""UPDATE clientes SET numero=%s,nombre=%s,empresa=%s,direccion_entrega=%s,observaciones=%s,
                       dias_credito=%s,contacto1=%s,revision=%s WHERE empresa=%s AND numero=%s""", (payload.numero.strip(), payload.nombre.strip(), payload.empresa.strip(), payload.direccion_entrega.strip(), payload.observaciones.strip(), payload.dias_credito.strip(), payload.contacto1.strip(), payload.revision.strip(), anterior_empresa, anterior_numero))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        conn.commit(); return {"ok": True}
    except HTTPException:
        conn.rollback(); raise
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=500, detail=f"No se pudo actualizar el cliente: {exc}") from exc
    finally:
        cur.close(); conn.close()


@router.delete("/clientes")
def eliminar_cliente(empresa: str, numero: str, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM clientes WHERE empresa=%s AND numero=%s", (empresa.strip(), numero.strip()))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Cliente no encontrado.")
        conn.commit(); return {"ok": True}
    except HTTPException:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.post("/productos")
def crear_producto(payload: ProductoIn, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute("INSERT INTO productos (cip,descripcion,unidad) VALUES (%s,%s,%s)", (payload.cip.strip(), payload.descripcion.strip(), payload.unidad.strip()))
        conn.commit(); return {"ok": True}
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=409, detail=f"No se pudo agregar el producto: {exc}") from exc
    finally:
        cur.close(); conn.close()


@router.put("/productos")
def editar_producto(payload: ProductoIn, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        original = payload.cip_original.strip() or payload.cip.strip()
        cur.execute("UPDATE productos SET cip=%s, descripcion=%s, unidad=%s WHERE cip=%s", (payload.cip.strip(), payload.descripcion.strip(), payload.unidad.strip(), original))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Producto no encontrado.")
        conn.commit(); return {"ok": True}
    except HTTPException:
        conn.rollback(); raise
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=409, detail=f"No se pudo actualizar el producto: {exc}") from exc
    finally:
        cur.close(); conn.close()


@router.delete("/productos")
def eliminar_producto(cip: str, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM productos WHERE cip=%s", (cip.strip(),))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Producto no encontrado.")
        conn.commit(); return {"ok": True}
    except HTTPException:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


def _tabla_catalogo(tipo: str) -> str:
    tablas = {"vendedores": "vendedores", "repartidores": "repartidores"}
    if tipo not in tablas: raise HTTPException(status_code=404, detail="Catálogo no soportado.")
    return tablas[tipo]


@router.get("/catalogos/{tipo}")
def listar_catalogo(tipo: str, user: dict = Depends(require_user)):
    tabla = _tabla_catalogo(tipo); conn = get_legacy_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute(f"SELECT id,nombre FROM {tabla} WHERE TRIM(COALESCE(nombre,''))<>'' ORDER BY id")
        return cur.fetchall() or []
    finally:
        cur.close(); conn.close()


@router.post("/catalogos/{tipo}")
def crear_catalogo(tipo: str, payload: NombreCatalogoIn, user: dict = Depends(require_user)):
    tabla = _tabla_catalogo(tipo); conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute(f"INSERT INTO {tabla} (nombre) VALUES (%s)", (payload.nombre.strip(),))
        conn.commit(); return {"ok": True, "id": cur.lastrowid}
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=409, detail=f"No se pudo agregar: {exc}") from exc
    finally:
        cur.close(); conn.close()


@router.put("/catalogos/{tipo}")
def editar_catalogo(tipo: str, payload: NombreCatalogoIn, user: dict = Depends(require_user)):
    if not payload.id: raise HTTPException(status_code=400, detail="Falta el identificador.")
    tabla = _tabla_catalogo(tipo); conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute(f"UPDATE {tabla} SET nombre=%s WHERE id=%s", (payload.nombre.strip(), payload.id))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Registro no encontrado.")
        conn.commit(); return {"ok": True}
    except HTTPException:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.delete("/catalogos/{tipo}")
def eliminar_catalogo(tipo: str, id: int, user: dict = Depends(require_user)):
    tabla = _tabla_catalogo(tipo); conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute(f"DELETE FROM {tabla} WHERE id=%s", (id,))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Registro no encontrado.")
        conn.commit(); return {"ok": True}
    except HTTPException:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/pedidos-vendedor/{pedido_id}")
def detalle_pedido_vendedor(pedido_id: int, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM pedidos_vendedor WHERE id=%s LIMIT 1", (pedido_id,)); pedido = cur.fetchone()
        if not pedido: raise HTTPException(status_code=404, detail="Pedido no encontrado.")
        cur.execute("SELECT cip,descripcion,kgs,piezas,COALESCE(observaciones,'') AS observaciones FROM pedidos_vendedor_detalle WHERE pedido_id=%s ORDER BY id", (pedido_id,))
        pedido["productos"] = cur.fetchall() or []
        return pedido
    finally:
        cur.close(); conn.close()


@router.post("/pedidos-vendedor/procesar")
def procesar_pedido_vendedor(payload: PedidoVendedorIn, user: dict = Depends(require_user)):
    """Convierte el pedido móvil a Comanda oficial igual que Comandas06."""
    conn = get_legacy_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT * FROM pedidos_vendedor WHERE id=%s FOR UPDATE", (payload.id,)); pedido = cur.fetchone()
        if not pedido: raise HTTPException(status_code=404, detail="Pedido no encontrado.")
        if str(pedido.get("estado") or "").upper() == "CANCELADO": raise HTTPException(status_code=400, detail="El pedido está cancelado.")
        if pedido.get("comanda_id"):
            cur.execute("SELECT folio FROM comandas WHERE id=%s", (pedido["comanda_id"],)); actual = cur.fetchone()
            conn.commit(); return {"ok": True, "folio": (actual or {}).get("folio", pedido.get("folio_usado", "")), "existente": True}
        cur.execute("SELECT cip,descripcion,kgs,piezas,COALESCE(observaciones,'') AS observaciones FROM pedidos_vendedor_detalle WHERE pedido_id=%s ORDER BY id", (payload.id,)); detalle = cur.fetchall() or []
        if not detalle: raise HTTPException(status_code=400, detail="El pedido no tiene productos.")
        cur.execute("SELECT COALESCE(MAX(CAST(folio AS UNSIGNED)),0)+1 AS siguiente FROM comandas WHERE folio REGEXP '^[0-9]+$'"); folio = str((cur.fetchone() or {}).get("siguiente") or 1)
        ahora = datetime.now()
        cur.execute("""INSERT INTO comandas (folio,vendedor,empresa,cliente_numero,cliente_nombre,fecha,observaciones_pedido)
                       VALUES (%s,%s,%s,%s,%s,%s,%s)""", (folio,pedido.get("vendedor") or "",pedido.get("empresa") or "",pedido.get("cliente_numero") or "",pedido.get("cliente_nombre") or "",ahora,pedido.get("observaciones_pedido") or ""))
        comanda_id = cur.lastrowid
        for item in detalle:
            cur.execute("INSERT INTO productos_comanda (comanda_id,cip,descripcion,kgs,piezas,observaciones) VALUES (%s,%s,%s,%s,%s,%s)", (comanda_id,item.get("cip") or "",item.get("descripcion") or "",item.get("kgs") or 0,item.get("piezas") or 0,item.get("observaciones") or ""))
        cur.execute("""UPDATE pedidos_vendedor SET estado='PROCESADO',procesado_en=%s,procesado_por=%s,comanda_id=%s,folio_usado=%s,folio_usado_en=%s WHERE id=%s""", (ahora,str(user.get("username") or "usuario"),comanda_id,folio,ahora,payload.id))
        conn.commit(); return {"ok": True, "folio": folio, "comanda_id": comanda_id}
    except HTTPException:
        conn.rollback(); raise
    except Exception as exc:
        conn.rollback(); raise HTTPException(status_code=500, detail=f"No se pudo procesar el pedido: {exc}") from exc
    finally:
        cur.close(); conn.close()


@router.post("/pedidos-vendedor/estatus")
def cambiar_estatus_pedido_vendedor(payload: PedidoVendedorIn, user: dict = Depends(require_user)):
    estado = payload.estado.strip().upper()
    if estado not in {"PENDIENTE", "PROCESADO", "CANCELADO"}: raise HTTPException(status_code=400, detail="Estatus inválido.")
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute("UPDATE pedidos_vendedor SET estado=%s WHERE id=%s", (estado, payload.id))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Pedido no encontrado.")
        conn.commit(); return {"ok": True, "estado": estado}
    except HTTPException:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.delete("/pedidos-vendedor/{pedido_id}")
def eliminar_pedido_vendedor(pedido_id: int, user: dict = Depends(require_user)):
    conn = get_legacy_connection(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM pedidos_vendedor_detalle WHERE pedido_id=%s", (pedido_id,))
        cur.execute("DELETE FROM pedidos_vendedor WHERE id=%s", (pedido_id,))
        if cur.rowcount < 1: raise HTTPException(status_code=404, detail="Pedido no encontrado.")
        conn.commit(); return {"ok": True}
    except HTTPException:
        conn.rollback(); raise
    finally:
        cur.close(); conn.close()


@router.get("/catalogos/{tipo}/plantilla")
def plantilla_catalogo(tipo: str, user: dict = Depends(require_user)):
    columnas = {
        "clientes": ("Numero", "Nombre", "Empresa", "Direccion de Entrega", "Observaciones", "Dias de credito", "Contacto", "Revision"),
        "productos": ("CIP", "Descripcion", "Unidad"),
    }.get(tipo)
    if not columnas: raise HTTPException(status_code=404, detail="Plantilla no soportada.")
    libro = Workbook(); hoja = libro.active; hoja.title = tipo.title()
    for indice, valor in enumerate(columnas, 1):
        celda = hoja.cell(1, indice, valor); celda.font = Font(bold=True)
        hoja.column_dimensions[celda.column_letter].width = max(16, len(valor) + 5)
    salida = BytesIO(); libro.save(salida); salida.seek(0)
    return StreamingResponse(salida, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="plantilla_{tipo}.xlsx"'})


@router.post("/catalogos/{tipo}/importar")
async def importar_catalogo(tipo: str, archivo: UploadFile = File(...), user: dict = Depends(require_user)):
    if tipo not in {"clientes", "productos"}: raise HTTPException(status_code=404, detail="Importación no soportada.")
    if not (archivo.filename or "").lower().endswith(".xlsx"): raise HTTPException(status_code=400, detail="Selecciona un archivo .xlsx.")
    try:
        libro = load_workbook(BytesIO(await archivo.read()), data_only=True, read_only=True); hoja = libro.active
        filas = list(hoja.iter_rows(values_only=True))
        if not filas: raise HTTPException(status_code=400, detail="El archivo no contiene filas.")
        encabezados = {_etiqueta_excel(v): i for i, v in enumerate(filas[0]) if _etiqueta_excel(v)}
        conn = get_legacy_connection(); cur = conn.cursor(); cargados = 0
        try:
            if tipo == "clientes":
                requeridos = ("NUMERO", "NOMBRE", "EMPRESA")
                if any(c not in encabezados for c in requeridos): raise HTTPException(status_code=400, detail="La plantilla debe incluir Numero, Nombre y Empresa.")
                for fila in filas[1:]:
                    valor = lambda nombre: _texto_excel(fila[encabezados[nombre]] if encabezados.get(nombre) is not None and encabezados[nombre] < len(fila) else "")
                    numero, nombre, empresa = valor("NUMERO"), valor("NOMBRE"), valor("EMPRESA")
                    if not (numero and nombre and empresa): continue
                    direccion = valor("DIRECCION DE ENTREGA"); observaciones = valor("OBSERVACIONES"); pago = valor("DIAS DE CREDITO"); contacto = valor("CONTACTO"); revision = valor("REVISION")
                    cur.execute("SELECT id FROM clientes WHERE empresa=%s AND numero=%s", (empresa, numero)); existe = cur.fetchone()
                    if existe:
                        cur.execute("UPDATE clientes SET nombre=%s,direccion_entrega=%s,observaciones=%s,dias_credito=%s,contacto1=%s,revision=%s WHERE empresa=%s AND numero=%s", (nombre,direccion,observaciones,pago,contacto,revision,empresa,numero))
                    else:
                        cur.execute("INSERT INTO clientes (numero,nombre,empresa,direccion_entrega,observaciones,dias_credito,contacto1,revision) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)", (numero,nombre,empresa,direccion,observaciones,pago,contacto,revision))
                    cargados += 1
            else:
                if "CIP" not in encabezados or "DESCRIPCION" not in encabezados: raise HTTPException(status_code=400, detail="La plantilla debe incluir CIP y Descripcion.")
                for fila in filas[1:]:
                    cip = _texto_excel(fila[encabezados["CIP"]] if encabezados["CIP"] < len(fila) else ""); descripcion = _texto_excel(fila[encabezados["DESCRIPCION"]] if encabezados["DESCRIPCION"] < len(fila) else "")
                    unidad = _texto_excel(fila[encabezados["UNIDAD"]] if encabezados.get("UNIDAD") is not None and encabezados["UNIDAD"] < len(fila) else "")
                    if not (cip and descripcion): continue
                    cur.execute("SELECT cip FROM productos WHERE cip=%s", (cip,))
                    if cur.fetchone(): cur.execute("UPDATE productos SET descripcion=%s,unidad=%s WHERE cip=%s", (descripcion,unidad,cip))
                    else: cur.execute("INSERT INTO productos (cip,descripcion,unidad) VALUES (%s,%s,%s)", (cip,descripcion,unidad))
                    cargados += 1
            conn.commit()
        except HTTPException:
            conn.rollback(); raise
        finally:
            cur.close(); conn.close()
        return {"ok": True, "cargados": cargados}
    except HTTPException: raise
    except Exception as exc: raise HTTPException(status_code=400, detail=f"No se pudo importar el archivo: {exc}") from exc


@router.get("/{folio}")
def obtener_comanda(folio: str):
    conn = None
    cursor = None
    try:
        folio = str(folio or "").strip()
        if not folio:
            raise HTTPException(status_code=400, detail="Folio vacío")

        conn = get_editor_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT
                id,
                folio,
                empresa,
                cliente_numero,
                cliente_nombre,
                IFNULL(observaciones_pedido, '') AS observaciones_pedido
            FROM comandas_editables
            WHERE folio = %s
            ORDER BY id DESC
            LIMIT 1
        """,
            (folio,),
        )
        cab = cursor.fetchone()

        if not cab:
            raise HTTPException(
                status_code=404,
                detail=f"Comanda {folio} no encontrada en el editor",
            )

        comanda_id = cab["id"]

        cursor.execute(
            """
            SELECT
                IFNULL(cip, '') AS cip,
                IFNULL(kgs, 0) AS kgs,
                IFNULL(piezas, 0) AS piezas,
                IFNULL(descripcion, '') AS descripcion,
                IFNULL(observaciones, '') AS observaciones
            FROM comandas_editables_detalle
            WHERE comanda_id = %s
            ORDER BY id ASC
        """,
            (comanda_id,),
        )
        detalle = cursor.fetchall() or []

        productos = []
        for r in detalle:
            productos.append(
                {
                    "cip": str(r.get("cip") or "").strip(),
                    "kgs": float(r.get("kgs") or 0),
                    "piezas": float(r.get("piezas") or 0),
                    "descripcion": str(r.get("descripcion") or "").strip(),
                    "observaciones": str(r.get("observaciones") or "").strip(),
                }
            )

        return {
            "folio": str(cab.get("folio") or "").strip(),
            "empresa": str(cab.get("empresa") or "").strip(),
            "cliente_numero": str(cab.get("cliente_numero") or "").strip(),
            "cliente_nombre": str(cab.get("cliente_nombre") or "").strip(),
            "observaciones_pedido": str(cab.get("observaciones_pedido") or "").strip(),
            "productos": productos,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            if cursor:
                cursor.close()
        except Exception:
            pass
        try:
            if conn:
                conn.close()
        except Exception:
            pass
