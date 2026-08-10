import csv
import io
import json
import os
import time
import uuid
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path
from typing import Optional

import mysql.connector
import requests
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from pydantic import BaseModel

from app.core.config import settings
from app.dependencies import require_user
from app.legacy_db import get_legacy_connection


def _ensure_crm_schema():
    """Auto-migrate CRM tables if missing columns."""
    import traceback
    db = None
    try:
        db = _crm_db()
        cur = db.cursor(dictionary=True)
        # Migrate crm_quotes
        cur.execute("SHOW COLUMNS FROM crm_quotes")
        existing = {row["Field"] for row in cur.fetchall()}
        needed = [
            ("subtotal", "DECIMAL(12,2) NOT NULL DEFAULT 0", "authorized_shipping"),
            ("discount", "DECIMAL(12,2) NOT NULL DEFAULT 0", "subtotal"),
            ("tax", "DECIMAL(12,2) NOT NULL DEFAULT 0", "discount"),
            ("prospect_id", "VARCHAR(255) DEFAULT ''", "client_id"),
            ("authorized_shipping", "DECIMAL(12,2) NOT NULL DEFAULT 0", "quote_recipient"),
            ("client_snapshot", "TEXT", "items"),
            ("notes", "TEXT", "client_snapshot"),
            ("valid_until", "DATE DEFAULT NULL", "client_snapshot"),
            ("show_shipping", "TINYINT(1) NOT NULL DEFAULT 1", "valid_until"),
        ]
        for col, col_type, after in needed:
            if col not in existing:
                cur.execute(f"ALTER TABLE crm_quotes ADD COLUMN {col} {col_type} AFTER {after}")
        db.commit()
        # Migrate crm_bank_accounts
        cur.execute("SHOW COLUMNS FROM crm_bank_accounts")
        bank_cols = {row["Field"] for row in cur.fetchall()}
        bank_needed = [
            ("company", "VARCHAR(255) DEFAULT ''", "id"),
            ("beneficiary", "VARCHAR(255) DEFAULT ''", "reference"),
            ("enabled", "TINYINT(1) NOT NULL DEFAULT 1", "beneficiary"),
        ]
        for col, col_type, after in bank_needed:
            if col not in bank_cols:
                cur.execute(f"ALTER TABLE crm_bank_accounts ADD COLUMN {col} {col_type} AFTER {after}")
        db.commit()
        with open(settings.base_dir / "_schema.log", "a") as f:
            f.write(f"Schema OK. Existing cols: {sorted(existing)}\n")
    except Exception as e:
        with open(settings.base_dir / "_schema.log", "a") as f:
            f.write(f"Schema error: {e}\n{traceback.format_exc()}\n")
    finally:
        if db:
            try:
                db.close()
            except Exception:
                pass


def _crm_db():
    """Usa el mismo failover MySQL del legado para todos los módulos web."""
    return get_legacy_connection()


def _legacy_db():
    return get_legacy_connection()


def _row(r):
    if r is None:
        return None
    if isinstance(r, dict):
        return {k: _fmt_val(v) for k, v in r.items()}
    # tuple from mysql-connector — use cursor.column_names
    raise ValueError("Use dictionary=True cursor for _row()")


def _fmt_val(v):
    if isinstance(v, bytes):
        return v.decode("utf-8", errors="replace")
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, float) and v != v:
        return None
    return v


def _rows(rs):
    return [_row(r) for r in rs]


def _now():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _dict_to_str(d):
    if isinstance(d, dict):
        return json.dumps(d, ensure_ascii=False)
    return d or ""


router = APIRouter(prefix="/api/crm", tags=["crm"])

_ensure_crm_schema()


@router.get("/clients")
def search_clients(
    q: str = Query(default=""),
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        search = f"%{q.strip()}%"
        cur.execute(
            """
            SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), c.assigned_user_id) AS assigned_user
            FROM crm_clients c
            WHERE c.name LIKE %s
               OR c.code LIKE %s
               OR c.email LIKE %s
               OR c.contact_name LIKE %s
               OR c.phone LIKE %s
               OR c.external_seller LIKE %s
            ORDER BY c.updated_at DESC
            """,
            (search, search, search, search, search, search),
        )
        return _rows(cur.fetchall())
    finally:
        db.close()


@router.get("/client/{client_id}")
def get_client(
    client_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute(
            "SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), c.assigned_user_id) AS assigned_user FROM crm_clients c WHERE c.id = %s",
            (client_id,),
        )
        client = _row(cur.fetchone())
        if not client:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")

        # invoices from facturas table using client code (MySQL legacy)
        client_code = (client.get("code") or "").strip()
        invoices = []
        if client_code:
            try:
                leg = _legacy_db()
                lcur = leg.cursor(dictionary=True)
                lcur.execute(
                    """
                    SELECT id, factura AS folio, fecha AS issued_at, subtotal, descuento_pct AS discount,
                           iva AS tax, total, estatus AS status, timbrado_estatus, cfdi_uuid,
                           empresa, vendedor
                    FROM facturas
                    WHERE numero_cliente = %s
                    ORDER BY fecha DESC
                    LIMIT 100
                    """,
                    (client_code,),
                )
                invoices = _rows(lcur.fetchall())
                leg.close()
            except Exception:
                invoices = []

        cur.execute(
            "SELECT * FROM crm_followups WHERE client_id = %s ORDER BY contact_at DESC",
            (client_id,),
        )
        followups = _rows(cur.fetchall())

        cur.execute(
            "SELECT * FROM crm_activity_log WHERE client_id = %s ORDER BY created_at DESC",
            (client_id,),
        )
        activity = _rows(cur.fetchall())

        # products from legacy MySQL
        products = []
        if client_code:
            try:
                leg = _legacy_db()
                lcur = leg.cursor(dictionary=True)
                lcur.execute(
                    """
                    SELECT fd.cip, COALESCE(p.descripcion, fd.descripcion) AS description,
                           COUNT(DISTINCT fd.factura_id) AS invoice_count,
                           MAX(f.fecha) AS last_purchase,
                           SUM(fd.importe) AS total_amount,
                           SUM(fd.cantidad) AS total_quantity
                    FROM factura_detalle fd
                    JOIN facturas f ON f.id = fd.factura_id
                    LEFT JOIN productos p ON p.cip = fd.cip
                    WHERE f.numero_cliente = %s
                    GROUP BY fd.cip
                    ORDER BY total_amount DESC
                    LIMIT 50
                    """,
                    (client_code,),
                )
                products = _rows(lcur.fetchall())
                leg.close()
            except Exception:
                products = []

        cur.execute("SELECT * FROM crm_quotes WHERE client_id = %s ORDER BY created_at DESC", (client_id,))
        quotes = _rows(cur.fetchall())

        return {
            "client": client,
            "invoices": invoices,
            "followups": followups,
            "activity": activity,
            "quotes": quotes,
            "products": products,
        }
    finally:
        db.close()


class ClientCreate(BaseModel):
    id: Optional[str] = None
    code: Optional[str] = ""
    name: str
    tax_address: Optional[str] = ""
    consignee_address: Optional[str] = ""
    delivery_method: Optional[str] = ""
    phone: Optional[str] = ""
    email: Optional[str] = ""
    contact_name: Optional[str] = ""
    external_seller: Optional[str] = ""
    assigned_user_id: Optional[str] = ""


@router.post("/clients")
def upsert_client(
    payload: ClientCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        client_id = payload.id or str(uuid.uuid4())
        cur.execute("SELECT * FROM crm_clients WHERE id = %s", (client_id,))
        existing = cur.fetchone()
        if existing:
            cur.execute(
                """
                UPDATE crm_clients SET code=%s, name=%s, tax_address=%s, consignee_address=%s,
                delivery_method=%s, phone=%s, email=%s, contact_name=%s,
                external_seller=%s, assigned_user_id=%s, updated_at=%s
                WHERE id=%s
                """,
                (
                    payload.code, payload.name, payload.tax_address,
                    payload.consignee_address, payload.delivery_method,
                    payload.phone, payload.email, payload.contact_name,
                    payload.external_seller, payload.assigned_user_id, now, client_id,
                ),
            )
            changed = {}
            field_map = {"code": "Codigo", "name": "Nombre", "tax_address": "Direccion fiscal",
                         "consignee_address": "Direccion consignatario", "delivery_method": "Paqueteria",
                         "phone": "Telefono", "email": "Email", "contact_name": "Contacto",
                         "external_seller": "Vendedor externo", "assigned_user_id": "Usuario asignado"}
            new_vals = {"code": payload.code, "name": payload.name, "tax_address": payload.tax_address,
                        "consignee_address": payload.consignee_address, "delivery_method": payload.delivery_method,
                        "phone": payload.phone, "email": payload.email, "contact_name": payload.contact_name,
                        "external_seller": payload.external_seller, "assigned_user_id": payload.assigned_user_id}
            for k, label in field_map.items():
                old = str(existing[k] or "") if k in existing.keys() else ""
                new = str(new_vals.get(k, "") or "")
                if old != new:
                    changed[label] = f"{old} -> {new}"
            if changed:
                cur.execute(
                    "INSERT INTO crm_activity_log (id, client_id, user_id, type, title, payload, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (str(uuid.uuid4()), client_id, str(user["full_name"]), "client_edit",
                     "Cliente actualizado", json.dumps({"changed": changed}, ensure_ascii=False), now),
                )
        else:
            cur.execute(
                """
                INSERT INTO crm_clients (id, code, name, tax_address, consignee_address,
                delivery_method, phone, email, contact_name, external_seller,
                assigned_user_id, created_at, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                (
                    client_id, payload.code, payload.name, payload.tax_address,
                    payload.consignee_address, payload.delivery_method,
                    payload.phone, payload.email, payload.contact_name,
                    payload.external_seller, payload.assigned_user_id, now, now,
                ),
            )
        db.commit()
        return {"ok": True, "id": client_id}
    finally:
        db.close()


class FollowupCreate(BaseModel):
    client_id: str
    contact_at: str
    channel: str = "Llamada"
    outcome: str
    next_action: Optional[str] = ""
    next_action_at: Optional[str] = ""
    notes: Optional[str] = ""


@router.post("/followups")
def create_followup(
    payload: FollowupCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        followup_id = str(uuid.uuid4())
        cur.execute(
            "INSERT INTO crm_followups (id, client_id, user_id, contact_at, channel, outcome, next_action, next_action_at, notes, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (followup_id, payload.client_id, str(user["full_name"]),
             payload.contact_at, payload.channel, payload.outcome,
             payload.next_action, payload.next_action_at, payload.notes, now),
        )
        cur.execute(
            "INSERT INTO crm_activity_log (id, client_id, user_id, type, title, payload, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), payload.client_id, str(user["full_name"]),
             "followup", f"Seguimiento: {payload.outcome}", _dict_to_str(payload.model_dump()), now),
        )
        db.commit()
        return {"ok": True, "id": followup_id}
    finally:
        db.close()


@router.get("/client/{client_id}/invoices")
def client_invoices(
    client_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT code FROM crm_clients WHERE id = %s", (client_id,))
        client = _row(cur.fetchone())
        if not client:
            return []
        client_code = (client.get("code") or "").strip()
        if not client_code:
            return []
        leg = _legacy_db()
        try:
            lcur = leg.cursor(dictionary=True)
            lcur.execute(
                """
                SELECT id, factura AS folio, fecha AS issued_at, subtotal, descuento_pct AS discount,
                       iva AS tax, total, estatus AS status, timbrado_estatus, cfdi_uuid,
                       empresa, vendedor
                FROM facturas
                WHERE numero_cliente = %s
                ORDER BY fecha DESC
                LIMIT 100
                """,
                (client_code,),
            )
            return _rows(lcur.fetchall())
        finally:
            leg.close()
    finally:
        db.close()


@router.get("/client/activity/{activity_id}")
def get_activity(
    activity_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_activity_log WHERE id = %s", (activity_id,))
        return _row(cur.fetchone())
    finally:
        db.close()


@router.get("/client/{client_id}/activity")
def client_activity(
    client_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_activity_log WHERE client_id = %s ORDER BY created_at DESC", (client_id,))
        return _rows(cur.fetchall())
    finally:
        db.close()


@router.get("/client/{client_id}/quotes")
def client_quotes(
    client_id: str,
    user=Depends(require_user),
):
    return []


@router.get("/client/{client_id}/products")
def client_products(
    client_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT code FROM crm_clients WHERE id = %s", (client_id,))
        client = _row(cur.fetchone())
        if not client:
            return []
        client_code = (client.get("code") or "").strip()
        if not client_code:
            return []
        leg = _legacy_db()
        try:
            lcur = leg.cursor(dictionary=True)
            lcur.execute(
                """
                SELECT fd.cip, COALESCE(p.descripcion, fd.descripcion) AS description,
                       COUNT(DISTINCT fd.factura_id) AS invoice_count,
                       MAX(f.fecha) AS last_purchase,
                       SUM(fd.importe) AS total_amount,
                       SUM(fd.cantidad) AS total_quantity
                FROM factura_detalle fd
                JOIN facturas f ON f.id = fd.factura_id
                LEFT JOIN productos p ON p.cip = fd.cip
                WHERE f.numero_cliente = %s
                GROUP BY fd.cip
                ORDER BY total_amount DESC
                LIMIT 50
                """,
                (client_code,),
            )
            return _rows(lcur.fetchall())
        finally:
            leg.close()
    finally:
        db.close()


@router.get("/client/{client_id}/product-purchases")
def product_purchases(
    client_id: str,
    cip: str = "",
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT code FROM crm_clients WHERE id = %s", (client_id,))
        client = _row(cur.fetchone())
        if not client:
            return []
        client_code = (client.get("code") or "").strip()
        if not client_code or not cip:
            return []
        leg = _legacy_db()
        try:
            lcur = leg.cursor(dictionary=True)
            lcur.execute(
                """
                SELECT f.factura AS folio, f.fecha AS issued_at,
                       fd.cantidad AS quantity, fd.importe AS amount
                FROM factura_detalle fd
                JOIN facturas f ON f.id = fd.factura_id
                WHERE f.numero_cliente = %s AND fd.cip = %s
                ORDER BY f.fecha DESC
                """,
                (client_code, cip),
            )
            return _rows(lcur.fetchall())
        finally:
            leg.close()
    finally:
        db.close()


@router.get("/followups")
def list_followups(
    all: str = Query(default=""),
    q: str = Query(default=""),
    id: str = Query(default=""),
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        if id:
            cur.execute(
                """
                SELECT f.*, c.name AS client_name, c.code AS client_code, c.phone AS client_phone
                FROM crm_followups f
                LEFT JOIN crm_clients c ON c.id = f.client_id
                WHERE f.id = %s
                """,
                (id,),
            )
            return _row(cur.fetchone())
        elif all:
            search = f"%{q.strip()}%" if q.strip() else None
            if search:
                cur.execute(
                    """
                    SELECT f.*, c.name AS client_name, c.code AS client_code, c.phone AS client_phone,
                           q.folio AS quote_folio
                    FROM crm_followups f
                    LEFT JOIN crm_clients c ON c.id = f.client_id
                    LEFT JOIN crm_quotes q ON q.id = f.quote_id
                    WHERE c.name LIKE %s OR f.outcome LIKE %s OR f.channel LIKE %s OR f.notes LIKE %s
                    ORDER BY f.contact_at DESC
                    LIMIT 200
                    """,
                    (search, search, search, search),
                )
            else:
                cur.execute(
                    """
                    SELECT f.*, c.name AS client_name, c.code AS client_code, c.phone AS client_phone,
                           q.folio AS quote_folio
                    FROM crm_followups f
                    LEFT JOIN crm_clients c ON c.id = f.client_id
                    LEFT JOIN crm_quotes q ON q.id = f.quote_id
                    ORDER BY f.contact_at DESC
                    LIMIT 200
                    """
                )
        else:
            return []
        return _rows(cur.fetchall())
    finally:
        db.close()


# ── Prospector ──

PLACES_TEXT_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"


def _google_places_key() -> str:
    cfg_path = settings.base_dir / "app" / "comandas_legacy" / "crm_ventas" / "app_config.json"
    cfg_key = ""
    try:
        if cfg_path.exists():
            cfg_key = (json.loads(cfg_path.read_text(encoding="utf-8")).get("google_places_api_key") or "").strip()
    except Exception:
        cfg_key = ""
    return (os.environ.get("GOOGLE_API_KEY") or cfg_key).strip()


def _google_places_headers(api_key: str) -> dict:
    return {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": (
            "places.id,"
            "places.displayName,"
            "places.formattedAddress,"
            "places.location,"
            "places.rating,"
            "places.userRatingCount,"
            "places.businessStatus,"
            "places.nationalPhoneNumber,"
            "places.internationalPhoneNumber,"
            "places.websiteUri,"
            "places.primaryTypeDisplayName,"
            "nextPageToken"
        ),
    }


def _normalize_place(place: dict) -> dict:
    display = place.get("displayName") or {}
    category = place.get("primaryTypeDisplayName") or {}
    location = place.get("location") or {}
    return {
        "google_place_id": place.get("id", ""),
        "name": display.get("text", ""),
        "category": category.get("text", ""),
        "phone": place.get("nationalPhoneNumber") or place.get("internationalPhoneNumber") or "",
        "website": place.get("websiteUri", ""),
        "address": place.get("formattedAddress", ""),
        "latitude": location.get("latitude"),
        "longitude": location.get("longitude"),
        "rating": place.get("rating"),
        "total_reviews": place.get("userRatingCount"),
        "business_status": place.get("businessStatus", ""),
    }


def _rectangle_restriction(zone: dict) -> dict:
    return {
        "rectangle": {
            "low": {"latitude": float(zone["sur"]), "longitude": float(zone["oeste"])},
            "high": {"latitude": float(zone["norte"]), "longitude": float(zone["este"])},
        }
    }


def _search_google_places(text_query: str, limit_count: int = 20, location_restriction: Optional[dict] = None) -> list[dict]:
    api_key = _google_places_key()
    if not api_key:
        raise RuntimeError("Falta configurar GOOGLE_API_KEY o google_places_api_key.")
    query = (text_query or "").strip()
    if not query:
        return []
    limit_count = min(max(int(limit_count or 20), 1), 60)
    payload = {
        "textQuery": query,
        "pageSize": min(limit_count, 20),
        "languageCode": "es",
        "regionCode": "MX",
    }
    if location_restriction:
        payload["locationRestriction"] = location_restriction
    results = []
    next_page_token = None
    while len(results) < limit_count:
        if next_page_token:
            payload["pageToken"] = next_page_token
        response = requests.post(
            PLACES_TEXT_SEARCH_URL,
            headers=_google_places_headers(api_key),
            json=payload,
            timeout=30,
        )
        if response.status_code != 200:
            raise RuntimeError(f"Error Google Places {response.status_code}: {response.text[:300]}")
        data = response.json()
        results.extend(_normalize_place(place) for place in data.get("places", []))
        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break
        time.sleep(2)
    return [p for p in results[:limit_count] if p.get("name")]


def _upsert_google_prospect(cur, prospect: dict, source_query: str, zone_name: str = "") -> str:
    now = _now()
    google_place_id = (prospect.get("google_place_id") or "").strip()
    existing_id = None
    if google_place_id:
        cur.execute("SELECT id FROM crm_prospects WHERE google_place_id = %s LIMIT 1", (google_place_id,))
        row = cur.fetchone()
        existing_id = row["id"] if row else None
    if existing_id:
        cur.execute(
            """
            UPDATE crm_prospects
            SET name=%s, address=%s, phone=%s, category=%s, source=%s, google_place_id=%s,
                rating=%s, total_reviews=%s, website=%s, latitude=%s, longitude=%s,
                business_status=%s, source_query=%s, zone_name=COALESCE(NULLIF(%s, ''), zone_name), updated_at=%s
            WHERE id=%s
            """,
            (
                prospect.get("name") or "",
                prospect.get("address") or "",
                prospect.get("phone") or "",
                prospect.get("category") or "",
                "google_places",
                google_place_id,
                prospect.get("rating"),
                prospect.get("total_reviews"),
                prospect.get("website") or "",
                prospect.get("latitude"),
                prospect.get("longitude"),
                prospect.get("business_status") or "",
                source_query,
                zone_name or "",
                now,
                existing_id,
            ),
        )
        return existing_id
    prospect_id = str(uuid.uuid4())
    cur.execute(
        """
        INSERT INTO crm_prospects
            (id, name, address, phone, zone, status, category, source, notes, client_id,
             created_at, updated_at, google_place_id, rating, total_reviews, website,
             latitude, longitude, business_status, source_query, zone_name)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            prospect_id,
            prospect.get("name") or "",
            prospect.get("address") or "",
            prospect.get("phone") or "",
            zone_name or "",
            "nuevo",
            prospect.get("category") or "",
            "google_places",
            "",
            None,
            now,
            now,
            google_place_id,
            prospect.get("rating"),
            prospect.get("total_reviews"),
            prospect.get("website") or "",
            prospect.get("latitude"),
            prospect.get("longitude"),
            prospect.get("business_status") or "",
            source_query,
            zone_name or "",
        ),
    )
    return prospect_id


@router.get("/prospector/zones")
def list_prospector_zones(user=Depends(require_user)):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_zones WHERE enabled = 1 ORDER BY name")
        rows = _rows(cur.fetchall())
        return rows
    finally:
        db.close()


@router.post("/prospector/search")
def prospector_search(
    query: str = Query(default=""),
    business_name: str = Query(default=""),
    zone_name: str = Query(default=""),
    limit: int = Query(default=20),
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        query = (query or "").strip()
        business_name = (business_name or "").strip()
        words = [w.strip() for w in query.replace(",", " ").split() if w.strip()]
        placeholders = []
        params = []
        if business_name:
            placeholders.append("p.name LIKE %s")
            params.append(f"%{business_name}%")
        for w in words[:5]:
            placeholders.append("(p.name LIKE %s OR p.address LIKE %s OR p.category LIKE %s OR p.source_query LIKE %s OR p.phone LIKE %s OR p.website LIKE %s)")
            params.extend([f"%{w}%", f"%{w}%", f"%{w}%", f"%{w}%", f"%{w}%", f"%{w}%"])
        rows = []
        if placeholders:
            conditions = ["p.status = 'nuevo'", *placeholders]
            if zone_name:
                conditions.append("p.zone_name = %s")
                params.append(zone_name)
            sql = "SELECT p.*, c.name AS client_name FROM crm_prospects p LEFT JOIN crm_clients c ON c.id = p.client_id WHERE " + " AND ".join(conditions) + f" ORDER BY p.updated_at DESC LIMIT {min(limit, 60)}"
            cur.execute(sql, params)
            existing = _rows(cur.fetchall())
        else:
            existing = []
        if existing:
            rows = existing
        else:
            search_text = " ".join(part for part in (business_name, query) if part).strip()
            location_restriction = None
            if zone_name:
                cur.execute("SELECT * FROM crm_zones WHERE name = %s AND enabled = 1 LIMIT 1", (zone_name,))
                zone = _row(cur.fetchone())
                if zone:
                    location_restriction = _rectangle_restriction(zone)
            try:
                places = _search_google_places(search_text, limit, location_restriction)
            except Exception as exc:
                return {"prospects": [], "count": 0, "message": str(exc)}
            ids = [_upsert_google_prospect(cur, place, search_text, zone_name or "") for place in places]
            db.commit()
            if ids:
                id_params = ",".join(["%s"] * len(ids))
                cur.execute(
                    f"""
                    SELECT p.*, c.name AS client_name
                    FROM crm_prospects p
                    LEFT JOIN crm_clients c ON c.id = p.client_id
                    WHERE p.id IN ({id_params})
                    ORDER BY p.rating DESC, p.total_reviews DESC
                    """,
                    ids,
                )
                rows = _rows(cur.fetchall())
        return {"prospects": rows, "count": len(rows)}
    finally:
        db.close()


@router.post("/prospector/scan")
def prospector_scan(
    body: dict,
    user=Depends(require_user),
):
    zones = body.get("zones", [])
    rescan = body.get("rescan", False)
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        prospects = []
        for zone_name in zones:
            cur.execute("SELECT p.*, c.name AS client_name FROM crm_prospects p LEFT JOIN crm_clients c ON c.id = p.client_id WHERE p.zone_name = %s AND p.status = 'nuevo' LIMIT 50", (zone_name,))
            existing = _rows(cur.fetchall())
            if existing and not rescan:
                prospects.extend(existing)
                continue
            categories = ["Restaurante", "Tienda", "Oficina", "Taller", "Clinica"]
            for cat in categories:
                for i in range(3):
                    pid = str(uuid.uuid4())
                    name = f"{cat} {zone_name} #{i + 1}"
                    addr = f"Col. Centro, {zone_name}"
                    cur.execute(
                        "REPLACE INTO crm_prospects (id, name, address, zone_name, status, category, source_query, created_at, updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (pid, name, addr, zone_name, "nuevo", cat, "scanner", _now(), _now()),
                    )
                    prospects.append({"id": pid, "name": name, "address": addr, "zone_name": zone_name, "status": "nuevo", "category": cat, "source_query": "scanner"})
        db.commit()
        return {"prospects": prospects, "count": len(prospects)}
    finally:
        db.close()


@router.get("/prospector/prospects")
def list_prospects(
    q: str = Query(default=""),
    status: str = Query(default="todos"),
    zone: str = Query(default="todas"),
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        conditions = []
        params = []
        if q:
            conditions.append("(p.name LIKE %s OR p.address LIKE %s OR p.category LIKE %s OR p.source_query LIKE %s OR p.phone LIKE %s OR p.website LIKE %s)")
            params.extend([f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"])
        if status != "todos":
            conditions.append("p.status = %s")
            params.append(status)
        if zone != "todas":
            conditions.append("p.zone_name = %s")
            params.append(zone)
        where = " WHERE " + " AND ".join(conditions) if conditions else ""
        cur.execute(f"SELECT p.*, c.name AS client_name FROM crm_prospects p LEFT JOIN crm_clients c ON c.id = p.client_id{where} ORDER BY p.updated_at DESC LIMIT 200", params)
        return {"prospects": _rows(cur.fetchall())}
    finally:
        db.close()


@router.get("/prospector/prospects/{prospect_id}")
def get_prospect(
    prospect_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT p.*, c.name AS client_name FROM crm_prospects p LEFT JOIN crm_clients c ON c.id = p.client_id WHERE p.id = %s", (prospect_id,))
        prospect = _row(cur.fetchone())
        if not prospect:
            raise HTTPException(status_code=404, detail="Prospecto no encontrado")
        cur.execute("SELECT * FROM crm_prospect_followups WHERE prospect_id = %s ORDER BY contact_at DESC", (prospect_id,))
        followups = _rows(cur.fetchall())
        cur.execute("SELECT * FROM crm_prospect_phones WHERE prospect_id = %s", (prospect_id,))
        phones = _rows(cur.fetchall())
        cur.execute("SELECT * FROM crm_prospect_quotes WHERE prospect_id = %s ORDER BY created_at DESC", (prospect_id,))
        quotes = _rows(cur.fetchall())
        cur.execute("SELECT * FROM crm_prospect_activity WHERE prospect_id = %s ORDER BY created_at DESC", (prospect_id,))
        activity = _rows(cur.fetchall())
        prospect["followups"] = followups
        prospect["phones"] = phones
        prospect["quotes"] = quotes
        prospect["activity"] = activity
        return prospect
    finally:
        db.close()


class ProspectFollowupCreate(BaseModel):
    prospect_id: Optional[str] = ""
    contact_at: Optional[str] = ""
    channel: str = "Llamada"
    outcome: str = ""
    next_action: Optional[str] = ""
    next_action_at: Optional[str] = ""
    notes: Optional[str] = ""


@router.post("/prospector/prospects/{prospect_id}/followups")
def create_prospect_followup(
    prospect_id: str,
    payload: ProspectFollowupCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        cur.execute(
            "INSERT INTO crm_prospect_followups (id, prospect_id, user_id, contact_at, channel, outcome, next_action, next_action_at, notes, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), prospect_id, str(user["full_name"]), payload.contact_at or now, payload.channel, payload.outcome, payload.next_action, payload.next_action_at, payload.notes, now),
        )
        cur.execute(
            "INSERT INTO crm_prospect_activity (id, prospect_id, user_id, type, title, payload, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), prospect_id, str(user["full_name"]), "prospect_followup", f"Seguimiento: {payload.outcome}", _dict_to_str(payload.model_dump()), now),
        )
        db.commit()
        return {"ok": True}
    finally:
        db.close()


class ProspectPhoneCreate(BaseModel):
    prospect_id: Optional[str] = ""
    label: Optional[str] = ""
    phone: str = ""
    notes: Optional[str] = ""


@router.post("/prospector/prospects/{prospect_id}/phones")
def create_prospect_phone(
    prospect_id: str,
    payload: ProspectPhoneCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        cur.execute(
            "INSERT INTO crm_prospect_phones (id, prospect_id, label, phone, notes, created_at) VALUES (%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), prospect_id, payload.label or "General", payload.phone, payload.notes, now),
        )
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.post("/prospector/prospects/{prospect_id}/convert")
def convert_prospect(
    prospect_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_prospects WHERE id = %s", (prospect_id,))
        p = _row(cur.fetchone())
        if not p:
            raise HTTPException(status_code=404, detail="Prospecto no encontrado")
        now = _now()
        client_id = str(uuid.uuid4())
        cur.execute(
            "INSERT INTO crm_clients (id, code, name, phone, tax_address, created_at, updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (client_id, f"PRO-{prospect_id[:8]}", p.get("name", "Prospecto convertido"), p.get("phone", ""), p.get("address", ""), now, now),
        )
        cur.execute("UPDATE crm_prospects SET status = 'convertido', updated_at = %s WHERE id = %s", (now, prospect_id))
        cur.execute(
            "INSERT INTO crm_prospect_activity (id, prospect_id, user_id, type, title, payload, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), prospect_id, str(user["full_name"]), "convert", "Prospecto convertido a cliente", "", now),
        )
        db.commit()
        return {"ok": True, "client_id": client_id}
    finally:
        db.close()


@router.put("/prospector/prospects/{prospect_id}")
def update_prospect(
    prospect_id: str,
    body: dict,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        fields = []
        params = []
        for key in ("status", "name", "address", "zone_name", "phone", "category"):
            if key in body:
                fields.append(f"{key} = %s")
                params.append(body[key])
        if fields:
            params.append(now)
            params.append(prospect_id)
            cur.execute(f"UPDATE crm_prospects SET {', '.join(fields)}, updated_at = %s WHERE id = %s", params)
            db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.delete("/prospector/prospects/{prospect_id}")
def delete_prospect(
    prospect_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT id FROM crm_prospects WHERE id = %s", (prospect_id,))
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Prospecto no encontrado")
        for table in (
            "crm_prospect_followups",
            "crm_prospect_phones",
            "crm_prospect_quotes",
            "crm_prospect_activity",
        ):
            cur.execute(f"DELETE FROM {table} WHERE prospect_id = %s", (prospect_id,))
        cur.execute("DELETE FROM crm_quotes WHERE prospect_id = %s", (prospect_id,))
        cur.execute("DELETE FROM crm_prospects WHERE id = %s", (prospect_id,))
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.post("/prospector/check-clients")
def check_prospect_clients(user=Depends(require_user)):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_prospects WHERE status NOT IN ('cliente_existente', 'convertido', 'descartado')")
        prospects = _rows(cur.fetchall())
        updated = 0
        for p in prospects:
            name = (p.get("name") or "").strip()
            if not name:
                continue
            cur.execute("SELECT id FROM crm_clients WHERE name LIKE %s LIMIT 1", (f"%{name[:20]}%",))
            if cur.fetchone():
                cur.execute("UPDATE crm_prospects SET status = 'cliente_existente', updated_at = %s WHERE id = %s", (_now(), p["id"]))
                updated += 1
        db.commit()
        return {"ok": True, "updated": updated}
    finally:
        db.close()


# ── Quotes ──


@router.get("/quotes")
def list_quotes(
    status: str = Query(default=""),
    client_id: str = Query(default=""),
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        conditions = []
        params = []
        if status:
            conditions.append("status = %s")
            params.append(status)
        if client_id:
            conditions.append("client_id = %s")
            params.append(client_id)
        where = " WHERE " + " AND ".join(conditions) if conditions else ""
        cur.execute(f"SELECT * FROM crm_quotes{where} ORDER BY created_at DESC LIMIT 100", params)
        return _rows(cur.fetchall())
    finally:
        db.close()


# ── Quote helpers ──


def _parse_external_client_id(client_id):
    """Return (numero, empresa) from a CRM client_id or direct code."""
    if not client_id:
        return None, None
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT code, name FROM crm_clients WHERE id = %s LIMIT 1", (client_id,))
        row = cur.fetchone()
        if not row:
            return None, None
        return row["code"], row["name"]
    finally:
        db.close()


def _product_quote_info(cip):
    """Look up a product by CIP in the legacy productos table."""
    code = str(cip or "").strip()
    if not code:
        return None
    db = _legacy_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT cip, descripcion, unidad, tipo_lista, iva FROM productos WHERE cip = %s LIMIT 1", (code,))
        producto = cur.fetchone()
        if not producto:
            # try partial match
            cur.execute(
                "SELECT cip, descripcion, unidad, tipo_lista, iva FROM productos WHERE descripcion LIKE %s OR cip LIKE %s ORDER BY cip LIMIT 1",
                (f"%{code}%", f"%{code}%"),
            )
            producto = cur.fetchone()
        if not producto:
            return None
        # Get price from Lista General / L GENERAL
        cur.execute(
            """
            SELECT pp.precio
            FROM precios_productos pp
            JOIN listas_precios lp ON lp.id = pp.lista_id
            WHERE UPPER(TRIM(lp.nombre)) IN (UPPER(TRIM('Lista General')), UPPER(TRIM('L GENERAL'))) AND pp.cip = %s
              AND COALESCE(pp.precio, 0) > 0
            ORDER BY CASE WHEN UPPER(TRIM(lp.nombre)) = UPPER(TRIM('L GENERAL')) THEN 0 ELSE 1 END
            LIMIT 1
            """,
            (producto["cip"],),
        )
        precio = cur.fetchone()
        tax_rate = 16 if str(producto.get("iva") or "").strip().lower() in ("si", "sí", "s", "sÃ­") else 0
        return {
            "cip": producto["cip"],
            "description": producto.get("descripcion") or "",
            "unit": producto.get("unidad") or "",
            "unit_price": float(precio["precio"]) if precio else 0,
            "discount_rate": 0,
            "tax_rate": tax_rate,
        }
    finally:
        db.close()


def _price_for_client(client_id, cip):
    """Resolve price for a client based on their price list."""
    # client_id is a CRM UUID; look up the client code
    db_crm = _crm_db()
    try:
        cur = db_crm.cursor(dictionary=True)
        cur.execute("SELECT code FROM crm_clients WHERE id = %s LIMIT 1", (client_id,))
        row = cur.fetchone()
        if not row:
            return None
        code = row["code"]
    finally:
        db_crm.close()
    db = _legacy_db()
    try:
        cur = db.cursor(dictionary=True)
        # Look up the client by numero (code) only — empresa match is optional
        cur.execute("SELECT IFNULL(especial, 'Lista General') AS lista_nombre FROM clientes WHERE numero = %s LIMIT 1", (code,))
        cliente = cur.fetchone()
        if not cliente:
            return None
        lista = (cliente.get("lista_nombre") or "Lista General").strip()
        cur.execute("SELECT descripcion, unidad, tipo_lista, iva FROM productos WHERE cip = %s", (cip,))
        producto = cur.fetchone()
        if not producto:
            return None
        # Try client's list first, fallback to Lista General
        cur.execute(
            "SELECT pp.precio FROM precios_productos pp JOIN listas_precios lp ON lp.id = pp.lista_id WHERE UPPER(TRIM(lp.nombre)) = UPPER(TRIM(%s)) AND pp.cip = %s",
            (lista, cip),
        )
        precio = cur.fetchone()
        if not precio or float(precio.get("precio") or 0) <= 0:
            cur.execute(
                """
                SELECT pp.precio FROM precios_productos pp
                JOIN listas_precios lp ON lp.id = pp.lista_id
                WHERE UPPER(TRIM(lp.nombre)) IN (UPPER(TRIM('Lista General')), UPPER(TRIM('L GENERAL'))) AND pp.cip = %s
                  AND COALESCE(pp.precio, 0) > 0
                ORDER BY CASE WHEN UPPER(TRIM(lp.nombre)) = UPPER(TRIM('L GENERAL')) THEN 0 ELSE 1 END
                LIMIT 1
                """,
                (cip,),
            )
            precio = cur.fetchone()
        tax_rate = 16 if str(producto.get("iva") or "").strip().lower() in ("si", "sí", "s", "sÃ­") else 0
        return {
            "cip": cip,
            "description": producto["descripcion"],
            "unit": producto.get("unidad") or "",
            "list": lista,
            "unit_price": float(precio["precio"]) if precio else 0.0,
            "discount_rate": 0,
            "tax_rate": tax_rate,
        }
    finally:
        db.close()


def _calculate_quote(items):
    """Calculate per-line and overall quote totals."""
    subtotal = 0
    discount = 0
    tax = 0
    prepared = []
    for item in items:
        quantity = float(item.get("quantity") or 0)
        unit_price = float(item.get("unit_price") or item.get("price") or 0)
        discount_rate = float(item.get("discount_rate") or 0)
        tax_rate = float(item.get("tax_rate") if item.get("tax_rate") not in (None, "") else 16)
        gross = quantity * unit_price
        line_discount = gross * (discount_rate / 100)
        taxable = gross - line_discount
        line_tax = taxable * (tax_rate / 100)
        line_total = taxable + line_tax
        subtotal += gross
        discount += line_discount
        tax += line_tax
        prepared.append({
            "id": str(uuid.uuid4()),
            "cip": item.get("cip", ""),
            "description": item.get("description", ""),
            "quantity": quantity,
            "unit_price": unit_price,
            "discount_rate": discount_rate,
            "tax_rate": tax_rate,
            "line_total": round(line_total, 2),
        })
    total = subtotal - discount + tax
    return {
        "items": prepared,
        "subtotal": round(subtotal, 2),
        "discount": round(discount, 2),
        "tax": round(tax, 2),
        "total": round(total, 2),
        "authorized_shipping": round(total * 0.08, 2),
    }


def _quote_title_for_client(client_id, folio, quote_recipient=None):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT code FROM crm_clients WHERE id = %s LIMIT 1", (client_id,))
        row = cur.fetchone()
        code = row["code"] if row else None
    finally:
        db.close()
    if code != "100000":
        return folio
    name = (quote_recipient or "").strip() or "PUBLICO EN GENERAL"
    return f"{name} - {folio}"


def _quote_client_snapshot_if_needed(client_id, quote_recipient=None):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT code FROM crm_clients WHERE id = %s LIMIT 1", (client_id,))
        row = cur.fetchone()
        if not row or row["code"] != "100000":
            return None
        cur.execute("SELECT * FROM crm_clients WHERE id = %s LIMIT 1", (client_id,))
        client = cur.fetchone()
        if not client:
            return None
        display_name = (quote_recipient or "").strip() or client.get("name") or ""
        snapshot = {
            "client_name": display_name,
            "client_code": client.get("code") or "",
            "phone": client.get("phone") or "",
            "email": client.get("email") or "",
            "tax_address": client.get("tax_address") or "",
            "consignee_address": client.get("consignee_address") or "",
            "delivery_method": client.get("delivery_method") or "",
            "contact_name": client.get("contact_name") or "",
        }
        return json.dumps(snapshot, ensure_ascii=False)
    finally:
        db.close()


class QuoteCreate(BaseModel):
    id: Optional[str] = None
    client_id: Optional[str] = ""
    prospect_id: Optional[str] = ""
    client_identifier: Optional[str] = ""
    quote_recipient: Optional[str] = ""
    title: Optional[str] = ""
    notes: Optional[str] = ""
    authorized_shipping: float = 0
    show_shipping: bool = True
    items: list = []


@router.post("/quotes")
def create_quote(
    payload: QuoteCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        qid = payload.id or str(uuid.uuid4())
        today = datetime.now().strftime("%Y%m%d")
        cur.execute("SELECT COUNT(*) AS cnt FROM crm_quotes WHERE folio LIKE %s", (f"COT-{today}-%",))
        row = cur.fetchone()
        seq = (row["cnt"] if row else 0) + 1
        folio = f"COT-{today}-{seq:04d}"

        # Calculate quote
        calc = _calculate_quote(payload.items)
        items = calc["items"]
        total = calc["total"]
        subtotal = calc["subtotal"]
        discount = calc["discount"]
        tax = calc["tax"]

        # Auto shipping if not provided
        authorized_shipping = payload.authorized_shipping or calc["authorized_shipping"]

        # Title and recipient
        title = payload.title or _quote_title_for_client(payload.client_id, folio, payload.quote_recipient)
        quote_recipient = payload.quote_recipient or ""

        # Client snapshot for client 100000
        client_snapshot = _quote_client_snapshot_if_needed(payload.client_id, quote_recipient) if payload.client_id else None

        # Valid until (15 days)
        valid_until = (datetime.now(timezone.utc) + timedelta(days=15)).strftime("%Y-%m-%d")

        cur.execute(
            """INSERT INTO crm_quotes
                (id, client_id, prospect_id, user_id, folio, title, quote_recipient,
                 authorized_shipping, show_shipping, subtotal, discount, tax, total, items,
                 client_snapshot, notes, valid_until, status, created_at, updated_at)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (qid, payload.client_id, payload.prospect_id, str(user["full_name"]),
             folio, title, quote_recipient,
             authorized_shipping, payload.show_shipping, subtotal, discount, tax, total,
             json.dumps(items, ensure_ascii=False),
             client_snapshot, payload.notes, valid_until, "active", now, now),
        )
        if payload.client_id:
            cur.execute(
                "INSERT INTO crm_activity_log (id, client_id, user_id, type, title, payload, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (str(uuid.uuid4()), payload.client_id, str(user["full_name"]),
                 "quote", f"Cotizacion {folio} creada",
                 json.dumps({"total": total, "items_count": len(items)}, ensure_ascii=False), now),
            )
        db.commit()
        return {"ok": True, "id": qid, "folio": folio, "total": total}
    except Exception as e:
        import traceback
        err_msg = traceback.format_exc()
        with open(r"E:\Proyectos\Facturacion 150426 casa\Proyecto facturacion\MigracionWeb\_quote_error.log", "a") as f:
            f.write(f"Quote create error: {e}\n{err_msg}\n")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.put("/quotes/{quote_id}")
def update_quote(
    quote_id: str,
    payload: QuoteCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()

        # Verify quote exists
        cur.execute("SELECT * FROM crm_quotes WHERE id = %s", (quote_id,))
        existing = cur.fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="Cotizacion no encontrada")

        # Calculate quote
        calc = _calculate_quote(payload.items)
        items = calc["items"]
        total = calc["total"]
        subtotal = calc["subtotal"]
        discount = calc["discount"]
        tax = calc["tax"]

        authorized_shipping = payload.authorized_shipping or calc["authorized_shipping"]
        title = payload.title or existing.get("title") or f"COT-{quote_id[:8].upper()}"
        quote_recipient = payload.quote_recipient or existing.get("quote_recipient") or ""

        # Client snapshot for client 100000
        client_snapshot = _quote_client_snapshot_if_needed(payload.client_id, quote_recipient) if payload.client_id else existing.get("client_snapshot")

        cur.execute(
            """UPDATE crm_quotes SET
               client_id=%s, prospect_id=%s, title=%s, quote_recipient=%s,
               authorized_shipping=%s, show_shipping=%s, subtotal=%s, discount=%s, tax=%s, total=%s,
               items=%s, client_snapshot=%s, notes=%s, updated_at=%s
               WHERE id=%s""",
            (payload.client_id, payload.prospect_id, title, quote_recipient,
             authorized_shipping, payload.show_shipping, subtotal, discount, tax, total,
             json.dumps(items, ensure_ascii=False),
             client_snapshot, payload.notes, now, quote_id),
        )

        if payload.client_id:
            cur.execute(
                "INSERT INTO crm_activity_log (id, client_id, user_id, type, title, payload, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                (str(uuid.uuid4()), payload.client_id, str(user["full_name"]),
                 "quote", f"Cotizacion actualizada",
                 json.dumps({"total": total, "items_count": len(items)}, ensure_ascii=False), now),
            )
        db.commit()
        return {"ok": True, "id": quote_id, "total": total}
    finally:
        db.close()


@router.get("/quotes/{quote_id}")
def get_quote(quote_id: str, user=Depends(require_user)):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("""SELECT q.*, c.name AS client_name, c.code AS client_code,
                       c.phone, c.email, c.tax_address, c.consignee_address, c.delivery_method
                       FROM crm_quotes q
                       LEFT JOIN crm_clients c ON c.id = q.client_id
                       WHERE q.id = %s""", (quote_id,))
        quote = _row(cur.fetchone())
        if not quote:
            raise HTTPException(status_code=404, detail="Cotizacion no encontrada")
        if isinstance(quote.get("items"), str):
            quote["items"] = json.loads(quote["items"])
        # If client_snapshot exists, overlay it
        if quote.get("client_snapshot"):
            try:
                snap = json.loads(quote["client_snapshot"]) if isinstance(quote["client_snapshot"], str) else quote["client_snapshot"]
                for key, val in snap.items():
                    quote[key] = val
            except Exception:
                pass
        return quote
    finally:
        db.close()


class ProspectLinkCreate(BaseModel):
    client_code: str


@router.post("/prospector/prospects/{prospect_id}/link")
def link_prospect(
    prospect_id: str,
    payload: ProspectLinkCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT id FROM crm_clients WHERE code = %s LIMIT 1", (payload.client_code,))
        client = cur.fetchone()
        if not client:
            raise HTTPException(status_code=404, detail="Cliente no encontrado")
        client_id = client["id"]
        cur.execute("UPDATE crm_prospects SET client_id = %s, status = 'cliente_existente', updated_at = %s WHERE id = %s",
                    (client_id, _now(), prospect_id))
        cur.execute("INSERT INTO crm_prospect_activity (id, prospect_id, user_id, type, title, payload, created_at) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                    (str(uuid.uuid4()), prospect_id, str(user["full_name"]), "prospect_link", f"Asociado a cliente {payload.client_code}", json.dumps({"client_id": client_id}, ensure_ascii=False), _now()))
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.get("/product-list")
def get_product_list(user=Depends(require_user)):
    db = _legacy_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT cip, descripcion FROM productos ORDER BY cip LIMIT 500")
        return cur.fetchall()
    finally:
        db.close()


@router.get("/product-search")
def product_search(q: str = Query(default=""), user=Depends(require_user)):
    db = _legacy_db()
    try:
        cur = db.cursor(dictionary=True)
        search = f"%{q.strip()}%"
        cur.execute(
            "SELECT cip, descripcion FROM productos WHERE cip LIKE %s OR descripcion LIKE %s ORDER BY cip LIMIT 50",
            (search, search),
        )
        return cur.fetchall()
    finally:
        db.close()


@router.get("/product-info")
def get_product_info(cip: str = Query(...), user=Depends(require_user)):
    result = _product_quote_info(cip)
    if not result:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return result


@router.get("/price")
def get_price(client_id: str = Query(...), cip: str = Query(...), user=Depends(require_user)):
    result = _price_for_client(client_id, cip)
    if not result:
        raise HTTPException(status_code=404, detail="Precio no encontrado para este cliente")
    return result


@router.delete("/quotes/{quote_id}")
def delete_quote(
    quote_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT client_id FROM crm_quotes WHERE id = %s", (quote_id,))
        quote = cur.fetchone()
        cur.execute("DELETE FROM crm_quotes WHERE id = %s", (quote_id,))
        if quote and quote.get("client_id"):
            cur.execute(
                "INSERT INTO crm_activity_log (id, client_id, user_id, type, title, created_at) VALUES (%s,%s,%s,%s,%s,%s)",
                (str(uuid.uuid4()), quote["client_id"], str(user["full_name"]), "quote", "Cotizacion eliminada", _now()),
            )
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.get("/quotes/{quote_id}/export")
def export_quote(
    quote_id: str,
    format: str = Query("pdf"),
    user=Depends(require_user),
):
    from fastapi.responses import FileResponse, Response
    import os, tempfile, textwrap, traceback
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("""SELECT q.*, c.name AS client_name, c.code AS client_code,
                       c.phone, c.email, c.tax_address, c.consignee_address, c.delivery_method
                       FROM crm_quotes q
                       LEFT JOIN crm_clients c ON c.id = q.client_id
                       WHERE q.id = %s""", (quote_id,))
        quote = _row(cur.fetchone())
        if not quote:
            raise HTTPException(status_code=404, detail="Cotizacion no encontrada")
        if isinstance(quote.get("items"), str):
            quote["items"] = json.loads(quote["items"])
        if quote.get("client_snapshot"):
            try:
                snap = json.loads(quote["client_snapshot"]) if isinstance(quote["client_snapshot"], str) else quote["client_snapshot"]
                for key, val in snap.items():
                    quote[key] = val
            except Exception:
                pass

        # Determine company from client code
        company = "EZA2007"
        if quote.get("client_code"):
            code = quote["client_code"]
            try:
                ldb = _legacy_db()
                lcur = ldb.cursor(dictionary=True)
                lcur.execute("SELECT empresa FROM clientes WHERE numero = %s LIMIT 1", (code,))
                row = lcur.fetchone()
                if row and row.get("empresa", "").strip():
                    company = row["empresa"].strip()
                lcur.close()
                ldb.close()
            except Exception:
                pass
        quote["company"] = company

        # Resolve bank account
        bank_account = None
        try:
            bcur = db.cursor(dictionary=True)
            bcur.execute("SELECT * FROM crm_bank_accounts WHERE UPPER(TRIM(company)) = UPPER(TRIM(%s)) AND enabled = 1 LIMIT 1", (company,))
            bank_account = _row(bcur.fetchone())
        except Exception:
            pass
        quote["bank_account"] = bank_account

        if format == "xlsx":
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "Cotización"
                # Template layout matching original
                ws["E2"] = quote.get("folio", "")
                ws["E3"] = datetime.now().date()
                ws["E4"] = quote.get("valid_until") or "15 días"
                ws["E5"] = quote.get("user_id") or ""
                ws["B9"] = quote.get("client_name") or ""
                ws["B10"] = company
                ws["B11"] = quote.get("phone") or ""
                ws["B12"] = quote.get("email") or ""
                ws["B13"] = _quote_delivery_address(quote)
                # Clear old items rows
                for r in range(16, 43):
                    for c in "ABCDE":
                        ws[f"{c}{r}"] = None
                # Write items (max 27 rows)
                for idx, item in enumerate(quote.get("items", [])[:27], start=16):
                    ws[f"A{idx}"] = item.get("cip", "")
                    ws[f"B{idx}"] = item.get("description", "")
                    ws[f"C{idx}"] = item.get("quantity", 0)
                    ws[f"D{idx}"] = item.get("unit_price", 0)
                    ws[f"E{idx}"] = item.get("line_total", 0)
                ws["E44"] = quote.get("total", 0)
                ws["E46"] = f"Envio autorizado 8%: ${float(quote.get('authorized_shipping') or 0):,.2f}"
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                wb.save(tmp.name)
                tmp.close()
                return FileResponse(tmp.name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    filename=f"cotizacion_{quote_id[:8]}.xlsx",
                                    headers={"Access-Control-Expose-Headers": "Content-Disposition"})
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl no disponible")
        else:
            # PDF — identical layout to original build_quote_pdf_like_reference
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.lib import colors
                from reportlab.pdfgen import canvas as pdfcanvas
                try:
                    from PIL import Image as PILImage
                except ImportError:
                    PILImage = None
            except ImportError as e:
                raise HTTPException(status_code=500, detail=f"reportlab no disponible: {e}")

            def _logo_pdf_size(path, target_height=96, max_width=300):
                if not PILImage:
                    return target_height, target_height
                try:
                    img = PILImage.open(path).convert("RGBA")
                    bbox = img.getchannel("A").getbbox() or (0, 0, img.width, img.height)
                    w = max(1, bbox[2] - bbox[0])
                    h = max(1, bbox[3] - bbox[1])
                    width = target_height * (w / h)
                    if width > max_width:
                        return max_width, max_width * (h / w)
                    return width, target_height
                except Exception:
                    return target_height, target_height

            def _quote_delivery_address(d):
                for key in ("consignee_address", "delivery_method", "tax_address"):
                    value = " ".join(str(d.get(key) or "").split())
                    marker = value.replace(" ", "").replace("-", "")
                    if value and marker and set(marker) != {"0"}:
                        return value
                return ""

            def _draw_wrapped_pdf_text(cv, text, x, y, max_chars, max_lines, line_height=8):
                clean = " ".join(str(text or "").split())
                lines = textwrap.wrap(clean, width=max_chars)[:max_lines]
                for idx, line in enumerate(lines):
                    cv.drawString(x, y - (idx * line_height), line)

            # Logo path
            logo_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", "logos")
            if not os.path.isdir(logo_dir):
                logo_dir = r"E:\Proyectos\Facturacion 150426 casa\Proyecto facturacion\AspelAPI\logos"
            if not os.path.isdir(logo_dir):
                logo_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
            comp = (company or "").lower()
            if "gourmet" in comp:
                logo_path = os.path.join(logo_dir, "gourmet.png")
            elif "ibersur" in comp:
                logo_path = os.path.join(logo_dir, "ibersur.png")
            elif "remision" in comp or "remisión" in comp:
                logo_path = os.path.join(logo_dir, "Remision.png")
            elif "eza" in comp:
                logo_path = os.path.join(logo_dir, "eza2007.png")
            else:
                logo_path = os.path.join(logo_dir, "eza2007.png")
                if not os.path.exists(logo_path):
                    logo_path = ""

            tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            c = pdfcanvas.Canvas(tmp.name, pagesize=letter)
            width, height = letter
            left = 55
            right = width - 55
            top = height - 70

            # Logo
            if logo_path and os.path.exists(logo_path):
                logo_w, logo_h = _logo_pdf_size(logo_path)
                c.drawImage(logo_path, left, top - 86, width=logo_w, height=logo_h, preserveAspectRatio=False, mask="auto")

            # Title
            c.setFont("Helvetica-Bold", 10)
            c.drawCentredString(440, top - 5, "COTIZACIÓN")
            c.setFont("Helvetica", 7)
            meta_x = 365
            value_x = 425
            y = top - 20
            today_text = datetime.now().strftime("%d/%m/%Y")
            c.drawString(meta_x, y, "No:")
            c.drawString(value_x, y, str(quote.get("folio") or ""))
            y -= 12
            c.drawString(meta_x, y, "Fecha:")
            c.drawString(value_x, y, today_text)
            y -= 12
            c.drawString(meta_x, y, "Vigencia:")
            vigencia = quote.get("valid_until") or "15 días"
            if vigencia and vigencia != "15 días":
                try:
                    vigencia = datetime.strptime(vigencia, "%Y-%m-%d").strftime("%d/%m/%Y")
                except Exception:
                    pass
            c.drawString(value_x, y, str(vigencia))
            y -= 12
            c.drawString(meta_x, y, "Usuario:")
            c.drawString(value_x, y, str(quote.get("user_id") or ""))

            # Client data
            client_y = top - 100
            c.setFont("Helvetica-Bold", 7)
            c.drawString(left, client_y, "DATOS DEL CLIENTE")
            c.setFont("Helvetica", 7)
            fields = [
                ("Cliente:", quote.get("client_name") or ""),
                ("Empresa:", company),
                ("Teléfono:", quote.get("phone") or ""),
                ("Email:", quote.get("email") or ""),
                ("Dirección:", _quote_delivery_address(quote)),
            ]
            fy = client_y - 12
            for label, value in fields:
                c.drawString(left, fy, label)
                if label == "Teléfono:":
                    c.drawString(left + 50, fy, str(value))
                elif label == "Email:":
                    c.setFillColor(colors.blue)
                    c.drawString(left + 50, fy, str(value))
                    c.setFillColor(colors.black)
                elif "Direcci" in label:
                    lines = textwrap.wrap(" ".join(str(value or "").split()), width=96)[:3]
                    if not lines:
                        lines = [""]
                    for idx, line in enumerate(lines):
                        c.drawString(left + 50, fy - (idx * 8), line)
                    fy -= 8 * (len(lines) - 1)
                else:
                    c.drawString(left + 50, fy, str(value)[:75])
                fy -= 11

            # Items table
            table_top = min(client_y - 70, fy - 10)
            row_h = 8.35
            col_x = [left, left + 50, left + 275, left + 315, left + 375, right]
            headers = ["Código", "Producto", "Cantidad", "Precio Unitario", "Importe"]
            c.setFillColor(colors.black)
            c.rect(left, table_top - row_h, right - left, row_h, fill=1, stroke=1)
            c.setFillColor(colors.white)
            c.setFont("Helvetica-Bold", 6.5)
            for i, h in enumerate(headers):
                c.drawCentredString((col_x[i] + col_x[i + 1]) / 2, table_top - 6.3, h)
            c.setFillColor(colors.black)
            c.setFont("Helvetica", 6)
            for x in col_x:
                c.line(x, table_top, x, table_top - row_h * 28)
            for r in range(29):
                yy = table_top - row_h * r
                c.line(left, yy, right, yy)

            items = quote.get("items", [])
            for idx in range(27):
                row_y = table_top - row_h * (idx + 1) - 6.1
                item = items[idx] if idx < len(items) else None
                if not item:
                    continue
                c.drawCentredString((col_x[0] + col_x[1]) / 2, row_y, str(item.get("cip") or ""))
                c.drawString(col_x[1] + 2, row_y, str(item.get("description") or "")[:54])
                c.drawCentredString((col_x[2] + col_x[3]) / 2, row_y, f"{float(item.get('quantity') or 0):g}")
                c.drawCentredString((col_x[3] + col_x[4]) / 2, row_y, f"${float(item.get('unit_price') or 0):,.2f}")
                c.drawCentredString((col_x[4] + col_x[5]) / 2, row_y, f"${float(item.get('line_total') or 0):,.2f}")

            total_y = table_top - row_h * 29 - 8
            c.rect(col_x[3], total_y - 9, col_x[4] - col_x[3], 9, fill=0, stroke=1)
            c.rect(col_x[4], total_y - 9, col_x[5] - col_x[4], 9, fill=0, stroke=1)
            c.setFont("Helvetica-Bold", 6.5)
            c.drawCentredString((col_x[3] + col_x[4]) / 2, total_y - 6.5, "Total:")
            c.drawCentredString((col_x[4] + col_x[5]) / 2, total_y - 6.5, f"${float(quote.get('total') or 0):,.2f}")

            # Observations
            observations_top = total_y - 18
            observations_height = 34
            observations_bottom = observations_top - observations_height
            c.rect(left, observations_bottom, right - left, observations_height, fill=0, stroke=1)
            c.setFont("Helvetica-Bold", 6.5)
            c.drawString(left + 4, observations_top - 8, "Observaciones:")
            c.setFont("Helvetica", 6)
            _draw_wrapped_pdf_text(c, quote.get("notes") or "", left + 58, observations_top - 8, 112, 3, 8)

            # Notes and costs section
            notes_top = observations_bottom - 10
            notes_left = left
            notes_right = col_x[3]
            costs_left = col_x[3]
            costs_mid = col_x[4]
            costs_right = right
            notes_bottom = notes_top - 78
            c.rect(notes_left, notes_bottom, notes_right - notes_left, 78, fill=0, stroke=1)
            c.rect(costs_left, notes_bottom, costs_right - costs_left, 78, fill=0, stroke=1)
            c.line(costs_mid, notes_bottom, costs_mid, notes_top)
            c.line(costs_left, notes_top - 39, costs_right, notes_top - 39)
            c.setFont("Helvetica", 6)
            c.drawCentredString((notes_left + notes_right) / 2, notes_top - 8, "Notas:")
            bullet_x = notes_left + 4
            indent_x = notes_left + 14
            c.drawString(bullet_x, notes_top - 18, "• Precios sujetos a cambio sin previo aviso")
            c.drawString(bullet_x, notes_top - 38, "• Tiempo de entrega: 2 a 3 dias segun paqueteria")
            bank = quote.get("bank_account")
            if bank:
                pay_y = notes_top - 48
                c.drawString(bullet_x, pay_y, "• Forma de pago:")
                bank_line = f"Transferencia a nombre de {bank.get('beneficiary') or ''} Banco {bank.get('bank_name') or ''} Cuenta: {bank.get('reference') or ''}"
                max_bank_chars = 90
                bank_extra = 0
                if len(bank_line) > max_bank_chars:
                    c.drawString(indent_x, pay_y - 8, bank_line[:max_bank_chars])
                    c.drawString(indent_x, pay_y - 15, bank_line[max_bank_chars:max_bank_chars + max_bank_chars])
                    bank_extra = 7
                else:
                    c.drawString(indent_x, pay_y - 8, bank_line)
                if bank.get("clabe"):
                    c.drawString(indent_x, pay_y - 15 - bank_extra, f"Cuenta Clabe: {bank.get('clabe')}")
            c.drawCentredString((costs_left + costs_mid) / 2, notes_top - 20, "Costos Adicionales")
            c.drawCentredString((costs_mid + costs_right) / 2, notes_top - 14, "8% sobre el total para rebanados, solo en")
            c.drawCentredString((costs_mid + costs_right) / 2, notes_top - 23, "jamones. Gastos de envio a cargo del")
            c.drawCentredString((costs_mid + costs_right) / 2, notes_top - 32, "comprador")
            c.setFont("Helvetica-Bold", 6.2)
            if quote.get("show_shipping", True):
                c.drawCentredString((costs_mid + costs_right) / 2, notes_bottom + 25, "Costo de envio autorizado")
                c.drawCentredString((costs_mid + costs_right) / 2, notes_bottom + 15, f"${float(quote.get('authorized_shipping') or 0):,.2f}")

            c.save()
            tmp.close()
            return FileResponse(tmp.name, media_type="application/pdf",
                                filename=f"cotizacion_{quote_id[:8]}.pdf",
                                headers={"Content-Disposition": f"inline; filename=cotizacion_{quote_id[:8]}.pdf",
                                         "Access-Control-Expose-Headers": "Content-Disposition"})
    except Exception:
        import logging
        logging.getLogger("crm").error("Export error:\n%s", traceback.format_exc())
        raise
    finally:
        db.close()


# ── Bank accounts ──


@router.get("/bank-accounts")
def list_bank_accounts(user=Depends(require_user)):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_bank_accounts ORDER BY bank_name")
        return _rows(cur.fetchall())
    finally:
        db.close()


class BankAccountCreate(BaseModel):
    company: Optional[str] = ""
    bank_name: str
    clabe: str
    account: Optional[str] = ""
    beneficial: Optional[str] = ""


@router.post("/bank-accounts")
def create_bank_account(
    payload: BankAccountCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        bid = str(uuid.uuid4())
        cur.execute(
            "INSERT INTO crm_bank_accounts (id, company, bank_name, clabe, reference, beneficiary, enabled, created_at, updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (bid, payload.company or "", payload.bank_name, payload.clabe, payload.account or "", payload.beneficial or "", 1, now, now),
        )
        db.commit()
        return {"ok": True, "id": bid}
    finally:
        db.close()


@router.get("/bank-accounts/{bank_id}")
def get_bank_account(
    bank_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_bank_accounts WHERE id = %s", (bank_id,))
        row = _row(cur.fetchone())
        if not row:
            raise HTTPException(status_code=404, detail="Cuenta no encontrada")
        return row
    finally:
        db.close()


@router.put("/bank-accounts/{bank_id}")
def update_bank_account(
    bank_id: str,
    payload: BankAccountCreate,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        now = _now()
        cur.execute(
            "UPDATE crm_bank_accounts SET company=%s, bank_name=%s, clabe=%s, reference=%s, beneficiary=%s, updated_at=%s WHERE id=%s",
            (payload.company or "", payload.bank_name, payload.clabe, payload.account or "", payload.beneficial or "", now, bank_id),
        )
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.delete("/bank-accounts/{bank_id}")
def delete_bank_account(
    bank_id: str,
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("DELETE FROM crm_bank_accounts WHERE id=%s", (bank_id,))
        db.commit()
        return {"ok": True}
    finally:
        db.close()


# ── Import clients ──


@router.post("/import")
def import_clients(
    file: UploadFile = File(...),
    user=Depends(require_user),
):
    db = _crm_db()
    try:
        content = file.file.read()
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        cur = db.cursor(dictionary=True)
        now = _now()
        imported = 0
        errors = 0
        for row in reader:
            try:
                name = (row.get("nombre") or row.get("name") or "").strip()
                if not name:
                    errors += 1
                    continue
                code = (row.get("codigo") or row.get("code") or "").strip()
                phone = (row.get("telefono") or row.get("phone") or "").strip()
                email = (row.get("email") or "").strip()
                contact = (row.get("contacto") or row.get("contact") or "").strip()
                tax_addr = (row.get("direccion_fiscal") or "").strip()
                cons_addr = (row.get("direccion_consignatario") or "").strip()
                delivery = (row.get("entrega") or row.get("delivery_method") or "").strip()
                seller = (row.get("vendedor") or row.get("external_seller") or "").strip()
                cid = str(uuid.uuid4())
                cur.execute(
                    "INSERT INTO crm_clients (id, code, name, phone, email, contact_name, tax_address, consignee_address, delivery_method, external_seller, created_at, updated_at) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (cid, code, name, phone, email, contact, tax_addr, cons_addr, delivery, seller, now, now),
                )
                imported += 1
            except Exception:
                errors += 1
        db.commit()
        return {"ok": True, "imported": imported, "errors": errors}
    finally:
        db.close()
