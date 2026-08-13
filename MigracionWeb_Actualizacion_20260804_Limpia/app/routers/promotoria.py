import io
import math
import re
import statistics
import time
import unicodedata
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher

import pandas as pd
from fastapi import APIRouter, Body, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.dependencies import require_user
from app.legacy_db import get_legacy_connection


router = APIRouter(prefix="/api/promotoria", tags=["promotoria"])


def _text(value) -> str:
    return str(value or "").strip()


def _client_number(value) -> str:
    """Preserva claves de cliente al venir de una celda numérica de Excel."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = _text(value)
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def _money(value) -> float:
    try:
        number = float(str(value or 0).replace("$", "").replace(",", "").strip() or 0)
        return round(number, 2) if math.isfinite(number) else 0.0
    except Exception:
        return 0.0


def _percent(value) -> float:
    try:
        number = float(str(value or 0).replace("%", "").strip() or 0)
        return round(number / 100, 4) if number > 1 else round(number, 4)
    except Exception:
        return 0.0


def _boolish(value) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    return 1 if str(value or "").strip().lower() in {"1", "true", "si", "sí", "activo", "activa", "on"} else 0


def _year_month(value):
    text = _text(value).lower()
    match = re.search(r"(20\d{2})\s*[/-]\s*(\d{1,2})", text)
    if match:
        month = int(match.group(2))
        if 1 <= month <= 12:
            return int(match.group(1)), month
    months = {
        "ene": 1, "enero": 1, "feb": 2, "febrero": 2, "mar": 3, "marzo": 3,
        "abr": 4, "abril": 4, "may": 5, "mayo": 5, "jun": 6, "junio": 6,
        "jul": 7, "julio": 7, "ago": 8, "agosto": 8, "sep": 9, "sept": 9,
        "septiembre": 9, "oct": 10, "octubre": 10, "nov": 11, "noviembre": 11,
        "dic": 12, "diciembre": 12,
    }
    clean = re.sub(r"[^a-z0-9]+", " ", text).strip()
    for token in clean.split():
        if token not in months:
            continue
        suffix = re.search(rf"\b{re.escape(token)}\b\s*(?:[-/]\s*)?(\d{{2}}|20\d{{2}})\b", clean)
        if suffix:
            year = int(suffix.group(1))
            return (2000 + year if year < 100 else year), months[token]
    for token in clean.split():
        if token in months:
            year_match = re.search(r"(20\d{2})", text)
            return (int(year_match.group(1)) if year_match else datetime.now().year), months[token]
    return None


def _key(value) -> str:
    text = unicodedata.normalize("NFKD", _text(value)).encode("ascii", "ignore").decode("ascii").upper()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\bDET\.?\s*\d+\b", " ", text)
    text = re.sub(r"\bWAL\s*-?\s*MART\s+EXPRESS\b", "SUPERAMA", text)
    text = re.sub(r"\bWALMART\s+EXPRESS\b", "SUPERAMA", text)
    text = re.sub(r"\bWME\b", "SUPERAMA", text)
    text = re.sub(r"^WE\b", "SUPERAMA", text)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return " ".join(text.split())


def _invalid_store_name(value) -> bool:
    text = _text(value).strip().upper()
    key = _key(text)
    return not key or key in {"N A", "NA", "NAN", "NULL", "NONE", "SIN DATO", "SIN DATOS", "NO APLICA"}


def _sheet_header(ws):
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = [_key(item) for item in row]
        if "PROMOTOR" in values and "TIENDA" in values:
            return row_number, {value: idx for idx, value in enumerate(values) if value}
    return None, {}


def _read_structure(content: bytes, filename: str = "") -> tuple[list[dict], list[dict]]:
    try:
        book = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo XLSX: {exc}")

    source = None
    header_row = None
    headers = {}
    for ws in book.worksheets:
        found_row, found_headers = _sheet_header(ws)
        if found_row:
            source, header_row, headers = ws, found_row, found_headers
            if _key(ws.title) in {"BSE GEN", "BASE GENERAL"}:
                break
    if source is None:
        raise HTTPException(status_code=400, detail="No se encontro una hoja con las columnas PROMOTOR y TIENDA.")

    required = ["PROMOTOR", "CADENA", "STATUS", "ACTIVOS INACTIVAS", "SUPERVISOR", "ESTADO", "TIENDA"]
    missing = [item for item in required if item not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan columnas en la base: {', '.join(missing)}")

    rows = []
    history_rows = []
    header_values = list(source.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    explicit_periods = []
    for idx, value in enumerate(header_values):
        period = _year_month(value)
        if period and re.search(r"\d{2}|20\d{2}", _text(value)):
            explicit_periods.append((idx, period))
    max_explicit_year = max((period[0] for _, period in explicit_periods), default=datetime.now().year)
    previous_periods = [(idx, period) for idx, period in explicit_periods if period[0] < max_explicit_year]
    for raw in source.iter_rows(min_row=header_row + 1, values_only=True):
        def value(name):
            index = headers.get(name)
            return _text(raw[index]) if index is not None and index < len(raw) else ""

        tienda = value("TIENDA")
        if _invalid_store_name(tienda):
            continue
        rows.append({
            "promotor": value("PROMOTOR"),
            "horario": value("HORARIO"),
            "cadena": value("CADENA"),
            "status": value("STATUS"),
            "activa": value("ACTIVOS INACTIVAS"),
            "supervisor": value("SUPERVISOR"),
            "estado": value("ESTADO"),
            "tienda": tienda,
        })
        for idx, (year, month) in previous_periods:
            amount = _money(raw[idx] if idx < len(raw) else 0)
            if amount == 0:
                continue
            history_rows.append({
                "fuente": "BASE ESTRUCTURA HISTORICA",
                "archivo": _text(filename or "Base Estructura Cuotas")[:255],
                "cadena": value("CADENA"),
                "tienda_codigo": "",
                "tienda": tienda,
                "tienda_key": _key(tienda),
                "anio": year,
                "mes": month,
                "venta": amount,
            })
    if not rows:
        raise HTTPException(status_code=400, detail="La base no contiene asignaciones de promotoría.")
    if not history_rows:
        history_rows = _read_structure_history(book, filename)
    return rows, history_rows


def _read_structure_history(book, filename: str = "") -> list[dict]:
    """Lee los meses históricos explícitos de la base de cuotas, sin reemplazar la estructura."""
    for ws in book.worksheets:
        found_row, found_headers = None, {}
        for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [_key(item) for item in row]
            has_store = any(item in values for item in ("TIENDA", "TIENDAS", "TIENDAS TOTALES"))
            if "PROMOTOR" in values and has_store:
                found_row = row_number
                found_headers = {value: idx for idx, value in enumerate(values) if value}
                break
        if not found_row:
            continue
        header_values = list(ws.iter_rows(min_row=found_row, max_row=found_row, values_only=True))[0]
        explicit_periods = []
        for idx, value in enumerate(header_values):
            period = _year_month(value)
            if period and re.search(r"\d{2}|20\d{2}", _text(value)):
                explicit_periods.append((idx, period))
        max_explicit_year = max((period[0] for _, period in explicit_periods), default=0)
        previous_periods = [(idx, period) for idx, period in explicit_periods if period[0] < max_explicit_year]
        if not previous_periods:
            continue
        tienda_idx = found_headers.get("TIENDA") if found_headers.get("TIENDA") is not None else found_headers.get("TIENDAS TOTALES")
        cadena_idx = found_headers.get("CADENA")
        if tienda_idx is None:
            continue
        rows = []
        for raw in ws.iter_rows(min_row=found_row + 1, values_only=True):
            tienda = _text(raw[tienda_idx]) if tienda_idx < len(raw) else ""
            if _invalid_store_name(tienda):
                continue
            cadena = _text(raw[cadena_idx]) if cadena_idx is not None and cadena_idx < len(raw) else ""
            for idx, (year, month) in previous_periods:
                amount = _money(raw[idx] if idx < len(raw) else 0)
                if amount == 0:
                    continue
                rows.append({
                    "fuente": "BASE ESTRUCTURA HISTORICA",
                    "archivo": _text(filename or "Base Estructura Cuotas")[:255],
                    "cadena": cadena,
                    "tienda_codigo": "",
                    "tienda": tienda,
                    "tienda_key": _key(tienda),
                    "anio": year,
                    "mes": month,
                    "venta": amount,
                })
        if rows:
            return rows
    return []


def _ensure_tables(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS promotoria_asignaciones (
            id INT AUTO_INCREMENT PRIMARY KEY,
            promotor VARCHAR(180) NOT NULL DEFAULT '',
            horario VARCHAR(100) NOT NULL DEFAULT '',
            cadena VARCHAR(100) NOT NULL DEFAULT '',
            status VARCHAR(100) NOT NULL DEFAULT '',
            activa VARCHAR(100) NOT NULL DEFAULT '',
            supervisor VARCHAR(180) NOT NULL DEFAULT '',
            estado VARCHAR(100) NOT NULL DEFAULT '',
            tienda VARCHAR(255) NOT NULL,
            tienda_key VARCHAR(255) NOT NULL,
            cliente_numero VARCHAR(80) NOT NULL DEFAULT '',
            cliente_nombre VARCHAR(255) NOT NULL DEFAULT '',
            cliente_empresa VARCHAR(120) NOT NULL DEFAULT '',
            match_score DECIMAL(6,4) NOT NULL DEFAULT 0,
            actualizado_en DATETIME NOT NULL,
            INDEX idx_promotoria_tienda_key (tienda_key),
            INDEX idx_promotoria_cliente (cliente_numero),
            INDEX idx_promotoria_promotor (promotor)
        ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS promotoria_promotores (
            id INT AUTO_INCREMENT PRIMARY KEY,
            nombre VARCHAR(180) NOT NULL,
            supervisor VARCHAR(180) NOT NULL DEFAULT '',
            activo TINYINT(1) NOT NULL DEFAULT 1,
            cuota_objetivo DECIMAL(14,2) NOT NULL DEFAULT 0,
            porcentaje_objetivo DECIMAL(8,4) NOT NULL DEFAULT 0.10,
            notas TEXT NULL,
            creado_en DATETIME NOT NULL,
            actualizado_en DATETIME NOT NULL,
            UNIQUE KEY ux_promotoria_promotor_nombre (nombre),
            INDEX idx_promotoria_promotores_supervisor (supervisor),
            INDEX idx_promotoria_promotores_activo (activo)
        ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS promotoria_supervisores (
            id INT AUTO_INCREMENT PRIMARY KEY,
            nombre VARCHAR(180) NOT NULL,
            activo TINYINT(1) NOT NULL DEFAULT 1,
            creado_en DATETIME NOT NULL,
            actualizado_en DATETIME NOT NULL,
            UNIQUE KEY ux_promotoria_supervisor_nombre (nombre),
            INDEX idx_promotoria_supervisor_activo (activo)
        ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS promotoria_sellout (
            id INT AUTO_INCREMENT PRIMARY KEY,
            fuente VARCHAR(120) NOT NULL DEFAULT '',
            archivo VARCHAR(255) NOT NULL DEFAULT '',
            cadena VARCHAR(100) NOT NULL DEFAULT '',
            tienda_codigo VARCHAR(80) NOT NULL DEFAULT '',
            tienda VARCHAR(255) NOT NULL,
            tienda_key VARCHAR(255) NOT NULL,
            anio INT NOT NULL,
            mes INT NOT NULL,
            venta DECIMAL(14,2) NOT NULL DEFAULT 0,
            actualizado_en DATETIME NOT NULL,
            UNIQUE KEY ux_promotoria_sellout_mes (fuente, anio, mes, tienda_key, tienda_codigo),
            INDEX idx_promotoria_sellout_tienda (tienda_key),
            INDEX idx_promotoria_sellout_periodo (anio, mes)
        ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """
    )
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS promotoria_comisiones_porcentajes (
            porcentaje DECIMAL(8,4) PRIMARY KEY,
            monto DECIMAL(14,2) NOT NULL DEFAULT 0,
            activo TINYINT(1) NOT NULL DEFAULT 1,
            actualizado_en DATETIME NOT NULL
        ) CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci
        """
    )
    cursor.execute("SHOW COLUMNS FROM promotoria_asignaciones")
    existing = {row.get("Field") for row in cursor.fetchall()}
    alter_columns = {
        "cuota_objetivo": "ALTER TABLE promotoria_asignaciones ADD COLUMN cuota_objetivo DECIMAL(14,2) NOT NULL DEFAULT 0",
        "porcentaje_objetivo": "ALTER TABLE promotoria_asignaciones ADD COLUMN porcentaje_objetivo DECIMAL(8,4) NOT NULL DEFAULT 0",
        "tienda_codigo": "ALTER TABLE promotoria_asignaciones ADD COLUMN tienda_codigo VARCHAR(80) NOT NULL DEFAULT ''",
    }
    for column, sql in alter_columns.items():
        if column not in existing:
            cursor.execute(sql)


def _ensure_commission_percentages(cursor):
    now = datetime.now()
    cursor.executemany(
        """
        INSERT IGNORE INTO promotoria_comisiones_porcentajes
        (porcentaje, monto, activo, actualizado_en)
        VALUES (%s, %s, %s, %s)
        """,
        [(0.03, 0, 1, now), (0.08, 0, 1, now), (0.12, 0, 1, now), (0.15, 0, 1, now)],
    )


def _customers(cursor):
    cursor.execute("SELECT numero, nombre, empresa FROM clientes WHERE nombre IS NOT NULL AND TRIM(nombre) <> ''")
    return [dict(row, tienda_key=_key(row.get("nombre"))) for row in cursor.fetchall()]


def _match_customer(tienda: str, customers: list[dict]) -> tuple[dict | None, float]:
    wanted = _key(tienda)
    if not wanted:
        return None, 0.0
    exact = [item for item in customers if item["tienda_key"] == wanted]
    if len(exact) == 1:
        return exact[0], 1.0

    wanted_tokens = set(wanted.split())
    best, best_score = None, 0.0
    for item in customers:
        candidate = item["tienda_key"]
        tokens = set(candidate.split())
        if not tokens:
            continue
        overlap = len(wanted_tokens & tokens) / max(len(wanted_tokens), 1)
        ratio = SequenceMatcher(None, wanted, candidate).ratio()
        score = max(overlap, ratio)
        if wanted in candidate or candidate in wanted:
            score = max(score, 0.92)
        if score > best_score:
            best, best_score = item, score
    return (best, best_score) if best_score >= 0.72 else (None, best_score)


def _replace_structure(cursor, rows: list[dict]) -> dict:
    _ensure_tables(cursor)
    customers = _customers(cursor)
    cursor.execute("DELETE FROM promotoria_asignaciones")
    now = datetime.now()
    linked = 0
    for row in rows:
        customer, score = _match_customer(row["tienda"], customers)
        if customer:
            linked += 1
        promotor_nombre = _canonical_promotor_name(cursor, row["promotor"])
        cursor.execute(
            """
            INSERT INTO promotoria_asignaciones
            (promotor, horario, cadena, status, activa, supervisor, estado, tienda, tienda_key,
             cliente_numero, cliente_nombre, cliente_empresa, match_score, actualizado_en)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                promotor_nombre, row["horario"], row["cadena"], row["status"], row["activa"],
                row["supervisor"], row["estado"], row["tienda"], _key(row["tienda"]),
                _text(customer.get("numero")) if customer else "",
                _text(customer.get("nombre")) if customer else "",
                _text(customer.get("empresa")) if customer else "",
                round(float(score), 4), now,
            ),
        )
        _upsert_promotor_catalogo(
            cursor,
            {
                "nombre": promotor_nombre,
                "supervisor": row["supervisor"],
                "activo": row["activa"].upper() != "INACTIVA",
            },
            commit=False,
        )
    return {"asignaciones": len(rows), "vinculadas": linked, "sin_vinculo": len(rows) - linked}


def _cell_text(value) -> str:
    if value is None:
        return ""
    try:
        if isinstance(value, float) and not math.isfinite(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _find_sellout_header(frame) -> tuple[int | None, dict]:
    for idx, raw in frame.iterrows():
        values = [_key(value) for value in raw.tolist()]
        if "PROMOTOR" in values or "SUPERVISOR" in values:
            continue
        has_store = any(value in values for value in ("NOMBRE TIENDA CLUB", "NOMBRE TIENDA", "TIENDA", "TIENDAS"))
        month_columns = [pos for pos, value in enumerate(raw.tolist()) if _year_month(value)]
        if has_store and month_columns:
            return idx, {value: pos for pos, value in enumerate(values) if value}
    return None, {}


def _read_sellout(content: bytes, filename: str) -> list[dict]:
    try:
        sheets = pd.read_excel(io.BytesIO(content), sheet_name=None, header=None, dtype=object)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel de sell-out: {exc}")

    rows = []
    for sheet_name, frame in sheets.items():
        sample_text = " ".join(_key(value) for value in frame.head(20).values.flatten().tolist())
        if "PROMOTOR" in sample_text or "SUPERVISOR" in sample_text:
            continue
        if "REPORT OPTIONS" in sample_text or "REPORT COLUMNS" in sample_text:
            continue
        header_idx, headers = _find_sellout_header(frame)
        if header_idx is None:
            rows.extend(_read_sellout_month_grid(frame, sheet_name, filename))
            continue
        if header_idx is None:
            continue
        header = frame.iloc[header_idx].tolist()
        name_idx = next((headers.get(key) for key in ("NOMBRE TIENDA CLUB", "NOMBRE TIENDA", "TIENDA", "TIENDAS") if headers.get(key) is not None), None)
        code_idx = next((headers.get(key) for key in ("NUM DE TIENDA", "NUMERO DE TIENDA", "NO DE TIENDA", "CODIGO", "SUCURSAL") if headers.get(key) is not None), None)
        month_cols = [(idx, _year_month(value)) for idx, value in enumerate(header)]
        month_cols = [(idx, period) for idx, period in month_cols if period]
        if name_idx is None or not month_cols:
            continue
        for _, raw in frame.iloc[header_idx + 1:].iterrows():
            tienda = _cell_text(raw.iloc[name_idx] if name_idx < len(raw) else "")
            if not tienda or tienda.lower() == "nan":
                continue
            tienda_codigo = _cell_text(raw.iloc[code_idx] if code_idx is not None and code_idx < len(raw) else "")
            for idx, (year, month) in month_cols:
                amount = _money(raw.iloc[idx] if idx < len(raw) else 0)
                if amount == 0:
                    continue
                rows.append({
                    "fuente": _text(sheet_name)[:120],
                    "archivo": _text(filename)[:255],
                    "cadena": "WALMART" if "WM" in _key(sheet_name) or "WALMART" in _key(tienda) else "",
                    "tienda_codigo": tienda_codigo,
                    "tienda": tienda,
                    "tienda_key": _key(tienda),
                    "anio": year,
                    "mes": month,
                    "venta": amount,
                })
    if not rows:
        raise HTTPException(status_code=400, detail="No se encontraron columnas de tienda y meses en el Excel de sell-out.")
    return rows


def _read_sellout_month_grid(frame, sheet_name: str, filename: str) -> list[dict]:
    rows = []
    default_year = int(re.search(r"(20\d{2})", _text(sheet_name) + " " + _text(filename)).group(1)) if re.search(r"(20\d{2})", _text(sheet_name) + " " + _text(filename)) else datetime.now().year
    for idx, raw in frame.iterrows():
        month_cols = []
        for col_idx, value in enumerate(raw.tolist()):
            period = _year_month(value)
            if period:
                month_cols.append((col_idx, period))
            else:
                token = _key(value)
                month_only = _year_month(f"{token} {default_year}") if token else None
                if month_only:
                    month_cols.append((col_idx, month_only))
        if len(month_cols) < 2:
            continue
        for _, item in frame.iloc[idx + 1:].iterrows():
            values = item.tolist()
            if not any(_text(value) for value in values):
                break
            tienda_idx = None
            for possible in range(min(len(values), month_cols[0][0])):
                text = _cell_text(values[possible])
                if text and not text.replace(".", "", 1).isdigit() and _key(text) not in {"COSTO", "OUT", "TOTAL"}:
                    tienda_idx = possible
            if tienda_idx is None:
                continue
            tienda = _cell_text(values[tienda_idx])
            if not tienda or _key(tienda) in {"COSTO", "OUT"}:
                continue
            code_idx = next((pos for pos in range(tienda_idx) if _cell_text(values[pos]).replace(".", "", 1).isdigit()), None)
            tienda_codigo = _cell_text(values[code_idx]) if code_idx is not None else ""
            for col_idx, (year, month) in month_cols:
                amount = _money(values[col_idx] if col_idx < len(values) else 0)
                if amount == 0:
                    continue
                rows.append({
                    "fuente": _text(sheet_name)[:120],
                    "archivo": _text(filename)[:255],
                    "cadena": "LA COMER" if "CITY" in _key(sheet_name) or "CITY MARKET" in _key(tienda) else "",
                    "tienda_codigo": tienda_codigo,
                    "tienda": tienda,
                    "tienda_key": _key(tienda),
                    "anio": year,
                    "mes": month,
                    "venta": amount,
                })
        break
    return rows


def _save_sellout(cursor, rows: list[dict], delete_by_source: bool = False) -> dict:
    _ensure_tables(cursor)
    now = datetime.now()
    if delete_by_source:
        for archivo, fuente in sorted({
            (_text(row.get("archivo")), _text(row.get("fuente")))
            for row in rows
            if _text(row.get("archivo")) and _text(row.get("fuente"))
        }):
            cursor.execute("DELETE FROM promotoria_sellout WHERE archivo=%s AND fuente=%s", (archivo, fuente))
    else:
        for archivo in sorted({_text(row.get("archivo")) for row in rows if _text(row.get("archivo"))}):
            cursor.execute("DELETE FROM promotoria_sellout WHERE archivo=%s", (archivo,))
    for row in rows:
        cursor.execute(
            """
            INSERT INTO promotoria_sellout
            (fuente, archivo, cadena, tienda_codigo, tienda, tienda_key, anio, mes, venta, actualizado_en)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                archivo=VALUES(archivo), cadena=VALUES(cadena), tienda=VALUES(tienda),
                venta=VALUES(venta), actualizado_en=VALUES(actualizado_en)
            """,
            (
                row["fuente"], row["archivo"], row["cadena"], row["tienda_codigo"], row["tienda"],
                row["tienda_key"], row["anio"], row["mes"], row["venta"], now,
            ),
        )
    return {
        "registros": len(rows),
        "tiendas": len({(row["tienda_key"], row["tienda_codigo"]) for row in rows}),
        "periodos": len({(row["anio"], row["mes"]) for row in rows}),
        "total": round(sum(float(row["venta"] or 0) for row in rows), 2),
    }


def _upsert_promotor_catalogo(cursor, payload: dict, commit: bool = False) -> dict:
    _ensure_tables(cursor)
    nombre = _text(payload.get("nombre")).upper()
    if not nombre:
        raise HTTPException(status_code=400, detail="Captura el nombre del promotor.")
    now = datetime.now()
    data = {
        "nombre": nombre,
        "supervisor": _text(payload.get("supervisor")).upper(),
        "activo": _boolish(payload.get("activo", True)),
        "cuota_objetivo": _money(payload.get("cuota_objetivo")),
        "porcentaje_objetivo": _percent(payload.get("porcentaje_objetivo") if payload.get("porcentaje_objetivo") not in (None, "") else 0.10),
        "notas": _text(payload.get("notas")),
    }
    promotor_id = int(payload.get("id") or 0)
    previous_name = ""
    if promotor_id:
        cursor.execute("SELECT nombre FROM promotoria_promotores WHERE id=%s", (promotor_id,))
        previous_name = _text((cursor.fetchone() or {}).get("nombre")).upper()
        cursor.execute(
            """
            UPDATE promotoria_promotores
            SET nombre=%s, supervisor=%s, activo=%s, cuota_objetivo=%s,
                porcentaje_objetivo=%s, notas=%s, actualizado_en=%s
            WHERE id=%s
            """,
            (
                data["nombre"], data["supervisor"], data["activo"], data["cuota_objetivo"],
                data["porcentaje_objetivo"], data["notas"], now, promotor_id,
            ),
        )
        if previous_name and previous_name != data["nombre"]:
            cursor.execute(
                """
                UPDATE promotoria_asignaciones
                SET promotor=%s, supervisor=%s, porcentaje_objetivo=%s, actualizado_en=%s
                WHERE UPPER(TRIM(promotor))=%s
                """,
                (data["nombre"], data["supervisor"], data["porcentaje_objetivo"], now, previous_name),
            )
    else:
        cursor.execute(
            """
            INSERT INTO promotoria_promotores
            (nombre, supervisor, activo, cuota_objetivo, porcentaje_objetivo, notas, creado_en, actualizado_en)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                supervisor=VALUES(supervisor), activo=VALUES(activo),
                cuota_objetivo=VALUES(cuota_objetivo), porcentaje_objetivo=VALUES(porcentaje_objetivo),
                notas=VALUES(notas), actualizado_en=VALUES(actualizado_en)
            """,
            (
                data["nombre"], data["supervisor"], data["activo"], data["cuota_objetivo"],
                data["porcentaje_objetivo"], data["notas"], now, now,
            ),
        )
    if commit:
        cursor.connection.commit()
    cursor.execute("SELECT * FROM promotoria_promotores WHERE nombre=%s", (data["nombre"],))
    return cursor.fetchone()


def _upsert_supervisor_catalogo(cursor, payload: dict) -> dict:
    _ensure_tables(cursor)
    nombre = _text(payload.get("nombre")).upper()
    if not nombre:
        raise HTTPException(status_code=400, detail="Captura el nombre del supervisor.")
    now = datetime.now()
    supervisor_id = int(payload.get("id") or 0)
    activo = _boolish(payload.get("activo", True))
    if supervisor_id:
        cursor.execute(
            "UPDATE promotoria_supervisores SET nombre=%s, activo=%s, actualizado_en=%s WHERE id=%s",
            (nombre, activo, now, supervisor_id),
        )
    else:
        cursor.execute(
            """
            INSERT INTO promotoria_supervisores (nombre, activo, creado_en, actualizado_en)
            VALUES (%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE activo=VALUES(activo), actualizado_en=VALUES(actualizado_en)
            """,
            (nombre, activo, now, now),
        )
    cursor.execute("SELECT * FROM promotoria_supervisores WHERE nombre=%s", (nombre,))
    return cursor.fetchone()


def _promotores_meta(cursor) -> dict[str, dict]:
    _ensure_tables(cursor)
    cursor.execute("SELECT * FROM promotoria_promotores")
    return {_text(row.get("nombre")).upper(): row for row in cursor.fetchall()}


def _canonical_promotor_name(cursor, nombre: str) -> str:
    original = _text(nombre).upper()
    if not original:
        return ""
    meta = _promotores_meta(cursor)
    if original in meta:
        return original
    matches = [name for name in meta if name.startswith(original + " ") or name.startswith(original)]
    return matches[0] if len(matches) == 1 else original


@router.post("/importar-base")
async def importar_base(file: UploadFile = File(...), user=Depends(require_user)):
    filename = _text(file.filename).lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Selecciona la Base Estructura Cuotas en formato XLSX.")
    rows, history_rows = _read_structure(await file.read(), file.filename)
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        result = _replace_structure(cursor, rows)
        history_result = _save_sellout(cursor, history_rows, delete_by_source=True) if history_rows else {"registros": 0, "periodos": 0}
        conn.commit()
        return {
            **result,
            "historico": history_result,
            "mensaje": "Base de promotoría actualizada correctamente.",
        }
    except HTTPException:
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo importar la base: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/sellout/importar")
async def importar_sellout(file: UploadFile = File(...), user=Depends(require_user)):
    filename = _text(file.filename)
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Selecciona un archivo Excel XLSX o XLS para sell-out.")
    rows = _read_sellout(await file.read(), filename)
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        result = _save_sellout(cursor, rows)
        conn.commit()
        return {**result, "mensaje": "Sell-out importado y guardado en MySQL."}
    except HTTPException:
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo importar el sell-out: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def _load_assignments(cursor, promotor: str = "", search: str = "") -> list[dict]:
    _ensure_tables(cursor)
    sql = "SELECT * FROM promotoria_asignaciones"
    where, params = [], []
    if promotor:
        where.append("UPPER(TRIM(promotor)) = UPPER(TRIM(%s))")
        params.append(promotor)
    if search:
        where.append("(tienda LIKE %s OR promotor LIKE %s OR supervisor LIKE %s OR cadena LIKE %s)")
        params.extend([f"%{search}%"] * 4)
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY supervisor, promotor, cadena, tienda"
    cursor.execute(sql, tuple(params))
    return cursor.fetchall()


def _admin_payload(cursor) -> dict:
    _ensure_tables(cursor)
    # El catálogo se inicializa con los supervisores que ya existían en la
    # estructura importada y en los promotores; nunca reemplaza asignaciones.
    cursor.execute("SELECT DISTINCT TRIM(supervisor) AS nombre FROM promotoria_asignaciones WHERE TRIM(COALESCE(supervisor, '')) <> ''")
    existentes_estructura = [row.get("nombre") for row in cursor.fetchall()]
    cursor.execute("SELECT DISTINCT TRIM(supervisor) AS nombre FROM promotoria_promotores WHERE TRIM(COALESCE(supervisor, '')) <> ''")
    existentes_promotores = [row.get("nombre") for row in cursor.fetchall()]
    now = datetime.now()
    for nombre in {*[_text(item).upper() for item in existentes_estructura], *[_text(item).upper() for item in existentes_promotores]}:
        if nombre and nombre not in {"SIN SERVICIO", "VACANTE", "N/A", "NA"}:
            cursor.execute(
                """
                INSERT INTO promotoria_supervisores (nombre, activo, creado_en, actualizado_en)
                VALUES (%s,1,%s,%s)
                ON DUPLICATE KEY UPDATE actualizado_en=actualizado_en
                """,
                (nombre, now, now),
            )
    cursor.execute("SELECT * FROM promotoria_promotores ORDER BY activo DESC, supervisor, nombre")
    promotores = cursor.fetchall()
    cursor.execute("SELECT * FROM promotoria_supervisores ORDER BY activo DESC, nombre")
    supervisores_catalogo = cursor.fetchall()
    cursor.execute(
        """
        SELECT id, cadena, status, activa, supervisor, estado, promotor, horario, tienda,
               cliente_numero, cliente_nombre, cliente_empresa, tienda_codigo, cuota_objetivo, porcentaje_objetivo
        FROM promotoria_asignaciones
        ORDER BY supervisor, promotor, cadena, tienda
        """
    )
    tiendas = cursor.fetchall()
    # Los clientes son la fuente de verdad para el vÃ­nculo comercial.  La
    # tienda conserva su nombre de promotorÃ­a, que es el que se presenta en
    # los reportes, mientras que esta lista sÃ³lo se usa para asociarla.
    clientes = _customers(cursor)
    supervisores = sorted({
        _text(row.get("supervisor"))
        for row in [*promotores, *tiendas, *supervisores_catalogo]
        if _text(row.get("supervisor"))
    } | {_text(row.get("nombre")) for row in supervisores_catalogo if _text(row.get("nombre"))})
    cursor.execute(
        """
        SELECT COUNT(*) AS registros,
               COUNT(DISTINCT CONCAT(tienda_key, '|', tienda_codigo)) AS tiendas,
               COUNT(DISTINCT CONCAT(anio, '-', mes)) AS periodos,
               COALESCE(SUM(venta), 0) AS total
        FROM promotoria_sellout
        """
    )
    sellout = cursor.fetchone() or {}
    return {
        "promotores": promotores,
        "tiendas": tiendas,
        "clientes": sorted(clientes, key=lambda item: (_text(item.get("empresa")), _text(item.get("nombre")), _text(item.get("numero")))),
        "supervisores": supervisores,
        "supervisores_catalogo": supervisores_catalogo,
        "summary": {
            "promotores": len(promotores),
            "activos": sum(1 for row in promotores if int(row.get("activo") or 0) == 1),
            "tiendas": len(tiendas),
            "tiendas_sin_promotor": sum(1 for row in tiendas if not _text(row.get("promotor")) or _text(row.get("promotor")).upper() == "SIN SERVICIO"),
            "tiendas_sin_cliente": sum(1 for row in tiendas if not _text(row.get("cliente_numero"))),
            "sellout_registros": int(sellout.get("registros") or 0),
            "sellout_tiendas": int(sellout.get("tiendas") or 0),
            "sellout_periodos": int(sellout.get("periodos") or 0),
            "sellout_total": round(float(sellout.get("total") or 0), 2),
        },
    }


@router.get("/admin")
def admin_promotoria(user=Depends(require_user)):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        return _admin_payload(cursor)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo cargar la administración: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/admin/asociaciones/xlsx")
def exportar_asociaciones_xlsx(user=Depends(require_user)):
    """Base editable: sólo se cambian número y empresa del cliente vinculado."""
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        cursor.execute(
            """
            SELECT id, cadena, tienda, promotor, supervisor, cliente_numero, cliente_empresa
            FROM promotoria_asignaciones
            ORDER BY cadena, tienda
            """
        )
        rows = cursor.fetchall()
        book = Workbook()
        sheet = book.active
        sheet.title = "Asociaciones"
        headers = ["ID_ASIGNACION", "CADENA", "TIENDA_PERSONALIZADA", "PROMOTOR_ACTUAL", "SUPERVISOR", "CLIENTE_NUMERO", "CLIENTE_EMPRESA"]
        sheet.append(headers)
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
        for row in rows:
            sheet.append([
                row.get("id"), row.get("cadena"), row.get("tienda"), row.get("promotor"), row.get("supervisor"),
                _client_number(row.get("cliente_numero")), row.get("cliente_empresa"),
            ])
        sheet.freeze_panes = "A2"
        widths = [16, 22, 42, 24, 24, 20, 22]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index)].width = width
        note = book.create_sheet("Instrucciones")
        note["A1"] = "Cómo actualizar vínculos"
        note["A1"].font = Font(bold=True, size=14)
        note["A3"] = "No cambies ID_ASIGNACION ni TIENDA_PERSONALIZADA. Sólo captura CLIENTE_NUMERO y CLIENTE_EMPRESA; después importa este mismo archivo."
        note.column_dimensions["A"].width = 125
        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="asociaciones_promotoria.xlsx"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo generar la base de asociaciones: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/admin/asociaciones/xlsx")
async def importar_asociaciones_xlsx(file: UploadFile = File(...), user=Depends(require_user)):
    if not _text(file.filename).lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Selecciona el archivo XLSX descargado de asociaciones de promotoría.")
    try:
        book = load_workbook(io.BytesIO(await file.read()), data_only=True)
        sheet = book["Asociaciones"] if "Asociaciones" in book.sheetnames else book.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el XLSX: {exc}")
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene registros.")
    headers = {_key(value): index for index, value in enumerate(rows[0]) if _key(value)}
    required = ["ID ASIGNACION", "CLIENTE NUMERO", "CLIENTE EMPRESA"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan columnas: {', '.join(missing)}.")
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        updated = skipped = 0
        errors = []
        for line, values in enumerate(rows[1:], start=2):
            assignment_id = int(values[headers["ID ASIGNACION"]] or 0)
            customer_number = _client_number(values[headers["CLIENTE NUMERO"]])
            customer_company = _text(values[headers["CLIENTE EMPRESA"]])
            if not assignment_id:
                skipped += 1
                continue
            cursor.execute("SELECT id FROM promotoria_asignaciones WHERE id=%s", (assignment_id,))
            if not cursor.fetchone():
                errors.append(f"Fila {line}: no existe el ID de asignación {assignment_id}.")
                continue
            customer = None
            if customer_number:
                sql = "SELECT numero, nombre, empresa FROM clientes WHERE TRIM(COALESCE(numero,''))=%s"
                params = [customer_number]
                if customer_company:
                    sql += " AND TRIM(COALESCE(empresa,''))=%s"
                    params.append(customer_company)
                sql += " ORDER BY empresa, nombre LIMIT 1"
                cursor.execute(sql, tuple(params))
                customer = cursor.fetchone()
                if not customer:
                    errors.append(f"Fila {line}: no se encontró el cliente {customer_number}{' en ' + customer_company if customer_company else ''}.")
                    continue
            cursor.execute(
                """
                UPDATE promotoria_asignaciones
                SET cliente_numero=%s, cliente_nombre=%s, cliente_empresa=%s,
                    match_score=%s, actualizado_en=%s
                WHERE id=%s
                """,
                (
                    _client_number(customer.get("numero")) if customer else "",
                    _text(customer.get("nombre")) if customer else "",
                    _text(customer.get("empresa")) if customer else "",
                    1.0 if customer else 0.0,
                    datetime.now(), assignment_id,
                ),
            )
            updated += 1
        conn.commit()
        return {"ok": True, "actualizadas": updated, "omitidas": skipped, "errores": errors[:30], "mensaje": f"{updated} asociación(es) actualizada(s)."}
    except HTTPException:
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudieron importar las asociaciones: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/admin/promotores")
def guardar_promotor(payload: dict = Body(...), user=Depends(require_user)):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        promotor = _upsert_promotor_catalogo(cursor, payload)
        conn.commit()
        return {"ok": True, "promotor": promotor, "mensaje": "Promotor guardado."}
    except HTTPException:
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo guardar el promotor: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/admin/promotores/{promotor_id}/baja")
def baja_promotor(promotor_id: int, user=Depends(require_user)):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        cursor.execute(
            "UPDATE promotoria_promotores SET activo=0, actualizado_en=%s WHERE id=%s",
            (datetime.now(), promotor_id),
        )
        conn.commit()
        return {"ok": True, "mensaje": "Promotor dado de baja."}
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo dar de baja el promotor: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/admin/supervisores")
def guardar_supervisor(payload: dict = Body(...), user=Depends(require_user)):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        supervisor = _upsert_supervisor_catalogo(cursor, payload)
        conn.commit()
        return {"ok": True, "supervisor": supervisor, "mensaje": "Supervisor guardado."}
    except HTTPException:
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo guardar el supervisor: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/admin/asignaciones")
def guardar_asignacion(payload: dict = Body(...), user=Depends(require_user)):
    assignment_id = int(payload.get("id") or payload.get("assignment_id") or 0)
    if not assignment_id:
        raise HTTPException(status_code=400, detail="Selecciona una tienda/asignación.")
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        promotor = _text(payload.get("promotor")).upper()
        supervisor = _text(payload.get("supervisor")).upper()
        cliente_numero = _text(payload.get("cliente_numero"))
        cliente_empresa = _text(payload.get("cliente_empresa"))
        cliente_nombre = ""
        if cliente_numero:
            customer_sql = "SELECT numero, nombre, empresa FROM clientes WHERE TRIM(COALESCE(numero, ''))=%s"
            customer_params = [cliente_numero]
            if cliente_empresa:
                customer_sql += " AND TRIM(COALESCE(empresa, ''))=%s"
                customer_params.append(cliente_empresa)
            customer_sql += " ORDER BY empresa, nombre LIMIT 1"
            cursor.execute(customer_sql, tuple(customer_params))
            customer = cursor.fetchone()
            if not customer:
                raise HTTPException(status_code=400, detail="El cliente seleccionado ya no existe en el catÃ¡logo de clientes.")
            cliente_numero = _text(customer.get("numero"))
            cliente_nombre = _text(customer.get("nombre"))
            cliente_empresa = _text(customer.get("empresa"))
        cursor.execute(
            """
            UPDATE promotoria_asignaciones
            SET promotor=%s, supervisor=%s, status=%s, activa=%s, horario=%s, tienda_codigo=%s,
                cuota_objetivo=%s, porcentaje_objetivo=%s, cliente_numero=%s, cliente_nombre=%s,
                cliente_empresa=%s, match_score=%s, actualizado_en=%s
            WHERE id=%s
            """,
            (
                promotor,
                supervisor,
                _text(payload.get("status")).upper(),
                _text(payload.get("activa")).upper(),
                _text(payload.get("horario")).upper(),
                _text(payload.get("tienda_codigo")),
                _money(payload.get("cuota_objetivo")),
                _percent(payload.get("porcentaje_objetivo")),
                cliente_numero,
                cliente_nombre,
                cliente_empresa,
                1.0 if cliente_numero else 0.0,
                datetime.now(),
                assignment_id,
            ),
        )
        if promotor:
            _upsert_promotor_catalogo(
                cursor,
                {
                    "nombre": promotor,
                    "supervisor": supervisor,
                    "activo": True,
                    "porcentaje_objetivo": payload.get("porcentaje_objetivo") or 0.10,
                },
            )
        conn.commit()
        return {"ok": True, "mensaje": "Asignación actualizada."}
    except HTTPException:
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo actualizar la asignación: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/sellout/sin-asignar")
def sellout_sin_asignar(year: int = Query(default_factory=lambda: datetime.now().year, ge=2020, le=2100), user=Depends(require_user)):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        rows = _sellout_unassigned(cursor, year)
        return {
            "year": year,
            "rows": rows,
            "summary": {
                "tiendas": len(rows),
                "total": round(sum(float(row.get("total") or 0) for row in rows), 2),
            },
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo cargar sell-out sin asignar: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/sellout/asignar")
def asignar_sellout(payload: dict = Body(...), user=Depends(require_user)):
    tienda = _text(payload.get("tienda"))
    if not tienda:
        raise HTTPException(status_code=400, detail="Indica la tienda a asignar.")
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        tienda_codigo = _text(payload.get("tienda_codigo"))
        tienda_key = _text(payload.get("tienda_key")) or _key(tienda)
        promotor = (_text(payload.get("promotor")) or "SIN SERVICIO").upper()
        supervisor = (_text(payload.get("supervisor")) or "SIN SERVICIO").upper()
        status = (_text(payload.get("status")) or "SIN SERVICIO").upper()
        activa = (_text(payload.get("activa")) or "ACTIVA").upper()
        customers = _customers(cursor)
        customer, score = _match_customer(tienda, customers)
        cursor.execute(
            """
            SELECT id FROM promotoria_asignaciones
            WHERE tienda_key=%s OR (%s <> '' AND tienda_codigo=%s)
            ORDER BY id LIMIT 1
            """,
            (tienda_key, tienda_codigo, tienda_codigo),
        )
        existing = cursor.fetchone()
        values = (
            promotor,
            _text(payload.get("horario")),
            _text(payload.get("cadena")).upper(),
            status,
            activa,
            supervisor,
            _text(payload.get("estado")).upper(),
            tienda,
            tienda_key,
            _text(customer.get("numero")) if customer else "",
            _text(customer.get("nombre")) if customer else "",
            _text(customer.get("empresa")) if customer else "",
            round(float(score), 4),
            tienda_codigo,
            _money(payload.get("cuota_objetivo")),
            _percent(payload.get("porcentaje_objetivo")),
            datetime.now(),
        )
        if existing:
            cursor.execute(
                """
                UPDATE promotoria_asignaciones
                SET promotor=%s, horario=%s, cadena=%s, status=%s, activa=%s,
                    supervisor=%s, estado=%s, tienda=%s, tienda_key=%s,
                    cliente_numero=%s, cliente_nombre=%s, cliente_empresa=%s,
                    match_score=%s, tienda_codigo=%s, cuota_objetivo=%s,
                    porcentaje_objetivo=%s, actualizado_en=%s
                WHERE id=%s
                """,
                (*values, existing["id"]),
            )
            assignment_id = existing["id"]
        else:
            cursor.execute(
                """
                INSERT INTO promotoria_asignaciones
                (promotor, horario, cadena, status, activa, supervisor, estado, tienda, tienda_key,
                 cliente_numero, cliente_nombre, cliente_empresa, match_score, tienda_codigo,
                 cuota_objetivo, porcentaje_objetivo, actualizado_en)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                values,
            )
            assignment_id = cursor.lastrowid
        if promotor:
            _upsert_promotor_catalogo(
                cursor,
                {
                    "nombre": promotor,
                    "supervisor": supervisor,
                    "activo": promotor != "SIN SERVICIO",
                    "porcentaje_objetivo": payload.get("porcentaje_objetivo") or 0.10,
                },
            )
        conn.commit()
        return {"ok": True, "id": assignment_id, "mensaje": "Tienda agregada a estructura de promotoría."}
    except HTTPException:
        raise
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo asignar la tienda: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/diagnostico")
def diagnostico(user=Depends(require_user)):
    last_error = None
    for attempt in range(3):
        conn = cursor = None
        try:
            conn = get_legacy_connection()
            conn.autocommit = True
            cursor = conn.cursor(dictionary=True)
            _ensure_tables(cursor)
            data = _admin_payload(cursor)
            return {
                "ok": True,
                "base_datos": [
                    {"dato": "Cadena, status, activa, supervisor, estado, promotor y tienda", "estado": "Disponible", "fuente": "promotoria_asignaciones"},
                    {"dato": "Alta/baja de promotores, supervisor, cuota y porcentaje", "estado": "Disponible", "fuente": "promotoria_promotores"},
                    {"dato": "Sell-out externo importado", "estado": f"{data.get('summary', {}).get('sellout_registros', 0)} registros", "fuente": "promotoria_sellout"},
                    {"dato": "Facturas internas", "estado": "Referencia", "fuente": "facturas / factura_detalle"},
                    {"dato": "Media y cuotas + porcentaje", "estado": "Calculado", "fuente": "sell-out + configuración"},
                ],
                "faltantes": [
                    {"dato": "Número de tienda POS o identificador de cadena", "accion": "Capturarlo en Administración para cada tienda cuando aplique."},
                    {"dato": "Seguimiento semanal por bloques", "accion": "Definir si se alimentará manualmente o desde ventas diarias facturadas."},
                    {"dato": "Objetivos mensuales oficiales por tienda", "accion": "Guardar objetivo/porcentaje en MySQL; el Excel solo debe servir como referencia inicial."},
                ],
                "summary": data.get("summary", {}),
            }
        except Exception as exc:
            last_error = exc
            if "1213" not in str(exc) or attempt == 2:
                break
            time.sleep(0.25 * (attempt + 1))
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
    raise HTTPException(status_code=500, detail=f"No se pudo validar el diagnóstico: {last_error}")


def _sales_by_customer(cursor, year: int, customer_numbers: list[str]):
    if not customer_numbers:
        return {}, {}, {}
    placeholders = ",".join(["%s"] * len(customer_numbers))
    cursor.execute(
        f"""
        SELECT f.numero_cliente, MONTH(f.fecha) AS mes, SUM(COALESCE(f.total, 0)) AS total,
               COUNT(DISTINCT f.id) AS facturas
        FROM facturas f
        WHERE YEAR(f.fecha) = %s
          AND UPPER(TRIM(COALESCE(f.estatus, 'ACTIVA'))) <> 'CANCELADA'
          AND UPPER(TRIM(COALESCE(f.empresa, ''))) LIKE 'GOURMET%'
          AND TRIM(COALESCE(f.numero_cliente, '')) IN ({placeholders})
        GROUP BY f.numero_cliente, MONTH(f.fecha)
        """,
        tuple([year] + customer_numbers),
    )
    months, invoices, total_by_customer = defaultdict(dict), defaultdict(int), defaultdict(float)
    for row in cursor.fetchall():
        customer = _text(row.get("numero_cliente"))
        month = int(row.get("mes") or 0)
        amount = float(row.get("total") or 0)
        months[customer][month] = amount
        invoices[customer] += int(row.get("facturas") or 0)
        total_by_customer[customer] += amount
    return months, invoices, total_by_customer


def _sellout_by_assignment(cursor, year: int, assignments: list[dict]):
    _ensure_tables(cursor)
    if not assignments:
        return {}, {}
    keys = sorted({_text(item.get("tienda_key")) for item in assignments if _text(item.get("tienda_key"))})
    codes = sorted({_text(item.get("tienda_codigo")) for item in assignments if _text(item.get("tienda_codigo"))})
    where, params = ["anio=%s"], [year]
    parts = []
    if keys:
        parts.append("tienda_key IN (" + ",".join(["%s"] * len(keys)) + ")")
        params.extend(keys)
    if codes:
        parts.append("tienda_codigo IN (" + ",".join(["%s"] * len(codes)) + ")")
        params.extend(codes)
    if not parts:
        return {}, {}
    where.append("(" + " OR ".join(parts) + ")")
    cursor.execute(
        f"""
        SELECT tienda_key, tienda_codigo, mes, SUM(COALESCE(venta, 0)) AS total
        FROM promotoria_sellout
        WHERE {" AND ".join(where)}
        GROUP BY tienda_key, tienda_codigo, mes
        """,
        tuple(params),
    )
    by_key, by_code = defaultdict(dict), defaultdict(dict)
    for row in cursor.fetchall():
        month = int(row.get("mes") or 0)
        if not 1 <= month <= 12:
            continue
        amount = float(row.get("total") or 0)
        key = _text(row.get("tienda_key"))
        code = _text(row.get("tienda_codigo"))
        if key:
            by_key[key][month] = by_key[key].get(month, 0) + amount
        if code:
            by_code[code][month] = by_code[code].get(month, 0) + amount
    return by_key, by_code


def _comparison_month(cursor, year: int) -> int:
    """Último mes disponible para comparar el año seleccionado con el anterior."""
    cursor.execute(
        "SELECT COALESCE(MAX(mes), 0) AS ultimo_mes FROM promotoria_sellout WHERE anio=%s",
        (year,),
    )
    last_month = int((cursor.fetchone() or {}).get("ultimo_mes") or 0)
    cursor.execute(
        "SELECT COALESCE(MAX(mes), 0) AS ultimo_mes FROM promotoria_sellout WHERE anio=%s",
        (year - 1,),
    )
    previous_last_month = int((cursor.fetchone() or {}).get("ultimo_mes") or 0)
    if previous_last_month:
        last_month = min(last_month or previous_last_month, previous_last_month)
    if year == datetime.now().year:
        last_month = min(last_month or datetime.now().month, datetime.now().month)
    return max(1, min(last_month or 12, 12))


def _sellout_unassigned(cursor, year: int):
    _ensure_tables(cursor)
    cursor.execute("SELECT tienda_key, tienda_codigo FROM promotoria_asignaciones")
    assigned = cursor.fetchall()
    assigned_keys = {_text(row.get("tienda_key")) for row in assigned if _text(row.get("tienda_key"))}
    assigned_codes = {_text(row.get("tienda_codigo")) for row in assigned if _text(row.get("tienda_codigo"))}
    cursor.execute(
        """
        SELECT fuente, cadena, tienda, tienda_key, tienda_codigo,
               COUNT(*) AS registros,
               GROUP_CONCAT(DISTINCT mes ORDER BY mes SEPARATOR ',') AS meses,
               SUM(COALESCE(venta, 0)) AS total
        FROM promotoria_sellout
        WHERE anio=%s
        GROUP BY fuente, cadena, tienda, tienda_key, tienda_codigo
        ORDER BY total DESC, tienda
        """,
        (year,),
    )
    rows = []
    for row in cursor.fetchall():
        key = _text(row.get("tienda_key"))
        code = _text(row.get("tienda_codigo"))
        matched = key in assigned_keys or (code and code in assigned_codes)
        if matched:
            continue
        rows.append({
            "fuente": row.get("fuente"),
            "cadena": row.get("cadena"),
            "tienda": row.get("tienda"),
            "tienda_key": key,
            "tienda_codigo": code,
            "registros": int(row.get("registros") or 0),
            "meses": row.get("meses") or "",
            "total": round(float(row.get("total") or 0), 2),
        })
    return rows


@router.get("/resumen")
def resumen(
    year: int = Query(default_factory=lambda: datetime.now().year, ge=2020, le=2100),
    promotor: str = "",
    search: str = "",
    user=Depends(require_user),
):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        assignments = _load_assignments(cursor, promotor, search)
        promotores_meta = _promotores_meta(cursor)
        numbers = sorted({_text(item.get("cliente_numero")) for item in assignments if _text(item.get("cliente_numero"))})
        _invoice_sales, invoice_count, _invoice_totals = _sales_by_customer(cursor, year, numbers)
        sellout_by_key, sellout_by_code = _sellout_by_assignment(cursor, year, assignments)
        sellout_previous_by_key, sellout_previous_by_code = _sellout_by_assignment(cursor, year - 1, assignments)
        comparison_month = _comparison_month(cursor, year)
        rows, sellout_total = [], 0.0
        for item in assignments:
            number = _text(item.get("cliente_numero"))
            tienda_key = _text(item.get("tienda_key"))
            tienda_codigo = _text(item.get("tienda_codigo"))
            source_months = sellout_by_code.get(tienda_codigo) if tienda_codigo and sellout_by_code.get(tienda_codigo) else sellout_by_key.get(tienda_key, {})
            previous_source_months = sellout_previous_by_code.get(tienda_codigo) if tienda_codigo and sellout_previous_by_code.get(tienda_codigo) else sellout_previous_by_key.get(tienda_key, {})
            monthly = {str(month): round(float(source_months.get(month, 0)), 2) for month in range(1, 13)}
            previous_monthly = {str(month): round(float(previous_source_months.get(month, 0)), 2) for month in range(1, 13)}
            current_period_total = round(sum(monthly[str(month)] for month in range(1, comparison_month + 1)), 2)
            previous_period_total = round(sum(previous_monthly[str(month)] for month in range(1, comparison_month + 1)), 2)
            comparable_month_values = [
                previous_monthly[str(month)] for month in range(1, comparison_month + 1)
            ] + [
                monthly[str(month)] for month in range(1, comparison_month + 1)
            ]
            # El Excel original nombra esta columna "MEDIA", pero usa MEDIAN(I:T).
            quota_base = round(float(statistics.median(comparable_month_values)) if comparable_month_values else 0, 2)
            promotor_key = _text(item.get("promotor")).upper()
            percent = _percent(item.get("porcentaje_objetivo"))
            if percent <= 0 and promotor_key in promotores_meta:
                percent = _percent(promotores_meta[promotor_key].get("porcentaje_objetivo"))
            if percent <= 0:
                percent = 0.10
            cuota_base = round(quota_base * (1 + percent), 2)
            sellout_total += current_period_total
            state = "SELL-OUT" if current_period_total > 0 else "SIN SELL-OUT"
            if not number:
                state = "SIN VINCULO"
            rows.append({
                "id": item["id"], "cadena": item["cadena"], "status": item["status"], "activa": item["activa"],
                "supervisor": item["supervisor"], "estado": item["estado"], "promotor": item["promotor"],
                "horario": item["horario"], "tienda": item["tienda"], "cliente_numero": number,
                "cliente_nombre": item["cliente_nombre"], "cliente_empresa": item["cliente_empresa"],
                "match_score": float(item.get("match_score") or 0), "meses": monthly, "meses_anio_anterior": previous_monthly,
                "mes_comparativo": comparison_month, "periodo_actual": current_period_total,
                "periodo_anterior": previous_period_total, "media": quota_base, "media_periodo": quota_base,
                "cuota_base": cuota_base, "cuota_10": cuota_base, "cuota_3": round(cuota_base * 1.03, 2),
                "cuota_8": round(cuota_base * 1.08, 2), "cuota_12": round(cuota_base * 1.12, 2),
                "cuota_15": round(cuota_base * 1.15, 2), "ventas_anuales": current_period_total,
                "cuota_objetivo": cuota_base, "porcentaje_objetivo": percent,
                "facturas": invoice_count.get(number, 0), "estado_datos": state,
            })
        return {
            "year": year, "periodo_comparativo_hasta_mes": comparison_month, "rows": rows,
            "promotores": sorted({_text(item.get("promotor")) for item in assignments if _text(item.get("promotor"))}),
            "summary": {
                "asignaciones": len(assignments), "vinculadas": sum(1 for item in rows if item["cliente_numero"]),
                "sin_vinculo": sum(1 for item in rows if not item["cliente_numero"]),
                "con_facturas": sum(1 for item in rows if item["ventas_anuales"] > 0),
                "ventas": round(sellout_total, 2),
                "facturas": sum(invoice_count.get(number, 0) for number in numbers),
            },
        }
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo generar el análisis: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.get("/comisiones/porcentajes")
def comisiones_porcentajes(user=Depends(require_user)):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        _ensure_commission_percentages(cursor)
        conn.commit()
        cursor.execute(
            """
            SELECT porcentaje, monto, activo
            FROM promotoria_comisiones_porcentajes
            ORDER BY porcentaje
            """
        )
        rows = [
            {
                "porcentaje": round(float(row.get("porcentaje") or 0), 4),
                "monto": round(float(row.get("monto") or 0), 2),
                "activo": int(row.get("activo") or 0),
            }
            for row in cursor.fetchall()
        ]
        return {"rows": rows}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudieron cargar los porcentajes de comisión: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


@router.post("/comisiones/porcentajes")
def guardar_comisiones_porcentajes(payload: dict = Body(...), user=Depends(require_user)):
    rows = payload.get("rows") if isinstance(payload, dict) else []
    if not isinstance(rows, list):
        raise HTTPException(status_code=422, detail="La lista de porcentajes no es válida.")
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        _ensure_commission_percentages(cursor)
        now = datetime.now()
        allowed = {0.03, 0.08, 0.12, 0.15}
        saved = 0
        for item in rows:
            if not isinstance(item, dict):
                continue
            porcentaje = _percent(item.get("porcentaje"))
            if porcentaje not in allowed:
                continue
            cursor.execute(
                """
                INSERT INTO promotoria_comisiones_porcentajes
                (porcentaje, monto, activo, actualizado_en)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                  monto=VALUES(monto),
                  activo=VALUES(activo),
                  actualizado_en=VALUES(actualizado_en)
                """,
                (porcentaje, _money(item.get("monto")), _boolish(item.get("activo", True)), now),
            )
            saved += 1
        conn.commit()
        return {"ok": True, "guardados": saved}
    except Exception as exc:
        if conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudieron guardar los porcentajes de comisión: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def _excel_number(value):
    try:
        number = float(value or 0)
        return round(number, 2) if math.isfinite(number) else 0
    except Exception:
        return 0


def _excel_text(value):
    return str(value or "").strip()


def _style_title_cell(cell, fill="0B5394", color="FFFFFF", size=16):
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=color, bold=True, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_panel(ws, row, col, title, value, fill):
    ws.cell(row, col, title)
    ws.cell(row + 1, col, value)
    for r in (row, row + 1):
        cell = ws.cell(r, col)
        cell.fill = PatternFill("solid", fgColor=fill if r == row else "F8FBFF")
        cell.font = Font(color="FFFFFF" if r == row else "17365D", bold=True, size=10 if r == row else 13)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style="thin", color="B7C9DC"),
            right=Side(style="thin", color="B7C9DC"),
            top=Side(style="thin", color="B7C9DC"),
            bottom=Side(style="thin", color="B7C9DC"),
        )


def _style_table(ws, table_name, start_row, end_row, end_col, style="TableStyleMedium2"):
    ref = f"{ws.cell(start_row, 1).coordinate}:{ws.cell(end_row, end_col).coordinate}"
    ws.freeze_panes = ws.cell(start_row + 1, 1).coordinate
    ws.auto_filter.ref = ref
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[start_row]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_index in range(start_row + 1, end_row + 1):
        fill = PatternFill("solid", fgColor="F8FBFF" if (row_index - start_row) % 2 else "FFFFFF")
        for col in range(1, end_col + 1):
            cell = ws.cell(row_index, col)
            cell.fill = fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _append_table(ws, headers, rows, start_row):
    for col, header in enumerate(headers, 1):
        ws.cell(start_row, col, header)
    for row_index, row in enumerate(rows, start_row + 1):
        for col, header in enumerate(headers, 1):
            ws.cell(row_index, col, row.get(header, ""))
    return start_row + max(len(rows), 1)


@router.post("/reporte-excel")
def reporte_excel(payload: dict = Body(...), user=Depends(require_user)):
    try:
        dashboard_rows = payload.get("dashboard_rows") or []
        objetivos = payload.get("objetivos") or []
        comisiones = payload.get("comisiones") or []
        porcentajes = payload.get("porcentajes") or []
        filtros = payload.get("filtros") or []
        year = _text(payload.get("year") or datetime.now().year)
        month_name = _text(payload.get("month_name") or "Mes")
        periodo = _text(payload.get("periodo") or "")

        book = Workbook()
        book.calculation.calcMode = "auto"
        book.calculation.calcId = 0
        book.calculation.calcOnSave = True
        book.calculation.fullCalcOnLoad = True
        book.calculation.forceFullCalc = True
        ws = book.active
        ws.title = "Dashboard"
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 28
        ws.merge_cells("A1:N1")
        ws["A1"] = "Dashboard promotoría - objetivos y comisiones"
        _style_title_cell(ws["A1"])
        ws["A2"] = f"Año {year} · Comisión {month_name} · Objetivos {periodo}"
        ws.merge_cells("A2:N2")
        ws["A2"].font = Font(color="44546A", bold=True)
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = [
            "Cadena", "Supervisor", "Status", "Promotor", "Tienda",
            "Periodo anterior", "Periodo actual", "Cuota media periodo", "Cuota base",
            f"Venta {month_name}", "Cuota alcanzada", "Meta alcanzada", "Comision", "Estado datos",
        ]
        normalized_dashboard = []
        for row in dashboard_rows:
            normalized_dashboard.append({
                header: (_excel_number(row.get(header)) if header in {"Periodo anterior", "Periodo actual", "Cuota media periodo", "Cuota base", f"Venta {month_name}", "Meta alcanzada", "Comision"} else _excel_text(row.get(header)))
                for header in headers
            })
        if not normalized_dashboard:
            normalized_dashboard = [{header: "" for header in headers}]

        widths = [15, 18, 13, 22, 34, 15, 15, 18, 14, 15, 22, 15, 14, 16]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        datos = book.create_sheet("Datos")
        datos.sheet_view.showGridLines = False
        datos_end = _append_table(datos, headers, normalized_dashboard, 1)
        _style_table(datos, "DatosPromotoria", 1, datos_end, len(headers), "TableStyleMedium4")
        datos.sheet_state = "hidden"
        for col, width in enumerate(widths, 1):
            datos.column_dimensions[get_column_letter(col)].width = width
        for row in datos.iter_rows(min_row=2, max_row=datos_end, min_col=6, max_col=13):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "$#,##0.00"
        helper_visible_col = len(headers) + 1
        helper_order_col = len(headers) + 2
        datos.cell(1, helper_visible_col, "Filtro visible")
        datos.cell(1, helper_order_col, "Orden visible")
        for row_idx in range(2, datos_end + 1):
            datos.cell(row_idx, helper_visible_col).value = (
                f'=--AND('
                f'OR(Dashboard!$B$8="Todos",Dashboard!$B$8="",Dashboard!$B$8="Ninguno",A{row_idx}=Dashboard!$B$8),'
                f'OR(Dashboard!$E$8="Todos",Dashboard!$E$8="",Dashboard!$E$8="Ninguno",B{row_idx}=Dashboard!$E$8),'
                f'OR(Dashboard!$H$8="Todos",Dashboard!$H$8="",Dashboard!$H$8="Ninguno",C{row_idx}=Dashboard!$H$8),'
                f'OR(Dashboard!$K$8="Todos",Dashboard!$K$8="",Dashboard!$K$8="Ninguno",D{row_idx}=Dashboard!$K$8)'
                f')'
            )
            datos.cell(row_idx, helper_order_col).value = f'=IF({get_column_letter(helper_visible_col)}{row_idx}=1,SUM(${get_column_letter(helper_visible_col)}$2:{get_column_letter(helper_visible_col)}{row_idx}),"")'

        listas = book.create_sheet("Listas")
        listas.sheet_state = "hidden"
        list_map = [
            ("Cadena", "A", "B8"),
            ("Supervisor", "B", "E8"),
            ("Status", "C", "H8"),
            ("Promotor", "D", "K8"),
        ]
        for title, col_letter, target in list_map:
            values = ["Todos"] + sorted({_excel_text(row.get(title)) for row in normalized_dashboard if _excel_text(row.get(title))})
            listas[f"{col_letter}1"] = title
            for idx, value in enumerate(values, 2):
                listas[f"{col_letter}{idx}"] = value
            dv = DataValidation(type="list", formula1=f"'Listas'!${col_letter}$2:${col_letter}${len(values) + 1}", allow_blank=False)
            ws.add_data_validation(dv)
            dv.add(ws[target])

        total_rows = max(len(normalized_dashboard), 1)
        table_start = 12
        first_data = table_start + 1
        last_data = table_start + total_rows
        panel_defs = [
            ("CADENA", "B8", 1, 3, "D9EAF7", "4472C4"),
            ("SUPERVISOR", "E8", 4, 6, "E2F0D9", "70AD47"),
            ("STATUS", "H8", 7, 9, "D9EAF7", "4472C4"),
            ("PROMOTOR", "K8", 10, 14, "FCE4D6", "ED7D31"),
        ]
        for title, selector, start_col, end_col, fill, accent in panel_defs:
            for row_idx in range(7, 10):
                for col_idx in range(start_col, end_col + 1):
                    panel_cell = ws.cell(row_idx, col_idx)
                    panel_cell.fill = PatternFill("solid", fgColor=fill)
                    panel_cell.border = Border(left=Side(style="thin", color=accent), right=Side(style="thin", color=accent), top=Side(style="thin", color=accent), bottom=Side(style="thin", color=accent))
            title_cell = ws.cell(7, start_col)
            title_cell.value = title
            title_cell.font = Font(color="000000", bold=True, size=10)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            hint_cell = ws.cell(8, start_col)
            hint_cell.value = "Seleccionar"
            hint_cell.font = Font(color="44546A", size=9)
            hint_cell.alignment = Alignment(horizontal="right", vertical="center")
            selector_cell = ws[selector]
            selector_cell.value = "Todos"
            selector_cell.fill = PatternFill("solid", fgColor="FFFFFF")
            selector_cell.font = Font(color="17365D", bold=True)
            selector_cell.alignment = Alignment(horizontal="center", vertical="center")
            selector_cell.border = Border(left=Side(style="thin", color=accent), right=Side(style="thin", color=accent), top=Side(style="thin", color=accent), bottom=Side(style="thin", color=accent))

        for col, header in enumerate(headers, 1):
            ws.cell(table_start, col, header)

        visible_range = f"Datos!${get_column_letter(helper_visible_col)}$2:${get_column_letter(helper_visible_col)}${datos_end}"
        order_range = f"Datos!${get_column_letter(helper_order_col)}$2:${get_column_letter(helper_order_col)}${datos_end}"
        _style_panel(ws, 4, 1, "Tiendas visibles", f"=SUM({visible_range})", "4472C4")
        _style_panel(ws, 4, 3, f"Venta {month_name}", f"=SUMPRODUCT({visible_range},Datos!$J$2:$J${datos_end})", "70AD47")
        _style_panel(ws, 4, 5, "Comisión visible", f"=SUMPRODUCT({visible_range},Datos!$M$2:$M${datos_end})", "ED7D31")
        _style_panel(ws, 4, 7, "Con cuota", f'=SUMPRODUCT({visible_range},--(Datos!$K$2:$K${datos_end}<>"Sin cuota"))', "7030A0")
        ws["J10"] = "Si no aparecen filas, la combinación seleccionada no existe en la base."
        ws["J10"].font = Font(color="9E480E", italic=True, size=9)
        ws.merge_cells("J10:N10")
        ws["J10"].alignment = Alignment(horizontal="right", vertical="center")

        for row_number in range(first_data, last_data + 1):
            nth = f"ROWS($A${first_data}:A{row_number})"
            for col_number in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_number)
                ws.cell(row_number, col_number).value = (
                    f'=IFERROR(INDEX(Datos!${col_letter}$2:${col_letter}${datos_end},'
                    f'MATCH({nth},{order_range},0)),"")'
                )

        _style_table(ws, "DashboardPromotoria", table_start, last_data, len(headers), "TableStyleMedium4")
        for row in ws.iter_rows(min_row=first_data, max_row=last_data, min_col=6, max_col=13):
            for cell in row:
                cell.number_format = "$#,##0.00"
        ws.freeze_panes = "A13"

        sheets = [
            ("Objetivos", objetivos, "TableStyleMedium2"),
            ("Comisiones", comisiones, "TableStyleMedium7"),
            ("Porcentajes", porcentajes, "TableStyleMedium3"),
            ("Filtros", filtros, "TableStyleMedium6"),
        ]
        for sheet_name, rows, style in sheets:
            sheet = book.create_sheet(sheet_name)
            sheet.sheet_view.showGridLines = False
            clean_rows = rows or [{"Mensaje": "Sin registros"}]
            sheet_headers = list(clean_rows[0].keys())
            end = _append_table(sheet, sheet_headers, clean_rows, 1)
            _style_table(sheet, re.sub(r"[^A-Za-z0-9]", "", sheet_name)[:20] + "Tabla", 1, end, len(sheet_headers), style)
            for col, header in enumerate(sheet_headers, 1):
                sheet.column_dimensions[get_column_letter(col)].width = min(max(len(str(header)) + 4, 14), 38)
            for row in sheet.iter_rows(min_row=2, max_row=end):
                for cell in row:
                    if isinstance(cell.value, (int, float)) and ("monto" in str(sheet.cell(1, cell.column).value).lower() or "cuota" in str(sheet.cell(1, cell.column).value).lower() or "venta" in str(sheet.cell(1, cell.column).value).lower() or "periodo" in str(sheet.cell(1, cell.column).value).lower() or "comision" in str(sheet.cell(1, cell.column).value).lower()):
                        cell.number_format = "$#,##0.00"

        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        filename = f"promotoria_dashboard_{year}_{month_name}.xlsx".replace(" ", "_")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo generar el reporte de promotoría: {exc}")


@router.get("/detalle/{assignment_id}")
def detalle(assignment_id: int, year: int = Query(default_factory=lambda: datetime.now().year), user=Depends(require_user)):
    conn = cursor = None
    try:
        conn = get_legacy_connection()
        cursor = conn.cursor(dictionary=True)
        _ensure_tables(cursor)
        cursor.execute("SELECT * FROM promotoria_asignaciones WHERE id=%s", (assignment_id,))
        assignment = cursor.fetchone()
        if not assignment:
            raise HTTPException(status_code=404, detail="No se encontró la asignación de promotoría.")
        customer = _text(assignment.get("cliente_numero"))
        if not customer:
            return {"assignment": assignment, "facturas": [], "productos": [], "mensaje": "La tienda aún no está vinculada a un cliente."}
        base_where = """
            YEAR(f.fecha)=%s AND UPPER(TRIM(COALESCE(f.estatus,'ACTIVA'))) <> 'CANCELADA'
            AND UPPER(TRIM(COALESCE(f.empresa,''))) LIKE 'GOURMET%'
            AND TRIM(COALESCE(f.numero_cliente,''))=%s
        """
        cursor.execute(f"""SELECT f.factura, f.fecha, f.total, f.empresa FROM facturas f
                           WHERE {base_where} ORDER BY f.fecha DESC, f.id DESC LIMIT 50""", (year, customer))
        invoices = cursor.fetchall()
        cursor.execute(f"""SELECT COALESCE(fd.cip,'') AS cip, COALESCE(fd.descripcion,'') AS descripcion,
                                  SUM(COALESCE(fd.cantidad,0)) AS cantidad, SUM(COALESCE(fd.piezas,0)) AS piezas,
                                  SUM(COALESCE(fd.importe,0)) AS importe
                           FROM factura_detalle fd JOIN facturas f ON f.id=fd.factura_id
                           WHERE {base_where}
                           GROUP BY fd.cip, fd.descripcion ORDER BY importe DESC LIMIT 20""", (year, customer))
        products = cursor.fetchall()
        return {"assignment": assignment, "facturas": invoices, "productos": products}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"No se pudo cargar el detalle: {exc}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
