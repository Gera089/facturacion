from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from urllib.parse import urlparse, parse_qs, unquote
import csv
import io
import json
import mimetypes
import os
import shutil
import sqlite3
import subprocess
import sys
import textwrap
import time
import unicodedata
import uuid
from datetime import datetime, timedelta
from decimal import Decimal

import bcrypt
import mysql.connector
import requests
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle

try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None

try:
    from num2words import num2words
except Exception:
    num2words = None


ROOT = os.path.dirname(os.path.abspath(__file__))
PUBLIC_DIR = os.path.join(ROOT, "public")


def crm_data_dir():
    custom = os.environ.get("CRM_VENTAS_DATA_DIR")
    if custom:
        return custom
    if getattr(sys, "frozen", False):
        base = os.environ.get("PROGRAMDATA") or os.path.expanduser("~")
        return os.path.join(base, "ComandasAPI", "crm_ventas")
    return ROOT


DATA_DIR = crm_data_dir()
DB_PATH = os.path.join(DATA_DIR, "ventas.db")
CONFIG_PATH = os.path.join(DATA_DIR, "app_config.json")
QUOTE_TEMPLATE = r"\\SERVER\12.-Programas Sistemas\06-Diseño e Inventario\TRABAJO\Escritorio\Cotizaciones queso y jamon\Cotizacion_Queso_Jamon 137.xlsx"
EXPORT_DIR = os.path.join(DATA_DIR, "exports")


def ensure_runtime_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(EXPORT_DIR, exist_ok=True)
    for filename in ("ventas.db", "app_config.json"):
        target = os.path.join(DATA_DIR, filename)
        if os.path.exists(target):
            continue
        seed = os.path.join(ROOT, filename)
        if os.path.exists(seed):
            shutil.copy2(seed, target)


ensure_runtime_data()


DEFAULT_CONFIG = {
    "facturacion_api_url": "http://127.0.0.1:8000",
    "mysql_host": "100.69.142.19",
    "mysql_user": "Facturacion",
    "mysql_pass": "ALD2013*",
    "mysql_port": 3307,
    "mysql_database": "comandas_db",
    "company_name": "EZA2007",
    "quote_template": QUOTE_TEMPLATE,
    "google_places_api_key": "",
}


def now_iso():
    return datetime.now().replace(microsecond=0).isoformat()


def quote_valid_until():
    return (datetime.now().date() + timedelta(days=15)).isoformat()


def load_config():
    cfg = dict(DEFAULT_CONFIG)
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            user_cfg = json.load(f) or {}
        cfg.update({k: v for k, v in user_cfg.items() if v not in (None, "")})
    return cfg


CFG = load_config()


def connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    with connect() as db:
        db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                email TEXT,
                role TEXT NOT NULL DEFAULT 'vendedor',
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS clients (
                id TEXT PRIMARY KEY,
                code TEXT,
                name TEXT NOT NULL,
                tax_address TEXT,
                consignee_address TEXT,
                delivery_method TEXT,
                phone TEXT,
                email TEXT,
                contact_name TEXT,
                external_seller TEXT,
                assigned_user_id TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (assigned_user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS invoices (
                id TEXT PRIMARY KEY,
                client_id TEXT NOT NULL,
                user_id TEXT,
                folio TEXT NOT NULL,
                issued_at TEXT NOT NULL,
                subtotal REAL NOT NULL DEFAULT 0,
                discount REAL NOT NULL DEFAULT 0,
                tax REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'facturada',
                detail TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS quotes (
                id TEXT PRIMARY KEY,
                client_id TEXT NOT NULL,
                user_id TEXT NOT NULL,
                folio TEXT NOT NULL,
                quote_title TEXT,
                quote_recipient TEXT,
                created_at TEXT NOT NULL,
                valid_until TEXT,
                subtotal REAL NOT NULL DEFAULT 0,
                discount REAL NOT NULL DEFAULT 0,
                tax REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL DEFAULT 0,
                authorized_shipping REAL NOT NULL DEFAULT 0,
                notes TEXT,
                status TEXT NOT NULL DEFAULT 'borrador',
                client_snapshot TEXT,
                FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id)
            );

            CREATE TABLE IF NOT EXISTS quote_items (
                id TEXT PRIMARY KEY,
                quote_id TEXT NOT NULL,
                cip TEXT,
                description TEXT NOT NULL,
                quantity REAL NOT NULL DEFAULT 1,
                unit_price REAL NOT NULL DEFAULT 0,
                discount_rate REAL NOT NULL DEFAULT 0,
                tax_rate REAL NOT NULL DEFAULT 16,
                line_total REAL NOT NULL DEFAULT 0,
                FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS followups (
                id TEXT PRIMARY KEY,
                client_id TEXT NOT NULL,
                user_id TEXT NOT NULL,
                contact_at TEXT NOT NULL,
                channel TEXT NOT NULL,
                outcome TEXT NOT NULL,
                next_action TEXT,
                next_action_at TEXT,
                notes TEXT,
                quote_id TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id),
                FOREIGN KEY (quote_id) REFERENCES quotes(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS activity_log (
                id TEXT PRIMARY KEY,
                client_id TEXT,
                user_id TEXT,
                type TEXT NOT NULL,
                title TEXT NOT NULL,
                payload TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE SET NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS company_bank_accounts (
                company TEXT PRIMARY KEY,
                beneficiary TEXT,
                bank TEXT,
                account TEXT,
                clabe TEXT,
                enabled INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS prospect_searches (
                id TEXT PRIMARY KEY,
                query TEXT NOT NULL,
                zone_name TEXT,
                limit_count INTEGER NOT NULL DEFAULT 20,
                results_count INTEGER NOT NULL DEFAULT 0,
                user_id TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS prospects (
                id TEXT PRIMARY KEY,
                google_place_id TEXT UNIQUE,
                name TEXT NOT NULL,
                category TEXT,
                phone TEXT,
                website TEXT,
                address TEXT,
                latitude REAL,
                longitude REAL,
                rating REAL,
                total_reviews INTEGER,
                business_status TEXT,
                source_query TEXT,
                zone_name TEXT,
                status TEXT NOT NULL DEFAULT 'nuevo',
                notes TEXT,
                assigned_user_id TEXT,
                client_id TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (assigned_user_id) REFERENCES users(id) ON DELETE SET NULL,
                FOREIGN KEY (client_id) REFERENCES clients(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS prospect_quotes (
                id TEXT PRIMARY KEY,
                prospect_id TEXT NOT NULL,
                user_id TEXT,
                folio TEXT,
                title TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                valid_until TEXT,
                subtotal REAL NOT NULL DEFAULT 0,
                discount REAL NOT NULL DEFAULT 0,
                tax REAL NOT NULL DEFAULT 0,
                total REAL NOT NULL DEFAULT 0,
                authorized_shipping REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'borrador',
                notes TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS prospect_quote_items (
                id TEXT PRIMARY KEY,
                quote_id TEXT NOT NULL,
                cip TEXT,
                description TEXT NOT NULL,
                quantity REAL NOT NULL DEFAULT 1,
                unit_price REAL NOT NULL DEFAULT 0,
                discount_rate REAL NOT NULL DEFAULT 0,
                tax_rate REAL NOT NULL DEFAULT 16,
                line_total REAL NOT NULL DEFAULT 0,
                FOREIGN KEY (quote_id) REFERENCES prospect_quotes(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS prospect_followups (
                id TEXT PRIMARY KEY,
                prospect_id TEXT NOT NULL,
                user_id TEXT,
                contact_at TEXT NOT NULL,
                channel TEXT NOT NULL,
                outcome TEXT NOT NULL,
                next_action TEXT,
                next_action_at TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS prospect_messages (
                id TEXT PRIMARY KEY,
                prospect_id TEXT NOT NULL,
                user_id TEXT,
                sent_at TEXT NOT NULL,
                channel TEXT NOT NULL,
                direction TEXT NOT NULL DEFAULT 'saliente',
                body TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS prospect_phones (
                id TEXT PRIMARY KEY,
                prospect_id TEXT NOT NULL,
                label TEXT,
                phone TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS prospect_activity (
                id TEXT PRIMARY KEY,
                prospect_id TEXT NOT NULL,
                user_id TEXT,
                type TEXT NOT NULL,
                title TEXT NOT NULL,
                payload TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (prospect_id) REFERENCES prospects(id) ON DELETE CASCADE,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE SET NULL
            );

            CREATE TABLE IF NOT EXISTS prospect_zones (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                sur REAL NOT NULL,
                oeste REAL NOT NULL,
                norte REAL NOT NULL,
                este REAL NOT NULL,
                enabled INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )
        cols = [r["name"] for r in db.execute("PRAGMA table_info(quote_items)").fetchall()]
        if "sku" in cols and "cip" not in cols:
            db.execute("ALTER TABLE quote_items RENAME COLUMN sku TO cip")
        client_cols = [r["name"] for r in db.execute("PRAGMA table_info(clients)").fetchall()]
        if "external_seller" not in client_cols:
            db.execute("ALTER TABLE clients ADD COLUMN external_seller TEXT")
        if "contact_name" not in client_cols:
            db.execute("ALTER TABLE clients ADD COLUMN contact_name TEXT")
        invoice_cols = [r["name"] for r in db.execute("PRAGMA table_info(invoices)").fetchall()]
        if "user_id" not in invoice_cols:
            db.execute("ALTER TABLE invoices ADD COLUMN user_id TEXT")
        quote_cols = [r["name"] for r in db.execute("PRAGMA table_info(quotes)").fetchall()]
        if "client_snapshot" not in quote_cols:
            db.execute("ALTER TABLE quotes ADD COLUMN client_snapshot TEXT")
        if "quote_title" not in quote_cols:
            db.execute("ALTER TABLE quotes ADD COLUMN quote_title TEXT")
        if "quote_recipient" not in quote_cols:
            db.execute("ALTER TABLE quotes ADD COLUMN quote_recipient TEXT")
        followup_cols = [r["name"] for r in db.execute("PRAGMA table_info(followups)").fetchall()]
        if "quote_id" not in followup_cols:
            db.execute("ALTER TABLE followups ADD COLUMN quote_id TEXT")
        prospect_cols = [r["name"] for r in db.execute("PRAGMA table_info(prospects)").fetchall()]
        if prospect_cols and "client_id" not in prospect_cols:
            db.execute("ALTER TABLE prospects ADD COLUMN client_id TEXT")
        if prospect_cols and "assigned_user_id" not in prospect_cols:
            db.execute("ALTER TABLE prospects ADD COLUMN assigned_user_id TEXT")
        if prospect_cols and "zone_name" not in prospect_cols:
            db.execute("ALTER TABLE prospects ADD COLUMN zone_name TEXT")
        prospect_quote_cols = [r["name"] for r in db.execute("PRAGMA table_info(prospect_quotes)").fetchall()]
        for col_name, col_def in [
            ("folio", "TEXT"),
            ("valid_until", "TEXT"),
            ("subtotal", "REAL NOT NULL DEFAULT 0"),
            ("discount", "REAL NOT NULL DEFAULT 0"),
            ("tax", "REAL NOT NULL DEFAULT 0"),
            ("total", "REAL NOT NULL DEFAULT 0"),
            ("authorized_shipping", "REAL NOT NULL DEFAULT 0"),
        ]:
            if prospect_quote_cols and col_name not in prospect_quote_cols:
                db.execute(f"ALTER TABLE prospect_quotes ADD COLUMN {col_name} {col_def}")
        search_cols = [r["name"] for r in db.execute("PRAGMA table_info(prospect_searches)").fetchall()]
        if search_cols and "zone_name" not in search_cols:
            db.execute("ALTER TABLE prospect_searches ADD COLUMN zone_name TEXT")
        user_count = db.execute("SELECT COUNT(*) AS c FROM users").fetchone()["c"]
        if user_count == 0:
            seller_id = str(uuid.uuid4())
            manager_id = str(uuid.uuid4())
            db.executemany(
                "INSERT INTO users (id, name, email, role, created_at) VALUES (?, ?, ?, ?, ?)",
                [
                    (seller_id, "Vendedor Demo", "ventas@empresa.local", "vendedor", now_iso()),
                    (manager_id, "Gerente Comercial", "gerencia@empresa.local", "admin", now_iso()),
                ],
            )
            client_id = str(uuid.uuid4())
            db.execute(
                """
                INSERT INTO clients
                (id, code, name, tax_address, consignee_address, delivery_method, phone, email, contact_name, external_seller, assigned_user_id, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    client_id,
                    "C-001",
                    "Cliente Demo S.A. de C.V.",
                    "Av. Fiscal 120, Col. Centro, Monterrey, N.L.",
                    "Bodega 4, Parque Industrial Norte",
                    "Paqueteria documentada / ocurre",
                    "818-000-0000",
                    "compras@clientedemo.mx",
                    "Compras",
                    "",
                    seller_id,
                    now_iso(),
                    now_iso(),
                ),
            )
            db.execute(
                """
                INSERT INTO invoices
                (id, client_id, folio, issued_at, subtotal, discount, tax, total, status, detail, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    str(uuid.uuid4()),
                    client_id,
                    "F-1001",
                    "2026-06-24",
                    12500,
                    500,
                    1920,
                    13920,
                    "facturada",
                    "Pedido demo: 10 piezas SKU-A, 5 piezas SKU-B",
                    now_iso(),
                ),
            )
        migrate_crm_sqlite_to_mysql(db)


def row_to_dict(row):
    return dict(row) if row else None


def read_json(handler):
    length = int(handler.headers.get("Content-Length", "0"))
    if length == 0:
        return {}
    raw = handler.rfile.read(length)
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError:
        text = raw.decode("cp1252")
    return json.loads(text)


def parse_external_client_id(client_id):
    if client_id.startswith("ext:"):
        parts = client_id.split(":", 2)
        if len(parts) == 3:
            return parts[1], parts[2]
    return None, None


def mysql_conn():
    return mysql.connector.connect(
        host=CFG["mysql_host"],
        user=CFG["mysql_user"],
        password=CFG["mysql_pass"],
        database=CFG["mysql_database"],
        port=int(CFG["mysql_port"]),
        charset="utf8mb4",
        use_pure=True,
        connection_timeout=5,
    )


def decimal_to_float(value):
    if isinstance(value, Decimal):
        return float(value)
    return value


def normalize_mysql_row(row):
    return {k: decimal_to_float(v) for k, v in row.items()}


def crm_mysql_available():
    try:
        conn = mysql_conn()
        conn.close()
        return True
    except Exception:
        return False


def crm_user_name(user_id):
    return (user_id or "").replace("ext:", "", 1) or ""


def init_crm_mysql_tables():
    conn = mysql_conn()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_clients (
            id VARCHAR(191) PRIMARY KEY,
            code VARCHAR(80),
            name VARCHAR(255) NOT NULL,
            tax_address TEXT,
            consignee_address TEXT,
            delivery_method TEXT,
            phone TEXT,
            email VARCHAR(255),
            contact_name VARCHAR(255),
            external_seller VARCHAR(255),
            assigned_user_id VARCHAR(191),
            created_at VARCHAR(32) NOT NULL,
            updated_at VARCHAR(32) NOT NULL
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_followups (
            id VARCHAR(191) PRIMARY KEY,
            client_id VARCHAR(191) NOT NULL,
            user_id VARCHAR(191) NOT NULL,
            contact_at VARCHAR(32) NOT NULL,
            channel VARCHAR(80) NOT NULL,
            outcome TEXT NOT NULL,
            next_action TEXT,
            next_action_at VARCHAR(32),
            notes TEXT,
            quote_id VARCHAR(191),
            created_at VARCHAR(32) NOT NULL,
            INDEX idx_crm_followups_client (client_id),
            INDEX idx_crm_followups_quote (quote_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS crm_activity_log (
            id VARCHAR(191) PRIMARY KEY,
            client_id VARCHAR(191),
            user_id VARCHAR(191),
            type VARCHAR(80) NOT NULL,
            title TEXT NOT NULL,
            payload LONGTEXT,
            created_at VARCHAR(32) NOT NULL,
            INDEX idx_crm_activity_client (client_id)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci
        """
    )
    conn.commit()
    cur.close()
    conn.close()


def migrate_crm_sqlite_to_mysql(db):
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor()
        for row in db.execute("SELECT * FROM clients").fetchall():
            data = row_to_dict(row)
            cur.execute(
                """
                INSERT INTO crm_clients
                (id, code, name, tax_address, consignee_address, delivery_method, phone, email, contact_name,
                 external_seller, assigned_user_id, created_at, updated_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    code = VALUES(code),
                    name = VALUES(name),
                    tax_address = VALUES(tax_address),
                    consignee_address = VALUES(consignee_address),
                    delivery_method = VALUES(delivery_method),
                    phone = VALUES(phone),
                    email = VALUES(email),
                    contact_name = VALUES(contact_name),
                    external_seller = VALUES(external_seller),
                    assigned_user_id = VALUES(assigned_user_id),
                    updated_at = VALUES(updated_at)
                """,
                (
                    data.get("id"),
                    data.get("code") or "",
                    data.get("name") or "",
                    data.get("tax_address") or "",
                    data.get("consignee_address") or "",
                    data.get("delivery_method") or "",
                    data.get("phone") or "",
                    data.get("email") or "",
                    data.get("contact_name") or "",
                    data.get("external_seller") or "",
                    data.get("assigned_user_id") or None,
                    data.get("created_at") or now_iso(),
                    data.get("updated_at") or now_iso(),
                ),
            )
        for row in db.execute("SELECT * FROM followups").fetchall():
            data = row_to_dict(row)
            cur.execute(
                """
                INSERT INTO crm_followups
                (id, client_id, user_id, contact_at, channel, outcome, next_action, next_action_at, notes, quote_id, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    client_id = VALUES(client_id),
                    user_id = VALUES(user_id),
                    contact_at = VALUES(contact_at),
                    channel = VALUES(channel),
                    outcome = VALUES(outcome),
                    next_action = VALUES(next_action),
                    next_action_at = VALUES(next_action_at),
                    notes = VALUES(notes),
                    quote_id = VALUES(quote_id)
                """,
                (
                    data.get("id"),
                    data.get("client_id"),
                    data.get("user_id"),
                    data.get("contact_at") or now_iso(),
                    data.get("channel") or "llamada",
                    data.get("outcome") or "",
                    data.get("next_action") or "",
                    data.get("next_action_at") or "",
                    data.get("notes") or "",
                    data.get("quote_id") or None,
                    data.get("created_at") or now_iso(),
                ),
            )
        for row in db.execute("SELECT * FROM activity_log").fetchall():
            data = row_to_dict(row)
            cur.execute(
                """
                INSERT INTO crm_activity_log
                (id, client_id, user_id, type, title, payload, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    client_id = VALUES(client_id),
                    user_id = VALUES(user_id),
                    type = VALUES(type),
                    title = VALUES(title),
                    payload = VALUES(payload),
                    created_at = VALUES(created_at)
                """,
                (
                    data.get("id"),
                    data.get("client_id") or None,
                    data.get("user_id") or None,
                    data.get("type") or "",
                    data.get("title") or "",
                    data.get("payload") or "",
                    data.get("created_at") or now_iso(),
                ),
            )
        conn.commit()
    except Exception as e:
        print(f"[CRM] No se pudo migrar CRM a MySQL: {e}", flush=True)
    finally:
        try:
            cur.close()
            conn.close()
        except Exception:
            pass


def external_available():
    try:
        conn = mysql_conn()
        conn.close()
        return True
    except Exception:
        return False


def external_users():
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT usuario, rol FROM usuarios ORDER BY usuario ASC")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [
        {
            "id": f"ext:{row['usuario']}",
            "name": row["usuario"],
            "email": "",
            "role": row.get("rol") or "vendedor",
            "active": 1,
            "created_at": "",
        }
        for row in rows
    ]


def verify_external_user(username, password):
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT usuario, password, rol FROM usuarios WHERE TRIM(usuario) = %s", ((username or "").strip(),))
    user = cur.fetchone()
    cur.close()
    conn.close()
    if not user or not user.get("password"):
        return None
    hash_db = user["password"].encode("utf-8") if isinstance(user["password"], str) else bytes(user["password"])
    if not bcrypt.checkpw((password or "").encode("utf-8"), hash_db):
        return None
    return {"id": f"ext:{user['usuario']}", "name": user["usuario"], "role": user.get("rol") or "vendedor"}


def map_external_client(row):
    row = normalize_mysql_row(row)
    tax_address = " ".join(
        str(row.get(k) or "").strip()
        for k in ["calle", "no_exterior", "no_interior", "colonia", "municipio", "estado", "codigo_postal"]
        if str(row.get(k) or "").strip()
    )
    consignee_address = row.get("direccion_entrega") or " ".join(
        str(row.get(k) or "").strip()
        for k in ["consignatario", "consig_calle", "consig_no_exterior", "consig_colonia", "consig_municipio", "consig_estado"]
        if str(row.get(k) or "").strip()
    )
    numero = str(row.get("numero") or "").strip()
    empresa = str(row.get("empresa") or "").strip()
    phone = normalize_phone_display(row.get("telefono") or "")
    return {
        "id": f"ext:{numero}:{empresa}",
        "code": numero,
        "name": row.get("nombre") or row.get("razon_social") or empresa or numero,
        "empresa": empresa,
        "tax_address": tax_address,
        "consignee_address": consignee_address,
        "delivery_method": row.get("direccion_entrega") or "",
        "phone": phone,
        "email": row.get("correo_electronico") or "",
        "contact_name": " / ".join(x for x in [row.get("contacto1"), row.get("contacto2")] if x),
        "assigned_user_id": "",
        "assigned_user": row.get("vendedor") or "",
        "external_seller": row.get("vendedor") or "",
        "external": True,
        "descuento": row.get("descuento") or 0,
        "especial": row.get("especial") or "Lista General",
    }


def normalize_phone_display(value):
    raw = str(value or "").strip()
    digits = "".join(ch for ch in raw if ch.isdigit())
    if len(digits) > 10 and len(digits) % 10 == 0:
        return " / ".join(digits[i:i + 10] for i in range(0, len(digits), 10))
    if len(digits) == 20:
        return f"{digits[:10]} / {digits[10:]}"
    return raw


def external_clients(search=""):
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    like = f"%{search}%"
    cur.execute(
        """
        SELECT numero, nombre, empresa, razon_social, calle, no_exterior, no_interior,
               colonia, municipio, estado, codigo_postal, telefono, correo_electronico,
               consignatario, consig_calle, consig_no_exterior, consig_colonia,
               consig_municipio, consig_estado, direccion_entrega, vendedor,
               descuento, especial
        FROM clientes
        WHERE nombre LIKE %s OR numero LIKE %s OR empresa LIKE %s OR correo_electronico LIKE %s OR vendedor LIKE %s
        ORDER BY nombre ASC
        LIMIT 300
        """,
        (like, like, like, like, like),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return [map_external_client(row) for row in rows]


PLACES_TEXT_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"


def google_places_key():
    return (os.environ.get("GOOGLE_API_KEY") or CFG.get("google_places_api_key") or "").strip()


def google_places_headers(api_key):
    return {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        "X-Goog-FieldMask": (
            "places.id,"
            "places.displayName,"
            "places.formattedAddress,"
            "places.location,"
            "places.rating,"
            "places.userRatingCount,"
            "places.businessStatus,"
            "places.nationalPhoneNumber,"
            "places.internationalPhoneNumber,"
            "places.websiteUri,"
            "places.primaryTypeDisplayName,"
            "nextPageToken"
        ),
    }


def normalize_place(place):
    display = place.get("displayName") or {}
    category = place.get("primaryTypeDisplayName") or {}
    location = place.get("location") or {}
    return {
        "google_place_id": place.get("id", ""),
        "name": display.get("text", ""),
        "category": category.get("text", ""),
        "phone": place.get("nationalPhoneNumber") or place.get("internationalPhoneNumber") or "",
        "website": place.get("websiteUri", ""),
        "address": place.get("formattedAddress", ""),
        "latitude": location.get("latitude"),
        "longitude": location.get("longitude"),
        "rating": place.get("rating"),
        "total_reviews": place.get("userRatingCount"),
        "business_status": place.get("businessStatus", ""),
    }


def rectangle_restriction(zone):
    return {
        "rectangle": {
            "low": {"latitude": float(zone["sur"]), "longitude": float(zone["oeste"])},
            "high": {"latitude": float(zone["norte"]), "longitude": float(zone["este"])},
        }
    }


def search_google_places(text_query, limit_count=20, location_restriction=None):
    api_key = google_places_key()
    if not api_key:
        raise ValueError("Falta configurar GOOGLE_API_KEY o google_places_api_key en app_config.json.")
    query = (text_query or "").strip()
    if not query:
        raise ValueError("Escribe una busqueda para prospectar.")
    limit_count = min(max(int(limit_count or 20), 1), 60)
    payload = {
        "textQuery": query,
        "pageSize": min(limit_count, 20),
        "languageCode": "es",
        "regionCode": "MX",
    }
    if location_restriction:
        payload["locationRestriction"] = location_restriction
    results = []
    next_page_token = None
    while len(results) < limit_count:
        if next_page_token:
            payload["pageToken"] = next_page_token
        response = requests.post(
            PLACES_TEXT_SEARCH_URL,
            headers=google_places_headers(api_key),
            json=payload,
            timeout=30,
        )
        if response.status_code != 200:
            raise RuntimeError(f"Error Google Places {response.status_code}: {response.text}")
        data = response.json()
        results.extend(normalize_place(place) for place in data.get("places", []))
        next_page_token = data.get("nextPageToken")
        if not next_page_token:
            break
        time.sleep(2)
    return results[:limit_count]


def upsert_zone(db, zone):
    name = (zone.get("nombre") or zone.get("name") or "").strip()
    rect = zone.get("rectangle") or zone
    if not name:
        return None
    zone_id = str(uuid.uuid4())
    db.execute(
        """
        INSERT INTO prospect_zones (id, name, sur, oeste, norte, este, enabled, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?)
        ON CONFLICT(name) DO UPDATE SET
            sur = excluded.sur,
            oeste = excluded.oeste,
            norte = excluded.norte,
            este = excluded.este,
            updated_at = excluded.updated_at
        """,
        (
            zone_id,
            name,
            float(rect.get("sur")),
            float(rect.get("oeste")),
            float(rect.get("norte")),
            float(rect.get("este")),
            now_iso(),
            now_iso(),
        ),
    )
    return name


def list_zones(db):
    rows = db.execute(
        "SELECT * FROM prospect_zones WHERE enabled = 1 ORDER BY name"
    ).fetchall()
    return [row_to_dict(row) for row in rows]


def get_zone(db, name):
    if not name:
        return None
    row = db.execute("SELECT * FROM prospect_zones WHERE name = ? AND enabled = 1", (name,)).fetchone()
    return row_to_dict(row) if row else None


def scan_already_done(db, query_text, zone_name):
    row = db.execute(
        "SELECT id FROM prospect_searches WHERE query = ? AND zone_name = ? LIMIT 1",
        (query_text, zone_name),
    ).fetchone()
    return bool(row)


def upsert_prospect(db, prospect, source_query, user_id=None, zone_name=None):
    prospect_id = None
    google_place_id = prospect.get("google_place_id") or ""
    if google_place_id:
        row = db.execute("SELECT id FROM prospects WHERE google_place_id = ?", (google_place_id,)).fetchone()
        prospect_id = row["id"] if row else None
    prospect_id = prospect_id or prospect.get("id") or str(uuid.uuid4())
    db.execute(
        """
        INSERT INTO prospects
        (id, google_place_id, name, category, phone, website, address, latitude, longitude,
         rating, total_reviews, business_status, source_query, zone_name, status, notes, assigned_user_id,
         client_id, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, COALESCE((SELECT status FROM prospects WHERE id = ?), 'nuevo'),
                COALESCE((SELECT notes FROM prospects WHERE id = ?), ''), ?, COALESCE((SELECT client_id FROM prospects WHERE id = ?), NULL),
                COALESCE((SELECT created_at FROM prospects WHERE id = ?), ?), ?)
        ON CONFLICT(google_place_id) DO UPDATE SET
            name = excluded.name,
            category = excluded.category,
            phone = excluded.phone,
            website = excluded.website,
            address = excluded.address,
            latitude = excluded.latitude,
            longitude = excluded.longitude,
            rating = excluded.rating,
            total_reviews = excluded.total_reviews,
            business_status = excluded.business_status,
            source_query = excluded.source_query,
            zone_name = COALESCE(excluded.zone_name, prospects.zone_name),
            assigned_user_id = COALESCE(excluded.assigned_user_id, prospects.assigned_user_id),
            updated_at = excluded.updated_at
        """,
        (
            prospect_id,
            google_place_id or None,
            prospect.get("name") or "Sin nombre",
            prospect.get("category", ""),
            prospect.get("phone", ""),
            prospect.get("website", ""),
            prospect.get("address", ""),
            prospect.get("latitude"),
            prospect.get("longitude"),
            prospect.get("rating"),
            prospect.get("total_reviews"),
            prospect.get("business_status", ""),
            source_query,
            zone_name,
            prospect_id,
            prospect_id,
            user_id or None,
            prospect_id,
            prospect_id,
            now_iso(),
            now_iso(),
        ),
    )
    return prospect_id


def normalize_match_text(value):
    text = unicodedata.normalize("NFD", str(value or "").lower())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    cleaned = []
    for ch in text:
        cleaned.append(ch if ch.isalnum() else " ")
    words = [w for w in "".join(cleaned).split() if w not in {"sa", "cv", "de", "la", "el", "los", "las", "y", "the"}]
    return " ".join(words)


def compact_match_text(value):
    return normalize_match_text(value).replace(" ", "")


def client_match_reason(prospect, client):
    prospect_name = normalize_match_text(prospect.get("name"))
    client_name = normalize_match_text(client.get("name"))
    prospect_address = normalize_match_text(prospect.get("address"))
    client_address = normalize_match_text(
        " ".join(
            str(client.get(k) or "")
            for k in ("tax_address", "consignee_address", "delivery_method")
        )
    )
    if prospect_name and client_name:
        pn = compact_match_text(prospect_name)
        cn = compact_match_text(client_name)
        if pn == cn:
            return "nombre exacto"
        if len(pn) >= 8 and len(cn) >= 8 and (pn in cn or cn in pn):
            return "nombre similar"
    if prospect_address and client_address:
        pa = compact_match_text(prospect_address)
        ca = compact_match_text(client_address)
        if len(pa) >= 18 and len(ca) >= 18 and (pa in ca or ca in pa):
            return "direccion similar"
    return ""


def find_local_client_match(db, prospect):
    rows = db.execute(
        """
        SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), u.name) AS assigned_user
        FROM clients c
        LEFT JOIN users u ON u.id = c.assigned_user_id
        ORDER BY c.updated_at DESC
        LIMIT 2000
        """
    ).fetchall()
    prospect_data = dict(prospect)
    for row in rows:
        client = row_to_dict(row)
        reason = client_match_reason(prospect_data, client)
        if reason:
            return {**client, "match_reason": reason}
    return None


def find_external_client_match(prospect):
    search_terms = []
    name = str(prospect.get("name") or "").strip()
    address = str(prospect.get("address") or "").strip()
    if name:
        search_terms.append(name)
    if address:
        search_terms.append(address.split(",")[0])
    seen = set()
    for term in search_terms:
        if not term or term in seen:
            continue
        seen.add(term)
        try:
            candidates = external_clients(term)
        except Exception:
            continue
        for client in candidates[:40]:
            reason = client_match_reason(prospect, client)
            if reason:
                return {**client, "match_reason": reason}
    return None


def attach_client_match(db, prospect_id, client_id, reason):
    db.execute(
        """
        UPDATE prospects
        SET client_id = ?, status = CASE WHEN status = 'convertido' THEN status ELSE 'cliente_existente' END,
            notes = CASE
                WHEN COALESCE(notes, '') LIKE '%Cliente existente detectado:%' THEN notes
                ELSE TRIM(COALESCE(notes, '') || CASE WHEN COALESCE(notes, '') = '' THEN '' ELSE CHAR(10) END || 'Cliente existente detectado: ' || ?)
            END,
            updated_at = ?
        WHERE id = ?
        """,
        (client_id, reason, now_iso(), prospect_id),
    )


def check_prospect_clients(db, search="", status="", zone_name="", include_external=True):
    prospects = list_prospects(db, search, status, zone_name)
    checked = 0
    matched = 0
    for prospect in prospects:
        checked += 1
        if prospect.get("client_id"):
            continue
        match = find_local_client_match(db, prospect)
        if not match and include_external:
            match = find_external_client_match(prospect)
            if match and str(match.get("id") or "").startswith("ext:"):
                ensure_local_client(db, match["id"])
        if match and match.get("id"):
            attach_client_match(db, prospect["id"], match["id"], match.get("match_reason") or "coincidencia")
            matched += 1
    return {"checked": checked, "matched": matched}


def start_prospect_followup(db, prospect_id, user_id=None):
    prospect = db.execute("SELECT * FROM prospects WHERE id = ?", (prospect_id,)).fetchone()
    if not prospect:
        raise ValueError("Prospecto no encontrado")
    if user_id:
        ensure_local_user(db, user_id)
    db.execute(
        """
        UPDATE prospects
        SET status = 'en_seguimiento',
            assigned_user_id = COALESCE(?, assigned_user_id),
            updated_at = ?
        WHERE id = ?
        """,
        (user_id or None, now_iso(), prospect_id),
    )
    log_prospect_activity(db, prospect_id, user_id, "seguimiento", "Seguimiento iniciado", row_to_dict(prospect))
    return prospect_id


def find_client_by_code(db, client_code):
    code = str(client_code or "").strip()
    if not code:
        return None
    row = db.execute(
        """
        SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), u.name) AS assigned_user
        FROM clients c
        LEFT JOIN users u ON u.id = c.assigned_user_id
        WHERE c.code = ? OR c.id = ?
        LIMIT 1
        """,
        (code, code),
    ).fetchone()
    if row:
        return row_to_dict(row)
    try:
        for client in external_clients(code):
            if str(client.get("code") or "").strip() == code or str(client.get("id") or "").strip() == code:
                ensure_local_client(db, client["id"])
                local = db.execute("SELECT * FROM clients WHERE id = ?", (client["id"],)).fetchone()
                return row_to_dict(local) if local else client
    except Exception:
        return None
    return None


def link_prospect_to_client(db, prospect_id, client_code, user_id=None):
    prospect = db.execute("SELECT * FROM prospects WHERE id = ?", (prospect_id,)).fetchone()
    if not prospect:
        raise ValueError("Prospecto no encontrado")
    client = find_client_by_code(db, client_code)
    if not client:
        raise ValueError("No encontre cliente con ese numero/codigo.")
    if user_id:
        ensure_local_user(db, user_id)
    db.execute(
        """
        UPDATE prospects
        SET client_id = ?, status = 'convertido',
            assigned_user_id = COALESCE(?, assigned_user_id),
            notes = CASE
                WHEN COALESCE(notes, '') LIKE '%Asociado al cliente:%' THEN notes
                ELSE TRIM(COALESCE(notes, '') || CASE WHEN COALESCE(notes, '') = '' THEN '' ELSE CHAR(10) END || 'Asociado al cliente: ' || ?)
            END,
            updated_at = ?
        WHERE id = ?
        """,
        (client["id"], user_id or None, client.get("code") or client["id"], now_iso(), prospect_id),
    )
    log_prospect_activity(
        db,
        prospect_id,
        user_id,
        "cliente",
        f"Prospecto asociado al cliente: {client.get('code') or client.get('name') or client['id']}",
        {"client_id": client["id"], "client_code": client.get("code"), "client_name": client.get("name")},
    )
    return client["id"]


def log_prospect_activity(db, prospect_id, user_id, entry_type, title, payload=None):
    db.execute(
        """
        INSERT INTO prospect_activity (id, prospect_id, user_id, type, title, payload, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(uuid.uuid4()),
            prospect_id,
            user_id or None,
            entry_type,
            title,
            json.dumps(payload or {}, ensure_ascii=False),
            now_iso(),
        ),
    )


def prospect_payload(db, prospect_id):
    prospect = db.execute(
        """
        SELECT p.*, u.name AS assigned_user, c.name AS client_name, c.code AS client_code
        FROM prospects p
        LEFT JOIN users u ON u.id = p.assigned_user_id
        LEFT JOIN clients c ON c.id = p.client_id
        WHERE p.id = ?
        """,
        (prospect_id,),
    ).fetchone()
    if not prospect:
        return None
    quotes = db.execute(
        """
        SELECT q.*, COALESCE(NULLIF(q.folio, ''), q.title) AS quote_title, u.name AS user_name
        FROM prospect_quotes q
        LEFT JOIN users u ON u.id = q.user_id
        WHERE q.prospect_id = ?
        ORDER BY q.created_at DESC
        """,
        (prospect_id,),
    ).fetchall()
    followups = db.execute(
        """
        SELECT f.*, u.name AS user_name
        FROM prospect_followups f
        LEFT JOIN users u ON u.id = f.user_id
        WHERE f.prospect_id = ?
        ORDER BY f.contact_at DESC
        """,
        (prospect_id,),
    ).fetchall()
    messages = db.execute(
        """
        SELECT m.*, u.name AS user_name
        FROM prospect_messages m
        LEFT JOIN users u ON u.id = m.user_id
        WHERE m.prospect_id = ?
        ORDER BY m.sent_at DESC
        """,
        (prospect_id,),
    ).fetchall()
    phones = db.execute(
        """
        SELECT *
        FROM prospect_phones
        WHERE prospect_id = ?
        ORDER BY updated_at DESC
        """,
        (prospect_id,),
    ).fetchall()
    activity = db.execute(
        """
        SELECT a.*, u.name AS user_name
        FROM prospect_activity a
        LEFT JOIN users u ON u.id = a.user_id
        WHERE a.prospect_id = ?
        ORDER BY a.created_at DESC
        LIMIT 80
        """,
        (prospect_id,),
    ).fetchall()
    return {
        "prospect": row_to_dict(prospect),
        "quotes": [row_to_dict(row) for row in quotes],
        "followups": [row_to_dict(row) for row in followups],
        "messages": [row_to_dict(row) for row in messages],
        "phones": [row_to_dict(row) for row in phones],
        "activity": [row_to_dict(row) for row in activity],
    }


def ensure_prospect_for_followup(db, prospect_id, user_id=None):
    prospect = db.execute("SELECT * FROM prospects WHERE id = ?", (prospect_id,)).fetchone()
    if not prospect:
        raise ValueError("Prospecto no encontrado")
    if user_id:
        ensure_local_user(db, user_id)
    if prospect["status"] == "nuevo":
        start_prospect_followup(db, prospect_id, user_id)
    return prospect


def add_prospect_quote(db, data):
    prospect_id = data.get("prospect_id") or ""
    user_id = data.get("user_id") or None
    ensure_prospect_for_followup(db, prospect_id, user_id)
    items = data.get("items", [])
    if not items:
        raise ValueError("Agrega al menos una partida")
    totals = calculate_quote(items)
    quote_id = str(uuid.uuid4())
    folio = f"PCOT-{datetime.now().strftime('%Y%m%d')}-{db.execute('SELECT COUNT(*) + 1 AS n FROM prospect_quotes').fetchone()['n']:04d}"
    title = (data.get("quote_recipient") or data.get("title") or "").strip()
    if not title:
        prospect = db.execute("SELECT name FROM prospects WHERE id = ?", (prospect_id,)).fetchone()
        title = prospect["name"] if prospect else folio
    quote_title = f"{title} - {folio}"
    db.execute(
        """
        INSERT INTO prospect_quotes
        (id, prospect_id, user_id, folio, title, amount, valid_until, subtotal, discount, tax, total, authorized_shipping, status, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            quote_id,
            prospect_id,
            user_id,
            folio,
            quote_title,
            totals["total"],
            quote_valid_until(),
            totals["subtotal"],
            totals["discount"],
            totals["tax"],
            totals["total"],
            totals["authorized_shipping"],
            data.get("status") or "emitida",
            data.get("notes") or "",
            now_iso(),
        ),
    )
    for item in totals["items"]:
        db.execute(
            """
            INSERT INTO prospect_quote_items
            (id, quote_id, cip, description, quantity, unit_price, discount_rate, tax_rate, line_total)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item["id"],
                quote_id,
                item["cip"],
                item["description"],
                item["quantity"],
                item["unit_price"],
                item["discount_rate"],
                item["tax_rate"],
                item["line_total"],
            ),
        )
    db.execute("UPDATE prospects SET updated_at = ? WHERE id = ?", (now_iso(), prospect_id))
    log_prospect_activity(db, prospect_id, user_id, "cotizacion", f"Cotizacion de prospecto: {quote_title}", totals)
    return {"id": quote_id, "folio": folio, "quote_title": quote_title, **totals}


def update_prospect_quote(db, data):
    quote_id = data.get("id") or ""
    quote = db.execute("SELECT * FROM prospect_quotes WHERE id = ?", (quote_id,)).fetchone()
    if not quote:
        raise ValueError("Cotizacion de prospecto no encontrada")
    prospect_id = data.get("prospect_id") or quote["prospect_id"]
    user_id = data.get("user_id") or quote["user_id"]
    ensure_prospect_for_followup(db, prospect_id, user_id)
    items = data.get("items", [])
    if not items:
        raise ValueError("Agrega al menos una partida")
    totals = calculate_quote(items)
    title = (data.get("quote_recipient") or data.get("title") or "").strip()
    if not title:
        prospect = db.execute("SELECT name FROM prospects WHERE id = ?", (prospect_id,)).fetchone()
        title = prospect["name"] if prospect else quote["folio"]
    quote_title = f"{title} - {quote['folio']}"
    db.execute(
        """
        UPDATE prospect_quotes
        SET prospect_id = ?, user_id = ?, title = ?, amount = ?, subtotal = ?, discount = ?,
            tax = ?, total = ?, authorized_shipping = ?, notes = ?, status = ?
        WHERE id = ?
        """,
        (
            prospect_id,
            user_id,
            quote_title,
            totals["total"],
            totals["subtotal"],
            totals["discount"],
            totals["tax"],
            totals["total"],
            totals["authorized_shipping"],
            data.get("notes") or "",
            data.get("status") or quote["status"] or "emitida",
            quote_id,
        ),
    )
    db.execute("DELETE FROM prospect_quote_items WHERE quote_id = ?", (quote_id,))
    for item in totals["items"]:
        db.execute(
            """
            INSERT INTO prospect_quote_items
            (id, quote_id, cip, description, quantity, unit_price, discount_rate, tax_rate, line_total)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                item["id"],
                quote_id,
                item["cip"],
                item["description"],
                item["quantity"],
                item["unit_price"],
                item["discount_rate"],
                item["tax_rate"],
                item["line_total"],
            ),
        )
    db.execute("UPDATE prospects SET updated_at = ? WHERE id = ?", (now_iso(), prospect_id))
    log_prospect_activity(db, prospect_id, user_id, "cotizacion", f"Cotizacion de prospecto editada: {quote_title}", totals)
    return {"id": quote_id, "folio": quote["folio"], "quote_title": quote_title, **totals}


def prospect_quote_payload(db, quote_id):
    quote = db.execute(
        """
        SELECT q.*, p.name AS prospect_name, p.phone, p.address, p.website, p.category,
               p.zone_name, u.name AS user_name
        FROM prospect_quotes q
        JOIN prospects p ON p.id = q.prospect_id
        LEFT JOIN users u ON u.id = q.user_id
        WHERE q.id = ?
        """,
        (quote_id,),
    ).fetchone()
    if not quote:
        return None
    items = db.execute("SELECT * FROM prospect_quote_items WHERE quote_id = ?", (quote_id,)).fetchall()
    data = row_to_dict(quote)
    data["items"] = [row_to_dict(row) for row in items]
    data["quote_title"] = data.get("title") or data.get("folio")
    data["client_name"] = data.get("prospect_name") or ""
    data["client_code"] = "PROSPECTO"
    data["email"] = ""
    data["tax_address"] = data.get("address") or ""
    data["consignee_address"] = data.get("address") or ""
    data["delivery_method"] = data.get("address") or ""
    data["company"] = CFG.get("company_name") or ""
    data["bank_account"] = bank_account_for_company(db, data["company"])
    return data


def add_prospect_followup(db, data):
    prospect_id = data.get("prospect_id") or ""
    user_id = data.get("user_id") or None
    ensure_prospect_for_followup(db, prospect_id, user_id)
    followup_id = str(uuid.uuid4())
    channel = data.get("channel") or "Llamada"
    outcome = (data.get("outcome") or "").strip()
    if not outcome:
        raise ValueError("Captura el resultado del seguimiento")
    db.execute(
        """
        INSERT INTO prospect_followups
        (id, prospect_id, user_id, contact_at, channel, outcome, next_action, next_action_at, notes, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            followup_id,
            prospect_id,
            user_id,
            data.get("contact_at") or now_iso(),
            channel,
            outcome,
            data.get("next_action") or "",
            data.get("next_action_at") or "",
            data.get("notes") or "",
            now_iso(),
        ),
    )
    db.execute("UPDATE prospects SET updated_at = ? WHERE id = ?", (now_iso(), prospect_id))
    log_prospect_activity(db, prospect_id, user_id, "seguimiento", f"Seguimiento registrado: {channel}", data)
    return followup_id


def add_prospect_message(db, data):
    prospect_id = data.get("prospect_id") or ""
    user_id = data.get("user_id") or None
    ensure_prospect_for_followup(db, prospect_id, user_id)
    body = (data.get("body") or "").strip()
    if not body:
        raise ValueError("Captura el mensaje")
    message_id = str(uuid.uuid4())
    channel = data.get("channel") or "WhatsApp"
    db.execute(
        """
        INSERT INTO prospect_messages (id, prospect_id, user_id, sent_at, channel, direction, body, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            message_id,
            prospect_id,
            user_id,
            data.get("sent_at") or now_iso(),
            channel,
            data.get("direction") or "saliente",
            body,
            now_iso(),
        ),
    )
    db.execute("UPDATE prospects SET updated_at = ? WHERE id = ?", (now_iso(), prospect_id))
    log_prospect_activity(db, prospect_id, user_id, "mensaje", f"Mensaje registrado: {channel}", data)
    return message_id


def add_prospect_phone(db, data):
    prospect_id = data.get("prospect_id") or ""
    ensure_prospect_for_followup(db, prospect_id, data.get("user_id") or None)
    phone = (data.get("phone") or "").strip()
    if not phone:
        raise ValueError("Captura el telefono")
    phone_id = str(uuid.uuid4())
    db.execute(
        """
        INSERT INTO prospect_phones (id, prospect_id, label, phone, notes, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (phone_id, prospect_id, data.get("label") or "", phone, data.get("notes") or "", now_iso(), now_iso()),
    )
    db.execute(
        """
        UPDATE prospects
        SET phone = CASE WHEN COALESCE(phone, '') = '' THEN ? ELSE phone END,
            updated_at = ?
        WHERE id = ?
        """,
        (phone, now_iso(), prospect_id),
    )
    log_prospect_activity(db, prospect_id, data.get("user_id") or None, "telefono", f"Telefono agregado: {phone}", data)
    return phone_id


def list_prospects(db, search="", status="", zone_name=""):
    params = []
    where = []
    if search:
        like = f"%{search}%"
        where.append("(p.name LIKE ? OR p.category LIKE ? OR p.address LIKE ? OR p.phone LIKE ? OR p.source_query LIKE ? OR p.zone_name LIKE ?)")
        params.extend([like, like, like, like, like, like])
    if status and status != "todos":
        where.append("p.status = ?")
        params.append(status)
    if zone_name and zone_name != "todas":
        where.append("p.zone_name = ?")
        params.append(zone_name)
    sql_where = "WHERE " + " AND ".join(where) if where else ""
    rows = db.execute(
        f"""
        SELECT p.*, u.name AS assigned_user, c.name AS client_name
        FROM prospects p
        LEFT JOIN users u ON u.id = p.assigned_user_id
        LEFT JOIN clients c ON c.id = p.client_id
        {sql_where}
        ORDER BY p.updated_at DESC
        LIMIT 300
        """,
        params,
    ).fetchall()
    return [row_to_dict(row) for row in rows]


def convert_prospect_to_client(db, prospect_id, user_id=None):
    prospect = db.execute("SELECT * FROM prospects WHERE id = ?", (prospect_id,)).fetchone()
    if not prospect:
        raise ValueError("Prospecto no encontrado")
    if prospect["client_id"]:
        return prospect["client_id"]
    if user_id:
        ensure_local_user(db, user_id)
    client_id = str(uuid.uuid4())
    db.execute(
        """
        INSERT INTO clients
        (id, code, name, tax_address, consignee_address, delivery_method, phone, email,
         contact_name, external_seller, assigned_user_id, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, '', '', '', ?, ?, ?)
        """,
        (
            client_id,
            f"PROS-{datetime.now().strftime('%y%m%d')}",
            prospect["name"],
            prospect["address"] or "",
            prospect["address"] or "",
            "Pendiente por definir",
            prospect["phone"] or "",
            user_id or prospect["assigned_user_id"],
            now_iso(),
            now_iso(),
        ),
    )
    db.execute(
        """
        UPDATE prospects
        SET status = 'convertido', client_id = ?, assigned_user_id = COALESCE(?, assigned_user_id), updated_at = ?
        WHERE id = ?
        """,
        (client_id, user_id or None, now_iso(), prospect_id),
    )
    log_activity(
        db,
        client_id,
        user_id,
        "prospector",
        f"Cliente creado desde prospector: {prospect['name']}",
        row_to_dict(prospect),
    )
    return client_id


def external_client(numero, empresa):
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM clientes WHERE numero = %s AND UPPER(TRIM(empresa)) = UPPER(TRIM(%s))", (numero, empresa))
    row = cur.fetchone()
    cur.close()
    conn.close()
    return map_external_client(row) if row else None


def external_products(search=""):
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    like = f"%{search}%"
    cur.execute(
        """
        SELECT cip, descripcion, unidad, tipo_lista, iva
        FROM productos
        WHERE cip LIKE %s OR descripcion LIKE %s
        ORDER BY CAST(cip AS UNSIGNED), cip
        LIMIT 1000
        """,
        (like, like),
    )
    rows = [normalize_mysql_row(row) for row in cur.fetchall()]
    cur.close()
    conn.close()
    return rows


def external_invoices(client_id):
    numero, empresa = parse_external_client_id(client_id)
    if not numero:
        return []
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT id, factura, fecha, subtotal, descuento, iva, total, estatus
        FROM facturas
        WHERE numero_cliente = %s AND UPPER(TRIM(empresa)) = UPPER(TRIM(%s))
        ORDER BY fecha DESC
        LIMIT 100
        """,
        (numero, empresa),
    )
    rows = [normalize_mysql_row(row) for row in cur.fetchall()]
    cur.close()
    conn.close()
    return [
        {
            "id": f"extfactid:{row['id']}",
            "client_id": client_id,
            "folio": row.get("factura") or "",
            "issued_at": str(row.get("fecha") or ""),
            "subtotal": row.get("subtotal") or 0,
            "discount": row.get("descuento") or 0,
            "tax": row.get("iva") or 0,
            "total": row.get("total") or 0,
            "status": row.get("estatus") or "",
            "detail": "",
        }
        for row in rows
    ]


def external_client_products(client_id):
    numero, empresa = parse_external_client_id(client_id)
    if not numero:
        return []
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT
            f.id AS factura_id,
            fd.cip,
            fd.descripcion,
            f.factura,
            f.fecha,
            fd.cantidad,
            fd.piezas,
            fd.precio,
            fd.importe
        FROM facturas f
        JOIN factura_detalle fd ON fd.factura_id = f.id
        WHERE f.numero_cliente = %s
          AND UPPER(TRIM(f.empresa)) = UPPER(TRIM(%s))
        ORDER BY fd.descripcion ASC, f.fecha DESC, f.factura DESC
        LIMIT 3000
        """,
        (numero, empresa),
    )
    rows = [normalize_mysql_row(row) for row in cur.fetchall()]
    cur.close()
    conn.close()

    products = {}
    for row in rows:
        cip = str(row.get("cip") or "").strip()
        desc = str(row.get("descripcion") or "").strip()
        key = cip or desc
        item = products.setdefault(
            key,
            {
                "cip": cip,
                "description": desc,
                "total_quantity": 0,
                "total_pieces": 0,
                "total_amount": 0,
                "last_purchase": "",
                "purchases": [],
            },
        )
        quantity = float(row.get("cantidad") or 0)
        pieces = float(row.get("piezas") or 0)
        amount = float(row.get("importe") or 0)
        date = str(row.get("fecha") or "")
        item["total_quantity"] += quantity
        item["total_pieces"] += pieces
        if amount <= 0:
            amount = quantity * float(row.get("precio") or 0)
        item["total_amount"] += amount
        if date > item["last_purchase"]:
            item["last_purchase"] = date
        item["purchases"].append(
            {
                "invoice_id": f"extfactid:{row.get('factura_id') or ''}",
                "invoice": row.get("factura") or "",
                "date": date,
                "quantity": quantity,
                "pieces": pieces,
                "price": float(row.get("precio") or 0),
                "amount": amount,
            }
        )
    return sorted(products.values(), key=lambda item: (item["description"], item["cip"]))


def external_invoice_detail(invoice_id):
    if not (invoice_id.startswith("extfact:") or invoice_id.startswith("extfactid:")):
        return None
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    if invoice_id.startswith("extfactid:"):
        factura_id = invoice_id.split(":", 1)[1]
        cur.execute("SELECT * FROM facturas WHERE id = %s", (factura_id,))
    else:
        folio = invoice_id.split(":", 1)[1]
        cur.execute("SELECT * FROM facturas WHERE factura = %s", (folio,))
    factura = normalize_mysql_row(cur.fetchone() or {})
    if not factura:
        cur.close()
        conn.close()
        return None
    cur.execute(
        """
        SELECT cip, descripcion, cantidad, piezas, precio, importe
        FROM factura_detalle
        WHERE factura_id = %s
        """,
        (factura["id"],),
    )
    productos = [normalize_mysql_row(row) for row in cur.fetchall()]
    cur.execute(
        "SELECT * FROM clientes WHERE numero = %s AND UPPER(TRIM(empresa)) = UPPER(TRIM(%s)) LIMIT 1",
        (factura.get("numero_cliente") or "", factura.get("empresa") or ""),
    )
    cliente = normalize_mysql_row(cur.fetchone() or {})
    cur.close()
    conn.close()
    detail = "; ".join(
        f"{p.get('cantidad') or 0} x {p.get('cip') or ''} {p.get('descripcion') or ''} = ${float(p.get('importe') or 0):,.2f}"
        for p in productos
    )
    return {
        "id": invoice_id,
        "client_id": "",
        "numero_cliente": factura.get("numero_cliente") or "",
        "cliente_nombre": factura.get("consignatario") or cliente.get("nombre") or "",
        "empresa": factura.get("empresa") or "",
        "vendedor": factura.get("vendedor") or cliente.get("vendedor") or "",
        "rfc": factura.get("rfc") or cliente.get("rfc") or "",
        "folio": factura.get("factura") or "",
        "issued_at": str(factura.get("fecha") or ""),
        "subtotal": factura.get("subtotal") or 0,
        "discount": factura.get("descuento") or 0,
        "tax": factura.get("iva") or 0,
        "total": factura.get("total") or 0,
        "status": factura.get("estatus") or "",
        "detail": detail,
        "products": productos,
        "cliente": cliente,
    }


def company_invoice_config(empresa):
    emp = (empresa or "").lower()
    base = crm_logo_dir()
    if "remision" in emp or "remisión" in emp:
        return "REMISION", "Texas N°100 - Nápoles - Benito Juárez - CDMX", "REMISION INTERNA / TEL. 5555439933", os.path.join(base, "Remision.png")
    if "ibersur" in emp:
        return "IBERSUR", "Dakota N°359 Int. 301 - Ampliación Nápoles - Benito Juárez - CDMX", "RFC IBE 090212 JV1 / TEL. 5555439933", os.path.join(base, "ibersur.png")
    if "gourmet" in emp:
        return "GOURMET ESPAÑA", "Texas N°100 - Nápoles - Benito Juárez - CDMX", "RFC GES 090312 DJ1 / TEL. 5555439933", os.path.join(base, "gourmet.png")
    return "EZA2007", "Dakota N°359 Int. 301 - Ampliación Nápoles - Benito Juárez - CDMX", "RFC EZA 070521 MT4 / TEL. 5555439933", os.path.join(base, "eza2007.png")


def product_barcode_for_company(cip, empresa, lista_nombre=None):
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    try:
        lista = (lista_nombre or "").strip()
        if lista:
            cur.execute(
                """
                SELECT COALESCE(pp.codigo_barras, '') AS codigo_barras
                FROM precios_productos pp
                JOIN listas_precios lp ON lp.id = pp.lista_id
                WHERE pp.cip = %s
                  AND lp.nombre COLLATE utf8mb4_unicode_ci = %s COLLATE utf8mb4_unicode_ci
                  AND COALESCE(pp.codigo_barras, '') <> ''
                LIMIT 1
                """,
                (cip, lista),
            )
            row = cur.fetchone()
            if row and row.get("codigo_barras"):
                return clean_barcode(row["codigo_barras"])
        cur.execute(
            """
            SELECT COALESCE(pp.codigo_barras, '') AS codigo_barras
            FROM precios_productos pp
            WHERE pp.cip = %s
              AND COALESCE(pp.codigo_barras, '') <> ''
            LIMIT 1
            """,
            (cip,),
        )
        row = cur.fetchone()
        if row and row.get("codigo_barras"):
            return clean_barcode(row["codigo_barras"])
        cur.execute("SELECT COALESCE(codigo_barras, '') AS codigo_barras FROM productos WHERE cip = %s LIMIT 1", (cip,))
        row = cur.fetchone()
        return clean_barcode((row or {}).get("codigo_barras") or "")
    finally:
        cur.close()
        conn.close()


def clean_barcode(value):
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def amount_to_spanish_words(total):
    entero = int(float(total or 0))
    centavos = int(round((float(total or 0) - entero) * 100))
    if num2words:
        words = num2words(entero, lang="es").upper()
    else:
        words = str(entero)
    words = words.replace("EUROS", "").strip()
    return f"{words} PESOS {centavos:02d}/100 M.N."


def export_invoice_pdf(invoice_id):
    data = external_invoice_detail(invoice_id)
    if not data:
        raise ValueError("Factura no encontrada")
    os.makedirs(EXPORT_DIR, exist_ok=True)
    path = os.path.join(EXPORT_DIR, f"FACTURA-{data['folio']}.pdf")
    empresa_nombre, direccion, rfc_tel, logo_path = company_invoice_config(data.get("empresa"))
    cliente = data.get("cliente") or {}

    doc = SimpleDocTemplate(path, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    small = ParagraphStyle(name="small", parent=styles["Normal"], fontSize=8, leading=9)
    elements = []
    logo = Image(logo_path, width=100, height=80) if os.path.exists(logo_path) else Spacer(1, 60)
    info_empresa = f"<b>{empresa_nombre}</b><br/>{direccion}<br/>{rfc_tel}"
    header = Table([[logo, Paragraph(info_empresa, ParagraphStyle(name="empresa", fontSize=8, leftIndent=20))]], colWidths=[120, 420])
    header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (1, 0), (1, 0), 15)]))
    elements.append(header)
    elements.append(Spacer(1, 6))
    elements.append(Paragraph(f"<para alignment='right'><b>FOLIO: {data['folio']}</b></para>", ParagraphStyle(name="folio", fontSize=10, alignment=2)))

    def cell(value):
        return Paragraph(str(value or ""), small)

    cliente_table = Table([
        ["Cliente:", cell(cliente.get("razon_social") or data.get("cliente_nombre"))],
        ["RFC:", cell(data.get("rfc"))],
        ["Calle:", cell(f"{cliente.get('calle','')} {cliente.get('no_exterior','')} {cliente.get('no_interior','')}")],
        ["Colonia:", cell(cliente.get("colonia"))],
        ["Delegación:", cell(cliente.get("alcaldia") or cliente.get("municipio"))],
        ["Población:", cell(f"{cliente.get('poblacion','')} C.P. {cliente.get('codigo_postal','')}")],
        ["Estado:", cell(cliente.get("estado"))],
    ], colWidths=[60, 260])
    consignatario_table = Table([
        ["Consignatario:", cell(cliente.get("consignatario") or data.get("cliente_nombre"))],
        ["Calle:", cell(f"{cliente.get('consig_calle','')} {cliente.get('consig_no_exterior','')} {cliente.get('consig_no_interior','')}")],
        ["Colonia:", cell(cliente.get("consig_colonia"))],
        ["Delegación:", cell(cliente.get("consig_delegacion") or cliente.get("consig_municipio"))],
        ["Población:", cell(f"{cliente.get('consig_poblacion','')} C.P. {cliente.get('consig_codigo_postal','')}")],
        ["Estado:", cell(cliente.get("consig_estado"))],
    ], colWidths=[70, 250])
    for table in (cliente_table, consignatario_table):
        table.setStyle(TableStyle([("FONTNAME", (0, 0), (-1, -1), "Helvetica"), ("FONTSIZE", (0, 0), (-1, -1), 8), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
    elements.append(Table([[cliente_table, consignatario_table]], colWidths=[270, 270], style=[("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements.append(Spacer(1, 6))
    elements.append(Table([[""]], colWidths=[540], style=[("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.black)]))

    fecha = str(data.get("issued_at") or "")[:10]
    pago = f"{cliente.get('dias_credito')} días" if cliente.get("dias_credito") else "-"
    meta = Table([
        ["Ubicación", "Fecha", "Pago", "N° Proveedor", "Cliente N°", "Vendedor"],
        ["MEXICO DF", fecha, pago, cliente.get("no_proveedor") or "-", data.get("numero_cliente") or "-", data.get("vendedor") or "-"],
    ], colWidths=[100, 90, 80, 90, 90, 90])
    meta.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER"), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    elements.append(meta)
    elements.append(Table([[""]], colWidths=[540], style=[("LINEBELOW", (0, 0), (-1, 0), 0.8, colors.black)]))
    elements.append(Spacer(1, 5))

    is_gourmet = "gourmet" in (data.get("empresa") or "").lower()
    if is_gourmet:
        rows = [["Cantidad", "Unidad", "CIP", "Descripción", "Código", "Piezas", "Precio", "Total"]]
        col_widths = [55, 45, 35, 195, 70, 40, 50, 50]
    else:
        rows = [["Cantidad", "Unidad", "CIP", "Descripción", "Piezas", "Precio", "Total"]]
        col_widths = [60, 50, 45, 225, 55, 50, 55]
    for p in data.get("products", []):
        cantidad = float(p.get("cantidad") or 0)
        precio = float(p.get("precio") or 0)
        importe = float(p.get("importe") or (cantidad * precio))
        base_row = [
            f"{cantidad:,.2f}",
            p.get("unidad") or "PZA",
            p.get("cip") or "",
            Paragraph(str(p.get("descripcion") or ""), small),
        ]
        if is_gourmet:
            base_row.append(product_barcode_for_company(
                p.get("cip") or "",
                data.get("empresa"),
                (data.get("cliente") or {}).get("especial"),
            ) or "-")
        base_row.extend([
            str(p.get("piezas") or ""),
            f"${precio:,.2f}",
            f"${importe:,.2f}",
        ])
        rows.append(base_row)
    products = Table(rows, repeatRows=1, colWidths=col_widths)
    products.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 1), (-1, -1), 8), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 1), (-1, -1), "CENTER"), ("ALIGN", (3, 0), (3, -1), "LEFT")]))
    elements.append(products)
    elements.append(Spacer(1, 10))

    subtotal = float(data.get("subtotal") or 0)
    discount = float(data.get("discount") or 0)
    tax = float(data.get("tax") or 0)
    total = float(data.get("total") or 0)
    discount_pct = (discount / subtotal * 100) if subtotal else 0
    totals = Table([
        ["", "", "", "", "SUMA", f"${subtotal:,.2f}"],
        ["", "", "", "", f"Descuento ({discount_pct:.2f}%)", f"-${discount:,.2f}"],
        ["", "", "", "", "I.V.A.", f"${tax:,.2f}"],
        ["", "", "", "", "GRAN TOTAL", f"${total:,.2f}"],
    ], colWidths=[60, 60, 100, 120, 100, 80])
    totals.setStyle(TableStyle([("ALIGN", (-2, 0), (-1, -1), "CENTER"), ("FONTNAME", (-2, 0), (-1, -1), "Helvetica-Bold")]))
    elements.append(totals)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(amount_to_spanish_words(total), ParagraphStyle(name="total_letra", fontSize=10, alignment=0)))
    doc.build(elements)
    return path


def price_for_client(client_id, cip):
    numero, empresa = parse_external_client_id(client_id)
    if not numero:
        return None
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT IFNULL(especial, 'Lista General') AS lista_nombre
        FROM clientes
        WHERE numero = %s AND UPPER(TRIM(empresa)) = UPPER(TRIM(%s))
        """,
        (numero, empresa),
    )
    cliente = cur.fetchone()
    if not cliente:
        cur.close()
        conn.close()
        return None
    lista = (cliente.get("lista_nombre") or "Lista General").strip()
    cur.execute("SELECT descripcion, unidad, tipo_lista, iva FROM productos WHERE cip = %s", (cip,))
    producto = cur.fetchone()
    if not producto:
        cur.close()
        conn.close()
        return None
    cur.execute(
        """
        SELECT pp.precio
        FROM precios_productos pp
        JOIN listas_precios lp ON lp.id = pp.lista_id
        WHERE UPPER(TRIM(lp.nombre)) = UPPER(TRIM(%s)) AND pp.cip = %s
        """,
        (lista, cip),
    )
    precio = cur.fetchone()
    if not precio or float(precio.get("precio") or 0) <= 0:
        cur.execute(
            """
            SELECT pp.precio
            FROM precios_productos pp
            JOIN listas_precios lp ON lp.id = pp.lista_id
            WHERE UPPER(TRIM(lp.nombre)) IN (UPPER(TRIM('Lista General')), UPPER(TRIM('L GENERAL'))) AND pp.cip = %s
              AND COALESCE(pp.precio, 0) > 0
            ORDER BY CASE WHEN UPPER(TRIM(lp.nombre)) = UPPER(TRIM('L GENERAL')) THEN 0 ELSE 1 END
            LIMIT 1
            """,
            (cip,),
        )
        precio = cur.fetchone()
    cur.close()
    conn.close()
    tax_rate = 16 if str(producto.get("iva") or "").strip().lower() in ("si", "sí", "s") else 0
    return {
        "cip": cip,
        "description": producto["descripcion"],
        "unit": producto.get("unidad") or "",
        "list": lista,
        "unit_price": float(precio["precio"]) if precio else 0.0,
        "discount_rate": 0,
        "tax_rate": tax_rate,
    }


def product_quote_info(cip):
    code = str(cip or "").strip()
    if not code:
        return None
    conn = mysql_conn()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT cip, descripcion, unidad, tipo_lista, iva FROM productos WHERE cip = %s LIMIT 1", (code,))
        producto = cur.fetchone()
        if not producto:
            cur.execute(
                """
                SELECT cip, descripcion, unidad, tipo_lista, iva
                FROM productos
                WHERE descripcion LIKE %s OR cip LIKE %s
                ORDER BY cip
                LIMIT 1
                """,
                (f"%{code}%", f"%{code}%"),
            )
            producto = cur.fetchone()
        if not producto:
            return None
        cur.execute(
            """
            SELECT pp.precio
            FROM precios_productos pp
            JOIN listas_precios lp ON lp.id = pp.lista_id
            WHERE UPPER(TRIM(lp.nombre)) IN (UPPER(TRIM('Lista General')), UPPER(TRIM('L GENERAL'))) AND pp.cip = %s
              AND COALESCE(pp.precio, 0) > 0
            ORDER BY CASE WHEN UPPER(TRIM(lp.nombre)) = UPPER(TRIM('L GENERAL')) THEN 0 ELSE 1 END
            LIMIT 1
            """,
            (producto["cip"],),
        )
        precio = cur.fetchone()
        tax_rate = 16 if str(producto.get("iva") or "").strip().lower() in ("si", "sí", "s", "sÃ­") else 0
        return {
            "cip": producto["cip"],
            "description": producto.get("descripcion") or "",
            "unit": producto.get("unidad") or "",
            "list": "Lista General",
            "unit_price": float(precio["precio"]) if precio else 0.0,
            "discount_rate": 0,
            "tax_rate": tax_rate,
        }
    finally:
        cur.close()
        conn.close()


def ensure_local_client(db, client_id):
    if not client_id.startswith("ext:"):
        return
    existing_crm = crm_client_row(client_id)
    if existing_crm:
        mirror_client_to_sqlite(db, existing_crm)
        return
    numero, empresa = parse_external_client_id(client_id)
    client = external_client(numero, empresa)
    if not client:
        return
    server_client = {
        **client,
        "id": client_id,
        "assigned_user_id": None,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    crm_save_client(db, server_client)
    exists = db.execute("SELECT id FROM clients WHERE id = ?", (client_id,)).fetchone()
    if exists:
        db.execute(
            """
            UPDATE clients
            SET external_seller = COALESCE(NULLIF(?, ''), external_seller),
                contact_name = COALESCE(NULLIF(?, ''), contact_name),
                updated_at = ?
            WHERE id = ?
            """,
            (
                client.get("external_seller") or "",
                client.get("contact_name") or "",
                now_iso(),
                client_id,
            ),
        )
        return
    db.execute(
        """
        INSERT INTO clients
        (id, code, name, tax_address, consignee_address, delivery_method, phone, email, contact_name, external_seller, assigned_user_id, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?)
        """,
        (
            client["id"],
            client["code"],
            client["name"],
            client["tax_address"],
            client["consignee_address"],
            client["delivery_method"],
            client["phone"],
            client["email"],
            client.get("contact_name") or "",
            client.get("external_seller") or "",
            now_iso(),
            now_iso(),
        ),
    )


def ensure_local_user(db, user_id):
    if not user_id:
        return
    exists = db.execute("SELECT id FROM users WHERE id = ?", (user_id,)).fetchone()
    if exists:
        return
    name = user_id.replace("ext:", "", 1)
    db.execute(
        "INSERT INTO users (id, name, email, role, active, created_at) VALUES (?, ?, '', 'vendedor', 1, ?)",
        (user_id, name, now_iso()),
    )


def mirror_client_to_sqlite(db, data):
    client_id = data.get("id") or ""
    if not client_id:
        return
    existing = db.execute("SELECT id FROM clients WHERE id = ?", (client_id,)).fetchone()
    values = (
        data.get("code") or "",
        data.get("name") or "",
        data.get("tax_address") or "",
        data.get("consignee_address") or "",
        data.get("delivery_method") or "",
        data.get("phone") or "",
        data.get("email") or "",
        data.get("contact_name") or "",
        data.get("external_seller") or "",
        data.get("assigned_user_id") or None,
        data.get("updated_at") or now_iso(),
        client_id,
    )
    if existing:
        db.execute(
            """
            UPDATE clients
            SET code = ?, name = ?, tax_address = ?, consignee_address = ?,
                delivery_method = ?, phone = ?, email = ?, contact_name = ?,
                external_seller = ?, assigned_user_id = ?, updated_at = ?
            WHERE id = ?
            """,
            values,
        )
    else:
        db.execute(
            """
            INSERT INTO clients
            (id, code, name, tax_address, consignee_address, delivery_method, phone, email,
             contact_name, external_seller, assigned_user_id, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                client_id,
                data.get("code") or "",
                data.get("name") or "",
                data.get("tax_address") or "",
                data.get("consignee_address") or "",
                data.get("delivery_method") or "",
                data.get("phone") or "",
                data.get("email") or "",
                data.get("contact_name") or "",
                data.get("external_seller") or "",
                data.get("assigned_user_id") or None,
                data.get("created_at") or now_iso(),
                data.get("updated_at") or now_iso(),
            ),
        )


def crm_save_client(db, data):
    client_id = data.get("id") or str(uuid.uuid4())
    now = now_iso()
    existing = None
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT created_at, contact_name, external_seller FROM crm_clients WHERE id = %s", (client_id,))
        existing = cur.fetchone()
        created_at = (existing or {}).get("created_at") or data.get("created_at") or now
        contact_name = data.get("contact_name")
        if contact_name is None:
            contact_name = (existing or {}).get("contact_name") or ""
        external_seller = data.get("external_seller")
        if external_seller is None:
            external_seller = (existing or {}).get("external_seller") or ""
        payload = {
            "id": client_id,
            "code": data.get("code") or "",
            "name": data.get("name") or "",
            "tax_address": data.get("tax_address") or "",
            "consignee_address": data.get("consignee_address") or "",
            "delivery_method": data.get("delivery_method") or "",
            "phone": data.get("phone") or "",
            "email": data.get("email") or "",
            "contact_name": contact_name or "",
            "external_seller": external_seller or "",
            "assigned_user_id": data.get("assigned_user_id") or None,
            "created_at": created_at,
            "updated_at": now,
        }
        cur.execute(
            """
            INSERT INTO crm_clients
            (id, code, name, tax_address, consignee_address, delivery_method, phone, email,
             contact_name, external_seller, assigned_user_id, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                code = VALUES(code),
                name = VALUES(name),
                tax_address = VALUES(tax_address),
                consignee_address = VALUES(consignee_address),
                delivery_method = VALUES(delivery_method),
                phone = VALUES(phone),
                email = VALUES(email),
                contact_name = VALUES(contact_name),
                external_seller = VALUES(external_seller),
                assigned_user_id = VALUES(assigned_user_id),
                updated_at = VALUES(updated_at)
            """,
            (
                payload["id"],
                payload["code"],
                payload["name"],
                payload["tax_address"],
                payload["consignee_address"],
                payload["delivery_method"],
                payload["phone"],
                payload["email"],
                payload["contact_name"],
                payload["external_seller"],
                payload["assigned_user_id"],
                payload["created_at"],
                payload["updated_at"],
            ),
        )
        conn.commit()
        cur.close()
        conn.close()
        mirror_client_to_sqlite(db, payload)
        return client_id
    except Exception as e:
        print(f"[CRM] Guardado de cliente en MySQL no disponible, usando SQLite: {e}", flush=True)
        data = dict(data)
        data["id"] = client_id
        data.setdefault("created_at", now)
        data["updated_at"] = now
        mirror_client_to_sqlite(db, data)
        return client_id


def crm_client_row(client_id):
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT * FROM crm_clients WHERE id = %s", (client_id,))
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row:
            data = normalize_mysql_row(row)
            data["assigned_user"] = data.get("external_seller") or crm_user_name(data.get("assigned_user_id"))
            return data
    except Exception as e:
        print(f"[CRM] No se pudo leer cliente CRM desde MySQL: {e}", flush=True)
    return None


def crm_search_clients(search=""):
    try:
        init_crm_mysql_tables()
        like = f"%{search or ''}%"
        conn = mysql_conn()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT *
            FROM crm_clients
            WHERE name LIKE %s
               OR code LIKE %s
               OR email LIKE %s
               OR external_seller LIKE %s
               OR phone LIKE %s
            ORDER BY updated_at DESC
            LIMIT 300
            """,
            (like, like, like, like, like),
        )
        rows = []
        for row in cur.fetchall():
            data = normalize_mysql_row(row)
            data["assigned_user"] = data.get("external_seller") or crm_user_name(data.get("assigned_user_id"))
            rows.append(data)
        cur.close()
        conn.close()
        return rows
    except Exception as e:
        print(f"[CRM] Busqueda CRM MySQL no disponible: {e}", flush=True)
        return []


def company_from_client_id(client_id):
    _, empresa = parse_external_client_id(client_id or "")
    return empresa or CFG.get("company_name") or "EZA2007"


def bank_account_for_company(db, company):
    row = db.execute(
        "SELECT * FROM company_bank_accounts WHERE UPPER(TRIM(company)) = UPPER(TRIM(?)) AND enabled = 1",
        (company or "",),
    ).fetchone()
    return row_to_dict(row) if row else None


def apply_local_client_overrides(db, clients):
    if not clients:
        return clients
    ids = [client["id"] for client in clients if client.get("id")]
    if not ids:
        return clients
    overrides = {}
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor(dictionary=True)
        placeholders = ",".join(["%s"] * len(ids))
        cur.execute(f"SELECT * FROM crm_clients WHERE id IN ({placeholders})", ids)
        for row in cur.fetchall():
            data = normalize_mysql_row(row)
            data["assigned_user"] = data.get("external_seller") or crm_user_name(data.get("assigned_user_id"))
            overrides[data["id"]] = data
        cur.close()
        conn.close()
    except Exception as e:
        print(f"[CRM] No se pudieron aplicar overrides desde MySQL: {e}", flush=True)
    if not overrides:
        placeholders = ",".join("?" for _ in ids)
        rows = db.execute(
            f"""
            SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), u.name) AS assigned_user
            FROM clients c
            LEFT JOIN users u ON u.id = c.assigned_user_id
            WHERE c.id IN ({placeholders})
            """,
            ids,
        ).fetchall()
        overrides = {row["id"]: row_to_dict(row) for row in rows}
    merged = []
    editable_fields = [
        "code",
        "name",
        "tax_address",
        "consignee_address",
        "delivery_method",
        "phone",
        "email",
        "contact_name",
        "assigned_user",
        "external_seller",
    ]
    for client in clients:
        local = overrides.get(client.get("id"))
        if local:
            updated = dict(client)
            for field in editable_fields:
                if local.get(field) not in (None, ""):
                    updated[field] = local[field]
            merged.append(updated)
        else:
            merged.append(client)
    return merged


def send_json(handler, payload, status=200):
    encoded = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(encoded)))
    handler.end_headers()
    handler.wfile.write(encoded)


def log_activity(db, client_id, user_id, entry_type, title, payload=None):
    activity_id = str(uuid.uuid4())
    payload_text = json.dumps(payload or {}, ensure_ascii=False)
    created_at = now_iso()
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO crm_activity_log (id, client_id, user_id, type, title, payload, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (activity_id, client_id, user_id, entry_type, title, payload_text, created_at),
        )
        conn.commit()
        cur.close()
        conn.close()
        return
    except Exception as e:
        print(f"[CRM] Bitacora MySQL no disponible, usando SQLite: {e}", flush=True)
    db.execute(
        """
        INSERT INTO activity_log (id, client_id, user_id, type, title, payload, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (activity_id, client_id, user_id, entry_type, title, payload_text, created_at),
    )


def crm_followups_for_client(db, client_id):
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT *
            FROM crm_followups
            WHERE client_id = %s
            ORDER BY contact_at DESC
            """,
            (client_id,),
        )
        rows = [normalize_mysql_row(row) for row in cur.fetchall()]
        cur.close()
        conn.close()
        quote_ids = [row.get("quote_id") for row in rows if row.get("quote_id")]
        quote_map = {}
        if quote_ids:
            placeholders = ",".join("?" for _ in quote_ids)
            for quote in db.execute(f"SELECT id, folio, quote_title FROM quotes WHERE id IN ({placeholders})", quote_ids).fetchall():
                quote_map[quote["id"]] = row_to_dict(quote)
        for row in rows:
            quote = quote_map.get(row.get("quote_id")) or {}
            row["quote_folio"] = quote.get("folio") or ""
            row["quote_title"] = quote.get("quote_title") or ""
            row["user_name"] = crm_user_name(row.get("user_id"))
        return rows
    except Exception as e:
        print(f"[CRM] Seguimientos MySQL no disponibles, usando SQLite: {e}", flush=True)
        return [
            row_to_dict(row)
            for row in db.execute(
                """
                SELECT f.*, u.name AS user_name, q.folio AS quote_folio, q.quote_title
                FROM followups f
                JOIN users u ON u.id = f.user_id
                LEFT JOIN quotes q ON q.id = f.quote_id
                WHERE f.client_id = ?
                ORDER BY f.contact_at DESC
                """,
                (client_id,),
            ).fetchall()
        ]


def crm_followups_done(db, user_id="", search=""):
    params = []
    where = []
    if user_id:
        where.append("f.user_id = %s")
        params.append(user_id)
    if search:
        like = f"%{search}%"
        where.append("(COALESCE(c.name, '') LIKE %s OR COALESCE(c.code, '') LIKE %s OR COALESCE(f.channel, '') LIKE %s OR COALESCE(f.outcome, '') LIKE %s OR COALESCE(f.notes, '') LIKE %s)")
        params.extend([like, like, like, like, like])
    sql_where = "WHERE " + " AND ".join(where) if where else ""
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            f"""
            SELECT f.*, c.name AS client_name, c.code AS client_code, c.phone AS client_phone
            FROM crm_followups f
            LEFT JOIN crm_clients c ON c.id = f.client_id
            {sql_where}
            ORDER BY f.contact_at DESC, f.created_at DESC
            LIMIT 300
            """,
            tuple(params),
        )
        rows = [normalize_mysql_row(row) for row in cur.fetchall()]
        cur.close()
        conn.close()
        quote_ids = [row.get("quote_id") for row in rows if row.get("quote_id")]
        quote_map = {}
        if quote_ids:
            placeholders = ",".join("?" for _ in quote_ids)
            for quote in db.execute(f"SELECT id, folio, quote_title FROM quotes WHERE id IN ({placeholders})", quote_ids).fetchall():
                quote_map[quote["id"]] = row_to_dict(quote)
        for row in rows:
            quote = quote_map.get(row.get("quote_id")) or {}
            row["quote_folio"] = quote.get("folio") or ""
            row["quote_title"] = quote.get("quote_title") or ""
            row["user_name"] = crm_user_name(row.get("user_id"))
        return rows
    except Exception as e:
        print(f"[CRM] Seguimientos realizados MySQL no disponibles, usando SQLite: {e}", flush=True)

    sqlite_params = []
    sqlite_where = []
    if user_id:
        sqlite_where.append("f.user_id = ?")
        sqlite_params.append(user_id)
    if search:
        like = f"%{search}%"
        sqlite_where.append("(COALESCE(c.name, '') LIKE ? OR COALESCE(c.code, '') LIKE ? OR COALESCE(f.channel, '') LIKE ? OR COALESCE(f.outcome, '') LIKE ? OR COALESCE(f.notes, '') LIKE ?)")
        sqlite_params.extend([like, like, like, like, like])
    sql_where = "WHERE " + " AND ".join(sqlite_where) if sqlite_where else ""
    rows = db.execute(
        f"""
        SELECT f.*, c.name AS client_name, c.code AS client_code, c.phone AS client_phone,
               u.name AS user_name, q.folio AS quote_folio, q.quote_title
        FROM followups f
        LEFT JOIN clients c ON c.id = f.client_id
        LEFT JOIN users u ON u.id = f.user_id
        LEFT JOIN quotes q ON q.id = f.quote_id
        {sql_where}
        ORDER BY f.contact_at DESC, f.created_at DESC
        LIMIT 300
        """,
        sqlite_params,
    ).fetchall()
    return [row_to_dict(row) for row in rows]


def crm_activity_for_client(db, client_id):
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT *
            FROM crm_activity_log
            WHERE client_id = %s
            ORDER BY created_at DESC
            LIMIT 50
            """,
            (client_id,),
        )
        rows = [normalize_mysql_row(row) for row in cur.fetchall()]
        cur.close()
        conn.close()
        for row in rows:
            row["user_name"] = crm_user_name(row.get("user_id"))
        return rows
    except Exception as e:
        print(f"[CRM] Bitacora MySQL no disponible, usando SQLite: {e}", flush=True)
        return [
            row_to_dict(row)
            for row in db.execute(
                """
                SELECT a.*, u.name AS user_name
                FROM activity_log a
                LEFT JOIN users u ON u.id = a.user_id
                WHERE a.client_id = ?
                ORDER BY a.created_at DESC
                LIMIT 50
                """,
                (client_id,),
            ).fetchall()
        ]


def crm_insert_followup(db, data):
    followup_id = str(uuid.uuid4())
    created_at = now_iso()
    try:
        init_crm_mysql_tables()
        conn = mysql_conn()
        cur = conn.cursor()
        cur.execute(
            """
            INSERT INTO crm_followups
            (id, client_id, user_id, contact_at, channel, outcome, next_action, next_action_at, notes, quote_id, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                followup_id,
                data["client_id"],
                data["user_id"],
                data.get("contact_at") or created_at,
                data.get("channel", "llamada"),
                data.get("outcome", ""),
                data.get("next_action", ""),
                data.get("next_action_at", ""),
                data.get("notes", ""),
                data.get("quote_id") or None,
                created_at,
            ),
        )
        conn.commit()
        cur.close()
        conn.close()
        return followup_id
    except Exception as e:
        print(f"[CRM] Guardado de seguimiento en MySQL no disponible, usando SQLite: {e}", flush=True)
    db.execute(
        """
        INSERT INTO followups
        (id, client_id, user_id, contact_at, channel, outcome, next_action, next_action_at, notes, quote_id, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            followup_id,
            data["client_id"],
            data["user_id"],
            data.get("contact_at") or created_at,
            data.get("channel", "llamada"),
            data.get("outcome", ""),
            data.get("next_action", ""),
            data.get("next_action_at", ""),
            data.get("notes", ""),
            data.get("quote_id") or None,
            created_at,
        ),
    )
    return followup_id


def calculate_quote(items):
    subtotal = 0
    discount = 0
    tax = 0
    prepared = []
    for item in items:
        quantity = float(item.get("quantity") or 0)
        unit_price = float(item.get("unit_price") or 0)
        discount_rate = float(item.get("discount_rate") or 0)
        tax_rate = float(item.get("tax_rate") if item.get("tax_rate") not in (None, "") else 16)
        gross = quantity * unit_price
        line_discount = gross * (discount_rate / 100)
        taxable = gross - line_discount
        line_tax = taxable * (tax_rate / 100)
        line_total = taxable + line_tax
        subtotal += gross
        discount += line_discount
        tax += line_tax
        prepared.append(
            {
                "id": str(uuid.uuid4()),
                "cip": item.get("cip", ""),
                "description": item.get("description", ""),
                "quantity": quantity,
                "unit_price": unit_price,
                "discount_rate": discount_rate,
                "tax_rate": tax_rate,
                "line_total": round(line_total, 2),
            }
        )
    total = subtotal - discount + tax
    return {
        "items": prepared,
        "subtotal": round(subtotal, 2),
        "discount": round(discount, 2),
        "tax": round(tax, 2),
        "total": round(total, 2),
        "authorized_shipping": round(total * 0.08, 2),
    }


def quote_payload(db, quote_id):
    quote = db.execute(
        """
        SELECT q.*, c.name AS client_name, c.code AS client_code, c.phone, c.email,
               c.tax_address, c.consignee_address, c.delivery_method, u.name AS user_name
        FROM quotes q
        JOIN clients c ON c.id = q.client_id
        JOIN users u ON u.id = q.user_id
        WHERE q.id = ?
        """,
        (quote_id,),
    ).fetchone()
    if not quote:
        return None
    items = db.execute("SELECT * FROM quote_items WHERE quote_id = ?", (quote_id,)).fetchall()
    data = row_to_dict(quote)
    data["items"] = [row_to_dict(row) for row in items]
    data["company"] = company_from_client_id(data.get("client_id"))
    data["bank_account"] = bank_account_for_company(db, data["company"])
    if data.get("client_snapshot"):
        try:
            snapshot = json.loads(data["client_snapshot"])
            for key, value in snapshot.items():
                if key == "user_name":
                    continue
                data[key] = value
        except Exception:
            pass
    return data


def quote_client_snapshot_if_needed(db, client_id, quote_recipient=None):
    numero, empresa = parse_external_client_id(client_id or "")
    if numero != "100000":
        return None
    client = db.execute(
        """
        SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), u.name) AS assigned_user
        FROM clients c LEFT JOIN users u ON u.id = c.assigned_user_id
        WHERE c.id = ?
        """,
        (client_id,),
    ).fetchone()
    if not client:
        return None
    client = row_to_dict(client)
    display_name = (quote_recipient or "").strip() or client.get("name") or ""
    snapshot = {
        "client_name": display_name,
        "client_code": client.get("code") or "",
        "phone": client.get("phone") or "",
        "email": client.get("email") or "",
        "tax_address": client.get("tax_address") or "",
        "consignee_address": client.get("consignee_address") or "",
        "delivery_method": client.get("delivery_method") or "",
        "contact_name": client.get("contact_name") or "",
        "company": empresa or CFG.get("company_name") or "",
    }
    account = bank_account_for_company(db, snapshot["company"])
    if account:
        snapshot["bank_account"] = account
    return json.dumps(snapshot, ensure_ascii=False)


def quote_title_for_client(client_id, folio, quote_recipient=None):
    numero, _ = parse_external_client_id(client_id or "")
    if numero != "100000":
        return folio
    name = (quote_recipient or "").strip() or "PUBLICO EN GENERAL"
    safe_name = " ".join(name.split())
    return f"{safe_name} - {folio}"


def export_quote_xlsx(data):
    os.makedirs(EXPORT_DIR, exist_ok=True)
    template = CFG.get("quote_template") or QUOTE_TEMPLATE
    wb = load_workbook(template)
    ws = wb["Cotización"] if "Cotización" in wb.sheetnames else wb.active
    ws["E2"] = data["folio"]
    ws["E3"] = datetime.now().date()
    ws["E4"] = data.get("valid_until") or "15 días"
    ws["E5"] = data.get("user_name") or ""
    ws["B9"] = data.get("client_name") or ""
    ws["B10"] = CFG.get("company_name") or ""
    ws["B11"] = data.get("phone") or ""
    ws["B12"] = data.get("email") or ""
    ws["B13"] = quote_delivery_address(data)
    for row in range(16, 43):
        for col in "ABCDE":
            ws[f"{col}{row}"] = None
    for idx, item in enumerate(data["items"][:27], start=16):
        ws[f"A{idx}"] = item.get("cip") or ""
        ws[f"B{idx}"] = item.get("description") or ""
        ws[f"C{idx}"] = item.get("quantity") or 0
        ws[f"D{idx}"] = item.get("unit_price") or 0
        ws[f"E{idx}"] = item.get("line_total") or 0
    ws["E44"] = data.get("total") or 0
    ws["E46"] = f"Envio autorizado 8%: ${float(data.get('authorized_shipping') or 0):,.2f}"
    out = os.path.join(EXPORT_DIR, f"{safe_filename(data.get('quote_title') or data['folio'])}.xlsx")
    wb.save(out)
    return out


def export_quote_pdf(data):
    os.makedirs(EXPORT_DIR, exist_ok=True)
    pdf_path = os.path.join(EXPORT_DIR, f"{safe_filename(data.get('quote_title') or data['folio'])}.pdf")
    build_quote_pdf_like_reference(data, pdf_path)
    return pdf_path


def safe_filename(value):
    text = str(value or "archivo").strip()
    for ch in '<>:"/\\|?*':
        text = text.replace(ch, "-")
    return text[:140] or "archivo"


def quote_delivery_address(data):
    for key in ("consignee_address", "delivery_method", "tax_address"):
        value = " ".join(str(data.get(key) or "").split())
        marker = value.replace(" ", "").replace("-", "")
        if value and marker and set(marker) != {"0"}:
            return value
    return ""


def draw_wrapped_pdf_text(canvas, text, x, y, max_chars, max_lines, line_height=8):
    clean = " ".join(str(text or "").split())
    lines = textwrap.wrap(clean, width=max_chars)[:max_lines]
    for idx, line in enumerate(lines):
        canvas.drawString(x, y - (idx * line_height), line)


def logo_pdf_size(path, target_height=96, max_width=300):
    if not PILImage:
        return target_height, target_height
    try:
        img = PILImage.open(path).convert("RGBA")
        bbox = img.getchannel("A").getbbox() or (0, 0, img.width, img.height)
        w = max(1, bbox[2] - bbox[0])
        h = max(1, bbox[3] - bbox[1])
        width = target_height * (w / h)
        if width > max_width:
            return max_width, max_width * (h / w)
        return width, target_height
    except Exception:
        return target_height, target_height


def build_quote_pdf_like_reference(data, pdf_path):
    c = pdfcanvas.Canvas(pdf_path, pagesize=letter)
    width, height = letter
    left = 55
    right = width - 55
    top = height - 70
    logo_path = quote_logo_for_company(data.get("company") or data.get("empresa") or CFG.get("company_name"))

    if os.path.exists(logo_path):
        logo_w, logo_h = logo_pdf_size(logo_path)
        c.drawImage(logo_path, left, top - 86, width=logo_w, height=logo_h, preserveAspectRatio=False, mask="auto")

    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(440, top - 5, "COTIZACIÓN")
    c.setFont("Helvetica", 7)
    meta_x = 365
    value_x = 425
    y = top - 20
    today_text = datetime.now().strftime("%d/%m/%Y")
    c.drawString(meta_x, y, "No:")
    c.drawString(value_x, y, str(data.get("folio") or ""))
    y -= 12
    c.drawString(meta_x, y, "Fecha:")
    c.drawString(value_x, y, today_text)
    y -= 12
    c.drawString(meta_x, y, "Vigencia:")
    c.drawString(value_x, y, "15 días")
    y -= 12
    c.drawString(meta_x, y, "Usuario:")
    c.drawString(value_x, y, str(data.get("user_name") or ""))

    client_y = top - 100
    c.setFont("Helvetica-Bold", 7)
    c.drawString(left, client_y, "DATOS DEL CLIENTE")
    c.setFont("Helvetica", 7)
    fields = [
        ("Cliente:", data.get("client_name") or ""),
        ("Empresa:", data.get("company") or CFG.get("company_name") or ""),
        ("Teléfono:", data.get("phone") or ""),
        ("Email:", data.get("email") or ""),
        ("Dirección:", quote_delivery_address(data)),
    ]
    y = client_y - 12
    for label, value in fields:
        c.drawString(left, y, label)
        if label == "Teléfono:":
            c.drawString(left + 50, y, str(value))
        elif label == "Email:":
            c.setFillColor(colors.blue)
            c.drawString(left + 50, y, str(value))
            c.setFillColor(colors.black)
        elif "Direcci" in label:
            lines = textwrap.wrap(" ".join(str(value or "").split()), width=96)[:3]
            if not lines:
                lines = [""]
            for idx, line in enumerate(lines):
                c.drawString(left + 50, y - (idx * 8), line)
            y -= 8 * (len(lines) - 1)
        else:
            c.drawString(left + 50, y, str(value)[:75])
        y -= 11

    table_top = min(client_y - 70, y - 10)
    row_h = 8.35
    col_x = [left, left + 50, left + 275, left + 315, left + 375, right]
    headers = ["Código", "Producto", "Cantidad", "Precio Unitario", "Importe"]
    c.setFillColor(colors.black)
    c.rect(left, table_top - row_h, right - left, row_h, fill=1, stroke=1)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 6.5)
    for i, h in enumerate(headers):
        c.drawCentredString((col_x[i] + col_x[i + 1]) / 2, table_top - 6.3, h)
    c.setFillColor(colors.black)
    c.setFont("Helvetica", 6)
    for x in col_x:
        c.line(x, table_top, x, table_top - row_h * 28)
    for r in range(29):
        yy = table_top - row_h * r
        c.line(left, yy, right, yy)

    for idx in range(27):
        row_y = table_top - row_h * (idx + 1) - 6.1
        item = data["items"][idx] if idx < len(data["items"]) else None
        if not item:
            continue
        c.drawCentredString((col_x[0] + col_x[1]) / 2, row_y, str(item.get("cip") or ""))
        c.drawString(col_x[1] + 2, row_y, str(item.get("description") or "")[:54])
        c.drawCentredString((col_x[2] + col_x[3]) / 2, row_y, f"{float(item.get('quantity') or 0):g}")
        c.drawCentredString((col_x[3] + col_x[4]) / 2, row_y, f"${float(item.get('unit_price') or 0):,.2f}")
        c.drawCentredString((col_x[4] + col_x[5]) / 2, row_y, f"${float(item.get('line_total') or 0):,.2f}")

    total_y = table_top - row_h * 29 - 8
    c.rect(col_x[3], total_y - 9, col_x[4] - col_x[3], 9, fill=0, stroke=1)
    c.rect(col_x[4], total_y - 9, col_x[5] - col_x[4], 9, fill=0, stroke=1)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawCentredString((col_x[3] + col_x[4]) / 2, total_y - 6.5, "Total:")
    c.drawCentredString((col_x[4] + col_x[5]) / 2, total_y - 6.5, f"${float(data.get('total') or 0):,.2f}")

    observations_top = total_y - 18
    observations_height = 34
    observations_bottom = observations_top - observations_height
    c.rect(left, observations_bottom, right - left, observations_height, fill=0, stroke=1)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawString(left + 4, observations_top - 8, "Observaciones:")
    c.setFont("Helvetica", 6)
    draw_wrapped_pdf_text(c, data.get("notes") or "", left + 58, observations_top - 8, 112, 3, 8)

    notes_top = observations_bottom - 10
    notes_left = left
    notes_right = col_x[3]
    costs_left = col_x[3]
    costs_mid = col_x[4]
    costs_right = right
    notes_bottom = notes_top - 78
    c.rect(notes_left, notes_bottom, notes_right - notes_left, 78, fill=0, stroke=1)
    c.rect(costs_left, notes_bottom, costs_right - costs_left, 78, fill=0, stroke=1)
    c.line(costs_mid, notes_bottom, costs_mid, notes_top)
    c.line(costs_left, notes_top - 39, costs_right, notes_top - 39)
    c.setFont("Helvetica", 6)
    c.drawCentredString((notes_left + notes_right) / 2, notes_top - 8, "Notas:")
    c.drawString(notes_left + 2, notes_top - 18, "• Precios sujetos a cambio sin previo aviso")
    c.drawString(notes_left + 10, notes_top - 38, "• Tiempo de entrega: 2 a 3 dias segun paqueteria")
    bank = data.get("bank_account")
    if bank:
        c.drawString(notes_left + 2, notes_bottom + 20, "• Forma de pago:")
        bank_line = f"Transferencia a nombre de {bank.get('beneficiary') or ''} Banco {bank.get('bank') or ''} Cuenta: {bank.get('account') or ''}"
        c.drawCentredString((notes_left + notes_right) / 2, notes_bottom + 20, bank_line[:112])
        if bank.get("clabe"):
            c.drawCentredString((notes_left + notes_right) / 2, notes_bottom + 10, f"Cuenta Clabe: {bank.get('clabe')}")
    c.drawCentredString((costs_left + costs_mid) / 2, notes_top - 20, "Costos Adicionales")
    c.drawCentredString((costs_mid + costs_right) / 2, notes_top - 14, "8% sobre el total para rebanados, solo en")
    c.drawCentredString((costs_mid + costs_right) / 2, notes_top - 23, "jamones. Gastos de envio a cargo del")
    c.drawCentredString((costs_mid + costs_right) / 2, notes_top - 32, "comprador")
    c.setFont("Helvetica-Bold", 6.2)
    c.drawCentredString((costs_mid + costs_right) / 2, notes_bottom + 25, "Costo de envio autorizado")
    c.drawCentredString((costs_mid + costs_right) / 2, notes_bottom + 15, f"${float(data.get('authorized_shipping') or 0):,.2f}")
    c.save()


def quote_logo_for_company(company):
    comp = (company or "").lower()
    logo_dir = crm_logo_dir()
    if "gourmet" in comp:
        return os.path.join(logo_dir, "gourmet.png")
    if "ibersur" in comp:
        return os.path.join(logo_dir, "ibersur.png")
    if "remision" in comp or "remisión" in comp:
        return os.path.join(logo_dir, "Remision.png")
    if "eza" in comp:
        return os.path.join(logo_dir, "eza2007.png")
    return os.path.join(PUBLIC_DIR, "assets", "image1.png")


def crm_logo_dir():
    bundled = os.path.abspath(os.path.join(ROOT, "..", "logos"))
    if os.path.isdir(bundled):
        return bundled
    legacy = r"E:\Proyectos\Proyecto facturacion\AspelAPI\logos"
    if os.path.isdir(legacy):
        return legacy
    return os.path.join(PUBLIC_DIR, "assets")


class SalesHandler(SimpleHTTPRequestHandler):
    def translate_path(self, path):
        parsed = urlparse(path)
        rel_path = parsed.path.lstrip("/") or "index.html"
        return os.path.join(PUBLIC_DIR, rel_path)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_get(parsed)
        if parsed.path.startswith("/exports/"):
            return self.serve_export(parsed.path)
        target = self.translate_path(self.path)
        if os.path.isdir(target):
            target = os.path.join(target, "index.html")
        if not os.path.exists(target):
            target = os.path.join(PUBLIC_DIR, "index.html")
        mime = mimetypes.guess_type(target)[0] or "application/octet-stream"
        with open(target, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def serve_export(self, path):
        name = os.path.basename(unquote(path))
        target = os.path.join(EXPORT_DIR, name)
        if not os.path.exists(target):
            return send_json(self, {"error": "Archivo no encontrado"}, 404)
        mime = mimetypes.guess_type(target)[0] or "application/octet-stream"
        with open(target, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Content-Disposition", f'inline; filename="{name}"')
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_post(parsed)
        send_json(self, {"error": "Ruta no encontrada"}, 404)

    def do_PUT(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_put(parsed)
        send_json(self, {"error": "Ruta no encontrada"}, 404)

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_delete(parsed)
        send_json(self, {"error": "Ruta no encontrada"}, 404)

    def handle_api_get(self, parsed):
        query = parse_qs(parsed.query)
        with connect() as db:
            if parsed.path == "/api/users":
                try:
                    return send_json(self, external_users())
                except Exception:
                    rows = db.execute("SELECT * FROM users WHERE active = 1 ORDER BY role, name").fetchall()
                    return send_json(self, [row_to_dict(row) for row in rows])
            if parsed.path == "/api/integration/status":
                try:
                    return send_json(self, {"mysql": external_available(), "api_url": CFG["facturacion_api_url"], "mode": "facturacion"})
                except Exception as e:
                    return send_json(self, {"mysql": False, "api_url": CFG["facturacion_api_url"], "error": str(e), "mode": "local"})
            if parsed.path == "/api/settings/bank-accounts":
                rows = db.execute("SELECT * FROM company_bank_accounts ORDER BY company").fetchall()
                return send_json(self, [row_to_dict(row) for row in rows])
            if parsed.path == "/api/prospector/zones":
                return send_json(self, list_zones(db))
            if parsed.path == "/api/prospector/prospects":
                search = query.get("q", [""])[0]
                status = query.get("status", [""])[0]
                zone_name = query.get("zone", [""])[0]
                return send_json(self, list_prospects(db, search, status, zone_name))
            if parsed.path == "/api/prospector/prospect":
                prospect_id = query.get("id", [""])[0]
                data = prospect_payload(db, prospect_id)
                if not data:
                    return send_json(self, {"error": "Prospecto no encontrado"}, 404)
                return send_json(self, data)
            if parsed.path == "/api/prospector/quote":
                quote_id = query.get("id", [""])[0]
                data = prospect_quote_payload(db, quote_id)
                if not data:
                    return send_json(self, {"error": "Cotizacion de prospecto no encontrada"}, 404)
                return send_json(self, data)
            if parsed.path == "/api/prospector/quote/export":
                quote_id = query.get("id", [""])[0]
                fmt = query.get("format", ["pdf"])[0].lower()
                data = prospect_quote_payload(db, quote_id)
                if not data:
                    return send_json(self, {"error": "Cotizacion de prospecto no encontrada"}, 404)
                try:
                    path = export_quote_xlsx(data) if fmt == "xlsx" else export_quote_pdf(data)
                    return send_json(self, {"url": f"/exports/{os.path.basename(path)}", "path": path})
                except Exception as e:
                    return send_json(self, {"error": f"No se pudo exportar: {e}"}, 500)
            if parsed.path == "/api/clients":
                search = f"%{query.get('q', [''])[0]}%"
                raw_search = query.get("q", [""])[0]
                try:
                    external = apply_local_client_overrides(db, external_clients(raw_search))
                    seen = {client.get("id") for client in external}
                    server_clients = [client for client in crm_search_clients(raw_search) if client.get("id") not in seen]
                    return send_json(self, external + server_clients)
                except Exception:
                    pass
                server_clients = crm_search_clients(raw_search)
                if server_clients:
                    return send_json(self, server_clients)
                rows = db.execute(
                    """
                    SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), u.name) AS assigned_user
                    FROM clients c
                    LEFT JOIN users u ON u.id = c.assigned_user_id
                    WHERE c.name LIKE ?
                       OR c.code LIKE ?
                       OR c.email LIKE ?
                       OR c.external_seller LIKE ?
                       OR u.name LIKE ?
                    ORDER BY c.updated_at DESC
                    """,
                    (search, search, search, search, search),
                ).fetchall()
                return send_json(self, [row_to_dict(row) for row in rows])
            if parsed.path == "/api/followups-done":
                user_id = query.get("user_id", [""])[0]
                search = query.get("q", [""])[0]
                return send_json(self, crm_followups_done(db, user_id, search))
            if parsed.path == "/api/client":
                client_id = query.get("id", [""])[0]
                light = query.get("light", ["0"])[0] in ("1", "true", "si", "sí")
                if client_id.startswith("ext:"):
                    try:
                        ensure_local_client(db, client_id)
                    except Exception:
                        pass
                client = crm_client_row(client_id)
                if not client:
                    client = row_to_dict(
                        db.execute(
                            """
                            SELECT c.*, COALESCE(NULLIF(c.external_seller, ''), u.name) AS assigned_user
                            FROM clients c LEFT JOIN users u ON u.id = c.assigned_user_id
                            WHERE c.id = ?
                            """,
                            (client_id,),
                        ).fetchone()
                    )
                if not client:
                    return send_json(self, {"error": "Cliente no encontrado"}, 404)
                invoices = db.execute(
                    """
                    SELECT i.*, u.name AS user_name
                    FROM invoices i
                    LEFT JOIN users u ON u.id = i.user_id
                    WHERE i.client_id = ?
                    ORDER BY i.issued_at DESC
                    """,
                    (client_id,),
                ).fetchall()
                external_facturas = []
                products = []
                if not light:
                    try:
                        external_facturas = external_invoices(client_id)
                    except Exception:
                        external_facturas = []
                    try:
                        products = external_client_products(client_id)
                    except Exception:
                        products = []
                quotes = db.execute("SELECT q.*, u.name AS user_name FROM quotes q JOIN users u ON u.id = q.user_id WHERE client_id = ? ORDER BY created_at DESC", (client_id,)).fetchall()
                followups = crm_followups_for_client(db, client_id)
                activity = crm_activity_for_client(db, client_id)
                return send_json(
                    self,
                    {
                        "client": client,
                        "invoices": external_facturas + [row_to_dict(row) for row in invoices],
                        "products": products,
                        "quotes": [row_to_dict(row) for row in quotes],
                        "followups": followups,
                        "activity": activity,
                        "details_pending": light,
                    },
                )
            if parsed.path == "/api/products":
                search = query.get("q", [""])[0]
                try:
                    return send_json(self, external_products(search))
                except Exception as e:
                    return send_json(self, {"error": f"No se pudieron consultar productos: {e}"}, 503)
            if parsed.path == "/api/price":
                client_id = query.get("client_id", [""])[0]
                cip = query.get("cip", [""])[0]
                try:
                    price = price_for_client(client_id, cip)
                    if not price:
                        return send_json(self, {"error": "Precio no encontrado"}, 404)
                    return send_json(self, price)
                except Exception as e:
                    return send_json(self, {"error": f"No se pudo calcular precio: {e}"}, 503)
            if parsed.path == "/api/product-info":
                cip = query.get("cip", [""])[0]
                try:
                    info = product_quote_info(cip)
                    if not info:
                        return send_json(self, {"error": "Producto no encontrado"}, 404)
                    return send_json(self, info)
                except Exception as e:
                    return send_json(self, {"error": f"No se pudo consultar producto: {e}"}, 503)
            if parsed.path == "/api/invoice":
                invoice_id = query.get("id", [""])[0]
                if invoice_id.startswith("extfact:") or invoice_id.startswith("extfactid:"):
                    try:
                        invoice = external_invoice_detail(invoice_id)
                        if invoice:
                            return send_json(self, invoice)
                    except Exception as e:
                        return send_json(self, {"error": f"No se pudo consultar factura: {e}"}, 503)
                invoice = db.execute("SELECT * FROM invoices WHERE id = ?", (invoice_id,)).fetchone()
                if not invoice:
                    return send_json(self, {"error": "Factura no encontrada"}, 404)
                return send_json(self, row_to_dict(invoice))
            if parsed.path == "/api/invoice/export":
                invoice_id = query.get("id", [""])[0]
                if not (invoice_id.startswith("extfact:") or invoice_id.startswith("extfactid:")):
                    return send_json(self, {"error": "La reconstruccion PDF aplica para facturas de facturacion"}, 400)
                try:
                    path = export_invoice_pdf(invoice_id)
                    return send_json(self, {"url": f"/exports/{os.path.basename(path)}", "path": path})
                except Exception as e:
                    return send_json(self, {"error": f"No se pudo reconstruir factura: {e}"}, 500)
            if parsed.path == "/api/quote":
                quote_id = query.get("id", [""])[0]
                data = quote_payload(db, quote_id)
                if not data:
                    return send_json(self, {"error": "Cotizacion no encontrada"}, 404)
                return send_json(self, data)
            if parsed.path == "/api/quote/export":
                quote_id = query.get("id", [""])[0]
                fmt = query.get("format", ["pdf"])[0].lower()
                data = quote_payload(db, quote_id)
                if not data:
                    return send_json(self, {"error": "Cotizacion no encontrada"}, 404)
                try:
                    path = export_quote_xlsx(data) if fmt == "xlsx" else export_quote_pdf(data)
                    return send_json(self, {"url": f"/exports/{os.path.basename(path)}", "path": path})
                except Exception as e:
                    return send_json(self, {"error": f"No se pudo exportar: {e}"}, 500)
        return send_json(self, {"error": "Ruta no encontrada"}, 404)

    def handle_api_post(self, parsed):
        with connect() as db:
            if parsed.path == "/api/auth/login":
                data = read_json(self)
                try:
                    user = verify_external_user(data.get("username", ""), data.get("password", ""))
                    if user:
                        return send_json(self, user)
                except Exception:
                    pass
                row = db.execute("SELECT * FROM users WHERE name = ? AND active = 1", (data.get("username", ""),)).fetchone()
                if row and data.get("password") in ("demo", "admin"):
                    return send_json(self, {"id": row["id"], "name": row["name"], "role": row["role"]})
                return send_json(self, {"error": "Usuario o contraseña incorrectos"}, 401)
            if parsed.path == "/api/settings/bank-accounts":
                data = read_json(self)
                company = (data.get("company") or "").strip()
                if not company:
                    return send_json(self, {"error": "Empresa requerida"}, 400)
                db.execute(
                    """
                    INSERT INTO company_bank_accounts
                    (company, beneficiary, bank, account, clabe, enabled, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(company) DO UPDATE SET
                        beneficiary = excluded.beneficiary,
                        bank = excluded.bank,
                        account = excluded.account,
                        clabe = excluded.clabe,
                        enabled = excluded.enabled,
                        updated_at = excluded.updated_at
                    """,
                    (
                        company,
                        data.get("beneficiary", ""),
                        data.get("bank", ""),
                        data.get("account", ""),
                        data.get("clabe", ""),
                        1 if data.get("enabled", True) else 0,
                        now_iso(),
                    ),
                )
                return send_json(self, {"ok": True})
            if parsed.path == "/api/prospector/search":
                data = read_json(self)
                user_id = data.get("user_id") or None
                if user_id:
                    ensure_local_user(db, user_id)
                query_text = (data.get("query") or "").strip()
                limit_count = int(data.get("limit") or 20)
                zone_name = (data.get("zone_name") or "").strip()
                location_restriction = None
                if zone_name:
                    zone = get_zone(db, zone_name)
                    if not zone:
                        return send_json(self, {"error": "Zona no encontrada"}, 404)
                    location_restriction = rectangle_restriction(zone)
                try:
                    places = search_google_places(query_text, limit_count, location_restriction)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                search_id = str(uuid.uuid4())
                db.execute(
                    """
                    INSERT INTO prospect_searches (id, query, zone_name, limit_count, results_count, user_id, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (search_id, query_text, zone_name or None, limit_count, len(places), user_id, now_iso()),
                )
                ids = [upsert_prospect(db, place, query_text, user_id, zone_name or None) for place in places]
                rows = db.execute(
                    f"""
                    SELECT p.*, u.name AS assigned_user, c.name AS client_name
                    FROM prospects p
                    LEFT JOIN users u ON u.id = p.assigned_user_id
                    LEFT JOIN clients c ON c.id = p.client_id
                    WHERE p.id IN ({",".join("?" for _ in ids)})
                    ORDER BY p.rating DESC, p.total_reviews DESC
                    """,
                    ids,
                ).fetchall() if ids else []
                return send_json(self, {"search_id": search_id, "saved": len(ids), "prospects": [row_to_dict(row) for row in rows]})
            if parsed.path == "/api/prospector/scan-zones":
                data = read_json(self)
                user_id = data.get("user_id") or None
                if user_id:
                    ensure_local_user(db, user_id)
                query_text = (data.get("query") or "").strip()
                limit_count = int(data.get("limit") or 20)
                rescan = bool(data.get("rescan"))
                zone_names = data.get("zones") or []
                if not query_text:
                    return send_json(self, {"error": "Escribe una busqueda para escanear zonas."}, 400)
                if not zone_names:
                    return send_json(self, {"error": "Selecciona al menos una zona."}, 400)
                all_ids = []
                skipped = []
                scanned = []
                for zone_name in zone_names:
                    zone = get_zone(db, zone_name)
                    if not zone:
                        skipped.append({"zone": zone_name, "reason": "Zona no encontrada"})
                        continue
                    if not rescan and scan_already_done(db, query_text, zone_name):
                        skipped.append({"zone": zone_name, "reason": "Ya escaneada"})
                        continue
                    try:
                        places = search_google_places(query_text, limit_count, rectangle_restriction(zone))
                    except Exception as e:
                        skipped.append({"zone": zone_name, "reason": str(e)})
                        continue
                    search_id = str(uuid.uuid4())
                    db.execute(
                        """
                        INSERT INTO prospect_searches (id, query, zone_name, limit_count, results_count, user_id, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        """,
                        (search_id, query_text, zone_name, limit_count, len(places), user_id, now_iso()),
                    )
                    ids = [upsert_prospect(db, place, query_text, user_id, zone_name) for place in places]
                    all_ids.extend(ids)
                    scanned.append({"zone": zone_name, "saved": len(ids)})
                rows = db.execute(
                    f"""
                    SELECT p.*, u.name AS assigned_user, c.name AS client_name
                    FROM prospects p
                    LEFT JOIN users u ON u.id = p.assigned_user_id
                    LEFT JOIN clients c ON c.id = p.client_id
                    WHERE p.id IN ({",".join("?" for _ in all_ids)})
                    ORDER BY p.updated_at DESC
                    """,
                    all_ids,
                ).fetchall() if all_ids else []
                return send_json(self, {"saved": len(all_ids), "scanned": scanned, "skipped": skipped, "prospects": [row_to_dict(row) for row in rows]})
            if parsed.path == "/api/prospector/zones":
                data = read_json(self)
                zones = data.get("zones") or []
                names = [upsert_zone(db, zone) for zone in zones]
                return send_json(self, {"imported": len([name for name in names if name])})
            if parsed.path == "/api/prospector/prospects":
                data = read_json(self)
                user_id = data.get("user_id") or None
                if user_id:
                    ensure_local_user(db, user_id)
                prospect = {
                    "id": data.get("id"),
                    "google_place_id": data.get("google_place_id", ""),
                    "name": data.get("name", ""),
                    "category": data.get("category", ""),
                    "phone": data.get("phone", ""),
                    "website": data.get("website", ""),
                    "address": data.get("address", ""),
                    "latitude": data.get("latitude"),
                    "longitude": data.get("longitude"),
                    "rating": data.get("rating"),
                    "total_reviews": data.get("total_reviews"),
                    "business_status": data.get("business_status", ""),
                }
                prospect_id = upsert_prospect(db, prospect, data.get("source_query", "manual"), user_id, data.get("zone_name"))
                status = data.get("status")
                notes = data.get("notes")
                if status or notes is not None:
                    db.execute(
                        "UPDATE prospects SET status = COALESCE(?, status), notes = COALESCE(?, notes), updated_at = ? WHERE id = ?",
                        (status, notes, now_iso(), prospect_id),
                    )
                return send_json(self, {"id": prospect_id})
            if parsed.path == "/api/prospector/convert":
                data = read_json(self)
                user_id = data.get("user_id") or None
                try:
                    client_id = convert_prospect_to_client(db, data.get("prospect_id", ""), user_id)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, {"client_id": client_id})
            if parsed.path == "/api/prospector/start-followup":
                data = read_json(self)
                try:
                    prospect_id = start_prospect_followup(db, data.get("prospect_id", ""), data.get("user_id") or None)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, {"id": prospect_id})
            if parsed.path == "/api/prospector/link-client":
                data = read_json(self)
                try:
                    client_id = link_prospect_to_client(
                        db,
                        data.get("prospect_id", ""),
                        data.get("client_code", ""),
                        data.get("user_id") or None,
                    )
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, {"client_id": client_id})
            if parsed.path == "/api/prospector/quotes":
                data = read_json(self)
                try:
                    result = add_prospect_quote(db, data)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, result)
            if parsed.path == "/api/prospector/followups":
                data = read_json(self)
                try:
                    followup_id = add_prospect_followup(db, data)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, {"id": followup_id})
            if parsed.path == "/api/prospector/messages":
                data = read_json(self)
                try:
                    message_id = add_prospect_message(db, data)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, {"id": message_id})
            if parsed.path == "/api/prospector/phones":
                data = read_json(self)
                try:
                    phone_id = add_prospect_phone(db, data)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, {"id": phone_id})
            if parsed.path == "/api/prospector/check-clients":
                data = read_json(self)
                result = check_prospect_clients(
                    db,
                    data.get("q", ""),
                    data.get("status", "todos"),
                    data.get("zone", "todas"),
                    bool(data.get("include_external", True)),
                )
                return send_json(self, result)
            if parsed.path == "/api/users":
                data = read_json(self)
                user_id = data.get("id") or str(uuid.uuid4())
                db.execute(
                    "INSERT OR REPLACE INTO users (id, name, email, role, active, created_at) VALUES (?, ?, ?, ?, 1, COALESCE((SELECT created_at FROM users WHERE id = ?), ?))",
                    (user_id, data["name"], data.get("email", ""), data.get("role", "vendedor"), user_id, now_iso()),
                )
                return send_json(self, {"id": user_id})
            if parsed.path == "/api/clients":
                data = read_json(self)
                client_id = data.get("id") or str(uuid.uuid4())
                data["id"] = client_id
                if data.get("assigned_user_id"):
                    ensure_local_user(db, data.get("assigned_user_id"))
                crm_save_client(db, data)
                log_activity(db, client_id, data.get("assigned_user_id"), "cliente", f"Cliente actualizado: {data['name']}", data)
                return send_json(self, {"id": client_id})
            if parsed.path == "/api/import/clients":
                data = read_json(self)
                content = data.get("csv", "")
                reader = csv.DictReader(io.StringIO(content))
                imported = 0
                for row in reader:
                    if not row.get("name"):
                        continue
                    client_id = str(uuid.uuid4())
                    db.execute(
                        """
                        INSERT INTO clients
                        (id, code, name, tax_address, consignee_address, delivery_method, phone, email, contact_name, external_seller, assigned_user_id, created_at, updated_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, '', '', ?, ?, ?)
                        """,
                        (
                            client_id,
                            row.get("code", ""),
                            row["name"],
                            row.get("tax_address", ""),
                            row.get("consignee_address", ""),
                            row.get("delivery_method", ""),
                            row.get("phone", ""),
                            row.get("email", ""),
                            data.get("assigned_user_id") or None,
                            now_iso(),
                            now_iso(),
                        ),
                    )
                    imported += 1
                return send_json(self, {"imported": imported})
            if parsed.path == "/api/invoices":
                data = read_json(self)
                invoice_id = data.get("id") or str(uuid.uuid4())
                if data.get("user_id"):
                    ensure_local_user(db, data.get("user_id"))
                db.execute(
                    """
                    INSERT OR REPLACE INTO invoices
                    (id, client_id, user_id, folio, issued_at, subtotal, discount, tax, total, status, detail, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, COALESCE((SELECT created_at FROM invoices WHERE id = ?), ?))
                    """,
                    (
                        invoice_id,
                        data["client_id"],
                        data.get("user_id") or None,
                        data["folio"],
                        data.get("issued_at") or now_iso()[:10],
                        float(data.get("subtotal") or 0),
                        float(data.get("discount") or 0),
                        float(data.get("tax") or 0),
                        float(data.get("total") or 0),
                        data.get("status", "facturada"),
                        data.get("detail", ""),
                        invoice_id,
                        now_iso(),
                    ),
                )
                log_activity(db, data["client_id"], data.get("user_id"), "factura", f"Factura registrada: {data['folio']}", data)
                return send_json(self, {"id": invoice_id})
            if parsed.path == "/api/quotes":
                data = read_json(self)
                ensure_local_client(db, data["client_id"])
                ensure_local_user(db, data["user_id"])
                quote_id = str(uuid.uuid4())
                items = data.get("items", [])
                totals = calculate_quote(items)
                folio = f"COT-{datetime.now().strftime('%Y%m%d')}-{db.execute('SELECT COUNT(*) + 1 AS n FROM quotes').fetchone()['n']:04d}"
                quote_recipient = (data.get("quote_recipient") or "").strip()
                numero, _ = parse_external_client_id(data["client_id"])
                if numero == "100000" and not quote_recipient:
                    return send_json(self, {"error": "Captura empresa/persona cotizada para cliente 100000"}, 400)
                quote_title = quote_title_for_client(data["client_id"], folio, quote_recipient)
                client_snapshot = quote_client_snapshot_if_needed(db, data["client_id"], quote_recipient)
                db.execute(
                    """
                    INSERT INTO quotes
                    (id, client_id, user_id, folio, quote_title, quote_recipient, created_at, valid_until, subtotal, discount, tax, total, authorized_shipping, notes, status, client_snapshot)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        quote_id,
                        data["client_id"],
                        data["user_id"],
                        folio,
                        quote_title,
                        quote_recipient,
                        now_iso(),
                        quote_valid_until(),
                        totals["subtotal"],
                        totals["discount"],
                        totals["tax"],
                        totals["total"],
                        totals["authorized_shipping"],
                        data.get("notes", ""),
                        data.get("status", "emitida"),
                        client_snapshot,
                    ),
                )
                for item in totals["items"]:
                    db.execute(
                        """
                        INSERT INTO quote_items
                        (id, quote_id, cip, description, quantity, unit_price, discount_rate, tax_rate, line_total)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item["id"],
                            quote_id,
                            item["cip"],
                            item["description"],
                            item["quantity"],
                            item["unit_price"],
                            item["discount_rate"],
                            item["tax_rate"],
                            item["line_total"],
                        ),
                    )
                log_activity(db, data["client_id"], data["user_id"], "cotizacion", f"Cotizacion generada: {quote_title}", totals)
                return send_json(self, {"id": quote_id, "folio": folio, "quote_title": quote_title, **totals})
            if parsed.path == "/api/followups":
                data = read_json(self)
                ensure_local_client(db, data["client_id"])
                ensure_local_user(db, data["user_id"])
                quote_id = data.get("quote_id") or None
                if quote_id:
                    quote = db.execute("SELECT id FROM quotes WHERE id = ? AND client_id = ?", (quote_id, data["client_id"])).fetchone()
                    if not quote:
                        return send_json(self, {"error": "La cotizacion seleccionada no pertenece al cliente"}, 400)
                data["quote_id"] = quote_id
                followup_id = crm_insert_followup(db, data)
                log_activity(db, data["client_id"], data["user_id"], "seguimiento", f"Seguimiento registrado: {data.get('channel', 'llamada')}", data)
                return send_json(self, {"id": followup_id})
        return send_json(self, {"error": "Ruta no encontrada"}, 404)

    def handle_api_put(self, parsed):
        with connect() as db:
            if parsed.path == "/api/prospector/quotes":
                data = read_json(self)
                try:
                    result = update_prospect_quote(db, data)
                except Exception as e:
                    return send_json(self, {"error": str(e)}, 400)
                return send_json(self, result)
            if parsed.path == "/api/quotes":
                data = read_json(self)
                quote_id = data.get("id") or ""
                quote = db.execute("SELECT * FROM quotes WHERE id = ?", (quote_id,)).fetchone()
                if not quote:
                    return send_json(self, {"error": "Cotizacion no encontrada"}, 404)
                client_id = data.get("client_id") or quote["client_id"]
                user_id = data.get("user_id") or quote["user_id"]
                ensure_local_client(db, client_id)
                ensure_local_user(db, user_id)
                items = data.get("items", [])
                totals = calculate_quote(items)
                quote_recipient = (data.get("quote_recipient") or "").strip()
                numero, _ = parse_external_client_id(client_id)
                if numero == "100000" and not quote_recipient:
                    return send_json(self, {"error": "Captura empresa/persona cotizada para cliente 100000"}, 400)
                quote_title = quote_title_for_client(client_id, quote["folio"], quote_recipient)
                client_snapshot = quote_client_snapshot_if_needed(db, client_id, quote_recipient)
                db.execute(
                    """
                    UPDATE quotes
                    SET client_id = ?, user_id = ?, quote_title = ?, quote_recipient = ?,
                        subtotal = ?, discount = ?, tax = ?, total = ?, authorized_shipping = ?,
                        notes = ?, client_snapshot = ?
                    WHERE id = ?
                    """,
                    (
                        client_id,
                        user_id,
                        quote_title,
                        quote_recipient,
                        totals["subtotal"],
                        totals["discount"],
                        totals["tax"],
                        totals["total"],
                        totals["authorized_shipping"],
                        data.get("notes", ""),
                        client_snapshot,
                        quote_id,
                    ),
                )
                db.execute("DELETE FROM quote_items WHERE quote_id = ?", (quote_id,))
                for item in totals["items"]:
                    db.execute(
                        """
                        INSERT INTO quote_items
                        (id, quote_id, cip, description, quantity, unit_price, discount_rate, tax_rate, line_total)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            item["id"],
                            quote_id,
                            item["cip"],
                            item["description"],
                            item["quantity"],
                            item["unit_price"],
                            item["discount_rate"],
                            item["tax_rate"],
                            item["line_total"],
                        ),
                    )
                log_activity(db, client_id, user_id, "cotizacion", f"Cotizacion editada: {quote_title}", totals)
                return send_json(self, {"id": quote_id, "folio": quote["folio"], "quote_title": quote_title, **totals})
        return send_json(self, {"error": "Ruta no encontrada"}, 404)

    def handle_api_delete(self, parsed):
        query = parse_qs(parsed.query)
        with connect() as db:
            if parsed.path == "/api/prospector/quotes":
                quote_id = query.get("id", [""])[0]
                quote = db.execute("SELECT id, prospect_id, folio, title FROM prospect_quotes WHERE id = ?", (quote_id,)).fetchone()
                if not quote:
                    return send_json(self, {"error": "Cotizacion de prospecto no encontrada"}, 404)
                db.execute("DELETE FROM prospect_quotes WHERE id = ?", (quote_id,))
                log_prospect_activity(
                    db,
                    quote["prospect_id"],
                    None,
                    "cotizacion",
                    f"Cotizacion de prospecto eliminada: {quote['folio'] or quote['title']}",
                    {"quote_id": quote_id},
                )
                return send_json(self, {"ok": True})
            if parsed.path == "/api/quotes":
                quote_id = query.get("id", [""])[0]
                quote = db.execute("SELECT id, client_id, folio FROM quotes WHERE id = ?", (quote_id,)).fetchone()
                if not quote:
                    return send_json(self, {"error": "Cotizacion no encontrada"}, 404)
                db.execute("DELETE FROM quotes WHERE id = ?", (quote_id,))
                log_activity(db, quote["client_id"], None, "cotizacion", f"Cotizacion eliminada: {quote['folio']}", {"quote_id": quote_id})
                return send_json(self, {"ok": True})
        return send_json(self, {"error": "Ruta no encontrada"}, 404)


if __name__ == "__main__":
    init_db()
    host = "127.0.0.1"
    port = int(os.environ.get("PORT", "8010"))
    print(f"Sistema de ventas disponible en http://{host}:{port}")
    ThreadingHTTPServer((host, port), SalesHandler).serve_forever()
